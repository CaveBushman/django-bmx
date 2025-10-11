from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from rider.models import Rider
from event.models import Entry


def riders_on_events(club_pk):

    year = timezone.now().year

    wb = Workbook()
    ws = wb.active
    ws.title = "ÚČAST NA ZÁVODECH"

    # Styl záhlaví
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")  # indigo
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Příjmení", "Jméno", "Srartovní číslo", "UCI ID", year, year-1]
    ws.append(headers)

    # Styluj záhlaví
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Nastav šířky sloupců
    column_widths = [18, 18, 14, 14, 10, 10]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    i=2

    riders = Rider.objects.filter(club__pk=club_pk, is_active=True, is_approved=True)
    for rider in riders:
        ws.cell(i, 1, rider.last_name)
        ws.cell(i, 2, rider.first_name)
        ws.cell(i, 3, rider.plate)
        ws.cell(i, 4, rider.uci_id)

        events_now = Entry.objects.filter(rider=rider, checkout = False, event__date__year=year).count()
        ws.cell(i, 5, events_now)

        events_past = Entry.objects.filter(rider=rider, checkout=False, event__date__year=year-1).count()
        ws.cell(i, 6, events_past)

        i = i + 1

    file_path = f"media/riders_on_events/RIDERS_IN_EVENTS-{year}-{club_pk}.xlsx"
    wb.save(file_path)

    return file_path
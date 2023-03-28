from datetime import date, datetime
from openpyxl import load_workbook
from club.models import Club
from event.models import EntryClasses, Event, Entry
from rider.models import Rider
from django.utils import timezone


def expire_licence():
    year = date.today().year
    return f"{year}/12/31"

def rem_expire_licence():
    year = date.today().year
    return f"31.12.{year}"


def team_name_resolve(club):
    club = Club.objects.get(team_name=club)
    return club.team_name


def gender_resolve(rider):
    """ Set gender to BEM format """
    if rider == "Žena":
        return "F"
    else:
        return "M"


def excel_first_line(ws):
    """ set first line in BEM and Riders list excel file """
    ws.cell(1, 1, "Licence_num")
    ws.cell(1, 2, "UCI_ID")
    ws.cell(1, 3, "UCIcode")
    ws.cell(1, 4, "FederationID")
    ws.cell(1, 5, "International Licence Code")
    ws.cell(1, 6, "Expiry_date")
    ws.cell(1, 7, "Licence_type")
    ws.cell(1, 8, "Status")
    ws.cell(1, 9, "Dob")
    ws.cell(1, 10, "First_name")
    ws.cell(1, 11, "Surname")
    ws.cell(1, 12, "Email")
    ws.cell(1, 13, "Phone")
    ws.cell(1, 14, "Emergency Contact Person")
    ws.cell(1, 15, "Emergency Contact Number")
    ws.cell(1, 16, "Sex")
    ws.cell(1, 17, "CLUB")
    ws.cell(1, 18, "State")
    ws.cell(1, 19, "UCI_Country")
    ws.cell(1, 20, "Class")
    ws.cell(1, 21, "Class2")
    ws.cell(1, 22, "Class3")
    ws.cell(1, 23, "Class4")
    ws.cell(1, 24, "Plate")
    ws.cell(1, 25, "Plate2")
    ws.cell(1, 26, "Plate3")
    ws.cell(1, 27, "Plate4")
    ws.cell(1, 28, "Ranking")
    ws.cell(1, 29, "Ranking2")
    ws.cell(1, 30, "Ranking3")
    ws.cell(1, 31, "Ranking4")
    ws.cell(1, 32, "Transponder")
    ws.cell(1, 33, "Transponder2")
    ws.cell(1, 34, "Transponder3")
    ws.cell(1, 35, "Transponder4")
    ws.cell(1, 36, "Tlabel")
    ws.cell(1, 37, "Tlabel2")
    ws.cell(1, 38, "Tlabel3")
    ws.cell(1, 39, "Tlabel4")
    ws.cell(1, 40, "Reference")
    ws.cell(1, 41, "Team_No")
    ws.cell(1, 42, "Team2_No")
    ws.cell(1, 43, "Team3_No")
    ws.cell(1, 44, "Team4_No")
    ws.cell(1, 45, "Sponsor")
    ws.cell(1, 46, "Comment")
    ws.cell(1, 47, "Medical Suspension")
    ws.cell(1, 48, "Disciplinary Suspension")
    ws.cell(1, 49, "Other Suspension")
    ws.cell(1, 50, "POA Suspension")
    ws.cell(1, 51, "Suspension End Date")
    ws.cell(1, 52, "AdvancedRider")

    return ws


def excel_rem_first_line(ws):
    """ set first line in REM and Riders list excel file """
    ws.cell(1, 1, "CLUB_DESCRIPTION")
    ws.cell(1, 2, "TEAM_DESCRIPTION")
    ws.cell(1, 3, "RIDER_FIRST")
    ws.cell(1, 4, "RIDER_LAST")
    ws.cell(1, 5, "RIDER_SEX")
    ws.cell(1, 6, "RIDER_BIRTHDATE")
    ws.cell(1, 7, "RIDER_MAIL")
    ws.cell(1, 8, "RIDER_TYPE")
    ws.cell(1, 9, "RIDER_LICENCE_TYPE")
    ws.cell(1, 10, "RIDER_UCIID")
    ws.cell(1, 11, "RIDER_UCIID_EXP_DATE")
    ws.cell(1, 12, "RIDER_PLATE1")
    ws.cell(1, 13, "RIDER_CHAMP_PLATE1")
    ws.cell(1, 14, "RIDER_TRANSPONDER1")
    ws.cell(1, 15, "RIDER_PLATE2")
    ws.cell(1, 16, "RIDER_CHAMP_PLATE2")
    ws.cell(1, 17, "RIDER_TRANSPONDER2")
    ws.cell(1, 18, "RIDER_PLATE3")
    ws.cell(1, 19, "RIDER_CHAMP_PLATE3")
    ws.cell(1, 20, "RIDER_TRANSPONDER3")
    ws.cell(1, 21, "RIDER_IDENT")
    ws.cell(1, 22, "RIDER_ACTIVE")
    ws.cell(1, 23, "RIDER_LOCKED")

    return ws


def is_registration_open(event_id):
    """ Function for check, if registration is open"""
    event = Event.objects.get(id=event_id)
    now = timezone.now().strftime("%m/%d/%Y, %H:%M:%S")

    # if results is uploaded, registration is close
    if event.xml_results:
        return False
    
    # event registration is manually close
    if not event.reg_open:
        return False


    # check, id today is between reg_open_from and reg_open_to
    if (now >= event.reg_open_from.strftime("%m/%d/%Y, %H:%M:%S")) and (now <= event.reg_open_to.strftime("%m/%d/%Y, %H:%M:%S")):
        return True
    else:
        return False


def resolve_event_classes(event, gender, have_girl_bonus, rider_class, is_20):
    """ Function for resolve class in event by classes_code and xlsx file | is_20 = TRUE for 20" bike """
    event = Event.objects.get(id=event)
    event_classes = EntryClasses.objects.get(event=event.id)

    if is_20 and (gender == "Muž" or gender == "Ostatní" ):
        if rider_class == "Boys 6":
            return event_classes.boys_6
        elif rider_class == "Boys 7":
            return event_classes.boys_7 
        elif rider_class == "Boys 8":
            return event_classes.boys_8 
        elif rider_class == "Boys 9":
            return event_classes.boys_9 
        elif rider_class == "Boys 10":
            return event_classes.boys_10 
        elif rider_class == "Boys 11":
            return event_classes.boys_11 
        elif rider_class == "Boys 12":
            return event_classes.boys_12 
        elif rider_class == "Boys 13":
            return event_classes.boys_13 
        elif rider_class == "Boys 14":
            return event_classes.boys_14 
        elif rider_class == "Boys 15":
            return event_classes.boys_15
        elif rider_class == "Boys 16":
            return event_classes.boys_16
        elif rider_class == "Men 17-24":
            return event_classes.men_17_24 
        elif rider_class == "Men 25-29":
            return event_classes.men_25_29
        elif rider_class == "Men 30-34":
            return event_classes.men_30_34
        elif rider_class == "Men 35 and over":
            return event_classes.men_35_over
        elif rider_class == "Men Junior":
            return event_classes.men_junior
        elif rider_class == "Men Under 23":
            return event_classes.men_u23
        else: 
            return event_classes.men_elite 

    # Ženy s bonusem
    if is_20 and gender == "Žena" and have_girl_bonus:
        if rider_class == "Girls 7":
            return event_classes.girls_7
        elif rider_class == "Girls 8":
            return event_classes.girls_8
        elif rider_class == "Girls 9":
            return event_classes.girls_9
        elif rider_class == "Girls 10":
            return event_classes.girls_10
        elif rider_class == "Girls 11":
            return event_classes.girls_11
        elif rider_class == "Girls 12":
            return event_classes.girls_12
        elif rider_class == "Girls 13":
            return event_classes.girls_13
        elif rider_class == "Girls 14":
            return event_classes.girls_14
        elif rider_class == "Girls 15":
            return event_classes.girls_15
        elif rider_class == "Girls 16":
            return event_classes.girls_16
        elif rider_class == "Women 17-24":
            return event_classes.women_17_24
        elif rider_class == "Women 25 and over":
            return event_classes.women_25_over
        elif rider_class == "Women Junior":
            return event_classes.women_junior
        elif rider_class == "Women Under 23":
            return event_classes.women_u23
        else: 
            return event_classes.women_elite

    # Ženy bez bonusu
    if is_20 and gender == "Žena" and not have_girl_bonus:
        if rider_class == "Girls 7":
            return event_classes.girls_8
        elif rider_class == "Girls 8":
            return event_classes.girls_9
        elif rider_class == "Girls 9":
            return event_classes.girls_10
        elif rider_class == "Girls 10":
            return event_classes.girls_11
        elif rider_class == "Girls 11":
            return event_classes.girls_12
        elif rider_class == "Girls 12":
            return event_classes.girls_13
        elif rider_class == "Girls 13":
            return event_classes.girls_14
        elif rider_class == "Girls 14":
            return event_classes.girls_15
        elif rider_class == "Girls 15":
            return event_classes.girls_16
        elif rider_class == "Girls 16":
            return event_classes.girls_17_24
        elif rider_class == "Women 17-24":
            return event_classes.women_17_24
        elif rider_class == "Women 25 and over":
            return event_classes.girls_24_over
        elif rider_class == "Women Junior":
            return event_classes.women_junior
        elif rider_class == "Women Under 23":
            return event_classes.women_u23
        else: 
            return event_classes.women_elite

    if not is_20:
        if rider_class == "Boys 12 and under":
            return event_classes.cr_boys_12_and_under
        elif rider_class == "Boys 13 and 14":
            return event_classes.cr_boys_13_14
        elif rider_class == "Boys 15 and 16":
            return event_classes.cr_boys_15_16
        elif rider_class == "Men 17-24":
            return event_classes.cr_men_17_24
        elif rider_class == "Men 25-29":
            return event_classes.cr_men_25_29
        elif rider_class == "Men 30-34":
            return event_classes.cr_men_30_34
        elif rider_class == "Men 35-39":
            return event_classes.cr_men_35_39
        elif rider_class == "Men 40-49":
            return event_classes.cr_men_40_49
        elif rider_class == "Men 50 and over":
            return event_classes.cr_men_50_and_over
        elif rider_class == "Girls 12 and under":
            return event_classes. cr_girls_12_and_under
        elif rider_class == "Girls 13-16":
            return event_classes.cr_girls_13_16
        elif rider_class == "Women 17-29":
            return event_classes. cr_women_17_29
        elif rider_class == "Women 30-39":
            return event_classes.cr_women_30_39
        else:
            return event_classes.cr_women_40_and_over
        

def foreign_club_resolve(state):
    """ Function for setting Club based on state code """
    if state == "SVK":
        return "Slovakia - All Clubs"
    elif state == "GER":
        return "Germany - All Clubs"
    elif state == "POL":
        return "Poland - All Clubs"
    elif state =="HUN":
        return "Hungary - All Clubs"
    elif state == "AUT":
        return "Austria - All Clubs"
    elif state == "FRA":
        return "France - All Clubs"
    elif state == "BEL":
        return "Belgium - All Clubs"
    

def resolve_event_fee(event, gender, have_girl_bonus, rider_class, is_20):
    """ Function for resolve fees in event | is_20 = TRUE for 20" bike """

    event = Event.objects.get(id=event)  
    event_classes = EntryClasses.objects.get(event=event.id)

    if is_20 and (gender == "Muž" or gender == "Ostatní" ):
        if rider_class == "Boys 6":
            return event_classes.boys_6_fee
        elif rider_class == "Boys 7":
            return event_classes.boys_7_fee 
        elif rider_class == "Boys 8":
            return event_classes.boys_8_fee 
        elif rider_class == "Boys 9":
            return event_classes.boys_9_fee 
        elif rider_class == "Boys 10":
            return event_classes.boys_10_fee 
        elif rider_class == "Boys 11":
            return event_classes.boys_11_fee 
        elif rider_class == "Boys 12":
            return event_classes.boys_12_fee 
        elif rider_class == "Boys 13":
            return event_classes.boys_13_fee 
        elif rider_class == "Boys 14":
            return event_classes.boys_14_fee 
        elif rider_class == "Boys 15":
            return event_classes.boys_15_fee
        elif rider_class == "Boys 16":
            return event_classes.boys_16_fee
        elif rider_class == "Men 17-24":
            return event_classes.men_17_24_fee 
        elif rider_class == "Men 25-29":
            return event_classes.men_25_29_fee
        elif rider_class == "Men 30-34":
            return event_classes.men_30_34_fee
        elif rider_class == "Men 35 and over":
            return event_classes.men_35_over_fee
        elif rider_class == "Men Junior":
            return event_classes.men_junior_fee
        elif rider_class == "Men Under 23":
            return event_classes.men_u23_fee
        else: 
            return event_classes.men_elite_fee 

    # Ženy s bonusem
    if is_20 and gender == "Žena" and have_girl_bonus:
        if rider_class == "Girls 7":
            return event_classes.girls_7_fee
        elif rider_class == "Girls 8":
            return event_classes.girls_8_fee
        elif rider_class == "Girls 9":
            return event_classes.girls_9_fee
        elif rider_class == "Girls 10":
            return event_classes.girls_10_fee
        elif rider_class == "Girls 11":
            return event_classes.girls_11_fee
        elif rider_class == "Girls 12":
            return event_classes.girls_12_fee
        elif rider_class == "Girls 13":
            return event_classes.girls_13_fee
        elif rider_class == "Girls 14":
            return event_classes.girls_14_fee
        elif rider_class == "Girls 15":
            return event_classes.girls_15_fee
        elif rider_class == "Girls 16":
            return event_classes.girls_16_fee
        elif rider_class == "Women 17-24":
            return event_classes.women_17_24_fee
        elif rider_class == "Women 25 and over":
            return event_classes.women_25_over_fee
        elif rider_class == "Women Junior":
            return event_classes.women_junior_fee
        elif rider_class == "Women Under 23":
            return event_classes.women_u23_fee
        else: 
            return event_classes.women_elite_fee

    # Ženy bez bonusu
    if is_20 and gender == "Žena" and not have_girl_bonus:
        if rider_class == "Girls 7":
            return event_classes.girls_8_fee
        elif rider_class == "Girls 8":
            return event_classes.girls_9_fee
        elif rider_class == "Girls 9":
            return event_classes.girls_10_fee
        elif rider_class == "Girls 10":
            return event_classes.girls_11_fee
        elif rider_class == "Girls 11":
            return event_classes.girls_12_fee
        elif rider_class == "Girls 12":
            return event_classes.girls_13_fee
        elif rider_class == "Girls 13":
            return event_classes.girls_14_fee
        elif rider_class == "Girls 14":
            return event_classes.girls_15_fee
        elif rider_class == "Girls 15":
            return event_classes.girls_16_fee
        elif rider_class == "Girls 16":
            return event_classes.women_17_24_fee
        elif rider_class == "Women 17-24":
            return event_classes.women_17_24_fee
        elif rider_class == "Women 25 and over":
            return event_classes.women_25_over_fee
        elif rider_class == "Women Junior":
            return event_classes.women_junior_fee
        elif rider_class == "Women Under 23":
            return event_classes.women_u23_fee
        else: 
            return event_classes.women_elite_fee

    if not is_20:
        if rider_class == "Boys 12 and under":
            return event_classes.cr_boys_12_and_under_fee
        elif rider_class == "Boys 13 and 14":
            return event_classes.cr_boys_13_14_fee
        elif rider_class == "Boys 15 and 16":
            return event_classes.cr_boys_15_16_fee
        elif rider_class == "Men 17-24":
            return event_classes.cr_men_17_24_fee
        elif rider_class == "Men 25-29":
            return event_classes.cr_men_25_29_fee
        elif rider_class == "Men 30-34":
            return event_classes.cr_men_30_34_fee
        elif rider_class == "Men 35-39":
            return event_classes.cr_men_35_39_fee
        elif rider_class == "Men 40-49":
            return event_classes.cr_men_40_49_fee
        elif rider_class == "Men 50 and over":
            return event_classes.cr_men_50_and_over_fee
        elif rider_class == "Girls 12 and under":
            return event_classes. cr_girls_12_and_under_fee
        elif rider_class == "Girls 13-16":
            return event_classes.cr_girls_13_16_fee
        elif rider_class == "Women 17-29":
            return event_classes. cr_women_17_29_fee
        elif rider_class == "Women 30-39":
            return event_classes.cr_women_30_39_fee
        else:
            return event_classes.cr_women_40_and_over_fee
        


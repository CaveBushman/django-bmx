import logging
logger = logging.getLogger(__name__)
import json
from event.models import RaceRun
import pandas as pd
import requests
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from .models import (
    Event,
    Result,
    Entry,
    CreditTransaction,
    DebetTransaction, SeasonSettings,
)
from accounts.models import Account
from rider.models import Rider, ForeignRider
from club.models import Club
from django.shortcuts import render, reverse, HttpResponseRedirect
from django.conf import settings
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.cache import cache_control
from django.dispatch import receiver
from django.db.models.signals import post_save, post_delete
from rider.rider import get_api_token, generate_insurance_file
from .func import *
from .credit import *
from .entry import NumberInEvent, REMRiders
from datetime import datetime
from django.utils.timezone import now
import datetime
from ranking.ranking import RankingCount, RankPositionCount, Categories, SetRanking
from openpyxl import Workbook
from openpyxl import load_workbook
import stripe
from decouple import config
from django.utils import timezone
from event.func import update_cart
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.units import cm
from django.db import transaction
from django.db.models import F
from django.core.cache import cache
from datetime import date
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.db.models import Q
from rider.rider import get_api_token, resolve_api_category_code
from rider.models import Rider


# Create your views here.

def events_list_view(request):
    upcomming_events = (Event.objects.filter(date__year=date.today().year, date__gte=date.today()).order_by("date"))
    past_events = (Event.objects.filter(date__year=date.today().year, date__lt=date.today()).order_by("date"))

    for event in upcomming_events:
        if event.canceled:
            event.reg_open = False
        else:
            event.reg_open = is_registration_open(event)

        if event.classes_and_fees_like.event_name == "Dosud nenastaveno":
            event.reg_open = False
        event.save()
    year = date.today().year
    next_year = int(year) + 1
    last_year = int(year) - 1
    data = {
        "events": upcomming_events,
        "past_events": past_events,
        "year": year,
        "next_year": next_year,
        "last_year": last_year,
        "show_title": True,
    }

    return render(request, "event/events-list_new.html", data)


def events_list_by_year_view(request, pk):
    if pk == date.today().year:
        return redirect('event:events')
    events = Event.objects.filter(date__year=pk).order_by("date")
    for event in events:
        if (
            event.canceled
            or event.classes_and_fees_like.event_name == "Dosud nenastaveno"
        ):
            event.reg_open = False
        else:
            event.reg_open = is_registration_open(event)
        event.save()
    year = pk
    next_year = int(year) + 1
    last_year = int(year) - 1

    data = {
        "events": events,
        "year": year,
        "next_year": next_year,
        "last_year": last_year,
        "show_title": False,
    }



    return render(request, "event/events-list_new.html", data)


def event_detail_views(request, pk):
    event = get_object_or_404(Event, pk=pk)
    select_category = ""
    alert = False
    riders_sum = 0
    reg_open = is_registration_open(event)

    data = {
        "event": event,
        "alert": alert,
        "select_category": select_category,
        "riders_sum": riders_sum,
        "reg_open": reg_open,
    }
    return render(request, "event/event-detail.html", data)


def results_view(request, pk):
    event = get_object_or_404(Event, pk=pk)
    results = Result.objects.filter(event=pk).order_by("category", "place")
    data = {"results": results, "event": event}
    return render(request, "event/results.html", data)


# Signály pro invalidaci cache, když se změní Entry
@receiver(post_save, sender=Entry)
@receiver(post_delete, sender=Entry)
def invalidate_event_riders_cache(sender, instance, **kwargs):
    key = f"active_riders_{instance.event_id}"
    cache.delete(key)


@login_required(login_url="/event/not-reg")
def add_entries_view(request, pk):
    seasson = SeasonSettings.objects.filter(year=date.today().year)
    event = get_object_or_404(Event, id=pk)
    event.is_beginners_event = event.is_beginners_event() and seasson.beginners_allowed

    # Key pro cache na seznam jezdců relevantních pro registraci
    cache_key = f"active_riders_{event.id}"
    riders = cache.get(cache_key)
    if riders is None:
        riders = list(
            Rider.objects.filter(is_active=True, is_approved=True)
            .filter(Q(valid_licence=True) | Q(fix_valid_licence=True))
            .prefetch_related("entry_set")
        )
        cache.set(cache_key, riders, timeout=600)

    sum_fee = 0

    # Pokud registrace uzavřena nebo závod zrušený, nedovol vstup
    if event.canceled or not event.reg_open or (event.reg_open_to < timezone.now()):
        return render(request, "event/reg-close.html")

    if request.method == "POST":
        # Číst vybrané checkboxy
        selected_beginner = set(request.POST.getlist("checkbox_beginner"))
        selected_20 = set(request.POST.getlist("checkbox_20"))
        selected_24 = set(request.POST.getlist("checkbox_24"))

        # Načti jen jezdce, kteří byli označeni (úspora)
        selected_uci_ids = {int(val) for val in selected_beginner.union(selected_20).union(selected_24) if val.isdigit()}
        selected_riders = Rider.objects.filter(uci_id__in=selected_uci_ids)

        # Rozdělit jezdců do skupin podle označení
        riders_beginner = []
        riders_20 = []
        riders_24 = []

        for rider in selected_riders:
            uci_str = str(rider.uci_id)
            if uci_str in selected_beginner:
                riders_beginner.append(rider)
            if uci_str in selected_20:
                riders_20.append(rider)
            if uci_str in selected_24:
                riders_24.append(rider)

        # Spočítat částku
        for r in riders_beginner:
            sum_fee += resolve_event_fee(event, r, is_20=True, is_beginner=True)
        for r in riders_20:
            sum_fee += resolve_event_fee(event, r, is_20=True)
        for r in riders_24:
            sum_fee += resolve_event_fee(event, r, is_20=False)

        # Uložit do košíku (nebo do Entry) s transakcemi
        if "btn_add" in request.POST:
            user_account = request.user  # nebo Account objekt
            for rider in riders_beginner:
                if not Entry.objects.filter(rider=rider, event=event, is_beginner=True, payment_complete=True).exists():
                    e = Entry(
                        user_id=request.user.id,
                        event=event,
                        rider=rider,
                        is_beginner=True,
                        is_20=False,
                        is_24=False,
                        class_beginner=resolve_event_classes(event, rider, is_20=True, is_beginner=True),
                        fee_beginner=resolve_event_fee(event, rider, is_20=True, is_beginner=True),
                    )
                    e.save()
            for rider in riders_20:
                if not Entry.objects.filter(rider=rider, event=event, is_20=True, payment_complete=True).exists():
                    e = Entry(
                        user_id=request.user.id,
                        event=event,
                        rider=rider,
                        is_20=True,
                        is_beginner=False,
                        is_24=False,
                        class_20=resolve_event_classes(event, rider, is_20=True),
                        fee_20=resolve_event_fee(event, rider, is_20=True),
                    )
                    e.save()
            for rider in riders_24:
                if not Entry.objects.filter(rider=rider, event=event, is_24=True, payment_complete=True).exists():
                    e = Entry(
                        user_id=request.user.id,
                        event=event,
                        rider=rider,
                        is_24=True,
                        is_20=False,
                        is_beginner=False,
                        class_24=resolve_event_classes(event, rider, is_20=False),
                        fee_24=resolve_event_fee(event, rider, is_20=False),
                    )
                    e.save()

            update_cart(request)
            return redirect("event:events")

        # Pokud neuloženo, zobraz checkout stránku s výpočty atd.
        data = {
            "event": event,
            "riders_beginner": riders_beginner,
            "riders_20": riders_20,
            "riders_24": riders_24,
            "sum_fee": sum_fee,
        }
        return render(request, "event/checkout.html", data)

    # Nezahrnujeme SQL dotazy v cyklu — použij mapování Entry
    registered = Entry.objects.filter(event=event, payment_complete=True)
    reg_map = {e.rider_id: e for e in registered}

    # Přiřaď kategorie a označ registrované
    for rider in riders:
        if event.is_beginners_event and is_beginner(rider):
            rider.is_beginner = True
            rider.class_beginner = resolve_event_classes(event, rider, is_20=True, is_beginner=True)
            print(f"Jezdec {rider.last_name} splňuje podmínku a je v kategorii {rider.class_beginner}")
        rider.class_20 = resolve_event_classes(event, rider, is_20=True)
        rider.class_24 = resolve_event_classes(event, rider, is_20=False)
        if rider.is_elite:
            rider.class_24 = "NELZE PŘIHLÁSIT"

        entry = reg_map.get(rider.pk)
        if entry:
            if getattr(entry, "is_beginner", False):
                rider.class_beginner += "registered"
            if getattr(entry, "is_20", False):
                rider.class_20 += "registered"
            if getattr(entry, "is_24", False):
                rider.class_24 += "registered"

    data = {"event": event, "riders": riders, "sum_fee": sum_fee}
    return render(request, "event/entry.html", data)

def entry_riders_view(request, pk):
    """View for registered riders in event"""
    event = Event.objects.get(id=pk)
    entries = Entry.objects.filter(event=pk, payment_complete=1, checkout=0)
    checkout = Entry.objects.filter(event=pk, payment_complete=1, checkout=1)

    categories = []

    for entry in entries:
        if entry.class_20 not in categories:
            categories.append(entry.class_20)
        elif entry.class_24 not in categories:
            categories.append(entry.class_24)

    try:
        categories.sort()
        categories.remove("")
    except Exception as e:
        pass

    data = {
        "event": event,
        "entries": entries,
        "checkout": checkout,
        "categories": categories,
    }
    return render(request, "event/entry-list.html", data)


def confirm_view(request):
    this_event = json.loads(request.session["event"])
    event = Event.objects.get(id=this_event["event"])
    riders_beginner_list = json.loads(request.session["riders_beginner"])
    riders_20_list = json.loads(request.session["riders_20"])
    riders_24_list = json.loads(request.session["riders_24"])

    if "btn-add-event" in request.POST:
        pass

    if request.method == "POST":
        # add entries for 20" bikes
        line_items = []

        for rider_beginner in riders_beginner_list:
            rider = Rider.objects.get(uci_id=rider_beginner["fields"]["uci_id"])
            line_items += generate_stripe_line(
                event, rider, is_20=True, is_beginner=True
            )

        for rider_20_list in riders_20_list:
            rider = Rider.objects.get(uci_id=rider_20_list["fields"]["uci_id"])
            line_items += generate_stripe_line(event, rider, is_20=True)

        # add entries for cruiser
        for rider_24_list in riders_24_list:
            rider = Rider.objects.get(uci_id=rider_24_list["fields"]["uci_id"])
            line_items += generate_stripe_line(event, rider, is_20=False)

        try:
            checkout_session = stripe.checkout.Session.create(
                payment_method_types=["card"],
                line_items=line_items,
                mode="payment",
                success_url=settings.YOUR_DOMAIN + "/event/success/" + str(event.id),
                cancel_url=settings.YOUR_DOMAIN + "/event/cancel",
            )
            # TODO: Need last check for registration in the same time
            # save entries riders to database
            for rider in riders_beginner_list:
                current_rider = Rider.objects.get(uci_id=rider["fields"]["uci_id"])
                current_fee = resolve_event_fee(event, current_rider, 1)
                current_class = resolve_event_classes(
                    event, current_rider, is_20=True, is_beginner=True
                )
                entry = Entry(
                    transaction_id=checkout_session.id,
                    event=event,
                    rider=current_rider,
                    is_beginner=True,
                    is_20=False,
                    is_24=False,
                    class_beginner=current_class,
                    class_20="",
                    class_24="",
                    fee_beginner=current_fee,
                )
                entry.save()

            for rider in riders_20_list:
                current_rider = Rider.objects.get(uci_id=rider["fields"]["uci_id"])
                current_fee = resolve_event_fee(event, current_rider, 1)
                current_class = resolve_event_classes(event, current_rider, is_20=True)
                entry = Entry(
                    transaction_id=checkout_session.id,
                    event=event,
                    rider=current_rider,
                    is_20=True,
                    is_24=False,
                    is_beginner=False,
                    class_beginner="",
                    class_20=current_class,
                    class_24="",
                    fee_20=current_fee,
                )
                entry.save()

            for rider in riders_24_list:
                current_rider = Rider.objects.get(uci_id=rider["fields"]["uci_id"])
                current_fee = resolve_event_fee(event, current_rider, 0)
                current_class = resolve_event_classes(event, current_rider, is_20=False)
                entry = Entry(
                    transaction_id=checkout_session.id,
                    event=event,
                    rider=current_rider,
                    is_20=False,
                    is_24=True,
                    is_beginner=False,
                    class_beginner="",
                    class_24=current_class,
                    class_20="",
                    fee_24=current_fee,
                )
                entry.save()
            del entry

            return JsonResponse({"id": checkout_session.id})
        except Exception as e:
            return JsonResponse(error=str(e)), 403


def success_view(request, pk):
    transactions = Entry.objects.filter(
        Q(
            transaction_date__year=date.today().year,
            transaction_date__month=date.today().month,
            transaction_date__day=date.today().day,
            event=pk,
            payment_complete=False,
        )
        | (
            Q(
                transaction_date__year=date.today().year,
                transaction_date__month=date.today().month,
                transaction_date__day=date.today().day - 1,
                event=pk,
                payment_complete=False,
            )
        )
    )

    transactions_to_email = []
    # check, if fees was paid

    for transaction in transactions:
        try:
            confirm = stripe.checkout.Session.retrieve(transaction.transaction_id,)
            if confirm["payment_status"] == "paid":
                transaction.payment_complete = True
                transaction.customer_name = confirm["customer_details"]["name"]
                transaction.customer_email = confirm["customer_details"]["email"]
                transaction.save()
                # fill list for confirm transaction via email
                # if transaction.transaction_id not in transactions_to_email:
                #    transactions_to_email.append(transaction.transaction_id)
        except Exception as e:
            print(e)

    # clear duplitates
    transactions_to_email = set(transactions_to_email)

    # send e-mail about confirm registrations
    for transaction_to_email in transactions_to_email:
        # threading.Thread (target = SendConfirmEmail(transaction_to_email).send_email()).start()
        pass

    # vymaž sessions
    try:
        if request.session.get("sum_fee"):
            del request.session["sum_fee"]
        else:
            print("Session sum_fee neexistuje")

        if request.session.get("event"):
            del request.session["event"]
        else:
            print("Session event neexistuje")

        if request.session.get("riders_20"):
            del request.session["riders_20"]
        else:
            print("Session riders_20 neexistuje")

        if request.session.get("riders_24"):
            del request.session["riders_24"]
        else:
            print("Session riders_24 neexistuje")

    except Exception as e:
        print("Chyba " + str(e))

    data = {"event_id": pk}
    return render(request, "event/success.html", data)


def cancel_view(request):
    return render(request, "event/cancel.html")

#TODO: Dodělat webhooks
@csrf_exempt
def stripe_credit_webhook(request):
    payload = request.body
    sig_header = request.META.get("HTTP_STRIPE_SIGNATURE")
    event = None

    try:
        event = stripe.Webhook.construct_event(
            payload, sig_header, settings.STRIPE_ENDPOINT_SECRET
        )
    except ValueError as e:
        logger.error(f"Invalid payload: {e}")
        return HttpResponse(status=400)
    except stripe.error.SignatureVerificationError as e:
        logger.error(f"Invalid signature: {e}")
        return HttpResponse(status=400)

    if event["type"] == "checkout.session.completed":
        session = event["data"]["object"]
        session_id = session["id"]
        payment_intent = session.get("payment_intent")

        try:
            with transaction.atomic():
                credit_transaction = CreditTransaction.objects.select_for_update().get(
                    transaction_id=session_id
                )
                if not credit_transaction.payment_complete:
                    Account.objects.filter(id=credit_transaction.user.id).update(
                        credit=F("credit") + credit_transaction.amount
                    )
                    credit_transaction.payment_complete = True
                    credit_transaction.payment_intent = payment_intent
                    credit_transaction.save()
                    logger.info(
                        f"[Webhook] Kredit přičten uživateli {credit_transaction.user.id}: +{credit_transaction.amount} Kč"
                    )
        except CreditTransaction.DoesNotExist:
            logger.warning(f"[Webhook] Kreditní transakce s ID {session_id} nenalezena")
        except Exception as e:
            logger.error(f"[Webhook] Chyba při zpracování kreditu: {e}")

    return HttpResponse(status=200)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@staff_member_required
def event_admin_view(request, pk):
    event = Event.objects.get(id=pk)
    __LICENCE_USERNAME = config("LICENCE_USERNAME")
    __LICENCE_PASSWORD = config("LICENCE_PASSWORD")

    # Admin page for European Cup or European Championship
    if event.type_for_ranking in ["Evropský pohár", "Mistrovství Evropy"]:
        token = get_api_token()
        headers = {"Authorization": f"Bearer {token}"} if token else {}

        payments = 0
        entries = Entry.objects.filter(
            event=event.id, payment_complete=True, checkout=False
        ).select_related("rider")

        entries_20 = entries.filter(is_20=True)
        entries_24 = entries.filter(is_24=True)
        sum_entries = entries_20.count() + entries_24.count()

        file_name = f"media/ec-files/EC_RACE_ID-{event.id}-{event.name}.xlsx"
        if event.type_for_ranking == "Evropský pohár":
            wb = load_workbook(filename="media/ec-files/Entries example - UEC.xlsx")
        else:
            wb = load_workbook(
                filename="media/ec-files/Entries_upload_UEC_Champ_2024.xlsx"
            )
        ws = wb.active

        x = 3
        for entry in entries:
            rider = entry.rider
            try:
                ws.cell(x, 2, rider.uci_id)
                ws.cell(x, 3, rider.date_of_birth)
                ws.cell(x, 4, rider.first_name)
                ws.cell(x, 5, rider.last_name)
                ws.cell(x, 6, gender_resolve_small_letter(rider.gender))
                ws.cell(
                    x, 7, rider.transponder_20 if entry.is_20 else rider.transponder_24
                )
                if entry.is_24:
                    ws.cell(x, 8, "x")
                if entry.is_20:
                    if rider.is_elite:
                        ws.cell(x, 9, "x")
                    if rider.class_20 in ["Women Under 23", "Men Under 23"]:
                        ws.cell(x, 10, "x")
                x += 1
            except Exception as e:
                print(f"Chyba při zápisu do souboru: {e}")

            payments += entry.fee_20 if entry.is_20 else entry.fee_24

        wb.save(file_name)
        event.ec_file = file_name
        event.ec_file_created = timezone.now()
        event.save()

        # Pojišťovací soubor
        generate_insurance_file(event)

        data = {"event": event, "sum_entries": sum_entries, "payments": payments}
        return render(request, "event/event-admin-ec.html", data)

    # ... zbytek funkce zůstává beze změny ...
    if event.type_for_ranking == "Mistrovství světa":
        pass
    # Admin page for Czech events
    if "btn-upload-result" in request.POST:
        print("Stisknuto tlačítko nahrát výsledky v BEM")

        if "result-file" not in request.FILES:  # if xls file is not selected
            messages.error(request, "Musíš vybrat soubor s výsledky závodu")
            return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": pk}))

        else:
            print("Nahrávám výsledky")
            result_file = request.FILES.get("result-file")
            result_file_name = result_file.name
            fs = FileSystemStorage("media/xls_results")
            filename = fs.save(result_file_name, result_file)
            uploaded_file_url = fs.url(filename)[6:]
            event = Event.objects.get(id=pk)
            ranking_code = GetResult.ranking_code_resolve(type=event.type_for_ranking)
            data = pd.read_excel(
                "media/xls_results" + uploaded_file_url, sheet_name="Results"
            )
            for i in range(1, len(data.index)):
                uci_id = str(data.iloc[i][1])
                category = data.iloc[i][5]
                place = str(data.iloc[i][0])
                first_name = data.iloc[i][2]
                last_name = data.iloc[i][3]
                club = data.iloc[i][6]
                result = GetResult(
                    event.date,
                    event.id,
                    event.name,
                    ranking_code,
                    uci_id,
                    place,
                    category,
                    first_name,
                    last_name,
                    club,
                    event.organizer.team_name,
                    event.type_for_ranking,
                )
                result.write_result()
            event.xls_results = "xls_results" + uploaded_file_url
            event.save()

            SetRanking().start()
            return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": pk}))

    if "btn-delete-xls" in request.POST:
        print("Mažu XLS výsledky")

        Result.objects.filter(event=pk).delete()
        print("Výsledky vymazány")
        SetRanking().start()
        print("Ranking přepočítán")

        xls_file = event.xls_results
        print(f"Budu mazat xls_results {xls_file}")

        try:
            os.remove(f"{xls_file}")
        except Exception as e:
            print(f"Nebyl nalezen soubor {xls_file}")

        event.xls_results.delete(save=True)

        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": pk}))

    # ON LINE ENTRIES FOR BEM
    if "btn-bem-file" in request.POST:
        print("Vytvoř startovku")
        file_name = f"media/bem-files/BEM_FOR_RACE_ID-{event.id}-{event.name}.xlsx"
        wb = Workbook()
        wb.encoding = "utf-8"
        ws = wb.active
        ws.title = "BEM5_EXT"
        ws = excel_first_line(ws)

        # TODO: entries beginners classes

        entries_20 = Entry.objects.filter(
            event=event.id, is_20=True, payment_complete=1, checkout=0
        )
        x = 2
        for entry_20 in entries_20:
            try:
                rider = Rider.objects.get(uci_id=entry_20.rider.uci_id)
                ws.cell(x, 1, rider.uci_id)
                ws.cell(x, 2, rider.uci_id)
                ws.cell(x, 3, rider.uci_id)
                ws.cell(x, 4, rider.uci_id)
                ws.cell(x, 5, rider.uci_id)
                ws.cell(x, 6, expire_licence())
                ws.cell(x, 7, "BMX-RACE")
                ws.cell(x, 9, str(rider.date_of_birth).replace("-", "/"))
                ws.cell(x, 10, rider.first_name)
                ws.cell(x, 11, rider.last_name.upper())
                ws.cell(x, 12, rider.email)
                ws.cell(x, 13, rider.phone)
                ws.cell(x, 14, rider.emergency_contact)
                ws.cell(x, 15, rider.emergency_phone)
                ws.cell(x, 16, gender_resolve(rider))
                ws.cell(x, 17, team_name_resolve(rider.club))
                ws.cell(x, 18, "CZE")
                ws.cell(x, 19, "CZE")
                ws.cell(x, 20, resolve_event_classes(event, rider, is_20=True))
                ws.cell(x, 25, rider.plate)  # plate for cruiser

                if rider.plate_champ_20:
                    world_plate = "W" + str(rider.plate_champ_20)
                else:
                    world_plate = rider.plate

                ws.cell(x, 24, world_plate)
                ws.cell(x, 32, rider.transponder_20)
                ws.cell(x, 33, rider.transponder_24)
                ws.cell(x, 36, "T1")
                ws.cell(x, 37, "T2")
                ws.cell(x, 45, team_name_resolve(rider.club).upper())
                if rider.valid_licence:
                    ws.cell(x, 46, "")
                else:
                    ws.cell(x, 46, "NEPLATNÁ LICENCE")
            except Exception as E:
                pass
            x += 1

            # TODO: Dodělat zobrazení přihlášených jezdců s neplatnou licencí

        del entries_20

        entries_24 = Entry.objects.filter(
            event=event.id, is_24=True, payment_complete=1, checkout=0
        )
        for entry_24 in entries_24:
            rider = Rider.objects.get(uci_id=entry_24.rider.uci_id)
            ws.cell(x, 1, rider.uci_id)
            ws.cell(x, 2, rider.uci_id)
            ws.cell(x, 3, rider.uci_id)
            ws.cell(x, 4, rider.uci_id)
            ws.cell(x, 5, rider.uci_id)
            ws.cell(x, 6, expire_licence())
            ws.cell(x, 7, "BMX RACE")
            ws.cell(x, 9, str(rider.date_of_birth).replace("-", "/"))
            ws.cell(x, 10, rider.first_name)
            ws.cell(x, 11, rider.last_name.upper())
            ws.cell(x, 12, rider.email)
            ws.cell(x, 13, rider.phone)
            ws.cell(x, 14, rider.emergency_contact)
            ws.cell(x, 15, rider.emergency_phone)
            ws.cell(x, 16, gender_resolve(rider))
            ws.cell(x, 17, team_name_resolve(rider.club))
            ws.cell(x, 18, "CZE")
            ws.cell(x, 19, "CZE")
            ws.cell(x, 21, resolve_event_classes(event, rider, is_20=False))
            ws.cell(x, 24, rider.plate)

            if rider.plate_champ_24:
                world_plate = "W" + str(rider.plate_champ_24)
            else:
                world_plate = str(rider.plate)

            ws.cell(x, 25, world_plate)
            ws.cell(x, 32, rider.transponder_20)
            ws.cell(x, 33, rider.transponder_24)
            ws.cell(x, 36, "T1")
            ws.cell(x, 37, "T2")
            ws.cell(x, 45, team_name_resolve(rider.club).upper())
            if rider.valid_licence:
                ws.cell(x, 46, "")
            else:
                ws.cell(x, 46, "NEPLATNÁ LICENCE")

            x += 1
        del entries_24

        # TODO: Add foreign riders

        wb.save(file_name)
        event.bem_entries = file_name
        event.bem_entries_created = timezone.now()
        event.save()

    # ALL RIDERS FOR BEM
    if "btn-riders-list" in request.POST:
        file_name = f"media/riders-list/RIDERS_LIST_FOR_RACE_ID-{event.id}.xlsx"
        wb = Workbook()
        wb.encoding = "utf-8"
        ws = wb.active
        ws.title = "BEM5_EXT"
        ws = excel_first_line(ws)

        riders = Rider.objects.filter(is_active=True, is_approved=True)
        x = 2
        for rider in riders:
            ws.cell(x, 1, rider.uci_id)
            ws.cell(x, 2, rider.uci_id)
            ws.cell(x, 3, rider.uci_id)
            ws.cell(x, 4, rider.uci_id)
            is_approvedws.cell(x, 5, rider.uci_id)
            ws.cell(x, 6, expire_licence())
            ws.cell(x, 7, "BMX-RACE")
            ws.cell(x, 9, str(rider.date_of_birth).replace("-", "/"))
            ws.cell(x, 10, rider.first_name)
            ws.cell(x, 11, rider.last_name.upper())
            ws.cell(x, 12, rider.email)
            ws.cell(x, 13, rider.phone)
            ws.cell(x, 14, rider.emergency_contact)
            ws.cell(x, 15, rider.emergency_phone)
            ws.cell(x, 16, gender_resolve(rider))
            ws.cell(x, 17, team_name_resolve(rider.club))
            ws.cell(x, 18, "CZE")
            ws.cell(x, 19, "CZE")
            if rider.is_20:
                ws.cell(x, 20, resolve_event_classes(event, rider, is_20=True))
            if rider.is_24:
                ws.cell(x, 21, resolve_event_classes(event, rider, is_20=False))
            ws.cell(x, 28, "")
            ws.cell(x, 29, "")

            if rider.plate_champ_20:
                rider.plate = rider.plate_champ_20

            ws.cell(x, 24, rider.plate)

            if rider.plate_champ_24:
                rider.plate = rider.plate_champ_24

            ws.cell(x, 25, rider.plate)
            ws.cell(x, 32, rider.transponder_20)
            ws.cell(x, 33, rider.transponder_24)
            ws.cell(x, 36, "T1")
            ws.cell(x, 37, "T2")
            ws.cell(x, 45, team_name_resolve(rider.club).upper())
            if rider.valid_licence:
                ws.cell(x, 46, "")
            else:
                ws.cell(x, 46, "NEPLATNÁ LICENCE")
            x += 1

        del riders
        print("Čeští jezdci přidány")
        # foreign_riders = ForeignRider.objects.all()
        # for foreign_rider in foreign_riders:
        #     ws.cell(x, 1, foreign_rider.uci_id)
        #     ws.cell(x, 2, foreign_rider.uci_id)
        #     ws.cell(x, 3, foreign_rider.uci_id)
        #     ws.cell(x, 4, foreign_rider.uci_id)
        #     ws.cell(x, 5, foreign_rider.uci_id)
        #     ws.cell(x, 6, expire_licence())
        #     ws.cell(x, 7, "BMX-RACE")
        #     ws.cell(x, 9, str(foreign_rider.date_of_birth).replace('-', '/'))
        #     ws.cell(x, 10, foreign_rider.first_name)
        #     ws.cell(x, 11, foreign_rider.last_name.upper())
        #     ws.cell(x, 16, gender_resolve(foreign_rider))
        #     ws.cell(x, 17, foreign_club_resolve(foreign_rider.state))
        #     ws.cell(x, 18, foreign_rider.state)
        #     ws.cell(x, 19, foreign_rider.state)
        #     if foreign_rider.is_20:
        #         ws.cell(x, 20, resolve_event_classes(event.id, foreign_rider.gender, True, foreign_rider.class_20, 1))
        #     if foreign_rider.is_24:
        #         ws.cell(x, 21, resolve_event_classes(event.id, foreign_rider.gender, False, foreign_rider.class_24, 0))
        #     ws.cell(x, 28, "")
        #     ws.cell(x, 29, "")
        #     ws.cell(x, 24, foreign_rider.plate)
        #     ws.cell(x, 25, foreign_rider.plate)
        #     ws.cell(x, 32, foreign_rider.transponder_20)
        #     ws.cell(x, 33, foreign_rider.transponder_24)
        #     ws.cell(x, 36, "T1")
        #     ws.cell(x, 37, "T2")
        #     ws.cell(x, 45, foreign_rider.club.upper())
        #     x += 1
        # del foreign_riders

        wb.save(file_name)
        event.bem_riders_list = file_name
        event.bem_riders_created = timezone.now()
        event.save()

    # ON LINE ENTRIES FOR REM
    if "btn-rem-file" in request.POST:
        all_entries = REMRiders()
        all_entries.event = event
        all_entries.create_entries_list()

    # ALL RIDERS FOR REM
    if "btn-rem-riders-list" in request.POST:
        all_riders = REMRiders()
        all_riders.event = event
        all_riders.create_all_riders_list()

    if "btn-upload-txt" in request.POST:
        if "result-file-txt" not in request.FILES:  # if txt file is not selected
            messages.error(request, "Musíš vybrat soubor s výsledky závodu")
            return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": pk}))

        if request.POST["btn-upload-txt"] == "txt":
            print("Nahrávám výsledky z REM")
            result_file = request.FILES.get("result-file-txt")
            result_file_name = result_file.name
            fs = FileSystemStorage("media/rem_results")
            filename = fs.save(result_file_name, result_file)
            uploaded_file_url = fs.url(filename)[6:]
            event = Event.objects.get(id=pk)
            results = SetResults()
            results.setEvent(pk)
            results.setFile(uploaded_file_url)
            results.start()

            # ✨ Nová logika pro vytvoření RaceRun záznamů
            file_path = os.path.join(fs.location, filename)
            df = pd.read_csv(file_path, sep="\t")

            for _, row in df.iterrows():
                plate = row.get("PLATE")
                if pd.isna(plate):
                    continue

                # Odstraň písmena a ponech pouze čísla
                plate_digits = "".join(filter(str.isdigit, str(plate)))
                if not plate_digits:
                    continue

                result = Result.objects.filter(
                    event=event, rider=int(plate_digits)
                ).first()
                if not result:
                    continue

                # MOTO jízdy
                for i in range(1, 10):
                    if not pd.isna(row.get(f"MOTO{i}_PLACE")):
                        if not RaceRun.objects.filter(
                            result=result, round_type="MOTO", round_number=i
                        ).exists():
                            RaceRun.objects.create(
                                result=result,
                                round_type="MOTO",
                                round_number=i,
                                gate=row.get(f"MOTO{i}_GATE"),
                                lane=row.get(f"MOTO{i}_LANE"),
                                place=row.get(f"MOTO{i}_PLACE"),
                                race_points=row.get(f"MOTO{i}_RACE_POINTS"),
                                moto_points=row.get(f"MOTO{i}_MOTO_POINTS"),
                                finish_time=row.get(f"MOTO{i}_TIME"),
                                hill_time=None,
                                split_1=None,
                                split_2=None,
                                split_3=None,
                                split_4=None,
                            )

                # Finále a eliminace
                for phase in ["FINAL", "F2", "F4", "F8", "F16", "F32", "F64", "F128"]:
                    if not pd.isna(row.get(f"{phase}_PLACE")):
                        if not RaceRun.objects.filter(
                            result=result, round_type=phase
                        ).exists():
                            RaceRun.objects.create(
                                result=result,
                                round_type=phase,
                                round_number=None,
                                gate=row.get(f"{phase}_GATE"),
                                lane=row.get(f"{phase}_LANE"),
                                place=row.get(f"{phase}_PLACE"),
                                race_points=row.get(f"{phase}_RACE_POINTS"),
                                moto_points=row.get(f"{phase}_MOTO_POINTS"),
                                finish_time=row.get(f"{phase}_TIME"),
                                hill_time=None,
                                split_1=None,
                                split_2=None,
                                split_3=None,
                                split_4=None,
                            )

    if "btn-txt-delete" in request.POST:
        print("Mažu výsledky závodu")
        Result.objects.filter(event=pk).delete()
        event = Event.objects.get(id=pk)
        event.ccf_uploaded = False
        event.ccf_created = None
        event.save()
        print("Výsledky závodu vymazány")

        SetRanking().start()

        rem_file = event.rem_results
        print(f"Budu mazat rem_results {rem_file}")

        try:
            os.remove(f"{rem_file}")
        except Exception as e:
            print(f"Nebyl nalezen soubor {rem_file}")

        event.rem_results.delete(save=True)

    sum_of_fees: int = 0
    sum_of_riders: int = 0

    # check riders with invalid licences
    invalid_licences = invalid_licence_in_event(event)

    # sum fees on event
    entries = Entry.objects.filter(event=event.id, payment_complete=1, checkout=False)

    for entry in entries:
        sum_of_fees += entry.fee_beginner
        sum_of_fees += entry.fee_20
        sum_of_fees += entry.fee_24
        sum_of_riders += 1

    asociation_fee = int(sum_of_fees * event.commission_fee / 100)

    results_exist = Result.objects.filter(event=event).exists()
    print(f"Výsledky existují: {results_exist}")

    data = {
        "event": event,
        "invalid_licences": invalid_licences,
        "sum_of_fees": sum_of_fees,
        "sum_of_riders": sum_of_riders,
        "asociation_fee": asociation_fee,
        "results_exist": results_exist,
    }
    return render(request, "event/event-admin.html", data)


@staff_member_required
def find_payment_view(request):
    """Views for find e-mail address recorded in payment status"""

    events = Event.objects.filter()

    if "find-payment" in request.POST:
        rider = request.POST["rider"]
        event = request.POST["event"]
        try:
            entry = Entry.objects.get(
                event=event.id, rider=rider.uci_id, payment_complete=True
            )
            rider = Rider.objects.get(uci_id=rider)
            event = Event.objects.get(id=event)
        except Exception as e:
            pass

        data = {"event": event, "rider": rider, "entry": entry}

    data = {"events": events}
    return render(request, "event/find-payment.html", data)


def ranking_table_view(request):
    """Function for viewing ranking table of points"""
    data = {}
    return render(request, "event/ranking-table.html", data)


def entry_foreign_view(request, pk):
    """View for foreign riders registrations"""
    event = get_object_or_404(Event, pk=pk)
    data = {"event": event}
    views = render(request, "event/entry-foreign.html", data)
    return views


def ec_by_club_xls(request, pk):
    event = get_object_or_404(Event, pk=pk)
    clubs = Club.objects.filter(is_active=True).order_by("team_name")
    entries = Entry.objects.filter(event=pk, payment_complete=True).order_by("rider")

    file_name = f"media/ec-files/EC_RACE_ID_BY_CLUB-{event.id}-{event.name}.xlsx"

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    wb = load_workbook(filename="media/ec-files/Club example - UEC.xlsx")
    ws = wb.active

    ws.cell(3, 2, event.name)

    line = 6
    for club in clubs:
        for entry in entries:
            rider = Rider.objects.get(id=entry.rider.id)
            if rider.club.id == club.id:
                ws.cell(line, 1, rider.last_name)
                ws.cell(line, 2, rider.first_name)
                ws.cell(line, 3, rider.uci_id)
                ws.cell(line, 4, rider.club.team_name)

                line += 1
    wb.save(response)

    return response


def summary_riders_in_event(request, pk):
    event = Event.objects.get(id=pk)

    classes_20_24 = clean_classes_on_event(event)
    count_20_24 = []

    for class_20_24 in classes_20_24:
        sum_20_24 = NumberInEvent()
        sum_20_24.event = pk
        sum_20_24.category_name = class_20_24
        if class_20_24 is None:
            pass
        elif ("Cruiser" or "cruiser") in class_20_24:
            sum_20_24.count_riders_24()
        elif "Beginners" in class_20_24:
            sum_20_24.count_beginners()
        else:
            sum_20_24.count_riders_20()
        if (
            class_20_24 is not None
            and ("NENÍ VYPSÁNO" or "není vypsáno") not in class_20_24
        ):
            count_20_24.append(sum_20_24)

    data = {"count_riders": count_20_24}

    return render(request, "event/riders-sum-event.html", data)


@login_required(login_url="/login/")
def confirm_user_order(request):
    # aktualizuj košík
    update_cart(request)
    # vymaž propadnuté registrace, registrace již byla ukončena a nebyla zaplacena
    delete_reg = Entry.objects.filter(
        user__id=request.user.id,
        payment_complete=False,
        event__reg_open_to__lt=timezone.now(),
    )

    # TODO: Vymazat duplicitní položky v košíku

    if delete_reg:
        delete_reg.delete()
        return redirect("event:order")
    # načti platné registrace v nákupním košíku
    orders = Entry.objects.filter(
        user__id=request.user.id,
        payment_complete=False,
        event__date__gte=timezone.now(),
    ).order_by("event__date", "rider__last_name", "rider__first_name")

    # Odstranění duplicit v košíku
    duplicities = []
    for order in orders:
        if order.is_beginner:
            if (
                Entry.objects.filter(
                    event=order.event, rider=order.rider, is_beginner=True
                ).count()
                > 1
            ):
                duplicities.append(order)
                order.delete()
        elif order.is_20:
            if (
                Entry.objects.filter(
                    event=order.event, rider=order.rider, is_20=True
                ).count()
                > 1
            ):
                duplicities.append(order)
                order.delete()
        else:
            if (
                Entry.objects.filter(
                    event=order.event, rider=order.rider, is_24=True
                ).count()
                > 1
            ):
                duplicities.append(order)
                order.delete()
        if duplicities:
            print("Duplicitní položky v košíku:")
            print(duplicities)
            data = {"duplicities": duplicities}
            return render(request, "event/order.html", data)

    if "btn-del" in request.POST:
        order = Entry.objects.get(id=request.POST["btn-del"])
        order.delete()
        update_cart(request)
        return redirect("event:order")

    if request.POST:
        price: int = 0
        for order in orders:
            if order.is_beginner:
                price += order.fee_beginner
            elif order.is_20:
                price += order.fee_20
            else:
                price += order.fee_24

        user = Account.objects.get(id=request.user.id)

        if price > user.credit:
            data = {}
            return render(request, "event/order_error.html", data)

        try:
            # TODO: Dodělat debetní transakce
            for order in orders:
                amount = order.fee_beginner + order.fee_20 + order.fee_24
                debet_transaction = DebetTransaction(
                    user_id=request.user.id, amount=amount, entry=order
                )
                debet_transaction.save()

                user.credit = user.credit - amount
                user.save()

                order.payment_complete = True
                order.save()

                update_cart(request)

            return redirect("event:checkout")
        except Exception as e:
            return JsonResponse(error=str(e)), 403

    price: int = 0
    sum: int = orders.count()
    for order in orders:
        price += order.fee_beginner + order.fee_20 + order.fee_24
        if order.is_beginner:
            order.event_class = order.class_beginner
        elif order.is_20:
            order.event_class = order.class_20
        else:
            order.event_class = order.class_24
    data = {"orders": orders, "price": price, "sum": sum}
    return render(request, "event/order.html", data)


@login_required(login_url="/login")
def check_order_payments(request):
    orders = Entry.objects.filter(
        Q(
            updated__year=date.today().year,
            updated__month=date.today().month,
            updated__day=date.today().day,
            payment_complete=False,
        )
        | Q(
            updated__year=date.today().year,
            updated__month=date.today().month,
            updated__day=date.today().day - 1,
            payment_complete=False,
        )
    )
    for order in orders:
        try:
            confirm = stripe.checkout.Session.retrieve(order.stripe_payload,)
            if confirm["payment_status"] == "paid":
                order.confirmed = True
                order.save()
        except Exception as e:
            print(e)

    transactions = Entry.objects.filter(
        Q(
            transaction_date__year=date.today().year,
            transaction_date__month=date.today().month,
            transaction_date__day=date.today().day,
            payment_complete=False,
        )
        | (
            Q(
                transaction_date__year=date.today().year,
                transaction_date__month=date.today().month,
                transaction_date__day=date.today().day - 1,
                payment_complete=False,
            )
        )
    )
    for transaction in transactions:
        try:
            confirm = stripe.checkout.Session.retrieve(transaction.transaction_id,)
            if confirm["payment_status"] == "paid":
                transaction.payment_complete = True
                transaction.customer_name = confirm["customer_details"]["name"]
                transaction.customer_email = confirm["customer_details"]["email"]
                transaction.save()
        except:
            pass

    update_cart(request)
    messages.success(request, "Vaše přihláška byla úspěšně přijata.")
    data = {}
    return render(request, "event/success.html", data)


@login_required(login_url="/login/")
def checkout_view(request):
    user_id = request.user.id
    user = Account.objects.get(id=user_id)
    confirmed_events = Entry.objects.filter(
        user__id=user_id, payment_complete=True, event__date__gte=timezone.now(),
    ).order_by("event__date", "rider__last_name", "rider__first_name")
    for confirmed_event in confirmed_events:
        if is_registration_open(confirmed_event.event):
            confirmed_event.is_visible = True
        else:
            confirmed_event.is_visible = False
    # if user want change status
    if "btn-change" in request.POST:
        confirmed_event = Entry.objects.get(id=request.POST["btn-change"])

        debet_transaction = DebetTransaction.objects.filter(
            user=user, entry=confirmed_event
        )

        debet_transaction.delete()
        confirmed_event.delete()

        user.credit = calculate_user_balance(user.id)
        user.save()

        return redirect("event:checkout")
    else:
        data = {"confirmed_events": confirmed_events, "user": user}
        return render(request, "event/event-checkout.html", data)


@login_required(login_url="/login/")
def fees_on_event(request, pk):
    """Function for print fees in event by club"""
    event = Event.objects.get(pk=pk)
    entries = Entry.objects.filter(event=pk, checkout=False)
    clubs = Club.objects.filter(is_active=True).order_by("team_name")
    club_in_event = []
    for club in clubs:
        fee = 0
        for entry in entries:
            if club == entry.rider.club:
                fee += entry.fee_20 + entry.fee_24 + entry.fee_beginner
        if fee > 0:
            club.fee = fee
            club_in_event.append(club)
    data = {"clubs": club_in_event, "event": event}
    return render(request, "event/fees-on-event.html", data)


@login_required(login_url="/login")
def credit_view(request):
    user_id = request.user.id
    user = Account.objects.get(id=user_id)

    if request.POST:
        amount = request.POST["price"]
        amount = int(amount)

        if amount < 100:
            messages.error(request, "Minimální částka pro nákup kreditu je 100 Kč.")
            return redirect("event:credit")

        line_item = (
            {
                "price_data": {
                    "currency": "czk",
                    "unit_amount": amount * 100,
                    "product_data": {
                        "name": user.first_name + " " + user.last_name,
                        "images": [],
                        "description": "nabití kreditu pro registraci na závody BMX Racing",
                    },
                },
                "quantity": 1,
            },
        )
        # TODO: Dodělat stripe
        try:
            checkout_session = stripe.checkout.Session.create(
                payment_method_types=["card"],
                line_items=line_item,
                mode="payment",
                success_url=settings.YOUR_DOMAIN + "/event/success-credit",
                cancel_url=settings.YOUR_DOMAIN + "/event/cancel",
            )

            credit_transaction = CreditTransaction(
                transaction_id=checkout_session.id, amount=amount, user_id=user_id
            )
            credit_transaction.save()

            return redirect(checkout_session.url, code=303)
        except Exception as e:
            return JsonResponse(error=str(e)), 403

    else:
        credits = CreditTransaction.objects.filter(
            user__id=user_id,
            payment_complete=True,
            transaction_date__gte=timezone.now() - datetime.timedelta(days=365),
        ).order_by("-transaction_date")
        debets = DebetTransaction.objects.filter(
            user__id=user_id,
            transaction_date__gte=timezone.now() - datetime.timedelta(days=365),
        ).order_by("-entry__event__date")
        data = {"credits": credits, "debets": debets}
        return render(request, "event/credit.html", data)


@login_required(login_url="/login")
def success_credit_view(request):
    credit_transactions = CreditTransaction.objects.filter(
        Q(
            transaction_date__year=date.today().year,
            transaction_date__month=date.today().month,
            transaction_date__day=date.today().day,
            payment_complete=False,
        )
        | Q(
            transaction_date__year=date.today().year,
            transaction_date__month=date.today().month,
            transaction_date__day=date.today().day - 1,
            payment_complete=False,
        )
    )

    for credit_transaction in credit_transactions:
        try:
            # Zablokování řádku pro úpravu
            with transaction.atomic():
                credit_transaction = CreditTransaction.objects.select_for_update().get(
                    id=credit_transaction.id
                )

                # Kontrola, zda transakce již byla zpracována podle UUID
                if CreditTransaction.objects.filter(
                    uuid=credit_transaction.uuid, payment_complete=True
                ).exists():
                    continue

                # Ověření, že platba ještě není dokončena
                if credit_transaction.payment_complete:
                    continue

                confirm = stripe.checkout.Session.retrieve(
                    credit_transaction.transaction_id
                )

                if confirm["payment_status"] == "paid":
                    credit_transaction.payment_complete = True
                    credit_transaction.payment_intent = confirm["payment_intent"]
                    credit_transaction.save()

                    # Bezpečná kontrola, zda platba již nebyla zpracována, a pak přičtení kreditu
                    if not credit_transaction.payment_complete:
                        Account.objects.filter(id=credit_transaction.user.id).update(
                            credit=F("credit") + credit_transaction.amount
                        )
                        credit_transaction.payment_complete = True
                        credit_transaction.payment_intent = confirm["payment_intent"]
                        credit_transaction.save()
        except CreditTransaction.DoesNotExist:
            continue  # Pokud byl mezitím smazán, přeskočíme
        except stripe.error.StripeError as e:
            print(f"Stripe error: {e}")
        except Exception as e:
            print(f"Unexpected error: {e}")

    return redirect("event:success-credit-update")


@login_required(login_url="/login")
def success_credit_update_view(request):
    messages.success(request, "Váš kredit byl úspěšně navýšen.")
    data = {}
    return render(request, "event/success_credit.html", data)


def not_reg_view(request):
    data = {}
    return render(request, "event/not-reg.html", data)


def check_rider(request):
    uci_id = request.GET.get("uci_id", None)
    if uci_id:
        uci_id = "".join(filter(str.isdigit, uci_id))
        uci_id = int(uci_id)
        try:
            rider = ForeignRider.objects.get(uci_id=uci_id)
            data = {
                "first_name": rider.first_name,
                "last_name": rider.last_name,
                "date_of_birth": rider.date_of_birth,
                "sex": rider.gender,
                "plate": rider.plate,
                "transponder_20": rider.transponder_20,
                "transponder_24": rider.transponder_24,
                "nationality": rider.nationality,
            }
            return JsonResponse(data)
        except ForeignRider.DoesNotExist:
            return JsonResponse({"error": "Rider not found"}, status=404)
    return JsonResponse({"error": "UCI ID is required"}, status=400)


def generate_pdf(request, pk):
    try:
        event = Event.objects.get(pk=pk)
    except Event.DoesNotExist:
        return HttpResponse("Event not found", status=404)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="protokol_cipy.pdf"'

    # Vytvoření PDF souboru na šířku (landscape)
    p = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)

    # Okraje
    margin = 2 * cm
    content_width = width - 2 * margin
    content_height = height - 2 * margin

    # Cesta k fontům
    font_regular = os.path.join(settings.BASE_DIR, "static/fonts/DejaVuSans.ttf")
    font_bold = os.path.join(settings.BASE_DIR, "static/fonts/DejaVuSans-Bold.ttf")
    pdfmetrics.registerFont(TTFont("DejaVuSans", font_regular))
    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", font_bold))

    # Logo v pravém horním rohu
    logo_path = os.path.join(settings.BASE_DIR, "static/images/logo.png")
    if os.path.exists(logo_path):
        p.drawImage(
            logo_path,
            width - margin - 75,
            height - margin - 15,
            width=100,
            height=50,
            preserveAspectRatio=True,
            mask="auto",
        )

    # Název závodu v levém horním rohu
    p.setFont("DejaVuSans", 12)
    p.drawString(margin, height - margin, event.name.upper())

    # Nadpis
    p.setFont("DejaVuSans-Bold", 18)
    p.drawCentredString(width / 2, height - margin - 30, "PROTOKOL – ČIPY K PŮJČENÍ")
    p.line(
        margin, height - margin - 35, width - margin, height - margin - 35
    )  # Podtržení

    # Načtení dat z databáze
    entries = Entry.objects.filter(
        event=pk,
        payment_complete=True,
        rider__transponder_20__isnull=True,
        rider__transponder_24__isnull=True,
        is_beginner=False,
    ).order_by("rider__last_name", "rider__first_name")

    # Hlavička tabulky
    header = ["JEZDEC", "ČÍSLO", "KATEGORIE", "KLUB", "ZÁLOHA", "ČIP", "PŘED.", "VRÁC."]

    # Naplnění tabulky daty z databáze
    data = [header]  # Hlavička tabulky je první řádek
    for entry in entries:
        # Dynamický výběr kategorie
        category = (
            entry.rider.class_20
            if entry.is_20
            else entry.rider.class_24
            if entry.is_24
            else ""
        )

        data.append(
            [
                f"{entry.rider.last_name} {entry.rider.first_name}",  # Jezdec
                entry.rider.plate or "",  # Startovní číslo
                category,  # Dynamická kategorie
                entry.rider.club or "",  # Klub
                "",  # Záloha
                "",  # Čip
                "☐",  # Předáno
                "☐",  # Vráceno
            ]
        )

    for i in range(1, 10):
        data.append(["", "", "", "", "", "", "☐", "☐"])

    # Šířky sloupců (rozšířeno pro lepší čitelnost)
    col_widths = [
        content_width * 0.30,  # Jezdec (30 %)
        content_width * 0.12,  # Startovní číslo (12 %)
        content_width * 0.12,  # Kategorie (12 %)
        content_width * 0.18,  # Klub (18 %)
        content_width * 0.08,  # Záloha (8 %)
        content_width * 0.08,  # Čip (8 %)
        content_width * 0.05,  # Předáno (5 %)
        content_width * 0.05,  # Vráceno (5 %)
    ]

    # Pro stránkování
    rows_per_page = 10
    total_pages = (len(data) // rows_per_page) + (
        1 if len(data) % rows_per_page > 0 else 0
    )
    current_row = 0

    def draw_table_page(start_row, current_page):
        # Pokud je to první stránka, vykreslíme hlavičku
        if start_row == 0:
            page_data = data[start_row : start_row + rows_per_page]
        else:
            page_data = [header] + data[
                start_row : start_row + rows_per_page
            ]  # Přidáme hlavičku na každou stránku

        table = Table(page_data, colWidths=col_widths)
        table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.grey,
                    ),  # Šedý pozadí pro první řádek (hlavičku)
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        12,
                    ),  # Zvětšení spodního paddingu pro větší výšku řádku
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        12,
                    ),  # Zvětšení horního paddingu pro větší výšku řádku
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        # Umístění tabulky
        table_width, table_height = table.wrap(0, 0)
        x_position = margin
        y_position = height - margin - 60 - table_height  # 10px mezera pod nadpisem

        table.drawOn(p, x_position, y_position)

        # Přidání stránkování (Stránka 1 z X) do pravého dolního rohu
        p.setFont("DejaVuSans", 10)
        page_number_text = f"Stránka {current_page} z {total_pages}"
        p.drawRightString(width - margin - 10, margin + 10, page_number_text)

        # Přidání textu "VYTIŠTĚNO DNE:" do levého dolního rohu
        current_date = datetime.datetime.now().strftime("%d.%m.%Y")
        printed_text = f"VYTIŠTĚNO DNE: {current_date}"
        p.setFont("DejaVuSans", 10)
        p.drawString(margin, margin + 10, printed_text)

        # Pokud máme více řádků, přidáme stránkování
        if start_row + rows_per_page < len(data):
            p.showPage()  # Nová stránka
            draw_table_page(
                start_row + rows_per_page, current_page + 1
            )  # Rekurzivně vykreslíme další stránku

    # První stránka s hlavičkou tabulky
    draw_table_page(0, 1)  # Začneme od řádku 1, protože řádek 0 je hlavička

    p.showPage()
    p.save()

    return response


def generate_invoice_preparation_pdf(request, pk):
    event = Event.objects.get(pk=pk)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="podklad_pro_fakturaci.pdf"'

    # Vytvoření PDF souboru na výšku (portrait)
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Okraje
    margin = 2 * cm
    content_width = width - 2 * margin
    content_height = height - 2 * margin

    # Cesta k fontům
    font_regular = os.path.join(settings.BASE_DIR, "static/fonts/DejaVuSans.ttf")
    font_bold = os.path.join(settings.BASE_DIR, "static/fonts/DejaVuSans-Bold.ttf")
    pdfmetrics.registerFont(TTFont("DejaVuSans", font_regular))
    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", font_bold))

    # Logo v pravém horním rohu
    logo_path = os.path.join(settings.BASE_DIR, "static/images/logo.png")
    if os.path.exists(logo_path):
        p.drawImage(
            logo_path,
            width - margin - 75,
            height - margin - 15,
            width=100,
            height=50,
            preserveAspectRatio=True,
            mask="auto",
        )

    # Datum a název závodu v levém horním rohu
    p.setFont("DejaVuSans", 12)
    event_date = event.date.strftime("%d.%m.%Y")  # Datum konání
    p.drawString(margin, height - margin, f"{event_date} - {event.name.upper()}")

    # Nadpis
    p.setFont("DejaVuSans-Bold", 18)
    p.drawCentredString(width / 2, height - margin - 30, "PODKLAD PRO FAKTURACI")
    p.line(
        margin, height - margin - 35, width - margin, height - margin - 35
    )  # Podtržení

    # Hlavička tabulky
    header = ["JEZDEC A ST. ČÍSLO", "ČIP", "KLUB", "FAKTURA", "HOTOVOST"]

    # Naplnění tabulky daty z databáze
    data = [header]  # Hlavička tabulky je první řádek

    for i in range(1, 17):
        data.append(["", "", "", "☐", "☐"])

    # Šířky sloupců (rozšířeno pro lepší čitelnost)
    col_widths = [
        content_width * 0.30,  # Jezdec (30 %)
        content_width * 0.12,  # Startovní číslo (12 %)
        content_width * 0.30,  # Klub (30 %)
        content_width * 0.15,  # Faktura (15 %)
        content_width * 0.15,  # Hotovost (15 %)
    ]

    # Pro stránkování
    rows_per_page = 17
    total_pages = (len(data) // rows_per_page) + (
        1 if len(data) % rows_per_page > 0 else 0
    )
    current_row = 0

    def draw_table_page(start_row, current_page):
        # Pokud je to první stránka, vykreslíme hlavičku
        if start_row == 0:
            page_data = data[start_row : start_row + rows_per_page]
        else:
            page_data = [header] + data[
                start_row : start_row + rows_per_page
            ]  # Přidáme hlavičku na každou stránku

        table = Table(page_data, colWidths=col_widths)
        table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.grey,
                    ),  # Šedý pozadí pro první řádek (hlavičku)
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        12,
                    ),  # Zvětšení spodního paddingu pro větší výšku řádku
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        12,
                    ),  # Zvětšení horního paddingu pro větší výšku řádku
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        # Umístění tabulky
        table_width, table_height = table.wrap(0, 0)
        x_position = margin
        y_position = height - margin - 60 - table_height  # 10px mezera pod nadpisem

        table.drawOn(p, x_position, y_position)

        # Přidání stránkování (Stránka 1 z X) do pravého dolního rohu
        p.setFont("DejaVuSans", 10)
        page_number_text = f"Stránka {current_page} z {total_pages}"
        p.drawRightString(width - margin - 10, margin + 10, page_number_text)

        # Přidání textu "VYTIŠTĚNO DNE:" do levého dolního rohu
        current_date = datetime.datetime.now().strftime("%d.%m.%Y")
        printed_text = f"VYTIŠTĚNO DNE: {current_date}"
        p.setFont("DejaVuSans", 10)
        p.drawString(margin, margin + 10, printed_text)

        # Pokud máme více řádků, přidáme stránkování
        if start_row + rows_per_page < len(data):
            p.showPage()  # Nová stránka
            draw_table_page(
                start_row + rows_per_page, current_page + 1
            )  # Rekurzivně vykreslíme další stránku

    # První stránka s hlavičkou tabulky
    draw_table_page(0, 1)  # Začneme od řádku 1, protože řádek 0 je hlavička

    p.showPage()
    p.save()

    return response


def invoice_view(request, pk):
    # invoice_data = get_invoice_data(invoice_id)  # tvá funkce na získání dat z DB

    invoice_data = {
        "number": "20240001",
        "issue_date": "02.01.2024",
        "due_date": "09.01.2024",
        "payment_method": "Převodem",
        "vs": "20240001",
        "iban": "CZ32 0300 0000 0000 5051 1001",
        "supplier": [
            "Adventure Land s.r.o.",
            "Křenova 438/7",
            "162 00 Praha",
            "IČ: 25747908",
            "DIČ: CZ25747908",
        ],
        "customer": [
            "PROMAFIX s.r.o.",
            "Semčice 96",
            "294 46 Semčice",
            "IČ: 08554625",
            "DIČ: CZ08554625",
        ],
        "items": [
            {
                "description": "Rallye test - pronájem okruhu, zabezpečení TK, Com",
                "qty": 5,
                "unit_price": 3990.00,
                "vat": 21,
                "total": 24139.50,
            },
            {
                "description": "Zaokrouhlení",
                "qty": 1,
                "unit_price": 0.50,
                "vat": 0,
                "total": 0.50,
            },
        ],
        "summary": {"base": 19950.00, "vat": 4189.50, "total": 24140.00},
    }

    pdf_buffer = generate_invoice_pdf(invoice_data)

    return HttpResponse(
        pdf_buffer,
        content_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="faktura_{invoice_data["number"]}.pdf"'
        },
    )


@login_required(login_url="/login")
@staff_member_required
def recalculate_balances_view(request):
    # TODO: Dodělat hlášení na webovou stránku
    try:
        recalculate_all_balances()
        return JsonResponse(
            {"status": "success", "message": "Zůstatky byly úspěšně přepočítány."}
        )
    except Exception as e:
        return JsonResponse(
            {"status": "error", "message": f"Chyba při přepočtu: {e}"}, status=500
        )


@login_required(login_url="/login/")
@staff_member_required
def export_event_results(request, event_id):
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return render(request, "error.html", {"message": "Závod nebyl nalezen."})

    if event.ccf_uploaded:
        return render(
            request,
            "error.html",
            {
                "message": "Výsledky už byly odeslány.",
                "detail": f"Závod {event.name} byl již odeslán {event.ccf_created.strftime('%d.%m.%Y %H:%M:%S')}.",
            },
        )

    token = get_api_token()
    if not token:
        return render(
            request,
            "error.html",
            {"message": "Nepodařilo se získat token pro přihlášení k API ČSC."},
        )

    results = Result.objects.filter(event=event)
    payload = []

    for res in results:
        rider = Rider.objects.filter(uci_id=res.rider).first()
        if not rider:
            print(f"Rider with UCI ID {res.rider} not found, skipping.")
            continue

        if res.is_20 == False and res.is_beginner == False:
            cruiser = True
        else:
            cruiser = False

        category_code = resolve_api_category_code(
            rider=rider, is_20=res.is_20, is_24=cruiser, is_beginner=res.is_beginner
        )

        payload.append(
            {
                "category": category_code,
                "rank": res.place,
                "bib": rider.plate,
                "uciid": str(rider.uci_id),
                "lastName": rider.last_name,
                "firstName": rider.first_name,
                "country": res.country,
                "team": rider.club.team_name if rider.club else "",
                "gender": "F" if rider.gender == "Žena" else "M",
                "phase": "",
                "heat": "",
                "result": str(res.place),
                "irm": "",
                "sortOrder": res.place,
            }
        )

    api_url = f"https://portal.api.czechcyclingfederation.com/api/services/saveraceresults?raceId={event.ccf_id}&subDisciplineCode=BMX_RAC"
    headers = {"Authorization": f"Bearer {token}"}

    try:
        response = requests.post(api_url, json=payload, headers=headers)
        response.raise_for_status()
        with open(
            f"media/api-payloads/payload_event_{event.id}.json", "w", encoding="utf-8"
        ) as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as e:
        return render(
            request,
            "error.html",
            {"message": "Nepodařilo se odeslat výsledky na API.", "detail": str(e)},
        )

    # ✅ aktualizuj timestamp a flag
    event.ccf_created = now()
    event.ccf_uploaded = True
    event.save()

    return render(
        request, "event/results_sent.html", {"event": event, "sent_count": len(payload)}
    )


@login_required(login_url="/login")
@staff_member_required
def price_money_pdf(request, pk):
    #TODO: Dodělat list s price money
    pass


@login_required(login_url="/login/")
@staff_member_required
def send_invoices(request, pk):
    #TODO: Dodělat rozesílání faktur
    pass
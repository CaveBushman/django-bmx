"""
event/views/views_admin.py — admin a komisařské operace

Obsah:
  Veřejné view funkce (URL handlery):
    - event_admin_view       — hlavní admin panel závodu
    - ec_by_club_xls         — export přihlášek EP po klubech
    - summary_riders_in_event — počty jezdců v kategoriích
    - export_event_results   — odeslání výsledků do API ČSC
    - send_invoices          — TODO
    - price_money_pdf        — TODO

  Privátní helpery pro event_admin_view (prefix _handle_):
    - _handle_ec_event       — Evropský pohár / Mistrovství Evropy
    - _handle_upload_xls     — nahrání BEM XLS výsledků
    - _handle_delete_xls     — smazání BEM XLS výsledků
    - _handle_bem_entries    — generování BEM startovky (přihlášení jezdci)
    - _handle_bem_riders     — generování BEM seznamu všech jezdců
    - _handle_rem_entries    — generování REM online přihlášek
    - _handle_upload_txt     — nahrání REM TSV výsledků + RaceRun záznamy
    - _handle_delete_txt     — smazání REM výsledků
    - _write_bem_rider_row   — pomocný zápis jednoho jezdce do BEM XLS řádku
"""

import logging
import json
import datetime
import os
from collections import Counter
from io import BytesIO
import requests
import pikepdf
from datetime import date
from types import SimpleNamespace
from django.shortcuts import get_object_or_404, render, reverse, HttpResponseRedirect, redirect
from django.http import FileResponse, HttpResponse
from django.contrib import messages
from django.core.files.base import ContentFile
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from django.db.models import Count, Q
from django.template.defaultfilters import slugify
from django.utils import timezone
from django.utils.timezone import now
from django.views.decorators.cache import cache_control
from django.utils.translation import gettext as _
from event.models import Event, Entry, EntryForeign, Result, RaceRun, SeasonSettings
from club.models import Club, McrClubTeam, McrClubTeamMember
from commissar.models import Commissar
from rider.models import Rider
from event.func import (
    invalid_licence_in_event,
    excel_first_line, expire_licence, gender_resolve,
    gender_resolve_small_letter, team_name_resolve, resolve_event_classes,
    foreign_club_resolve,
    SetResults,
)
from event.entry import REMRiders
from event.result import GetResult
from ranking.ranking import schedule_ranking_recount
from rider.rider import (
    get_api_token,
    generate_insurance_file,
    resolve_api_category_code,
    trigger_cn_qualification_recount_if_needed,
)
from finance.invoices import generate_event_invoices, send_event_invoices
from event.prize_money import PrizeMoneyPdfService
from event.services.race_run_import import RaceRunImportService, START_KEYS, RESULT_KEYS
from event.services.rem_tsv_import import RemTsvRaceRunImportService
from event.services.uci_export import (
    generate_uci_export_zip,
    get_missing_uci_competition_codes,
)
from event.services.unpaid_moto_report import build_unpaid_moto_report
from event.services.unpaid_moto_report_pdf import UnpaidMotoReportPdfService
from accounts.push_notifications import send_to_users
from openpyxl import Workbook, load_workbook
import pandas as pd
import stripe
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import Image, KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

stripe.api_key = settings.STRIPE_SECRET_KEY
logger = logging.getLogger(__name__)
audit_logger = logging.getLogger("audit")
MAX_RESULTS_PDF_SIZE_BYTES = 20 * 1024 * 1024
CCF_API_TIMEOUT = (5, 20)
PDF_FONT_REGULAR_PATH = os.path.join(settings.BASE_DIR, "static/fonts/DejaVuSans.ttf")
PDF_FONT_BOLD_PATH = os.path.join(settings.BASE_DIR, "static/fonts/DejaVuSans-Bold.ttf")
PDF_LOGO_PATH = os.path.join(settings.BASE_DIR, "static/images/logo.png")

COMMISSAR_EXCLUDED_EVENT_TYPES = [
    "Evropský pohár",
    "Mistrovství Evropy",
    "Mistrovství světa",
    "Světový pohár",
]

COMMISSAR_PLACEHOLDER_LAST_NAME = "Bude upřesněno"

UCI_EXPORT_TEMPLATE = os.path.join(settings.MEDIA_ROOT, "uci-templates", "uci-results-template.xlsx")


def _render_error_page(request, message, *, detail=None, status=400):
    context = {"message": message}
    if detail:
        context["detail"] = detail
    return render(request, "error.html", context, status=status)


def _get_event_year_options():
    return (
        Event.objects.exclude(date__isnull=True)
        .dates("date", "year", order="DESC")
    )


def _resolve_selected_year(selected_year, year_options):
    if selected_year is None:
        return str(year_options[0].year) if year_options else str(date.today().year)
    if not selected_year.isdigit():
        return str(year_options[0].year) if year_options else str(date.today().year)
    return selected_year


def _is_mcr_club_teams_event(event):
    return event.type_for_ranking == "Mistrovství ČR družstev"


def _get_mcr_club_teams_year(event):
    return event.date.year if event.date else date.today().year


def _get_mcr_club_teams_for_event(event):
    return (
        McrClubTeam.objects.filter(year=_get_mcr_club_teams_year(event))
        .select_related("club")
        .prefetch_related("members__rider")
        .order_by("club__team_name", "name")
    )


def _get_mcr_club_score_table_clubs(event):
    return (
        Club.objects.filter(is_active=True, mcr_teams__year=_get_mcr_club_teams_year(event))
        .distinct()
        .order_by("team_name")
    )


def _get_mcr_club_teams_riders_count(event):
    return (
        McrClubTeamMember.objects.filter(team__year=_get_mcr_club_teams_year(event))
        .values("rider_id")
        .distinct()
        .count()
    )


def _get_mcr_club_team_entry_fee_total(teams_count):
    return teams_count * 2000


def _is_mcr_club_registration_open(year):
    season = SeasonSettings.objects.filter(year=year).first()
    if season is None:
        return True
    return bool(season.mcr_club_registration_open)


def _set_mcr_club_registration_open(year, *, is_open):
    season = SeasonSettings.objects.filter(year=year).first()
    if season is None:
        season = SeasonSettings.objects.create(year=year, mcr_club_registration_open=is_open)
    else:
        season.mcr_club_registration_open = is_open
        season.save(update_fields=["mcr_club_registration_open"])
    return season


def can_access_commissar_pages(user):
    return user.is_authenticated and (
        getattr(user, "is_staff", False)
        or getattr(user, "is_superuser", False)
        or getattr(user, "is_commission", False)
    )


def can_edit_commissar_assignments(user):
    return user.is_authenticated and (
        getattr(user, "is_superuser", False)
        or getattr(user, "is_commission", False)
    )


commissar_pages_required = user_passes_test(
    can_access_commissar_pages,
    login_url="/bmx-admin/login/",
)


def _build_media_storage(*parts):
    relative_dir = os.path.join(*parts)
    return FileSystemStorage(
        location=os.path.join(settings.MEDIA_ROOT, relative_dir),
        base_url=f"{settings.MEDIA_URL}{relative_dir}/",
    ), relative_dir


def _save_uploaded_file(uploaded_file, *storage_parts):
    storage, relative_dir = _build_media_storage(*storage_parts)
    if storage.exists(uploaded_file.name):
        storage.delete(uploaded_file.name)
    filename = storage.save(uploaded_file.name, uploaded_file)
    return storage, filename, os.path.join(relative_dir, filename)


def _merge_uploaded_pdfs(*uploaded_files):
    merged_buffer = BytesIO()
    appended_any_file = False

    with pikepdf.Pdf.new() as merged_pdf:
        for uploaded_file in uploaded_files:
            if not uploaded_file:
                continue
            uploaded_file.seek(0)
            with pikepdf.Pdf.open(BytesIO(uploaded_file.read())) as source_pdf:
                merged_pdf.pages.extend(source_pdf.pages)
                appended_any_file = True

        if not appended_any_file:
            raise ValueError("Nebyl nahrán žádný PDF soubor.")

        merged_pdf.save(merged_buffer)

    return merged_buffer.getvalue()


def _validate_uploaded_results_pdf(uploaded_file, field_label, required=False):
    if not uploaded_file:
        if required:
            raise ValueError(f"Nahraj {field_label}.")
        return

    filename = (uploaded_file.name or "").lower()
    if not filename.endswith(".pdf"):
        raise ValueError(f"{field_label} musí být PDF soubor.")

    content_type = (uploaded_file.content_type or "").lower()
    if content_type and content_type not in {"application/pdf", "application/x-pdf"}:
        raise ValueError(f"{field_label} nemá platný PDF content type.")

    if uploaded_file.size > MAX_RESULTS_PDF_SIZE_BYTES:
        raise ValueError(
            f"{field_label} je příliš velké. Maximální velikost je {MAX_RESULTS_PDF_SIZE_BYTES // (1024 * 1024)} MB."
        )


def _build_ec_results_pdf_name(event):
    event_slug = slugify(event.name) or f"event-{event.id}"
    timestamp = timezone.now().strftime("%Y%m%d%H%M%S")
    return f"ec-results-{event.id}-{event_slug}-{timestamp}.pdf"


def _handle_ec_results_pdf_upload(request, event):
    uploaded_pdfs = [uploaded_file for uploaded_file in request.FILES.getlist("results-pdf-files") if uploaded_file]

    if not uploaded_pdfs:
        messages.error(request, "Nahraj alespoň jeden PDF soubor s výsledky.")
        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": event.id}))

    try:
        for index, uploaded_pdf in enumerate(uploaded_pdfs, start=1):
            _validate_uploaded_results_pdf(uploaded_pdf, f"PDF s výsledky #{index}", required=True)
        merged_pdf_bytes = _merge_uploaded_pdfs(*uploaded_pdfs)
    except Exception as exc:
        logger.exception("Chyba při spojování PDF výsledků pro event_id=%s", event.id)
        messages.error(request, f"Nepodařilo se spojit PDF soubory: {exc}")
        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": event.id}))

    event.full_results.save(
        _build_ec_results_pdf_name(event),
        ContentFile(merged_pdf_bytes),
        save=False,
    )
    event.save(update_fields=["full_results"])

    messages.success(request, "Výsledky byly nahrány a zveřejněny jako jedno sloučené PDF.")
    audit_logger.info(
        "event_ec_results_pdf_uploaded admin_user_id=%s event_id=%s pdf_count=%s",
        request.user.id,
        event.id,
        len(uploaded_pdfs),
    )
    return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": event.id}))


def _handle_ec_results_pdf_delete(request, event):
    if event.full_results:
        event.full_results.delete(save=True)
        messages.success(request, "Zveřejněné PDF výsledků bylo smazáno.")
        audit_logger.info(
            "event_ec_results_pdf_deleted admin_user_id=%s event_id=%s",
            request.user.id,
            event.id,
        )
    else:
        messages.warning(request, "U závodu není nahrané žádné zveřejněné PDF výsledků.")
    return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": event.id}))


def _generate_ec_file(event):
    entries = Entry.objects.filter(
        event=event.id,
        payment_complete=True,
        checkout=False,
    ).select_related("rider")

    template = os.path.join(
        settings.MEDIA_ROOT,
        "ec-files",
        "Entries example - UEC.xlsx"
        if event.type_for_ranking == "Evropský pohár"
        else "Entries_upload_UEC_Champ_2024.xlsx",
    )
    target_path = os.path.join(
        settings.MEDIA_ROOT,
        "ec-files",
        f"EC_RACE_ID-{event.id}-{event.name}.xlsx",
    )
    os.makedirs(os.path.dirname(target_path), exist_ok=True)

    wb = load_workbook(filename=template)
    ws = wb.active

    row_index = 3
    for entry in entries:
        rider = entry.rider
        try:
            ws.cell(row_index, 2, rider.uci_id)
            ws.cell(row_index, 3, rider.date_of_birth)
            ws.cell(row_index, 4, rider.first_name)
            ws.cell(row_index, 5, rider.last_name)
            ws.cell(row_index, 6, gender_resolve_small_letter(rider.gender))
            ws.cell(row_index, 7, rider.transponder_20 if entry.is_20 else rider.transponder_24)
            if entry.is_24:
                ws.cell(row_index, 8, "x")
            if entry.is_20:
                if rider.is_elite:
                    ws.cell(row_index, 9, "x")
                if rider.class_20 in ["Women Under 23", "Men Under 23"]:
                    ws.cell(row_index, 10, "x")
            row_index += 1
        except Exception as exc:
            logger.error("Chyba při zápisu do EC souboru pro event_id=%s rider_id=%s: %s", event.id, rider.uci_id, exc)

    wb.save(target_path)
    event.ec_file = target_path
    event.ec_file_created = timezone.now()
    event.save(update_fields=["ec_file", "ec_file_created"])
    return target_path


def _get_ec_admin_context(event):
    entries = Entry.objects.filter(
        event=event.id,
        payment_complete=True,
        checkout=False,
    )
    payments = sum((entry.fee_20 if entry.is_20 else 0) + (entry.fee_24 if entry.is_24 else 0) for entry in entries)
    sum_entries = entries.filter(Q(is_20=True) | Q(is_24=True)).count()
    return {
        "event": event,
        "sum_entries": sum_entries,
        "payments": payments,
    }


def _delete_uploaded_files_by_prefix(*storage_parts, prefix):
    storage, _ = _build_media_storage(*storage_parts)
    location = getattr(storage, "location", "")
    if not location or not os.path.isdir(location):
        return 0

    deleted_count = 0
    filename_prefix = f"{prefix}__"
    for existing_name in os.listdir(location):
        existing_path = os.path.join(location, existing_name)
        if not os.path.isfile(existing_path):
            continue
        if not existing_name.startswith(filename_prefix):
            continue
        storage.delete(existing_name)
        deleted_count += 1
    return deleted_count


def _download_generated_file(file_path):
    file_handle = open(file_path, "rb")
    return FileResponse(file_handle, as_attachment=True, filename=os.path.basename(file_path))


def _save_workbook_to_project_path(workbook, relative_file_name):
    file_path = os.path.join(settings.BASE_DIR, relative_file_name)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    workbook.save(file_path)
    return file_path


def _resolve_association_fee(sum_of_fees, commission_fee):
    try:
        normalized_fee = int(commission_fee or 0)
    except (TypeError, ValueError):
        logger.warning("Neplatná commission_fee=%r při výpočtu asociační odměny.", commission_fee)
        normalized_fee = 0
    return int(sum_of_fees * normalized_fee / 100)


# ===========================================================================
# PRIVÁTNÍ HELPERY — volány z event_admin_view
# Každý helper zpracuje jeden btn-* blok a vrátí response nebo None.
# None = pokračuj dál (zobraz shrnutí), response = okamžitý redirect/render.
# ===========================================================================

def _handle_ec_event(request, event):
    """Admin stránka pro Evropský pohár / ME s explicitními akcemi."""
    if request.method == "POST":
        if "btn-upload-ec-results-pdf" in request.POST:
            return _handle_ec_results_pdf_upload(request, event)
        if "btn-delete-ec-results-pdf" in request.POST:
            return _handle_ec_results_pdf_delete(request, event)
        if "btn-generate-ec-file" in request.POST:
            try:
                file_path = _generate_ec_file(event)
                audit_logger.info(
                    "event_ec_file_generated_and_downloaded admin_user_id=%s event_id=%s",
                    request.user.id,
                    event.id,
                )
                return _download_generated_file(file_path)
            except Exception as exc:
                logger.exception("Chyba při generování UEC XLS pro event_id=%s", event.id)
                messages.error(request, f"Nepodařilo se vygenerovat UEC XLS: {exc}")
            return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": event.id}))
        if "btn-generate-ec-insurance-file" in request.POST:
            try:
                file_path = generate_insurance_file(event)
                audit_logger.info(
                    "event_ec_insurance_file_generated_and_downloaded admin_user_id=%s event_id=%s",
                    request.user.id,
                    event.id,
                )
                return _download_generated_file(file_path)
            except Exception as exc:
                logger.exception("Chyba při generování pojišťovacího XLS pro event_id=%s", event.id)
                messages.error(request, f"Nepodařilo se vygenerovat XLS pro pojišťovnu: {exc}")
            return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": event.id}))

    return render(request, "event/event-admin-ec.html", _get_ec_admin_context(event))


def _handle_upload_xls(request, event, pk):
    """Nahraje XLS výsledky z BEM a přepočítá ranking."""
    if "result-file" not in request.FILES:
        messages.error(request, "Musíš vybrat soubor s výsledky závodu")
        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": pk}))

    result_file = request.FILES["result-file"]
    storage, filename, relative_path = _save_uploaded_file(result_file, "xls_results")

    ranking_code = GetResult.ranking_code_resolve(type=event.type_for_ranking)
    data = pd.read_excel(storage.path(filename), sheet_name="Results")

    for i in range(1, len(data.index)):
        GetResult(
            event.date, event.id, event.name, ranking_code,
            str(data.iloc[i][1]), str(data.iloc[i][0]), data.iloc[i][5],
            data.iloc[i][2], data.iloc[i][3], data.iloc[i][6],
            event.organizer.team_name, event.type_for_ranking,
        ).write_result()

    event.xls_results = relative_path
    event.save()
    schedule_ranking_recount()
    trigger_cn_qualification_recount_if_needed(event)
    logger.info(f"BEM výsledky nahrány pro závod {event.id}")
    audit_logger.info(
        "event_bem_results_uploaded admin_user_id=%s event_id=%s filename=%s",
        request.user.id,
        event.id,
        filename,
    )
    return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": pk}))


def _handle_delete_xls(request, event, pk):
    """Smaže XLS výsledky a přepočítá ranking."""
    Result.objects.filter(event=pk).delete()
    schedule_ranking_recount()
    try:
        xls_path = event.xls_results.path
    except (ValueError, NotImplementedError):
        xls_path = None
    if xls_path:
        try:
            os.remove(xls_path)
        except FileNotFoundError:
            logger.warning(f"XLS soubor nenalezen při mazání: {xls_path}")
    event.xls_results.delete(save=True)
    logger.info(f"BEM výsledky smazány pro závod {event.id}")
    audit_logger.info(
        "event_bem_results_deleted admin_user_id=%s event_id=%s",
        request.user.id,
        event.id,
    )
    return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": pk}))


def _write_bem_rider_row(ws, row, rider, event, is_20):
    """Zapíše data jednoho jezdce do zadaného řádku BEM XLS souboru.

    Sdílený kód pro startovku (přihlášení jezdci) i seznam všech jezdců.
    """
    ws.cell(row, 1, rider.uci_id)
    ws.cell(row, 2, rider.uci_id)
    ws.cell(row, 3, rider.uci_id)
    ws.cell(row, 4, rider.uci_id)
    ws.cell(row, 5, rider.uci_id)
    ws.cell(row, 6, expire_licence())
    ws.cell(row, 7, "BMX-RACE")
    ws.cell(row, 9, str(rider.date_of_birth).replace("-", "/"))
    ws.cell(row, 10, rider.first_name)
    ws.cell(row, 11, rider.last_name.upper())
    ws.cell(row, 12, rider.email)
    ws.cell(row, 13, rider.phone)
    ws.cell(row, 14, rider.emergency_contact)
    ws.cell(row, 15, rider.emergency_phone)
    ws.cell(row, 16, gender_resolve(rider))
    ws.cell(row, 17, team_name_resolve(rider.club))
    ws.cell(row, 18, "CZE")
    ws.cell(row, 19, "CZE")

    if is_20:
        ws.cell(row, 20, resolve_event_classes(event, rider, is_20=True))
        world_plate = ("W" + str(rider.plate_champ_20)) if rider.plate_champ_20 else rider.plate_display
        ws.cell(row, 24, world_plate)
        ws.cell(row, 25, rider.plate_display)
    else:
        ws.cell(row, 21, resolve_event_classes(event, rider, is_20=False))
        world_plate = ("W" + str(rider.plate_champ_24)) if rider.plate_champ_24 else rider.plate_display
        ws.cell(row, 24, rider.plate_display)
        ws.cell(row, 25, world_plate)

    ws.cell(row, 32, rider.transponder_20)
    ws.cell(row, 33, rider.transponder_24)
    ws.cell(row, 36, "T1")
    ws.cell(row, 37, "T2")
    ws.cell(row, 45, team_name_resolve(rider.club).upper())
    ws.cell(row, 46, "" if rider.valid_licence else "NEPLATNÁ LICENCE")


def _write_bem_foreign_entry_row(ws, row, entry, is_20):
    """Zapíše zahraniční přihlášku do zadaného řádku BEM XLS souboru."""
    ws.cell(row, 1, entry.uci_id)
    ws.cell(row, 2, entry.uci_id)
    ws.cell(row, 3, entry.uci_id)
    ws.cell(row, 4, entry.uci_id)
    ws.cell(row, 5, entry.uci_id)
    ws.cell(row, 6, expire_licence())
    ws.cell(row, 7, "BMX-RACE")
    ws.cell(row, 9, str(entry.date_of_birth).replace("-", "/") if entry.date_of_birth else "")
    ws.cell(row, 10, entry.first_name)
    ws.cell(row, 11, (entry.last_name or "").upper())
    ws.cell(row, 12, entry.customer_email or "")
    ws.cell(row, 13, "")
    ws.cell(row, 14, "")
    ws.cell(row, 15, "")
    ws.cell(row, 16, "F" if entry.gender == "Žena" else "M")

    club_name = entry.club or foreign_club_resolve(entry.nationality or "")
    ws.cell(row, 17, club_name)
    ws.cell(row, 18, entry.nationality or "")
    ws.cell(row, 19, entry.nationality or "")

    if is_20:
        ws.cell(row, 20, entry.class_20 or "")
        ws.cell(row, 24, entry.plate)
        ws.cell(row, 25, entry.plate)
        ws.cell(row, 32, entry.transponder_20)
        ws.cell(row, 33, entry.transponder_24)
    else:
        ws.cell(row, 21, entry.class_24 or "")
        ws.cell(row, 24, entry.plate)
        ws.cell(row, 25, entry.plate)
        ws.cell(row, 32, entry.transponder_20)
        ws.cell(row, 33, entry.transponder_24)

    ws.cell(row, 36, "T1")
    ws.cell(row, 37, "T2")
    ws.cell(row, 45, (club_name or "").upper())
    ws.cell(row, 46, "")


def _handle_bem_entries(request, event):
    """Vygeneruje BEM startovku (XLS) pro přihlášené jezdce na závod."""
    file_name = f"media/bem-files/BEM_FOR_RACE_ID-{event.id}-{event.name}.xlsx"
    wb = Workbook()
    wb.encoding = "utf-8"
    ws = wb.active
    ws.title = "BEM5_EXT"
    ws = excel_first_line(ws)

    x = 2
    for entry in Entry.objects.filter(event=event.id, is_20=True, payment_complete=True, checkout=False):
        try:
            _write_bem_rider_row(ws, x, entry.rider, event, is_20=True)
        except Exception as e:
            logger.error(f"Chyba BEM startovka řádek {x}: {e}")
        x += 1

    for entry in Entry.objects.filter(event=event.id, is_24=True, payment_complete=True, checkout=False):
        try:
            _write_bem_rider_row(ws, x, entry.rider, event, is_20=False)
        except Exception as e:
            logger.error(f"Chyba BEM startovka řádek {x}: {e}")
        x += 1

    for entry in EntryForeign.objects.filter(event=event.id, is_20=True, payment_complete=True, checkout=False):
        try:
            _write_bem_foreign_entry_row(ws, x, entry, is_20=True)
        except Exception as e:
            logger.error(f"Chyba BEM startovka cizinec řádek {x}: {e}")
        x += 1

    for entry in EntryForeign.objects.filter(event=event.id, is_24=True, payment_complete=True, checkout=False):
        try:
            _write_bem_foreign_entry_row(ws, x, entry, is_20=False)
        except Exception as e:
            logger.error(f"Chyba BEM startovka cizinec řádek {x}: {e}")
        x += 1

    file_path = _save_workbook_to_project_path(wb, file_name)
    event.bem_entries = file_name
    event.bem_entries_created = timezone.now()
    event.save()
    logger.info(f"BEM startovka vygenerována: {file_name}")
    return file_path


def _handle_bem_riders(request, event):
    """Vygeneruje BEM seznam všech aktivních jezdců (bez ohledu na přihlášení)."""
    file_name = f"media/riders-list/RIDERS_LIST_FOR_RACE_ID-{event.id}.xlsx"
    wb = Workbook()
    wb.encoding = "utf-8"
    ws = wb.active
    ws.title = "BEM5_EXT"
    ws = excel_first_line(ws)

    x = 2
    for rider in Rider.objects.filter(is_active=True, is_approved=True):
        # Jezdec může jezdit 20" i 24" — zapíše se do obou sloupců
        if rider.is_20:
            _write_bem_rider_row(ws, x, rider, event, is_20=True)
        if rider.is_24:
            # Cruiser třída jde do sloupce 21 (přepíše se i když is_20 bylo True)
            ws.cell(x, 21, resolve_event_classes(event, rider, is_20=False))
            ws.cell(x, 24, rider.plate_champ_20 or rider.plate_display)
            ws.cell(x, 25, rider.plate_champ_24 or rider.plate_display)
        x += 1

    file_path = _save_workbook_to_project_path(wb, file_name)
    event.bem_riders_list = file_name
    event.bem_riders_created = timezone.now()
    event.save()
    logger.info(f"BEM seznam jezdců vygenerován: {file_name}")
    return file_path


def _handle_rem_entries(request, event):
    """Vygeneruje REM soubor s online přihláškami (přihlášení jezdci)."""
    all_entries = REMRiders()
    all_entries.event = event
    all_entries.create_entries_list()
    logger.info(f"REM přihlášky vygenerovány pro závod {event.id}")
    event.refresh_from_db(fields=["rem_entries"])
    return event.rem_entries.path if event.rem_entries else None


def _handle_mcr_club_rem_entries(request, event):
    """Vygeneruje REM soubor s jezdci přihlášenými přes MČR družstva."""
    all_entries = REMRiders()
    all_entries.event = event
    all_entries.create_mcr_club_entries_list()
    logger.info(f"REM přihlášky MČR družstev vygenerovány pro závod {event.id}")
    event.refresh_from_db(fields=["rem_entries"])
    return event.rem_entries.path if event.rem_entries else None


def _handle_rem_riders(request, event):
    """Vygeneruje REM soubor se seznamem všech aktivních jezdců."""
    all_riders = REMRiders()
    all_riders.event = event
    all_riders.create_all_riders_list()
    logger.info(f"REM seznam jezdců vygenerován pro závod {event.id}")
    event.refresh_from_db(fields=["rem_riders_list"])
    return event.rem_riders_list.path if event.rem_riders_list else None


def _handle_upload_txt(request, event, pk):
    """Nahraje TSV výsledky z REM a zapíše pouze Result."""
    if "result-file-txt" not in request.FILES:
        messages.error(request, "Musíš vybrat soubor s výsledky závodu")
        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": pk}))

    result_file = request.FILES["result-file-txt"]
    storage, filename, relative_path = _save_uploaded_file(result_file, "rem_results")

    results = SetResults()
    results.setEvent(pk)
    results.setFile(filename)
    results.run()

    event.rem_results = relative_path
    event.save(update_fields=["rem_results"])

    logger.info(f"REM výsledky nahrány pro závod {event.id}")
    audit_logger.info(
        "event_rem_results_uploaded admin_user_id=%s event_id=%s filename=%s",
        request.user.id,
        event.id,
        filename,
    )
    return None  # Pokračuj na shrnutí


def _handle_delete_txt(request, event, pk):
    """Smaže REM výsledky a přepočítá ranking."""
    Result.objects.filter(event=pk).delete()
    event.ccf_uploaded = False
    event.ccf_created = None
    event.save()
    schedule_ranking_recount()
    try:
        rem_path = event.rem_results.path
    except (ValueError, NotImplementedError):
        rem_path = None
    if rem_path:
        try:
            os.remove(rem_path)
        except FileNotFoundError:
            logger.warning(f"REM soubor nenalezen při mazání: {rem_path}")
    event.rem_results.delete(save=True)
    logger.info(f"REM výsledky smazány pro závod {event.id}")
    audit_logger.info(
        "event_rem_results_deleted admin_user_id=%s event_id=%s",
        request.user.id,
        event.id,
    )
    return None


# ===========================================================================
# VEŘEJNÉ VIEW FUNKCE
# ===========================================================================

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@staff_member_required
def event_admin_view(request, pk):
    """Hlavní admin panel závodu — dispečer akcí podle stisknutého tlačítka.

    Pro Evropský pohár/ME vrátí okamžitě EC stránku.
    Pro ostatní závodní typy zpracuje POST akci (btn-*) nebo zobrazí shrnutí.
    """
    event = get_object_or_404(Event.objects.select_related("organizer"), id=pk)

    # Evropský pohár a Mistrovství Evropy mají vlastní admin stránku
    if event.type_for_ranking in ["Evropský pohár", "Mistrovství Evropy"]:
        return _handle_ec_event(request, event)

    # Dispatch POST akcí — každý blok vrátí response nebo None
    if "btn-upload-result" in request.POST:
        response = _handle_upload_xls(request, event, pk)
        if response:
            return response

    elif "btn-delete-xls" in request.POST:
        return _handle_delete_xls(request, event, pk)

    elif "btn-bem-file" in request.POST:
        audit_logger.info("event_bem_entries_exported admin_user_id=%s event_id=%s", request.user.id, event.id)
        return _download_generated_file(_handle_bem_entries(request, event))

    elif "btn-riders-list" in request.POST:
        audit_logger.info("event_bem_riders_exported admin_user_id=%s event_id=%s", request.user.id, event.id)
        return _download_generated_file(_handle_bem_riders(request, event))

    elif "btn-rem-file" in request.POST:
        audit_logger.info("event_rem_entries_exported admin_user_id=%s event_id=%s", request.user.id, event.id)
        return _download_generated_file(_handle_rem_entries(request, event))

    elif "btn-rem-riders-list" in request.POST:
        audit_logger.info("event_rem_riders_exported admin_user_id=%s event_id=%s", request.user.id, event.id)
        return _download_generated_file(_handle_rem_riders(request, event))

    elif "btn-upload-txt" in request.POST:
        response = _handle_upload_txt(request, event, pk)
        if response:
            return response

    elif "btn-txt-delete" in request.POST:
        _handle_delete_txt(request, event, pk)

    elif "btn-generate-invoices" in request.POST:
        result = generate_event_invoices(event.id)
        if result["generated"]:
            messages.success(
                request,
                f"Vygenerováno {len(result['generated'])} faktur. Teď je můžeš upravit a následně odeslat.",
            )
        else:
            messages.warning(request, "Pro tento závod nebyly nalezeny žádné uhrazené registrace klubů k fakturaci.")

    elif "btn-send-notification" in request.POST:
        title = request.POST.get("notification_title", "").strip()
        body = request.POST.get("notification_body", "").strip()
        if not title or not body:
            messages.error(request, "Vyplň název i text notifikace.")
        else:
            user_ids = list(
                Entry.objects.filter(event=event, payment_complete=True)
                .exclude(user__isnull=True)
                .values_list("user_id", flat=True)
                .distinct()
            )
            result = (
                send_to_users(user_ids, title=title, body=body, path=f"/events/{event.pk}")
                if user_ids
                else {"success": 0}
            )
            messages.success(request, f"Notifikace odeslána na {result.get('success', 0)} zařízení.")

    elif "btn-send-invoices" in request.POST:
        generated = generate_event_invoices(event.id)
        result = send_event_invoices(event.id)
        if generated["generated"]:
            messages.success(
                request,
                (
                    f"Odesláno {len(result['sent'])} e-mailů, "
                    f"bez e-mailu zůstalo {len(result['skipped'])} klubů."
                ),
            )
        else:
            messages.warning(request, "Pro tento závod nebyly nalezeny žádné uhrazené registrace klubů k fakturaci.")

    # Shrnutí pro šablonu event-admin.html (zobrazí se vždy po akci nebo při GET)
    entries = Entry.objects.filter(event=event.id, payment_complete=True, checkout=False)
    foreign_entries = EntryForeign.objects.filter(event=event.id, payment_complete=True, checkout=False)
    sum_of_fees = sum((e.fee_beginner or 0) + (e.fee_20 or 0) + (e.fee_24 or 0) for e in entries)
    sum_of_fees += sum((e.fee_20 or 0) + (e.fee_24 or 0) for e in foreign_entries)

    notification_user_count = (
        Entry.objects.filter(event=event, payment_complete=True)
        .exclude(user__isnull=True)
        .values("user_id")
        .distinct()
        .count()
    )
    data = {
        "event": event,
        "invalid_licences": invalid_licence_in_event(event),
        "sum_of_fees": sum_of_fees,
        "sum_of_riders": entries.count() + foreign_entries.count(),
        "notification_user_count": notification_user_count,
        "asociation_fee": _resolve_association_fee(sum_of_fees, event.commission_fee),
        "results_exist": Result.objects.filter(event=event).exists(),
        "moto_runs_exist": RaceRun.objects.filter(event=event, round_type="MOTO").exists(),
        "prize_money_amount_toggle": PrizeMoneyPdfService().allows_amount_toggle(event),
        "is_mcr_club_teams_event": _is_mcr_club_teams_event(event),
    }
    return render(request, "event/event-admin.html", data)


@staff_member_required(login_url="/bmx-admin/login/")
def mcr_club_teams_admin_view(request, pk):
    event = get_object_or_404(Event.objects.select_related("organizer"), pk=pk)
    if not _is_mcr_club_teams_event(event):
        return _render_error_page(request, "Tato stránka je dostupná pouze pro závod Mistrovství ČR družstev.", status=404)

    year = _get_mcr_club_teams_year(event)
    if request.method == "POST":
        if request.FILES:
            return _handle_stats_file_upload(
                request,
                event,
                pk,
                redirect_to="event:mcr-club-teams-admin",
            )
        action = request.POST.get("action")
        if action == "open_registration":
            _set_mcr_club_registration_open(year, is_open=True)
            messages.success(request, _("Registrace družstev MČR byla otevřena."))
            return redirect("event:mcr-club-teams-admin", pk=event.pk)
        if action == "close_registration":
            _set_mcr_club_registration_open(year, is_open=False)
            messages.success(request, _("Registrace družstev MČR byla uzavřena."))
            return redirect("event:mcr-club-teams-admin", pk=event.pk)

    teams = _get_mcr_club_teams_for_event(event)
    teams_count = teams.count()
    entry_fee_total = _get_mcr_club_team_entry_fee_total(teams_count)
    return render(
        request,
        "event/mcr-club-teams-admin.html",
        {
            "event": event,
            "year": year,
            "teams": teams,
            "teams_count": teams_count,
            "riders_count": _get_mcr_club_teams_riders_count(event),
            "entry_fee_total": entry_fee_total,
            "entry_fee_total_display": f"{entry_fee_total:,}".replace(",", " "),
            "registration_open": _is_mcr_club_registration_open(year),
            "mcr_ops_results_file": _get_uploaded_stats_by_key(pk).get("overall_results"),
            "mcr_score_table_clubs": _get_mcr_club_score_table_clubs(event),
        },
    )


@staff_member_required(login_url="/bmx-admin/login/")
def mcr_club_teams_roster_view(request, pk):
    event = get_object_or_404(Event.objects.select_related("organizer"), pk=pk)
    if not _is_mcr_club_teams_event(event):
        return _render_error_page(request, "Soupis družstev je dostupný pouze pro závod Mistrovství ČR družstev.", status=404)

    _register_pdf_fonts()
    teams = list(_get_mcr_club_teams_for_event(event).prefetch_related("members__rider"))
    teams_count = len(teams)
    riders_count = _get_mcr_club_teams_riders_count(event)
    entry_fee_total = _get_mcr_club_team_entry_fee_total(teams_count)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "mcr_roster_title",
        parent=styles["Heading1"],
        fontName="DejaVuSans-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "mcr_roster_subtitle",
        parent=styles["BodyText"],
        fontName="DejaVuSans",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#475569"),
    )
    team_title_style = ParagraphStyle(
        "mcr_roster_team_title",
        parent=styles["BodyText"],
        fontName="DejaVuSans-Bold",
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
    )
    cell_style = ParagraphStyle(
        "mcr_roster_cell",
        parent=styles["BodyText"],
        fontName="DejaVuSans",
        fontSize=7.5,
        leading=9,
    )
    header_style = ParagraphStyle(
        "mcr_roster_header",
        parent=cell_style,
        fontName="DejaVuSans-Bold",
        textColor=colors.white,
    )

    story = []
    header_data = []
    if os.path.exists(PDF_LOGO_PATH):
        header_data.append(Image(PDF_LOGO_PATH, width=20 * mm, height=18.3 * mm))
    header_data.append(
        [
            Paragraph("Soupis přihlášených družstev", title_style),
            Paragraph(f"{event.name} · {_get_mcr_club_teams_year(event)}", subtitle_style),
        ]
    )
    story.append(
        Table(
            [header_data],
            colWidths=[24 * mm, 150 * mm] if len(header_data) == 2 else [174 * mm],
            style=TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]),
        )
    )
    story.append(Spacer(1, 5 * mm))
    story.append(
        Table(
            [[
                Paragraph(f"<b>{teams_count}</b><br/>družstev", subtitle_style),
                Paragraph(f"<b>{riders_count}</b><br/>jezdců", subtitle_style),
                Paragraph(f"<b>{entry_fee_total:,} Kč</b><br/>startovné".replace(",", " "), subtitle_style),
            ]],
            colWidths=[58 * mm, 58 * mm, 58 * mm],
            style=TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f1f5f9")),
                ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]),
        )
    )
    story.append(Spacer(1, 6 * mm))

    if not teams:
        story.append(Paragraph("Pro rok závodu zatím nejsou sestavená žádná družstva.", subtitle_style))
    for team in teams:
        team_rows = [[
            Paragraph("Jezdec", header_style),
            Paragraph("Kolo", header_style),
            Paragraph("Číslo", header_style),
            Paragraph("UCI ID", header_style),
            Paragraph("Kategorie", header_style),
            Paragraph("Startovné", header_style),
        ]]
        for member in team.members.all():
            rider = member.rider
            team_rows.append([
                Paragraph(f"{rider.first_name} {rider.last_name}", cell_style),
                Paragraph(f'{member.wheel}"', cell_style),
                Paragraph(str(rider.plate_display or ""), cell_style),
                Paragraph(str(rider.uci_id or ""), cell_style),
                Paragraph(_mcr_member_category(event, member), cell_style),
                Paragraph("□", cell_style),
            ])

        team_block = [
            Table(
                [[
                    Paragraph(team.name, team_title_style),
                    Paragraph(f"{team.club.team_name}<br/>Manager: {team.manager_name or '-'}", subtitle_style),
                ]],
                colWidths=[82 * mm, 92 * mm],
                style=TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#e0f2fe")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#38bdf8")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]),
            ),
            Table(
                team_rows,
                colWidths=[48 * mm, 12 * mm, 18 * mm, 30 * mm, 44 * mm, 22 * mm],
                repeatRows=1,
                style=TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ]),
            ),
            Spacer(1, 4 * mm),
        ]
        story.append(KeepTogether(team_block))

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="mcr-klubu-soupis-druzstev-{event.id}.pdf"'
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    audit_logger.info(
        "mcr_club_teams_roster_pdf_downloaded admin_user_id=%s event_id=%s",
        request.user.id,
        event.id,
    )
    return response


@staff_member_required(login_url="/bmx-admin/login/")
def mcr_club_teams_rem_entries_view(request, pk):
    event = get_object_or_404(Event.objects.select_related("organizer"), pk=pk)
    if not _is_mcr_club_teams_event(event):
        return _render_error_page(request, "Startovní listina REM je dostupná pouze pro závod Mistrovství ČR družstev.", status=404)

    year = _get_mcr_club_teams_year(event)
    if _is_mcr_club_registration_open(year):
        messages.warning(request, _("Startovní listinu pro REM lze stáhnout až po uzavření registrace."))
        return redirect("event:mcr-club-teams-admin", pk=event.pk)

    audit_logger.info("mcr_club_teams_rem_entries_exported admin_user_id=%s event_id=%s", request.user.id, event.id)
    return _download_generated_file(_handle_mcr_club_rem_entries(request, event))


def _register_pdf_fonts():
    if "DejaVuSans" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("DejaVuSans", PDF_FONT_REGULAR_PATH))
    if "DejaVuSans-Bold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", PDF_FONT_BOLD_PATH))


def _mcr_member_category(event, member):
    if event.classes_and_fees_like_id:
        return resolve_event_classes(
            event,
            member.rider,
            is_20=member.wheel != McrClubTeamMember.WHEEL_24,
        ) or ""
    return ""


def _draw_mcr_score_table_page(pdf, event, team):
    width, height = landscape(A4)
    margin_x = 12 * mm
    y = height - 16 * mm
    table_x = margin_x
    table_width = width - 2 * margin_x
    col_widths = [36, 20, 78, 14, 14, 14, 16, 16, 18, 24]
    scale = table_width / sum(col_widths)
    col_widths = [value * scale for value in col_widths]
    headers = ["Kategorie:", "číslo", "jméno", "1", "2", "3", "1/4", "1/2", "finále", "celkem"]
    members = list(team.members.all())
    row_heights = [12 * mm] + [15 * mm] * max(4, len(members))

    if os.path.exists(PDF_LOGO_PATH):
        logo = ImageReader(PDF_LOGO_PATH)
        logo_width = 28 * mm
        logo_height = 25.6 * mm
        pdf.drawImage(
            logo,
            width - margin_x - logo_width,
            height - 13 * mm - logo_height,
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True,
            mask="auto",
        )

    pdf.setFont("DejaVuSans-Bold", 12)
    pdf.drawString(margin_x, y, "MISTROVSTVÍ ČESKÉ REPUBLIKY KLUBŮ A DRUŽSTEV")
    y -= 8 * mm
    pdf.setFont("DejaVuSans-Bold", 11)
    event_date = event.date.strftime("%-d.%-m.%Y") if event.date else ""
    pdf.drawString(margin_x, y, f"{event.name.upper()} {event_date}".strip())

    y -= 13 * mm
    pdf.setFont("DejaVuSans-Bold", 10)
    pdf.drawString(margin_x, y, "NÁZEV:")
    pdf.setFont("DejaVuSans", 10)
    pdf.drawString(margin_x + 24 * mm, y, team.name)

    y -= 8 * mm
    table_top = y
    current_y = table_top
    member_rows = []
    for member in members:
        rider = member.rider
        member_rows.append([
            _mcr_member_category(event, member),
            rider.plate_display,
            f"{rider.first_name} {rider.last_name}",
            "", "", "", "", "", "", "",
        ])
    while len(member_rows) < 4:
        member_rows.append(["", "", "", "", "", "", "", "", "", ""])

    for row_index, row in enumerate([headers] + member_rows):
        row_height = row_heights[row_index]
        current_x = table_x
        for col_index, value in enumerate(row):
            col_width = col_widths[col_index]
            pdf.rect(current_x, current_y - row_height, col_width, row_height, stroke=1, fill=0)
            pdf.setFont("DejaVuSans-Bold" if row_index == 0 else "DejaVuSans", 8 if col_index >= 3 else 8.5)
            text = str(value or "")
            max_width = col_width - 4 * mm
            while pdf.stringWidth(text, pdf._fontname, pdf._fontsize) > max_width and len(text) > 3:
                text = text[:-4] + "..."
            pdf.drawString(current_x + 2 * mm, current_y - row_height + row_height / 2 - 2.2, text)
            current_x += col_width
        current_y -= row_height

    y = current_y - 10 * mm
    pdf.setFont("DejaVuSans-Bold", 10)
    pdf.drawString(margin_x, y, "MANAGER:")
    pdf.setFont("DejaVuSans", 10)
    pdf.drawString(margin_x + 24 * mm, y, team.manager_name)

    y -= 18 * mm
    key_rows = [
        ("ROZJÍŽĎKY:", "1-8, 2-7, 3-6, 4-5, 5-4, 6-3, 7-2, 8-1"),
        ("1/4", "1-5, 2-5, 3-5, 4-5, 5-4, 6-3, 7-2, 8-1"),
        ("1/2", "1-0, 2-0, 3-0, 4-0, 5-8, 6-6, 7-4, 8-2"),
        ("FINÁLE:", "1-22, 2-18, 3-15, 4-13, 5-12, 6-11, 7-10, 8-9"),
    ]
    for label, value in key_rows:
        pdf.setFont("DejaVuSans-Bold", 9)
        pdf.drawString(margin_x, y, label)
        pdf.setFont("DejaVuSans", 9)
        pdf.drawString(margin_x + 34 * mm, y, value)
        y -= 8 * mm

    y -= 3 * mm
    pdf.setFont("DejaVuSans-Bold", 9)
    pdf.drawString(margin_x, y, "Klíč:")
    pdf.setFont("DejaVuSans", 9)
    pdf.drawString(margin_x + 34 * mm, y, "první číslo je pořadí, v jakém jezdec dojel, za pomlčkou")
    y -= 7 * mm
    pdf.drawString(margin_x + 34 * mm, y, "je pak počet bodů, které za jízdu získal")


@staff_member_required(login_url="/bmx-admin/login/")
def mcr_club_teams_score_table_view(request, pk):
    event = get_object_or_404(Event.objects.select_related("organizer"), pk=pk)
    if not _is_mcr_club_teams_event(event):
        return _render_error_page(request, "Bodovací tabulka je dostupná pouze pro závod Mistrovství ČR družstev.", status=404)

    club_id = request.GET.get("club_id")
    club = _get_mcr_club_score_table_clubs(event).filter(pk=club_id).first()
    if club is None:
        messages.warning(request, _("Vyber klub, pro který chceš vytisknout bodovací tabulku."))
        return redirect("event:mcr-club-teams-admin", pk=event.pk)

    _register_pdf_fonts()
    teams = list(
        _get_mcr_club_teams_for_event(event)
        .filter(club=club)
        .select_related("club")
        .prefetch_related("members__rider")
    )
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
    for index, team in enumerate(teams):
        if index:
            pdf.showPage()
        _draw_mcr_score_table_page(pdf, event, team)
    pdf.save()
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="mcr-klubu-bodovaci-tabulka-{event.id}-{club.id}.pdf"'
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    audit_logger.info(
        "mcr_club_teams_score_table_downloaded admin_user_id=%s event_id=%s club_id=%s",
        request.user.id,
        event.id,
        club.id,
    )
    return response


_MCR_RESULTS_ROUND_ORDER = ["MOTO", "F4", "F2", "FINAL"]
_MCR_RESULTS_ROUND_LABELS = {
    "MOTO": _("rozjížďkách"),
    "F4": _("QF"),
    "F2": _("SF"),
    "FINAL": _("finále"),
}


def _mcr_result_included_rounds(round_type):
    if not round_type:
        return []
    try:
        return _MCR_RESULTS_ROUND_ORDER[: _MCR_RESULTS_ROUND_ORDER.index(round_type) + 1]
    except ValueError:
        return [round_type]


def _get_latest_mcr_results_round(event):
    existing_rounds = set(
        RaceRun.objects.filter(event=event, round_type__in=_MCR_RESULTS_ROUND_ORDER)
        .values_list("round_type", flat=True)
        .distinct()
    )
    for round_type in reversed(_MCR_RESULTS_ROUND_ORDER):
        if round_type in existing_rounds:
            return round_type
    return None


def _format_mcr_run_result(run, round_type):
    if not run:
        return ""
    parts = []
    if run.place:
        parts.append(str(run.place))
    if run.race_points is not None:
        parts.append(_("{points} b.").format(points=run.race_points))
    if run.qualified_to_next_round and round_type != "FINAL":
        parts.append(_("postup"))
    return ", ".join(parts)


def _format_mcr_run_points_cell(run):
    if not run:
        return ""
    if run.place and run.race_points is not None:
        return f"{run.place} / {run.race_points}"
    if run.race_points is not None:
        return str(run.race_points)
    return str(run.place or "")


def _mcr_runs_breakdown(rider_runs):
    by_round = {}
    for run in rider_runs:
        if run.round_type == "MOTO":
            key = f"moto_{run.round_number or 1}"
        else:
            key = run.round_type
        by_round[key] = run

    return {
        "moto_1": _format_mcr_run_points_cell(by_round.get("moto_1")),
        "moto_2": _format_mcr_run_points_cell(by_round.get("moto_2")),
        "moto_3": _format_mcr_run_points_cell(by_round.get("moto_3")),
        "f4": _format_mcr_run_points_cell(by_round.get("F4")),
        "f2": _format_mcr_run_points_cell(by_round.get("F2")),
        "final": _format_mcr_run_points_cell(by_round.get("FINAL")),
    }


def _get_mcr_club_results_rows(event, round_type):
    included_rounds = _mcr_result_included_rounds(round_type)
    runs_by_rider = {}
    latest_run_map = {}
    runs = (
        RaceRun.objects.filter(event=event, round_type__in=included_rounds)
        .select_related("rider")
        .order_by("rider_id", "round_type", "round_number", "id")
    )
    for run in runs:
        if run.rider_id:
            key = (run.rider_id, bool(run.is_20))
            runs_by_rider.setdefault(key, []).append(run)
            if run.round_type == round_type:
                latest_run_map[key] = run

    rows = []
    for team in _get_mcr_club_teams_for_event(event).prefetch_related("members__rider"):
        rider_results = []
        points_total = 0
        individual_points = []
        qualified_count = 0
        active_count = 0
        for member in team.members.all():
            rider = member.rider
            key = (rider.id, member.wheel == McrClubTeamMember.WHEEL_20)
            rider_runs = runs_by_rider.get(key, [])
            run = latest_run_map.get(key)
            if rider_runs:
                active_count += 1
                if any(item.qualified_to_next_round for item in rider_runs if item.round_type == round_type):
                    qualified_count += 1
                rider_points = sum(item.race_points or 0 for item in rider_runs)
                points_total += rider_points
                individual_points.append(rider_points)
            else:
                rider_points = 0
            rider_results.append(
                {
                    "rider": f"{rider.first_name} {rider.last_name}",
                    "plate": rider.plate_display,
                    "category": _mcr_member_category(event, member),
                    "runs": _mcr_runs_breakdown(rider_runs),
                    "points": rider_points,
                    "result": _format_mcr_run_result(run, round_type),
                }
            )
        rows.append(
            {
                "team": team,
                "rider_results": rider_results,
                "active_count": active_count,
                "qualified_count": qualified_count,
                "points_total": points_total,
                "individual_points": sorted(individual_points, reverse=True),
            }
        )

    return sorted(rows, key=lambda row: (-row["points_total"], [-points for points in row["individual_points"]], row["team"].name))


@staff_member_required(login_url="/bmx-admin/login/")
def mcr_club_teams_results_pdf_view(request, pk):
    event = get_object_or_404(Event.objects.select_related("organizer"), pk=pk)
    if not _is_mcr_club_teams_event(event):
        return _render_error_page(request, "Výsledky jsou dostupné pouze pro závod Mistrovství ČR družstev.", status=404)

    _register_pdf_fonts()
    round_type = _get_latest_mcr_results_round(event)
    is_final = round_type == "FINAL"
    title = _("Finální výsledky MČR klubů") if is_final else _("Průběžné výsledky MČR klubů")
    phase_label = _MCR_RESULTS_ROUND_LABELS.get(round_type, _("bez nahraných jízd"))

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "mcr_results_title",
        parent=styles["Heading1"],
        fontName="DejaVuSans-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "mcr_results_meta",
        parent=styles["BodyText"],
        fontName="DejaVuSans",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#475569"),
    )
    cell_style = ParagraphStyle(
        "mcr_results_cell",
        parent=styles["BodyText"],
        fontName="DejaVuSans",
        fontSize=6.7,
        leading=8,
    )
    center_cell_style = ParagraphStyle(
        "mcr_results_center_cell",
        parent=cell_style,
        alignment=1,
    )
    points_cell_style = ParagraphStyle(
        "mcr_results_points_cell",
        parent=center_cell_style,
        fontName="DejaVuSans-Bold",
    )
    header_style = ParagraphStyle(
        "mcr_results_header",
        parent=cell_style,
        fontName="DejaVuSans-Bold",
        textColor=colors.white,
        alignment=1,
    )
    team_title_style = ParagraphStyle(
        "mcr_results_team_title",
        parent=styles["BodyText"],
        fontName="DejaVuSans-Bold",
        fontSize=10.5,
        leading=12.5,
        textColor=colors.HexColor("#0f172a"),
    )
    points_style = ParagraphStyle(
        "mcr_results_points",
        parent=styles["BodyText"],
        fontName="DejaVuSans-Bold",
        fontSize=14,
        leading=16,
        alignment=1,
        textColor=colors.HexColor("#1d4ed8"),
    )

    header_cells = []
    if os.path.exists(PDF_LOGO_PATH):
        header_cells.append(Image(PDF_LOGO_PATH, width=19 * mm, height=17.4 * mm))
    header_cells.append([
        Paragraph(title, title_style),
        Paragraph(f"{event.name} · {_get_mcr_club_teams_year(event)} · {phase_label}", meta_style),
    ])
    story = [
        Table(
            [header_cells],
            colWidths=[23 * mm, 157 * mm] if len(header_cells) == 2 else [180 * mm],
            style=TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]),
        ),
        Spacer(1, 7 * mm),
    ]

    if not round_type:
        story.append(Paragraph(_("Zatím nejsou nahrané žádné výsledky jízd."), meta_style))
    else:
        for position, row in enumerate(_get_mcr_club_results_rows(event, round_type), start=1):
            rider_table = [[
                Paragraph(_("Startovní<br/>číslo"), header_style),
                Paragraph(_("Jméno a<br/>příjmení"), header_style),
                Paragraph(_("Kategorie"), header_style),
                Paragraph(_("1"), header_style),
                Paragraph(_("2"), header_style),
                Paragraph(_("3"), header_style),
                Paragraph(_("QF"), header_style),
                Paragraph(_("SF"), header_style),
                Paragraph(_("F"), header_style),
                Paragraph(_("Body"), header_style),
            ]]
            for item in row["rider_results"]:
                runs_breakdown = item["runs"]
                rider_table.append([
                    Paragraph(str(item["plate"] or ""), center_cell_style),
                    Paragraph(item["rider"], cell_style),
                    Paragraph(item["category"] or "", cell_style),
                    Paragraph(runs_breakdown["moto_1"], center_cell_style),
                    Paragraph(runs_breakdown["moto_2"], center_cell_style),
                    Paragraph(runs_breakdown["moto_3"], center_cell_style),
                    Paragraph(runs_breakdown["f4"], center_cell_style),
                    Paragraph(runs_breakdown["f2"], center_cell_style),
                    Paragraph(runs_breakdown["final"], center_cell_style),
                    Paragraph(str(item["points"]), points_cell_style),
                ])

            team_header = Table(
                [[
                    Paragraph(f"{position}. {row['team'].name}", team_title_style),
                    Paragraph(
                        f"{row['team'].club.team_name}<br/>Manager: {row['team'].manager_name or '-'}",
                        meta_style,
                    ),
                    Paragraph(f"{row['points_total']}<br/><font size='7'>bodů</font>", points_style),
                ]],
                colWidths=[86 * mm, 68 * mm, 36 * mm],
                style=TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#e0f2fe")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#38bdf8")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]),
            )
            members_table = Table(
                rider_table,
                colWidths=[20 * mm, 38 * mm, 35 * mm, 13 * mm, 13 * mm, 13 * mm, 14 * mm, 14 * mm, 14 * mm, 16 * mm],
                repeatRows=1,
                style=TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#94a3b8")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
                    ("FONTNAME", (0, 1), (-1, -1), "DejaVuSans"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]),
            )
            story.append(KeepTogether([team_header, members_table, Spacer(1, 4 * mm)]))

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="mcr-klubu-vysledky-{event.id}.pdf"'
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    audit_logger.info(
        "mcr_club_teams_results_pdf_downloaded admin_user_id=%s event_id=%s round_type=%s",
        request.user.id,
        event.id,
        round_type or "",
    )
    return response


@staff_member_required
def ec_by_club_xls(request, pk):
    """Export přihlášených jezdců na Evropský pohár seřazených po klubech (XLS)."""
    event = get_object_or_404(Event, pk=pk)
    clubs = Club.objects.filter(is_active=True).order_by("team_name")
    entries = Entry.objects.filter(
        event=pk,
        payment_complete=True,
        checkout=False,
    ).select_related("rider__club").order_by("rider")

    file_name = f"media/ec-files/EC_RACE_ID_BY_CLUB-{event.id}-{event.name}.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    wb = load_workbook(
        filename=os.path.join(settings.MEDIA_ROOT, "ec-files", "Club example - UEC.xlsx")
    )
    ws = wb.active
    ws.cell(3, 2, event.name)

    line = 6
    for club in clubs:
        for entry in entries:
            if entry.rider.club_id == club.id:
                ws.cell(line, 1, entry.rider.last_name)
                ws.cell(line, 2, entry.rider.first_name)
                ws.cell(line, 3, entry.rider.uci_id)
                ws.cell(line, 4, club.team_name)
                line += 1

    wb.save(response)
    return response


@login_required(login_url="/login/")
def summary_riders_in_event(request, pk):
    """Přehled počtu jezdců v každé třídě na závodě."""
    import json
    from collections import defaultdict
    event = get_object_or_404(Event, id=pk)
    category_counts = Counter()
    category_riders = defaultdict(list)

    czech_entries = Entry.objects.filter(event=pk, payment_complete=True, checkout=False).select_related("rider__club")
    foreign_entries = EntryForeign.objects.filter(event=pk, payment_complete=True, checkout=False)

    for entry in czech_entries:
        if entry.rider:
            name = f"{entry.rider.last_name} {entry.rider.first_name}"
            plate = entry.rider.plate_display or "—"
            club = entry.rider.club.team_name if entry.rider.club else "—"
        else:
            name = "—"
            plate = "—"
            club = "—"
        rider_item = {"plate": plate, "name": name, "club": club, "foreign": False}
        if entry.is_beginner and entry.class_beginner:
            category_counts[entry.class_beginner] += 1
            category_riders[entry.class_beginner].append(rider_item)
        if entry.is_20 and entry.class_20:
            category_counts[entry.class_20] += 1
            category_riders[entry.class_20].append(rider_item)
        if entry.is_24 and entry.class_24:
            category_counts[entry.class_24] += 1
            category_riders[entry.class_24].append(rider_item)

    for entry in foreign_entries:
        name = f"{entry.last_name} {entry.first_name}"
        plate = entry.plate or "—"
        club = entry.club or "—"
        rider_item = {"plate": plate, "name": name, "club": club, "foreign": True}
        if entry.is_20 and entry.class_20:
            category_counts[entry.class_20] += 1
            category_riders[entry.class_20].append(rider_item)
        if entry.is_24 and entry.class_24:
            category_counts[entry.class_24] += 1
            category_riders[entry.class_24].append(rider_item)

    count_20_24 = [
        SimpleNamespace(
            category_name=category_name,
            riders_in_category=riders_in_category,
        )
        for category_name, riders_in_category in sorted(category_counts.items())
        if category_name and "NENÍ VYPSÁNO" not in category_name and "není vypsáno" not in category_name
    ]

    riders_data = {
        cat: sorted(riders, key=lambda r: r["name"])
        for cat, riders in category_riders.items()
        if cat and "NENÍ VYPSÁNO" not in cat and "není vypsáno" not in cat
    }

    total_riders = sum(class_counter.riders_in_category for class_counter in count_20_24)
    return render(
        request,
        "event/riders-sum-event.html",
        {
            "event": event,
            "count_riders": count_20_24,
            "total_riders": total_riders,
            "riders_data_json": json.dumps(riders_data, ensure_ascii=False),
        },
    )


@commissar_pages_required
def commissar_assignments_view(request):
    """Přehled nasazení rozhodčích podle roku s volitelným editačním režimem."""
    year_options = list(_get_event_year_options())
    selected_year = _resolve_selected_year(request.GET.get("year"), year_options)
    can_edit = can_edit_commissar_assignments(request.user)
    edit_mode = can_edit and request.GET.get("edit") == "1"

    if request.method == "POST":
        if not can_edit:
            return HttpResponse(status=403)

        selected_year = _resolve_selected_year(request.POST.get("year"), year_options)
        events_to_update = list(
            Event.objects.filter(date__year=selected_year)
            .exclude(type_for_ranking__in=COMMISSAR_EXCLUDED_EVENT_TYPES)
            .select_related("pcp", "pcp_assist", "start_commissar")
            .order_by("date", "name")
        )

        active_commissars = {
            str(commissar.id): commissar
            for commissar in Commissar.objects.filter(is_active=True).order_by("last_name", "first_name")
        }

        validation_errors = []
        updated_events = []
        for event in events_to_update:
            original_values = {
                "pcp": event.pcp,
                "pcp_assist": event.pcp_assist,
                "start_commissar": event.start_commissar,
            }
            submitted_assignments = {}
            changed_fields = []
            has_conflict = False

            original_snapshot = {
                "pcp": str(event.pcp_id or ""),
                "pcp_assist": str(event.pcp_assist_id or ""),
                "start_commissar": str(event.start_commissar_id or ""),
            }
            for field_name, snapshot_value in original_snapshot.items():
                posted_snapshot = (request.POST.get(f"original_{field_name}_{event.id}") or "").strip()
                if posted_snapshot != snapshot_value:
                    validation_errors.append(
                        _("Závod %(event)s mezitím upravil někdo jiný. Obnov stránku a zkontroluj aktuální nasazení.")
                        % {"event": event.name}
                    )
                    has_conflict = True
                    break

            if has_conflict:
                continue

            for field_name in ("pcp", "pcp_assist", "start_commissar"):
                submitted_value = (request.POST.get(f"{field_name}_{event.id}") or "").strip()
                new_value = active_commissars.get(submitted_value) if submitted_value else None
                submitted_assignments[field_name] = new_value

            selected_ids = [
                commissar.id
                for commissar in submitted_assignments.values()
                if commissar is not None
            ]
            if len(selected_ids) != len(set(selected_ids)):
                validation_errors.append(
                    _("V závodu %(event)s nemůže být jeden rozhodčí nasazen do více rolí současně.")
                    % {"event": event.name}
                )
                continue

            for field_name, new_value in submitted_assignments.items():
                if getattr(event, field_name) != new_value:
                    setattr(event, field_name, new_value)
                    changed_fields.append(field_name)

            if changed_fields:
                event.save(update_fields=changed_fields)
                updated_events.append(event)
                for field_name in changed_fields:
                    previous = original_values[field_name] or "-"
                    current = getattr(event, field_name) or "-"
                    audit_logger.info(
                        "commissar_assignment_updated user_id=%s event_id=%s field=%s old=%s new=%s",
                        request.user.id,
                        event.id,
                        field_name,
                        previous,
                        current,
                    )

        if validation_errors:
            for error in validation_errors:
                messages.error(request, error)
            return HttpResponseRedirect(
                f"{reverse('event:commissar-assignments')}?year={selected_year}&edit=1"
            )

        if updated_events:
            messages.success(request, _("Nasazení rozhodčích bylo uloženo."))
        else:
            messages.info(request, _("Nebyla provedena žádná změna."))

        return HttpResponseRedirect(
            f"{reverse('event:commissar-assignments')}?year={selected_year}"
        )

    events = (
        Event.objects.filter(date__year=selected_year)
        .exclude(type_for_ranking__in=COMMISSAR_EXCLUDED_EVENT_TYPES)
        .select_related("pcp", "pcp_assist", "start_commissar")
        .order_by("date", "name")
    )

    pcp_events_count = events.exclude(pcp__isnull=True).exclude(
        pcp__last_name=COMMISSAR_PLACEHOLDER_LAST_NAME
    ).count()
    pcp_assist_events_count = events.exclude(pcp_assist__isnull=True).exclude(
        pcp_assist__last_name=COMMISSAR_PLACEHOLDER_LAST_NAME
    ).count()
    start_commissar_events_count = events.exclude(start_commissar__isnull=True).exclude(
        start_commissar__last_name=COMMISSAR_PLACEHOLDER_LAST_NAME
    ).count()

    return render(request, "event/commissar-assignments.html", {
        "events": events,
        "selected_year": int(selected_year),
        "year_options": year_options,
        "can_edit_assignments": can_edit,
        "edit_mode": edit_mode,
        "active_commissars": Commissar.objects.filter(is_active=True).order_by("last_name", "first_name"),
        "pcp_events_count": pcp_events_count,
        "pcp_assist_events_count": pcp_assist_events_count,
        "start_commissar_events_count": start_commissar_events_count,
    })


@commissar_pages_required
def commissar_statistics_view(request):
    """Admin-only statistika nasazení rozhodčích podle roku."""
    year_options = list(_get_event_year_options())
    selected_year = _resolve_selected_year(request.GET.get("year"), year_options)

    commissars = (
        Commissar.objects.filter(is_active=True)
        .exclude(last_name=COMMISSAR_PLACEHOLDER_LAST_NAME)
        .annotate(
            pcp_count=Count(
                "PCP",
                filter=Q(PCP__date__year=selected_year)
                & ~Q(PCP__type_for_ranking__in=COMMISSAR_EXCLUDED_EVENT_TYPES),
                distinct=True,
            ),
            pcp_assist_count=Count(
                "PCP_asist",
                filter=Q(PCP_asist__date__year=selected_year)
                & ~Q(PCP_asist__type_for_ranking__in=COMMISSAR_EXCLUDED_EVENT_TYPES),
                distinct=True,
            ),
            start_commissar_count=Count(
                "start_commissar_events",
                filter=Q(start_commissar_events__date__year=selected_year)
                & ~Q(start_commissar_events__type_for_ranking__in=COMMISSAR_EXCLUDED_EVENT_TYPES),
                distinct=True,
            ),
        )
        .order_by("last_name", "first_name")
    )

    active_commissars_count = commissars.count()
    pcp_total = sum(commissar.pcp_count for commissar in commissars)
    pcp_assist_total = sum(commissar.pcp_assist_count for commissar in commissars)
    start_commissar_total = sum(commissar.start_commissar_count for commissar in commissars)

    return render(request, "event/commissar-statistics.html", {
        "commissars": commissars,
        "selected_year": int(selected_year),
        "year_options": year_options,
        "active_commissars_count": active_commissars_count,
        "pcp_total": pcp_total,
        "pcp_assist_total": pcp_assist_total,
        "start_commissar_total": start_commissar_total,
    })


@login_required(login_url="/login/")
@staff_member_required
def export_event_results(request, event_id):
    """Odešle výsledky závodu do API ČSC (Czech Cycling Federation).

    Výsledky jsou odesílány jednorázově — po odeslání je závod označen ccf_uploaded=True.
    """
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return _render_error_page(request, "Závod nebyl nalezen.", status=404)

    if event.ccf_uploaded:
        sent_at = event.ccf_created.strftime("%d.%m.%Y %H:%M:%S") if event.ccf_created else "v minulosti"
        return _render_error_page(
            request,
            "Výsledky už byly odeslány.",
            detail=f"Závod {event.name} byl již odeslán {sent_at}.",
            status=409,
        )

    token = get_api_token()
    if not token:
        return _render_error_page(
            request,
            "Nepodařilo se získat token pro přihlášení k API ČSC.",
            status=502,
        )

    results = list(Result.objects.filter(event=event).select_related("rider__club"))
    payload = []
    category_breakdown = Counter()
    skipped_results = 0

    result_entries = []
    for res in results:
        rider = res.rider
        if not rider:
            logger.warning(f"Výsledek {res.pk} nemá přiřazeného jezdce, přeskakuji.")
            skipped_results += 1
            continue

        cruiser = not res.is_20 and not res.is_beginner
        category_code = resolve_api_category_code(
            rider=rider, is_20=res.is_20, is_24=cruiser, is_beginner=res.is_beginner
        )
        result_entries.append((res, rider, category_code))

    # U Českého poháru se kategorie často slučují (např. "Girls 11-12" nebo
    # "Boys 13-14" včetně dívek). Do API ČSC se ale odesílá umístění ve vlastní
    # kategorii jezdce (category_code), proto se zde přepočítá pořadí v rámci
    # každé vlastní kategorie podle dosaženého place ve sloučené kategorii.
    category_rank = {}
    if event.type_for_ranking == "Český pohár":
        groups = {}
        for res, rider, category_code in result_entries:
            groups.setdefault(category_code, []).append(res)
        for group_results in groups.values():
            ordered = sorted(
                group_results,
                key=lambda r: r.place if r.place > 0 else float("inf"),
            )
            for rank, ordered_res in enumerate(ordered, start=1):
                category_rank[ordered_res.pk] = rank

    for res, rider, category_code in result_entries:
        category_breakdown[category_code or _("Bez kódu")] += 1
        place = category_rank.get(res.pk, res.place)

        payload.append({
            "category": category_code,
            "rank": place,
            "bib": rider.plate_display,
            "uciid": str(rider.uci_id or ""),
            "lastName": rider.last_name,
            "firstName": rider.first_name,
            "country": res.country,
            "team": rider.club.team_name if rider.club else "",
            "gender": "F" if rider.gender == "Žena" else "M",
            "phase": "",
            "heat": "",
            "result": str(place),
            "irm": "",
            "sortOrder": place,
        })

    api_url = (
        f"https://portal.api.czechcyclingfederation.com/api/services/saveraceresults"
        f"?raceId={event.ccf_id}&subDisciplineCode=BMX_RAC"
    )

    try:
        logger.info(
            "Odesílám výsledky do API ČSC pro event_id=%s, race_id=%s, payload_items=%s",
            event.id,
            event.ccf_id,
            len(payload),
        )
        response = requests.post(
            api_url,
            json=payload,
            headers={"Authorization": f"Bearer {token}"},
            timeout=CCF_API_TIMEOUT,
        )
        response.raise_for_status()
        payload_dir = os.path.join(settings.MEDIA_ROOT, "api-payloads")
        os.makedirs(payload_dir, exist_ok=True)
        with open(
            os.path.join(payload_dir, f"payload_event_{event.id}.json"),
            "w",
            encoding="utf-8",
        ) as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        logger.info("Výsledky úspěšně odeslány do API ČSC pro event_id=%s", event.id)
    except requests.Timeout:
        logger.exception("Timeout při odesílání výsledků do API ČSC pro event_id=%s", event.id)
        return _render_error_page(
            request,
            "Nepodařilo se odeslat výsledky na API.",
            detail="API ČSC neodpovědělo včas.",
            status=504,
        )
    except requests.RequestException as e:
        logger.exception("Chyba při odesílání výsledků do API ČSC pro event_id=%s", event.id)
        return _render_error_page(
            request,
            "Nepodařilo se odeslat výsledky na API.",
            detail=str(e),
            status=502,
        )

    event.ccf_created = now()
    event.ccf_uploaded = True
    event.save()
    audit_logger.info(
        "event_results_exported_to_ccf admin_user_id=%s event_id=%s sent_count=%s",
        request.user.id,
        event.id,
        len(payload),
    )

    sent_count = len(payload)
    total_results = len(results)

    def _pct(value, total):
        if not total:
            return 0
        return round((value / total) * 100)

    export_summary = {
        "total_results": total_results,
        "sent_count": sent_count,
        "skipped_count": skipped_results,
        "sent_ratio": _pct(sent_count, total_results),
        "women_count": sum(1 for item in payload if item["gender"] == "F"),
        "men_count": sum(1 for item in payload if item["gender"] == "M"),
        "czech_count": sum(1 for item in payload if (item["country"] or "").upper() == "CZE"),
        "foreign_count": sum(1 for item in payload if (item["country"] or "").upper() != "CZE"),
        "bike20_count": sum(1 for res in results if res.rider and res.is_20 and not res.is_beginner),
        "cruiser_count": sum(1 for res in results if res.rider and not res.is_20 and not res.is_beginner),
        "beginner_count": sum(1 for res in results if res.rider and res.is_beginner),
        "with_uci_count": sum(1 for item in payload if item["uciid"]),
        "with_bib_count": sum(1 for item in payload if item["bib"]),
        "with_team_count": sum(1 for item in payload if item["team"]),
        "unique_teams_count": len({item["team"] for item in payload if item["team"]}),
        "unique_categories_count": len({item["category"] for item in payload if item["category"]}),
        "top_categories": [
            {"label": label, "count": count, "percent": _pct(count, sent_count)}
            for label, count in category_breakdown.most_common(5)
        ],
        "coverage_metrics": [
            {"label": _("Spárovaní jezdci"), "value": sent_count, "total": total_results, "percent": _pct(sent_count, total_results)},
            {"label": _("UCI ID"), "value": sum(1 for item in payload if item["uciid"]), "total": sent_count, "percent": _pct(sum(1 for item in payload if item["uciid"]), sent_count)},
            {"label": _("Startovní číslo"), "value": sum(1 for item in payload if item["bib"]), "total": sent_count, "percent": _pct(sum(1 for item in payload if item["bib"]), sent_count)},
            {"label": _("Klub / tým"), "value": sum(1 for item in payload if item["team"]), "total": sent_count, "percent": _pct(sum(1 for item in payload if item["team"]), sent_count)},
        ],
    }

    return render(request, "event/results_sent.html", {
        "event": event,
        "sent_count": sent_count,
        "export_summary": export_summary,
    })


@login_required(login_url="/login/")
@staff_member_required
def export_uci_results(request, event_id):
    event = get_object_or_404(Event.objects.select_related("classes_and_fees_like"), id=event_id)
    logger.info("Spouštím UCI export pro event_id=%s", event.id)

    if not event.is_uci_race:
        logger.warning("UCI export odmítnut: event_id=%s není UCI závod", event.id)
        messages.error(request, _("Tento závod není označen jako UCI závod."))
        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": event.id}))

    if not os.path.exists(UCI_EXPORT_TEMPLATE):
        logger.error("UCI export selhal: chybí šablona %s", UCI_EXPORT_TEMPLATE)
        messages.error(request, _("Chybí UCI XLSX šablona pro export."))
        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": event.id}))

    if not event.uci_event_code.strip():
        logger.warning("UCI export odmítnut: event_id=%s nemá UCI_EVENT_CODE", event.id)
        messages.error(request, _("U závodu chybí UCI_EVENT_CODE."))
        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": event.id}))

    missing_competition_codes = get_missing_uci_competition_codes(event)
    if missing_competition_codes:
        logger.warning(
            "UCI export odmítnut: event_id=%s chybí competition codes %s",
            event.id,
            ", ".join(missing_competition_codes),
        )
        messages.error(
            request,
            _("U závodu chybí některé UCI kódy kategorií: %(fields)s.")
            % {"fields": ", ".join(missing_competition_codes)},
        )
        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": event.id}))

    zip_name, zip_bytes, export_metadata = generate_uci_export_zip(event, UCI_EXPORT_TEMPLATE)
    for item in export_metadata:
        logger.info(
            "UCI export kategorie vygenerován pro event_id=%s, slug=%s, rows=%s",
            event.id,
            item["slug"],
            item["rows"],
        )

    logger.info(
        "UCI export ZIP připraven pro event_id=%s, files=%s, zip_name=%s",
        event.id,
        len(export_metadata),
        zip_name,
    )
    audit_logger.info(
        "event_uci_export_generated admin_user_id=%s event_id=%s files=%s zip_name=%s",
        request.user.id,
        event.id,
        len(export_metadata),
        zip_name,
    )
    response = HttpResponse(zip_bytes, content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{zip_name}"'
    return response


@login_required(login_url="/login/")
@staff_member_required
def unpaid_moto_riders_report(request, pk):
    event = get_object_or_404(Event, pk=pk)

    if not RaceRun.objects.filter(event=event, round_type="MOTO").exists():
        messages.error(request, _("Pro tento závod zatím nejsou nahrané žádné MOTO jízdy v RaceRun."))
        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": pk}))

    report = build_unpaid_moto_report(event)
    audit_logger.info(
        "event_unpaid_moto_report_opened admin_user_id=%s event_id=%s unmatched=%s no_uci=%s",
        request.user.id,
        event.id,
        len(report.get("unpaid_riders", [])),
        len(report.get("riders_without_uci", [])),
    )
    return render(
        request,
        "event/unpaid-moto-riders-report.html",
        {
            "event": event,
            **report,
        },
    )


@login_required(login_url="/login/")
@staff_member_required
def unpaid_moto_riders_report_pdf(request, pk):
    event = get_object_or_404(Event, pk=pk)

    if not RaceRun.objects.filter(event=event, round_type="MOTO").exists():
        messages.error(request, _("Pro tento závod zatím nejsou nahrané žádné MOTO jízdy v RaceRun."))
        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": pk}))

    report = build_unpaid_moto_report(event)
    service = UnpaidMotoReportPdfService()
    pdf_bytes = service.build_pdf(event, report)
    audit_logger.info(
        "event_unpaid_moto_report_pdf_generated admin_user_id=%s event_id=%s unmatched=%s no_uci=%s",
        request.user.id,
        event.id,
        len(report.get("unpaid_riders", [])),
        len(report.get("riders_without_uci", [])),
    )

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{service.build_filename(event)}"'
    return response


@login_required(login_url="/login/")
@staff_member_required
def send_invoices(request, pk):
    """Rozesílání faktur klubům pro daný závod."""
    generated = generate_event_invoices(pk)
    result = send_event_invoices(pk)
    audit_logger.info(
        "event_invoices_generated admin_user_id=%s event_id=%s sent_count=%s",
        request.user.id,
        pk,
        len(result["sent"]),
    )
    return render(request, "event/results_sent.html", {
        "event": get_object_or_404(Event, pk=pk),
        "sent_count": len(result["sent"]),
        "export_summary": {
            "total_results": len(generated["generated"]),
            "sent_count": len(result["sent"]),
            "skipped_count": len(result["skipped"]),
            "sent_ratio": round((len(result["sent"]) / len(generated["generated"])) * 100) if generated["generated"] else 0,
            "women_count": 0,
            "men_count": 0,
            "czech_count": 0,
            "foreign_count": 0,
            "bike20_count": 0,
            "cruiser_count": 0,
            "beginner_count": 0,
            "unique_teams_count": 0,
            "unique_categories_count": 0,
            "coverage_metrics": [],
            "top_categories": [],
        },
    })


@login_required(login_url="/login/")
@staff_member_required
def price_money_pdf(request, pk):
    """PDF potvrzení převzetí prize money."""
    event = get_object_or_404(Event.objects.select_related("organizer"), pk=pk)
    service = PrizeMoneyPdfService()
    include_amounts = request.GET.get("amounts", "1") != "0"
    if not service.allows_amount_toggle(event):
        include_amounts = True

    try:
        pdf_bytes = service.build_pdf(event, include_amounts=include_amounts)
    except ValueError as exc:
        messages.error(request, str(exc))
        return HttpResponseRedirect(reverse("event:event-admin", kwargs={"pk": pk}))
    audit_logger.info(
        "event_prize_money_pdf_generated admin_user_id=%s event_id=%s include_amounts=%s",
        request.user.id,
        event.id,
        include_amounts,
    )

    suffix = "with-amounts" if include_amounts else "without-amounts"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="prize-money-{event.id}-{suffix}-{timezone.now().strftime("%Y%m%d%H%M%S")}.pdf"'
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


_VALID_STATS_KEYS = START_KEYS | RESULT_KEYS | {"overall_results"}

_STATS_SECTION_LABELS = {
    "motos": "Motos – listina",
    "motos_results": "Motos – výsledky",
    "1_16": "1/16 – listina",
    "1_16_results": "1/16 – výsledky",
    "1_8": "1/8 – listina",
    "1_8_results": "1/8 – výsledky",
    "1_4": "1/4 – listina",
    "1_4_results": "1/4 – výsledky",
    "1_2": "1/2 – listina",
    "1_2_results": "1/2 – výsledky",
    "final": "Finále – listina",
    "final_results": "Finále – výsledky",
    "overall_results": "Celkové výsledky",
}

_ROUND_TYPE_LABELS = {
    "MOTO": "Motos",
    "F16": "1/16",
    "F8": "1/8",
    "F4": "1/4",
    "F2": "1/2",
    "FINAL": "Finále",
}
_ROUND_TYPE_ORDER = ["MOTO", "F16", "F8", "F4", "F2", "FINAL"]
_MAX_STATS_FILE_SIZE = 2 * 1024 * 1024  # 2 MB


def _format_import_message(import_result):
    total = import_result["created"]
    counts = import_result.get("counts_by_round", {})
    parts = [
        f"{_ROUND_TYPE_LABELS.get(rt, rt)}: {counts[rt]}"
        for rt in _ROUND_TYPE_ORDER
        if rt in counts
    ]
    msg = _("RaceRun aktualizován, zapsáno {count} jízd").format(count=total)
    if parts:
        msg += " (" + ", ".join(parts) + ")"
    return msg + "."


def _warn_unmatched_riders(request, unmatched):
    if not unmatched:
        return
    MAX_SHOWN = 10
    shown = unmatched[:MAX_SHOWN]
    rest = len(unmatched) - MAX_SHOWN
    parts = [
        f"{item['category']} — {item['plate']} {item['name']}".strip(" —")
        for item in shown
    ]
    msg = _("Tito závodníci nebyli nalezeni v databázi: ") + "; ".join(parts)
    if rest > 0:
        msg += _(" (a {n} dalších)").format(n=rest)
    messages.warning(request, msg)


def _handle_stats_delete_key(request, event, pk):
    delete_key = request.POST.get("delete_key", "").strip()
    if delete_key not in _VALID_STATS_KEYS:
        messages.error(request, _("Neplatný klíč pro smazání."))
        return redirect("event:import-stats", pk=pk)

    deleted_count = _delete_uploaded_files_by_prefix("event_stats", str(pk), prefix=delete_key)
    import_result = RaceRunImportService().import_event_runs(event)
    audit_logger.info(
        "event_stats_file_deleted admin_user_id=%s event_id=%s delete_key=%s deleted_files=%s imported_runs=%s",
        request.user.id, event.id, delete_key, deleted_count, import_result["created"],
    )
    if deleted_count > 0:
        messages.success(request, _("Soubor pro vybranou sekci byl smazán."))
        messages.info(request, _format_import_message(import_result))
        _warn_unmatched_riders(request, import_result["unmatched"])
    else:
        messages.info(request, _("Pro vybranou sekci nebyl nalezen žádný soubor."))
    return redirect("event:import-stats", pk=pk)


def _handle_stats_delete_all(request, event, pk):
    stats_dir = os.path.join(settings.MEDIA_ROOT, "event_stats", str(pk))
    deleted_count = 0
    if os.path.exists(stats_dir):
        for f in os.listdir(stats_dir):
            file_path = os.path.join(stats_dir, f)
            if os.path.isfile(file_path):
                try:
                    os.remove(file_path)
                    deleted_count += 1
                except Exception as e:
                    logger.error("Nepodařilo se smazat soubor %s: %s", file_path, e)
    RaceRun.objects.filter(event=event).delete()
    audit_logger.info(
        "event_stats_deleted admin_user_id=%s event_id=%s deleted_files=%s",
        request.user.id, event.id, deleted_count,
    )
    if deleted_count > 0:
        messages.success(request, _("Všechny statistiky byly smazány."))
    else:
        messages.info(request, _("Žádné soubory k smazání."))
    return redirect("event:import-stats", pk=pk)


def _handle_stats_file_upload(request, event, pk, redirect_to="event:import-stats"):
    count = 0
    uploaded_labels = []
    parseable_upload = False
    tsv_upload = False
    for key in request.FILES:
        file = request.FILES[key]
        lower_name = file.name.lower()
        is_tsv_results = key == "overall_results" and lower_name.endswith(".txt")
        is_html = lower_name.endswith(".html") or lower_name.endswith(".htm")
        if not is_html and not is_tsv_results:
            continue
        if file.size > _MAX_STATS_FILE_SIZE:
            messages.warning(
                request,
                _("Soubor {name} je příliš velký ({size} kB), maximum je 2 MB.").format(
                    name=file.name, size=file.size // 1024
                ),
            )
            continue
        _delete_uploaded_files_by_prefix("event_stats", str(pk), prefix=key)
        original_name = os.path.basename(file.name)
        file.name = f"{key}__{original_name}"
        _save_uploaded_file(file, "event_stats", str(pk))
        count += 1
        uploaded_labels.append(_STATS_SECTION_LABELS.get(key, key))
        parseable_upload = parseable_upload or key in START_KEYS or key in RESULT_KEYS
        tsv_upload = tsv_upload or is_tsv_results

    if count > 0:
        messages.success(
            request,
            _("Nahráno ({count}): {names}.").format(count=count, names=", ".join(uploaded_labels)),
        )
        if tsv_upload:
            latest_tsv = _get_uploaded_stats_by_key(pk).get("overall_results")
            import_result = {"created": 0, "counts_by_round": {}, "unmatched": []}
            if latest_tsv:
                import_result = RemTsvRaceRunImportService().import_file(event, latest_tsv["path"])
            audit_logger.info(
                "event_rem_tsv_stats_imported admin_user_id=%s event_id=%s uploaded_files=%s imported_runs=%s",
                request.user.id, event.id, count, import_result["created"],
            )
            messages.info(request, _format_import_message(import_result))
            _warn_unmatched_riders(request, import_result["unmatched"])
        elif parseable_upload:
            results_count = Result.objects.filter(event=event).count()
            import_result = RaceRunImportService().import_event_runs(event)
            audit_logger.info(
                "event_stats_imported admin_user_id=%s event_id=%s uploaded_files=%s imported_runs=%s results_count=%s",
                request.user.id, event.id, count, import_result["created"], results_count,
            )
            messages.info(request, _format_import_message(import_result))
            if results_count == 0:
                messages.warning(
                    request,
                    _(
                        "Pro tento závod nejsou v databázi žádné výsledky (Result), takže statistické jízdy nešlo s nikým spárovat. Nejdřív nahraj výsledky závodu do Result a potom spusť import statistik znovu."
                    ),
                )
            elif import_result["created"] == 0:
                messages.warning(
                    request,
                    _(
                        "Statistické soubory se nahrály, ale nevznikl žádný RaceRun. Zkontroluj, že kategorie, jména a čísla v HTML odpovídají výsledkům uloženým v Result."
                    ),
                )
            _warn_unmatched_riders(request, import_result["unmatched"])
        else:
            audit_logger.info(
                "event_stats_file_uploaded admin_user_id=%s event_id=%s uploaded_files=%s",
                request.user.id,
                event.id,
                count,
            )
    else:
        messages.warning(request, _("Nebyly nahrány žádné soubory (povoleny jsou pouze .html)."))
    return redirect(redirect_to, pk=pk)


def _get_uploaded_stats_by_key(pk):
    uploaded_files_by_key = {}
    stats_dir = os.path.join(settings.MEDIA_ROOT, "event_stats", str(pk))
    if not os.path.exists(stats_dir):
        return uploaded_files_by_key

    for f in os.listdir(stats_dir):
        file_path = os.path.join(stats_dir, f)
        if os.path.isfile(file_path):
            parts = f.split("__", 1)
            key = parts[0]
            original_name = parts[1] if len(parts) > 1 else f
            uploaded_files_by_key[key] = {
                "key": key,
                "filename": f,
                "original_name": original_name,
                "label": _STATS_SECTION_LABELS.get(key, _("Neznámá kategorie")),
                "path": file_path,
                "url": f"{settings.MEDIA_URL}event_stats/{pk}/{f}",
                "created": datetime.datetime.fromtimestamp(os.path.getctime(file_path)),
            }
    return uploaded_files_by_key


@login_required(login_url="/login/")
@staff_member_required
def import_event_stats(request, pk):
    """Stránka pro import statistických údajů závodu."""
    event = get_object_or_404(Event, pk=pk)

    if request.method == "POST":
        if "delete_key" in request.POST:
            return _handle_stats_delete_key(request, event, pk)
        if "delete_all" in request.POST:
            return _handle_stats_delete_all(request, event, pk)
        if request.FILES:
            return _handle_stats_file_upload(request, event, pk)
        return redirect("event:import-stats", pk=pk)

    # Načtení existujících souborů
    uploaded_files = []
    uploaded_files_by_key = _get_uploaded_stats_by_key(pk)

    upload_sections = [
        {
            "title": _("Motos (Rozjížďky)"),
            "fields": [
                {"name": "motos", "label": _("Startovní listina")},
                {"name": "motos_results", "label": _("Výsledky")},
            ],
        },
        {
            "title": _("1/16 Finále"),
            "fields": [
                {"name": "1_16", "label": _("Startovní listina")},
                {"name": "1_16_results", "label": _("Výsledky")},
            ],
        },
        {
            "title": _("1/8 Finále"),
            "fields": [
                {"name": "1_8", "label": _("Startovní listina")},
                {"name": "1_8_results", "label": _("Výsledky")},
            ],
        },
        {
            "title": _("1/4 Finále"),
            "fields": [
                {"name": "1_4", "label": _("Startovní listina")},
                {"name": "1_4_results", "label": _("Výsledky")},
            ],
        },
        {
            "title": _("1/2 Finále"),
            "fields": [
                {"name": "1_2", "label": _("Startovní listina")},
                {"name": "1_2_results", "label": _("Výsledky")},
            ],
        },
        {
            "title": _("Finále"),
            "fields": [
                {"name": "final", "label": _("Startovní listina")},
                {"name": "final_results", "label": _("Výsledky")},
            ],
        },
    ]
    overall_results_field = {
        "name": "overall_results",
        "section_title": _("Celkové výsledky"),
        "label": _("Soubor s celkovými výsledky"),
    }

    category_map = {}
    for section in upload_sections:
        for field in section["fields"]:
            category_map[field["name"]] = _("%(section)s - %(label)s") % {
                "section": section["title"],
                "label": field["label"],
            }
    category_map["overall_results"] = _("Celkové výsledky")

    for file_info in uploaded_files_by_key.values():
        file_info["label"] = category_map.get(file_info["key"], _("Neznámá kategorie"))
        uploaded_files.append(file_info)
    
    uploaded_files.sort(key=lambda x: x['label'])

    for section in upload_sections:
        for field in section["fields"]:
            field["uploaded_file"] = uploaded_files_by_key.get(field["name"])
    overall_results_field["uploaded_file"] = uploaded_files_by_key.get(overall_results_field["name"])

    phase_warnings = []
    for section in upload_sections:
        keys = [field["name"] for field in section["fields"]]
        has_start = keys[0] in uploaded_files_by_key
        has_results = keys[1] in uploaded_files_by_key
        if has_start and not has_results:
            phase_warnings.append(
                _("{section}: nahrána startovní listina, chybí výsledky").format(section=section["title"])
            )
        elif has_results and not has_start:
            phase_warnings.append(
                _("{section}: nahrány výsledky, chybí startovní listina").format(section=section["title"])
            )

    return render(
        request,
        "event/import-stats.html",
        {
            "event": event,
            "uploaded_files": uploaded_files,
            "upload_sections": upload_sections,
            "overall_results_field": overall_results_field,
            "phase_warnings": phase_warnings,
        },
    )

import logging
from django.conf import settings
import os
import event

logger = logging.getLogger(__name__)
from .models import Entry, Event, EntryForeign as EntryForeignModel
from rider.models import Rider, ForeignRider
from django.utils import timezone
import stripe
import json
from openpyxl import Workbook


class EntryClass:
    """ Class for saving entries to the Entry table in database """

    def __init__(self):
        self.transaction_id = None
        self.event = None
        self.rider: Rider = None
        self.is_beginner: bool = False
        self.is_20: bool = False
        self.is_24: bool = False
        self.class_beginner: str = None
        self.class_20: str = None
        self.class_24: str = None
        self.fee_beginner: int = 0
        self.fee_20: int = 0
        self.fee_24: int = 0

    def save(self):
        new_czech_entry = Entry.objects.create(
            transaction_id=self.transaction_id,
            event=self.event,
            rider=self.rider,
            is_beginner=self.is_beginner,
            is_20=self.is_20,
            is_24=self.is_24,
            fee_beginner=self.fee_beginner,
            fee_20=self.fee_20,
            fee_24=self.fee_24,
            class_beginner=self.class_beginner,
            class_20=self.class_20,
            class_24=self.class_24,
        )
        new_czech_entry.save()


class EntryForeign:
    """ Class for saving foreign entries to the Foreign Entry table in database """

    def __init__(self):
        self.event=None
        self.first_name=None
        self.last_name=None
        self.uci_id=None
        self.gender=None
        self.nationality=None
        self.club=None
        self.transponder=None
        self.is_20: bool =None
        self.is_24: bool=None
        self.class_20=None
        self.class_24=None
        self.fee_20=None
        self.fee_24=None
        self.checkout: bool=False
        self.transaction_id=None
        self.customer_name=None
        self.customer_email=None
        self.payment_complete:bool=False
        self.transaction_date=None

    def save(self):
        new_foreign_entry = EntryForeignModel.objects.create(
            event = self.event,
            first_name=self.first_name,
            last_name=self.last_name,
            uci_id=self.uci_id,
            gender=self.gender,
            nationality=self.nationality,
            club=self.club,
            transponder=self.transponder,
            is_20=self.is_20,
            is_24=self.is_24,
            class_20=self.class_20,
            class_24=self.class_24,
            fee_20=self.fee_20,
            fee_24=self.fee_24,
            checkout=self.checkout,
            transaction_id=self.transaction_id,
            customer_name=self.customer_name,
            customer_email=self.customer_email,
            payment_complete=self.payment_complete,
            transaction_date=self.transaction_date,
        )
        new_foreign_entry.save()
        

class SendConfirmEmail:
    """ Class for sending e-mail about registration """
    stripe.api_key = settings.STRIPE_SECRET_KEY

    def __init__(self, transaction_id):
        self.transaction_id = transaction_id

    def get_customers_email(self):
        """ Method for getting customer e-mail from stripe transaction """
        transaction_detail = stripe.checkout.Session.retrieve(self.transaction_id, )
        transaction_json = json.loads(str(transaction_detail))
        return transaction_json['customer_details']['email']

    def get_event_id(self):
        """ Method for getting event ID from Entry database table"""
        transaction = Entry.objects.filter(transaction_id=self.transaction_id)
        first_entry = transaction.first()
        return first_entry.event if first_entry else None

    def get_message_body(self):
        """ Method for setting e-mail MESSAGE_BODY """
        entries_beginner = Entry.objects.filter(transaction_id=self.transaction_id, is_beginner=True)
        entries_20 = Entry.objects.filter(transaction_id=self.transaction_id, is_20=True)
        entries_24 = Entry.objects.filter(transaction_id=self.transaction_id, is_24=True)

        riders_beginner = Rider.objects.filter(id__in=entries_beginner.values_list('rider_id', flat=True))
        riders_20 = Rider.objects.filter(id__in=entries_20.values_list('rider_id', flat=True))
        riders_24 = Rider.objects.filter(id__in=entries_24.values_list('rider_id', flat=True))

        message_body = ""

        if riders_beginner:
            message_body += " \r\n"
            message_body += " \r\n"
            message_body += "- do kategorie Příchozích byly přihlášeni tito jezdci: "
            for rider_beginner in riders_beginner:
                message_body += f"{rider_beginner.last_name.upper()} {rider_beginner.first_name}, UCI ID: {rider_beginner.uci_id}, v kategorii {rider_beginner.class_beginner}, "
        del entries_beginner

        if riders_20:
            message_body += " \r\n"
            message_body += " \r\n"
            message_body += "- do kategorie Challenge, Junior, Under a Elite byly přihlášeni tito jezdci: "
            for rider_20 in riders_20:
                message_body += f"{rider_20.last_name.upper()} {rider_20.first_name}, UCI ID: {rider_20.uci_id}, v kategorii {rider_20.class_20}, "
        del entries_20

        if riders_24:
            message_body += " \r\n"
            message_body += " \r\n"
            message_body += "- do kategorie Cruiser byly přihlášeni tito jezdci: "
            for rider_24 in riders_24:
                message_body += f"{rider_24.last_name.upper()} {rider_24.first_name}, UCI ID: {rider_24.uci_id}, v kategorii {rider_24.class_24}, "
        del entries_24
        return message_body

    def send_email(self):
        """ Method for sending e-mail with confirm registration - transaction ID required at creating instance """
        recipient = self.get_customers_email()
        message = self.get_message_body()
        event_id = self.get_event_id()
        event = Event.objects.filter(id=event_id).first()
        if not event:
            return
        MESSAGE_SUBJECT = f"TEST!!! Potvrzení o registraci jezdců na závod BMX race - {event.name}"
        MESSAGE_BODY = f"Do závodu -- {event.name} -- konaného dne {event.date} byly registrováni: " + message + "\r\n \r\n Komise BMX Českého svazu cyklistiky "

        # TODO: Dodělat pdf potvrzení přílohou
        # TODO: Dodělat MESSAGE_BODY v HTML

        # send an email
        # send_mail (
        #      subject = MESSAGE_SUBJECT,
        #      message = MESSAGE_BODY,
        #      from_email = "bmx@ceskysvazcyklistiky.cz",
        #      recipient_list = [recipient],)
        del event


class NumberInEvent:
    """ Class for number on-line registration riders in event """

    def __init__(self):
        self.riders_in_category = 0
        self.event = 0
        self.category_name = ""

    def count_beginners(self):
        """ function for count riders in class Beginners """
        self.riders_in_category = Entry.objects.filter(event=self.event, class_beginner=self.category_name,
                                                       is_beginner=True, payment_complete=True, checkout=False).count()

    def count_riders_20(self):
        """ function for count riders in class Challenge and Championschip """
        self.riders_in_category = Entry.objects.filter(event=self.event, class_20=self.category_name, is_20=True,
                                                       payment_complete=True, checkout=False, is_beginner=False).count()

    def count_riders_24(self):
        """ function for count riders in class Cruiser """
        self.riders_in_category = Entry.objects.filter(event=self.event, class_24=self.category_name, is_24=True,
                                                       payment_complete=True, checkout=False).count()


class REMRiders:
    """ Class for create riders lists in xlsx file for REM """

    def __init__(self):
        self.event = None
        self.file_name = None
        self.template_name = os.path.join(settings.MEDIA_ROOT, "rem_entries", "Rider_and_Registration.xlsx")
        # self.wb = load_workbook(self.template_name)
        self.wb = Workbook()
        self.wb.encoding = "utf-8"
        self.ws = self.wb.active
        self.ws.title = "Rider-Registration"
        self.riders = Rider.objects.filter(is_active=True, is_approved=True)
        self.foreign_riders = ForeignRider.objects.filter()

    def _save_workbook(self, file_name):
        os.makedirs(os.path.dirname(file_name), exist_ok=True)
        self.wb.save(file_name)

    def first_line(self):
        """ set first line in REM online entries excel file """
        self.ws.cell(1, 1, "Event")
        self.ws.cell(1, 2, "First")
        self.ws.cell(1, 3, "Last")
        self.ws.cell(1, 4, "Email")
        self.ws.cell(1, 5, "Club")
        self.ws.cell(1, 6, "Team")
        self.ws.cell(1, 7, "Country")
        self.ws.cell(1, 8, "Birthdate")
        self.ws.cell(1, 9, "Sex")
        self.ws.cell(1, 10, "UCIID")
        self.ws.cell(1, 11, "Rider_type")
        self.ws.cell(1, 12, "Rider_licence_type")
        self.ws.cell(1, 13, "Rider_ident")
        self.ws.cell(1, 14, "Paid")
        self.ws.cell(1, 15, "Event_price")
        self.ws.cell(1, 16, "Admin_fee")
        self.ws.cell(1, 17, "Transponder_hire_price")
        self.ws.cell(1, 18, "Class")
        self.ws.cell(1, 19, "Plate")
        self.ws.cell(1, 20, "Transponder")
        self.ws.cell(1, 21, "Plate_1")  # cruiser
        self.ws.cell(1, 22, "Transponder_1")  # cruiser
        self.ws.cell(1, 23, "Transponder_hire_flag")

    def create_all_riders_list(self):
        self.file_name = os.path.join(settings.MEDIA_ROOT, "rem_riders", f"REM_ALL_RIDERS_FOR_RACE_ID-{self.event.id}.xlsx")
        self.first_line()

        row: int = 2

        for rider in self.riders:
            self.ws.cell(row, 1, )
            self.ws.cell(row, 2, rider.first_name)
            self.ws.cell(row, 3, rider.last_name)
            self.ws.cell(row, 4, rider.email)
            self.ws.cell(row, 5, event.func.team_name_resolve(rider.club))
            self.ws.cell(row, 6, )
            self.ws.cell(row, 7, "CZE")
            self.ws.cell(row, 8, event.func.date_of_birth_resolve(rider))
            self.ws.cell(row, 9, event.func.gender_resolve(rider))
            self.ws.cell(row, 10, rider.uci_id)
            if rider.is_elite:
                self.ws.cell(row, 11, "E")
            else:
                self.ws.cell(row, 11, "C")
            self.ws.cell(row, 12, "U")
            self.ws.cell(row, 13, )
            self.ws.cell(row, 14, )
            self.ws.cell(row, 15, )
            self.ws.cell(row, 16, )
            self.ws.cell(row, 17, )
            self.ws.cell(row, 18, )
            self.ws.cell(row, 19, rider.plate_display)
            self.ws.cell(row, 20, rider.transponder_20)
            self.ws.cell(row, 21, rider.plate_display)
            self.ws.cell(row, 22, rider.transponder_24)
            self.ws.cell(row, 23, )
            row += 1

        # Add foreign riders
        for rider in self.foreign_riders:
            self.ws.cell(row, 1, )
            self.ws.cell(row, 2, rider.first_name)
            self.ws.cell(row, 3, rider.last_name)
            self.ws.cell(row, 4, )
            self.ws.cell(row, 5, rider.club)
            self.ws.cell(row, 6, )
            self.ws.cell(row, 7, rider.state)
            self.ws.cell(row, 8, event.func.date_of_birth_resolve(rider))
            self.ws.cell(row, 9, event.func.gender_resolve(rider))
            self.ws.cell(row, 10, rider.uci_id)
            if rider.is_elite:
                self.ws.cell(row, 11, "E")
            else:
                self.ws.cell(row, 11, "C")
            self.ws.cell(row, 12, "U")
            self.ws.cell(row, 13, )
            self.ws.cell(row, 14, )
            self.ws.cell(row, 15, )
            self.ws.cell(row, 16, )
            self.ws.cell(row, 17, )
            self.ws.cell(row, 18, )
            self.ws.cell(row, 19, rider.plate_display)
            self.ws.cell(row, 20, rider.transponder_20)
            self.ws.cell(row, 21, rider.plate_display)
            self.ws.cell(row, 22, rider.transponder_24)
            self.ws.cell(row, 23, )
            row += 1

        self._save_workbook(self.file_name)

        self.event.rem_riders_list = self.file_name
        self.event.rem_riders_created = timezone.now()
        self.event.save()

        # self.remove_temp_file()

    def create_entries_list(self):
        file_name = os.path.join(settings.MEDIA_ROOT, "rem_entries", f"REM_ENTRIES_FOR_RACE_ID-{self.event.id}.xlsx")
        self.first_line()
        czech_entries = Entry.objects.filter(event=self.event.id, payment_complete=True, checkout=False).select_related('rider')
        foreign_entries = EntryForeignModel.objects.filter(
            event=self.event.id,
            payment_complete=True,
            checkout=False,
        )

        uci_ids = list(foreign_entries.values_list('uci_id', flat=True))
        foreign_rider_clubs = {
            str(fr_uci_id): club
            for fr_uci_id, club in ForeignRider.objects.filter(uci_id__in=uci_ids, club__gt='').values_list('uci_id', 'club')
        }

        row: int = 2
        for entry in czech_entries:
            try:
                self.ws.cell(row, 1, self.event.name)
                self.ws.cell(row, 2, entry.rider.first_name)
                self.ws.cell(row, 3, entry.rider.last_name)
                self.ws.cell(row, 4, entry.rider.email)
                self.ws.cell(row, 5, event.func.team_name_resolve(entry.rider.club))
                self.ws.cell(row, 6, )
                self.ws.cell(row, 7, entry.rider.nationality)
                self.ws.cell(row, 8, event.func.date_of_birth_resolve_rem_online(entry.rider.date_of_birth))
                self.ws.cell(row, 9, event.func.gender_resolve_small_letter(entry.rider.gender))
                self.ws.cell(row, 10, entry.rider.uci_id)
                if entry.rider.is_elite:
                    self.ws.cell(row, 11, "E")
                else:
                    self.ws.cell(row, 11, "C")
                self.ws.cell(row, 12, "U")
                self.ws.cell(row, 13, )
                self.ws.cell(row, 14, "true")
                if entry.is_beginner:
                    self.ws.cell(row, 15, entry.fee_beginner)
                elif entry.is_20:
                    self.ws.cell(row, 15, entry.fee_20)
                else:
                    self.ws.cell(row, 15, entry.fee_24)
                self.ws.cell(row, 16, )
                self.ws.cell(row, 17, )
                if entry.is_beginner:
                    self.ws.cell(row, 18, entry.class_beginner)
                elif entry.is_20:
                    self.ws.cell(row, 18, entry.class_20)
                else:
                    self.ws.cell(row, 18, entry.class_24)
                if entry.is_20 and entry.rider.plate_champ_20:
                    world_plate = "W" + str(entry.rider.plate_champ_20)
                    self.ws.cell(row, 19, world_plate)
                elif entry.is_20 or entry.is_beginner:
                    self.ws.cell(row, 19, entry.rider.plate_display)
                elif entry.is_24 and entry.rider.plate_champ_24:
                    world_plate = "W" + str(entry.rider.plate_champ_24)
                    self.ws.cell(row, 21, world_plate)
                else:
                    self.ws.cell(row, 21, entry.rider.plate_display)
                if entry.is_24:
                    self.ws.cell(row, 22, entry.rider.transponder_24)
                else:
                    self.ws.cell(row, 20, entry.rider.transponder_20)
            except Exception as E:
                logger.error(f"Chyba při ukládání jezdce do REM: {E}")
            row += 1
        del czech_entries

        for entry in foreign_entries:
            try:
                self.ws.cell(row, 1, self.event.name)
                self.ws.cell(row, 2, entry.first_name)
                self.ws.cell(row, 3, entry.last_name)
                self.ws.cell(row, 4, entry.customer_email)
                club = foreign_rider_clubs.get(entry.uci_id) or entry.club or event.func.foreign_club_resolve(entry.nationality or "")
                self.ws.cell(row, 5, club)
                self.ws.cell(row, 6, )
                self.ws.cell(row, 7, entry.nationality)
                self.ws.cell(row, 8, event.func.date_of_birth_resolve_rem_online(entry.date_of_birth))
                self.ws.cell(row, 9, event.func.gender_resolve_small_letter(entry.gender))
                self.ws.cell(row, 10, entry.uci_id)
                if entry.is_elite:
                    self.ws.cell(row, 11, "E")
                else:
                    self.ws.cell(row, 11, "C")
                self.ws.cell(row, 12, "U")
                self.ws.cell(row, 13, )
                self.ws.cell(row, 14, "true")
                if entry.is_20:
                    self.ws.cell(row, 15, entry.fee_20)
                else:
                    self.ws.cell(row, 15, entry.fee_24)
                self.ws.cell(row, 16, )
                self.ws.cell(row, 17, )
                if entry.is_20:
                    self.ws.cell(row, 18, entry.class_20)
                else:
                    self.ws.cell(row, 18, entry.class_24)
                if entry.is_20:
                    self.ws.cell(row, 19, entry.plate)
                    self.ws.cell(row, 20, entry.transponder_20)
                else:
                    self.ws.cell(row, 21, entry.plate)
                    self.ws.cell(row, 22, entry.transponder_24)
            except Exception as E:
                logger.error(f"Chyba při ukládání jezdce do REM: {E}")
            row += 1
        del foreign_entries

        self._save_workbook(file_name)
        self.event.rem_entries = file_name
        self.event.rem_entries_created = timezone.now()
        self.event.save()

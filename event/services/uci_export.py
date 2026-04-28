import os
import re
import tempfile
import zipfile

from openpyxl import load_workbook

from event.models import RaceRun
from rider.plates import display_plate


UCI_EXPORT_CATEGORY_CONFIG = (
    ("women_elite", "Women Elite", "uci_code_women_elite"),
    ("men_elite", "Men Elite", "uci_code_men_elite"),
    ("women_u23", "Women Under 23", "uci_code_women_under_23"),
    ("men_u23", "Men Under 23", "uci_code_men_under_23"),
    ("women_junior", "Women Junior", "uci_code_women_junior"),
    ("men_junior", "Men Junior", "uci_code_men_junior"),
)


def sanitize_export_filename(value):
    safe = "".join(char if char.isalnum() or char in ("-", "_") else "_" for char in str(value or ""))
    return safe.strip("_") or "export"


def get_missing_uci_competition_codes(event):
    return [
        competition_code_field
        for _, _, competition_code_field in UCI_EXPORT_CATEGORY_CONFIG
        if not (getattr(event, competition_code_field, "") or "").strip()
    ]


def format_uci_rank_suffix(rank):
    try:
        rank_int = int(rank)
    except (TypeError, ValueError):
        return str(rank or "")

    if 10 <= rank_int % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(rank_int % 10, "th")
    return f"{rank_int}{suffix}"


def _parse_place_int(place_str):
    """Převede "1st", "2nd", "DNS" apod. na číslo pro řazení. Nečíselné hodnoty jdou na konec."""
    match = re.match(r"(\d+)", str(place_str or ""))
    return (0, int(match.group(1))) if match else (1, 0)


def build_uci_export_rows(event, rider_class_name):
    """Sestaví řádky UCI exportu výhradně z RaceRun dat (Import statistik).

    Zdroj je záměrně RaceRun, nikoli Result — v *.txt souboru nejsou oddělovány
    sloučené kategorie (např. Elite + U23 v jednom závodě), zatímco HTML soubory
    z BEMu mají každou kategorii ve vlastní tabulce s přesným názvem.
    """
    final_runs = (
        RaceRun.objects.filter(
            event=event,
            category__iexact=rider_class_name,
            round_type="FINAL",
        )
        .select_related("rider__club", "result")
        .exclude(place__isnull=True)
        .exclude(place="")
    )

    sorted_runs = sorted(final_runs, key=lambda run: _parse_place_int(run.place))

    rows = []
    subgroup_rank = 0
    for run in sorted_runs:
        rider = run.rider
        result = run.result

        if not rider and not result:
            continue

        subgroup_rank += 1

        uci_id = ""
        if rider:
            uci_id = rider.uci_id
        elif result and result.rider_id:
            uci_id = result.rider_id  # to_field="uci_id" — rider_id IS the UCI ID

        last_name = (result.last_name if result else None) or (rider.last_name if rider else "") or ""
        first_name = (result.first_name if result else None) or (rider.first_name if rider else "") or ""
        country = (result.country if result else None) or (rider.nationality if rider else "") or "CZE"

        if rider and rider.club:
            team = rider.club.team_name
        elif result and result.club:
            team = result.club
        else:
            team = ""

        gender = "W" if (rider and rider.gender == "Žena") else "M"

        # Tabulka: primárně z RaceRun.plate (přesné, z BEM HTML), fallback z jezdce
        bib = display_plate(run.plate) if run.plate else (
            display_plate(rider.plate_text, rider.plate) if rider else ""
        )

        finish_time = run.finish_time
        rows.append({
            "rank": subgroup_rank,
            "bib": bib,
            "uci_id": uci_id,
            "last_name": last_name,
            "first_name": first_name,
            "country": country,
            "team": team,
            "gender": gender,
            "phase": "Final",
            "heat": 1,
            "result": (
                f"{format_uci_rank_suffix(subgroup_rank)}, {finish_time:.3f}"
                if finish_time is not None
                else format_uci_rank_suffix(subgroup_rank)
            ),
            "irm": "",
        })

    return rows


def write_uci_export_workbook(template_path, destination_path, competition_code, event_code, rows):
    wb = load_workbook(template_path)
    general_ws = wb["General"]
    results_ws = wb["Results"]

    general_ws["B4"] = competition_code
    general_ws["B5"] = event_code

    row_index = 2
    for row in rows:
        results_ws.cell(row_index, 1, row["rank"])
        results_ws.cell(row_index, 2, row["bib"])
        results_ws.cell(row_index, 3, row["uci_id"])
        results_ws.cell(row_index, 4, row["last_name"])
        results_ws.cell(row_index, 5, row["first_name"])
        results_ws.cell(row_index, 6, row["country"])
        results_ws.cell(row_index, 7, row["team"])
        results_ws.cell(row_index, 8, row["gender"])
        results_ws.cell(row_index, 9, row["phase"])
        results_ws.cell(row_index, 10, row["heat"])
        results_ws.cell(row_index, 11, row["result"])
        results_ws.cell(row_index, 12, row["irm"])
        row_index += 1

    wb.save(destination_path)


def generate_uci_export_zip(event, template_path):
    generated_files = []
    export_metadata = []

    with tempfile.TemporaryDirectory(prefix="uci-export-") as tmp_dir:
        for slug, rider_class_name, competition_code_field in UCI_EXPORT_CATEGORY_CONFIG:
            competition_code = getattr(event, competition_code_field, "") or ""
            rows = build_uci_export_rows(event, rider_class_name)

            export_name = (
                f"uci_results_{event.id}_"
                f"{sanitize_export_filename(slug)}_"
                f"{sanitize_export_filename(competition_code)}.xlsx"
            )
            destination_path = os.path.join(tmp_dir, export_name)
            write_uci_export_workbook(
                template_path=template_path,
                destination_path=destination_path,
                competition_code=competition_code,
                event_code=event.uci_event_code,
                rows=rows,
            )
            generated_files.append(destination_path)
            export_metadata.append({
                "slug": slug,
                "rows": len(rows),
                "competition_code": competition_code,
                "filename": export_name,
            })

        zip_name = f"uci_results_event_{event.id}_{sanitize_export_filename(event.name)}.zip"
        zip_path = os.path.join(tmp_dir, zip_name)
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
            for file_path in generated_files:
                zip_file.write(file_path, arcname=os.path.basename(file_path))

        with open(zip_path, "rb") as zip_handle:
            return zip_name, zip_handle.read(), export_metadata

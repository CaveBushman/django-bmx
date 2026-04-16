from datetime import date, timedelta
from io import BytesIO
import os
from pathlib import Path
import tempfile
from types import SimpleNamespace
import zipfile
from unittest.mock import patch

from django.conf import settings
from django.contrib import admin
from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from django.contrib.auth import get_user_model
from django.contrib.messages import get_messages
from django.core import mail
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import transaction
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from django_ckeditor_5.widgets import CKEditor5Widget
from openpyxl import Workbook, load_workbook
from PIL import Image
from commissar.models import Commissar
from reportlab.pdfgen import canvas

from club.models import Club
from event.forms import EventPropositionForm
from event.admin import CreditTransactionAdmin, DebetTransactionAdmin, EntryForeignAdmin
from event.credit import recalculate_all_balances
from event.models import CreditTransaction, DebetTransaction, Entry, EntryAuditLog, EntryClasses, EntryForeign, Event, FinanceAuditLog, SeasonSettings, RaceRun, Result
from event.func import SetResults
from event.services.race_run_import import RaceRunImportService
from event.services.unpaid_moto_report import build_unpaid_moto_report
from event.services.uci_export import build_uci_export_rows, generate_uci_export_zip
from event.services.checkout_refunds import apply_entry_checkout
from event.services.payments import (
    enrich_cart_entries,
    get_recent_pending_entries,
    remove_conflicting_cart_entries,
)
from event.views.entry_helpers import (
    build_public_entry_rows,
    enrich_foreign_summary_rows,
    sync_paid_foreign_riders,
    validate_foreign_summary_payload,
)
from accounts.models import Account
from rider.models import Rider
from rider.models import ForeignRider
from rider.rider import RiderQualifyToCNThread, should_recount_cn_qualification_for_event


User = get_user_model()


class EventEntryWorkflowTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            first_name="Entry",
            last_name="Tester",
            username="entry_tester",
            email="entry_tester@example.com",
            password="StrongPass123!",
        )
        self.user.is_active = True
        self.user.save()

        self.club = Club.objects.create(team_name="Test Club")
        self.entry_classes = EntryClasses.objects.create(
            event_name="Test classes",
            beginners_1="Beginners 1",
            beginners_2="Beginners 2",
            beginners_3="Beginners 3",
            beginners_4="Beginners 4",
            boys_6="Boys 6",
            girls_6="Girls 6",
            cr_boys_12_and_under="Boys 12 and under",
        )
        self.event = Event.objects.create(
            name="Test race",
            date=date.today() + timedelta(days=30),
            organizer=self.club,
            classes_and_fees_like=self.entry_classes,
            reg_open=True,
            reg_open_from=timezone.now() - timedelta(days=1),
            reg_open_to=timezone.now() + timedelta(days=1),
            type_for_ranking="Volný závod",
        )
        self.rider = Rider.objects.create(
            uci_id=12345678901,
            first_name="Rider",
            last_name="Test",
            gender="Muž",
            date_of_birth=date(2018, 1, 1),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            class_20="Boys 6",
            class_24="Boys 12 and under",
            class_beginner="Beginners 1",
        )

    def test_entry_page_loads_for_logged_in_user(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("event:entry", kwargs={"pk": self.event.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["event"].pk, self.event.pk)
        self.assertContains(response, "Přihlášení na závody")

    def test_entry_page_loads_without_season_settings(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("event:entry", kwargs={"pk": self.event.pk}))

        self.assertEqual(response.status_code, 200)

    def test_mcr_entry_rejects_unqualified_rider_for_20_and_24(self):
        self.client.force_login(self.user)
        SeasonSettings.objects.create(year=date.today().year, qualify_to_cn=2)
        self.event.type_for_ranking = "Mistrovství ČR jednotlivců"
        self.event.save(update_fields=["type_for_ranking"])

        response = self.client.post(
            reverse("event:entry", kwargs={"pk": self.event.pk}),
            {
                "btn_add": "1",
                "checkbox_20": str(self.rider.uci_id),
                "checkbox_24": str(self.rider.uci_id),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("event:events"))
        self.assertFalse(
            Entry.objects.filter(event=self.event, rider=self.rider, is_20=True).exists()
        )
        self.assertFalse(
            Entry.objects.filter(event=self.event, rider=self.rider, is_24=True).exists()
        )

    def test_entry_rejects_elite_rider_for_24_category(self):
        self.client.force_login(self.user)
        self.rider.is_elite = True
        self.rider.save(update_fields=["is_elite"])

        response = self.client.post(
            reverse("event:entry", kwargs={"pk": self.event.pk}),
            {
                "btn_add": "1",
                "checkbox_24": str(self.rider.uci_id),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("event:events"))
        self.assertFalse(
            Entry.objects.filter(event=self.event, rider=self.rider, is_24=True).exists()
        )

    def test_historical_czech_cup_does_not_recount_current_year_cn_qualification(self):
        previous_year = date.today().year - 1
        SeasonSettings.objects.create(year=previous_year, qualify_to_cn=1)

        old_championship = Event.objects.create(
            name="Old championship",
            date=date(previous_year, 9, 1),
            organizer=self.club,
            type_for_ranking="Mistrovství ČR jednotlivců",
        )
        old_cup = Event.objects.create(
            name="Old cup",
            date=date(previous_year, 5, 1),
            organizer=self.club,
            type_for_ranking="Český pohár",
        )

        self.rider.is_qualify_to_cn_20 = True
        self.rider.is_qualify_to_cn_24 = False
        self.rider.save(update_fields=["is_qualify_to_cn_20", "is_qualify_to_cn_24"])

        self.assertFalse(should_recount_cn_qualification_for_event(old_cup))

        RiderQualifyToCNThread(year=previous_year).run()

        self.rider.refresh_from_db()
        self.assertTrue(self.rider.is_qualify_to_cn_20)
        self.assertFalse(self.rider.is_qualify_to_cn_24)


class Custom404Tests(TestCase):
    @override_settings(DEBUG=False)
    def test_unknown_route_returns_custom_404_page(self):
        response = self.client.get("/this-page-does-not-exist/")

        self.assertEqual(response.status_code, 404)
        self.assertContains(response, "Stránka nebyla nalezena", status_code=404)


class UnpaidMotoReportTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(team_name="Report Club")
        self.event = Event.objects.create(
            name="Moto report event",
            date=date.today(),
            organizer=self.club,
            type_for_ranking="Volný závod",
        )

    def test_flagged_count_excludes_rows_without_uci(self):
        paid_rider = Rider.objects.create(
            uci_id=10000000001,
            first_name="Paid",
            last_name="Rider",
            gender="Muž",
            date_of_birth=date(2010, 1, 1),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
        )
        unpaid_rider = Rider.objects.create(
            uci_id=10000000002,
            first_name="Unpaid",
            last_name="Rider",
            gender="Muž",
            date_of_birth=date(2010, 1, 2),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
        )

        Entry.objects.create(event=self.event, rider=paid_rider, payment_complete=True)
        Result.objects.create(
            event=self.event,
            rider=paid_rider,
            first_name=paid_rider.first_name,
            last_name=paid_rider.last_name,
            category="Boys 10",
            place=1,
        )
        Result.objects.create(
            event=self.event,
            rider=unpaid_rider,
            first_name=unpaid_rider.first_name,
            last_name=unpaid_rider.last_name,
            category="Boys 10",
            place=2,
        )

        RaceRun.objects.create(
            event=self.event,
            rider=paid_rider,
            category="Boys 10",
            round_type="MOTO",
            round_number=1,
            plate="101",
        )
        RaceRun.objects.create(
            event=self.event,
            rider=unpaid_rider,
            category="Boys 10",
            round_type="MOTO",
            round_number=1,
            plate="102",
        )

        missing_uci_result = Result.objects.create(
            event=self.event,
            first_name="Manual",
            last_name="Check",
            category="Boys 10",
        )
        RaceRun.objects.create(
            event=self.event,
            result=missing_uci_result,
            category="Boys 10",
            round_type="MOTO",
            round_number=1,
            plate="103",
        )

        report = build_unpaid_moto_report(self.event)

        self.assertEqual(report["flagged_count"], 1)
        self.assertEqual(len(report["confirmed_unpaid"]), 1)
        self.assertEqual(report["confirmed_unpaid"][0].uci_id, str(unpaid_rider.uci_id))
        self.assertEqual(len(report["missing_uci"]), 1)

    def test_report_ignores_moto_riders_without_matching_event_result(self):
        matched_rider = Rider.objects.create(
            uci_id=10000000003,
            first_name="Matched",
            last_name="Rider",
            gender="Muž",
            date_of_birth=date(2010, 1, 3),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
        )
        stray_rider = Rider.objects.create(
            uci_id=10000000004,
            first_name="Stray",
            last_name="Rider",
            gender="Muž",
            date_of_birth=date(2010, 1, 4),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
        )

        Result.objects.create(
            event=self.event,
            rider=matched_rider,
            first_name=matched_rider.first_name,
            last_name=matched_rider.last_name,
            category="Boys 10",
            place=5,
        )

        RaceRun.objects.create(
            event=self.event,
            rider=matched_rider,
            category="Boys 10",
            round_type="MOTO",
            round_number=1,
            plate="201",
            place="3rd",
        )
        RaceRun.objects.create(
            event=self.event,
            rider=stray_rider,
            category="Boys 10",
            round_type="MOTO",
            round_number=1,
            plate="202",
            place="4th",
        )

        report = build_unpaid_moto_report(self.event)

        self.assertEqual(report["flagged_count"], 1)
        self.assertEqual(len(report["confirmed_unpaid"]), 1)
        self.assertEqual(report["confirmed_unpaid"][0].uci_id, str(matched_rider.uci_id))

    def test_report_uses_motos_start_file_and_event_paid_entries_only(self):
        paid_rider = Rider.objects.create(
            uci_id=10000000005,
            first_name="Paid",
            last_name="Starter",
            gender="Muž",
            date_of_birth=date(2010, 1, 5),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            class_20="Boys 10",
            plate_text="101",
        )
        stray_rider = Rider.objects.create(
            uci_id=10000000006,
            first_name="Stray",
            last_name="Starter",
            gender="Muž",
            date_of_birth=date(2010, 1, 6),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            class_20="Boys 10",
            plate_text="102",
        )
        Entry.objects.create(
            event=self.event,
            rider=paid_rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 10",
        )

        with tempfile.TemporaryDirectory() as temp_media:
            with override_settings(MEDIA_ROOT=temp_media):
                target_dir = os.path.join(temp_media, "event_stats", str(self.event.pk))
                os.makedirs(target_dir, exist_ok=True)
                with open(os.path.join(target_dir, "motos__sample.html"), "w", encoding="utf-8") as handle:
                    handle.write(
                        """<html><body>
                        <table class="gridtable">
                        <caption>Boys 10 (2 Riders)</caption>
                        <tr><th>Plate</th><th>Club</th><th>Name</th><th>Moto 1</th></tr>
                        <tr><td>101</td><td>Report Club</td><td>Paid Starter</td><td>12 / 1</td></tr>
                        <tr><td>102</td><td>Report Club</td><td>Stray Starter</td><td>13 / 2</td></tr>
                        </table>
                        </body></html>"""
                    )

                report = build_unpaid_moto_report(self.event)

        self.assertEqual(report["flagged_count"], 1)
        self.assertEqual(len(report["confirmed_unpaid"]), 1)
        self.assertEqual(report["confirmed_unpaid"][0].last_name, "Starter")
        self.assertEqual(report["confirmed_unpaid"][0].plate, "102")
        self.assertEqual(report["confirmed_unpaid"][0].uci_id, str(stray_rider.uci_id))


class PaymentServiceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            first_name="Payment",
            last_name="Tester",
            username="payment_tester",
            email="payment_tester@example.com",
            password="StrongPass123!",
        )
        self.user.is_active = True
        self.user.save()

        self.other_user = User.objects.create_user(
            first_name="Other",
            last_name="User",
            username="other_user",
            email="other_user@example.com",
            password="StrongPass123!",
        )
        self.other_user.is_active = True
        self.other_user.save()

        self.club = Club.objects.create(team_name="Payment Club")
        self.entry_classes = EntryClasses.objects.create(
            event_name="Payment classes",
            beginners_1="Beginners 1",
            boys_6="Boys 6",
            cr_boys_12_and_under="Boys 12 and under",
        )
        self.event = Event.objects.create(
            name="Payment race",
            date=date.today() + timedelta(days=30),
            organizer=self.club,
            classes_and_fees_like=self.entry_classes,
            reg_open=True,
            reg_open_from=timezone.now() - timedelta(days=1),
            reg_open_to=timezone.now() + timedelta(days=1),
            type_for_ranking="Volný závod",
        )
        self.second_event = Event.objects.create(
            name="Second payment race",
            date=date.today() + timedelta(days=40),
            organizer=self.club,
            classes_and_fees_like=self.entry_classes,
            reg_open=True,
            reg_open_from=timezone.now() - timedelta(days=1),
            reg_open_to=timezone.now() + timedelta(days=2),
            type_for_ranking="Volný závod",
        )
        self.rider = Rider.objects.create(
            uci_id=12345678999,
            first_name="Payment",
            last_name="Rider",
            gender="Muž",
            date_of_birth=date(2018, 1, 1),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            class_20="Boys 6",
            class_24="Boys 12 and under",
            class_beginner="Beginners 1",
        )

    def test_get_recent_pending_entries_uses_24_hour_window(self):
        fresh_entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )
        stale_entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            is_24=True,
            class_24="Boys 12 and under",
            fee_24=350,
        )
        old_timestamp = timezone.now() - timedelta(days=2)
        Entry.objects.filter(pk=stale_entry.pk).update(transaction_date=old_timestamp)

        recent_ids = list(get_recent_pending_entries().values_list("id", flat=True))

        self.assertIn(fresh_entry.id, recent_ids)
        self.assertNotIn(stale_entry.id, recent_ids)

    def test_remove_conflicting_cart_entries_deletes_duplicate_only_once(self):
        first_entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )
        duplicate_entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )

        duplicates = remove_conflicting_cart_entries(
            Entry.objects.filter(pk__in=[first_entry.pk, duplicate_entry.pk]).order_by("id")
        )

        self.assertEqual([entry.pk for entry in duplicates], [duplicate_entry.pk])
        self.assertTrue(Entry.objects.filter(pk=first_entry.pk).exists())
        self.assertFalse(Entry.objects.filter(pk=duplicate_entry.pk).exists())

    def test_remove_conflicting_cart_entries_removes_entry_when_other_user_has_same_slot(self):
        foreign_cart_entry = Entry.objects.create(
            user=self.other_user,
            event=self.event,
            rider=self.rider,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )
        user_entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )

        duplicates = remove_conflicting_cart_entries(Entry.objects.filter(pk=user_entry.pk))

        self.assertEqual([entry.pk for entry in duplicates], [user_entry.pk])
        self.assertTrue(Entry.objects.filter(pk=foreign_cart_entry.pk).exists())
        self.assertFalse(Entry.objects.filter(pk=user_entry.pk).exists())


    def test_enrich_cart_entries_sets_display_class_and_total(self):
        beginner_entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            is_beginner=True,
            class_beginner="Beginners 1",
            fee_beginner=200,
        )
        cruiser_entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            is_24=True,
            class_24="Boys 12 and under",
            fee_24=350,
        )
        orders = list(Entry.objects.filter(pk__in=[beginner_entry.pk, cruiser_entry.pk]).order_by("id"))

        total_price = enrich_cart_entries(orders)

        self.assertEqual(total_price, 550)
        self.assertEqual(orders[0].event_class, "Beginners 1")
        self.assertEqual(orders[1].event_class, "Boys 12 and under")

    @patch("event.views.payment_helpers.stripe.checkout.Session.retrieve")
    def test_success_view_confirms_exact_session_even_when_entry_is_older_than_24_hours(self, retrieve_mock):
        self.client.force_login(self.user)
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            transaction_id="sess_exact_entry",
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )
        Entry.objects.filter(pk=entry.pk).update(
            transaction_date=timezone.now() - timedelta(days=3)
        )
        retrieve_mock.return_value = {
            "payment_status": "paid",
            "customer_details": {
                "name": "Exact Buyer",
                "email": "exact@example.com",
            },
        }

        response = self.client.get(
            reverse("event:success", kwargs={"pk": self.event.pk})
            + "?session_id=sess_exact_entry"
        )

        self.assertEqual(response.status_code, 200)
        entry.refresh_from_db()
        self.assertTrue(entry.payment_complete)
        self.assertEqual(entry.customer_email, "exact@example.com")

    @patch("event.views.payment_helpers.stripe.checkout.Session.retrieve")
    def test_success_credit_view_confirms_exact_session(self, retrieve_mock):
        self.client.force_login(self.user)
        credit_transaction = CreditTransaction.objects.create(
            user=self.user,
            amount=700,
            transaction_id="sess_credit_exact",
        )
        retrieve_mock.return_value = {
            "payment_status": "paid",
            "payment_intent": "pi_credit_exact",
        }

        response = self.client.get(
            reverse("event:success-credit") + "?session_id=sess_credit_exact"
        )

        self.assertRedirects(response, reverse("event:success-credit-update"))
        credit_transaction.refresh_from_db()
        self.user.refresh_from_db()
        self.assertTrue(credit_transaction.payment_complete)
        self.assertEqual(credit_transaction.payment_intent, "pi_credit_exact")
        self.assertEqual(self.user.credit, 700)

    def test_credit_view_shows_credit_transaction_types(self):
        self.client.force_login(self.user)
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )
        CreditTransaction.objects.create(
            user=self.user,
            amount=700,
            payment_complete=True,
            payment_intent="pi_credit_history",
            kind=CreditTransaction.Kind.TOPUP,
        )
        CreditTransaction.objects.create(
            user=self.user,
            source_entry=entry,
            amount=300,
            payment_complete=True,
            payment_intent=f"Vrácení startovného za závod {self.event.name}",
            kind=CreditTransaction.Kind.CHECKOUT_REFUND,
        )

        response = self.client.get(reverse("event:credit"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Dobití kreditu")
        self.assertContains(response, "Vrácení startovného po checkoutu")
        self.assertContains(response, f"Vrácení startovného za závod {self.event.name}")

    def test_entry_checkout_creates_refund_credit_transaction(self):
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )

        entry.checkout = True
        entry.save(update_fields=["checkout"])

        refund = CreditTransaction.objects.get(
            kind=CreditTransaction.Kind.CHECKOUT_REFUND,
            source_entry=entry,
        )
        self.user.refresh_from_db()
        self.assertEqual(refund.amount, 300)
        self.assertTrue(refund.payment_complete)
        self.assertEqual(refund.kind, CreditTransaction.Kind.CHECKOUT_REFUND)
        self.assertEqual(refund.source_entry_id, entry.pk)
        self.assertEqual(
            refund.payment_intent,
            f"Vrácení startovného za závod {self.event.name}",
        )
        self.assertEqual(self.user.credit, 300)

    def test_entry_checkout_uncheck_deletes_refund_credit_transaction(self):
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
            checkout=True,
        )

        entry.checkout = False
        entry.save(update_fields=["checkout"])

        self.user.refresh_from_db()
        self.assertFalse(
            CreditTransaction.objects.filter(
                kind=CreditTransaction.Kind.CHECKOUT_REFUND,
                source_entry=entry,
            ).exists()
        )
        self.assertEqual(self.user.credit, 0)

    def test_entry_checkout_raises_validation_error_for_unpaid_entry(self):
        with self.assertRaises(ValidationError):
            Entry.objects.create(
                user=self.user,
                event=self.event,
                rider=self.rider,
                payment_complete=False,
                is_20=True,
                class_20="Boys 6",
                fee_20=300,
                checkout=True,
            )

    def test_entry_checkout_updates_refund_when_fee_changes(self):
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
            checkout=True,
        )

        entry.fee_20 = 450
        entry.save(update_fields=["fee_20"])

        refund = CreditTransaction.objects.get(
            kind=CreditTransaction.Kind.CHECKOUT_REFUND,
            source_entry=entry,
        )
        self.user.refresh_from_db()
        self.assertEqual(refund.amount, 450)
        self.assertEqual(self.user.credit, 450)

    def test_entry_checkout_reassigns_refund_when_user_changes(self):
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
            checkout=True,
        )

        entry.user = self.other_user
        entry.save(update_fields=["user"])

        refund = CreditTransaction.objects.get(
            kind=CreditTransaction.Kind.CHECKOUT_REFUND,
            source_entry=entry,
        )
        self.user.refresh_from_db()
        self.other_user.refresh_from_db()
        self.assertEqual(refund.user_id, self.other_user.id)
        self.assertEqual(self.user.credit, 0)
        self.assertEqual(self.other_user.credit, 300)

    def test_entry_checkout_updates_refund_when_event_changes(self):
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
            checkout=True,
        )

        entry.event = self.second_event
        entry.save(update_fields=["event"])

        refund = CreditTransaction.objects.get(
            kind=CreditTransaction.Kind.CHECKOUT_REFUND,
            source_entry=entry,
        )
        self.assertEqual(
            refund.payment_intent,
            f"Vrácení startovného za závod {self.second_event.name}",
        )

    def test_apply_entry_checkout_creates_persistent_audit_log(self):
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )

        changed = apply_entry_checkout(
            entry,
            checkout=True,
            actor=self.other_user,
            source="test_case",
            note="test audit note",
        )

        audit = EntryAuditLog.objects.get(entry=entry)
        self.assertTrue(changed)
        self.assertEqual(audit.action, EntryAuditLog.Action.CHECKOUT_CHANGED)
        self.assertEqual(audit.actor_id, self.other_user.id)
        self.assertEqual(audit.source, "test_case")
        self.assertEqual(audit.note, "test audit note")
        self.assertFalse(audit.old_checkout)
        self.assertTrue(audit.new_checkout)

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_entry_checkout_sends_notification_emails(self):
        self.other_user.is_admin = True
        self.other_user.save(update_fields=["is_admin"])
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )

        entry.checkout = True
        entry.save(update_fields=["checkout"])

        self.assertEqual(len(mail.outbox), 1)
        self.assertIn("Checkout refund vytvořen", mail.outbox[0].subject)
        self.assertIn(self.user.email, mail.outbox[0].to)
        self.assertIn(self.other_user.email, mail.outbox[0].to)

    def test_entry_checkout_logs_refund_amount_change(self):
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
            checkout=True,
        )

        entry.fee_20 = 450
        entry.save(update_fields=["fee_20"])

        self.assertTrue(
            EntryAuditLog.objects.filter(
                entry=entry,
                action=EntryAuditLog.Action.REFUND_CONTEXT_CHANGED,
                note="Změna částky refundu: 300 -> 450 Kč",
            ).exists()
        )

    def test_entry_checkout_logs_refund_user_change(self):
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
            checkout=True,
        )

        entry.user = self.other_user
        entry.save(update_fields=["user"])

        self.assertTrue(
            EntryAuditLog.objects.filter(
                entry=entry,
                action=EntryAuditLog.Action.REFUND_CONTEXT_CHANGED,
                note=f"Změna uživatele refundu: {self.user.id} -> {self.other_user.id}",
            ).exists()
        )

    def test_entry_checkout_logs_refund_event_change(self):
        entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
            checkout=True,
        )

        entry.event = self.second_event
        entry.save(update_fields=["event"])

        self.assertTrue(
            EntryAuditLog.objects.filter(
                entry=entry,
                action=EntryAuditLog.Action.REFUND_CONTEXT_CHANGED,
                note=f"Změna závodu refundu: {self.event.id} -> {self.second_event.id}",
            ).exists()
        )

    def test_cancel_view_redirects_entry_checkout_back_to_confirm(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["event"] = '{"event": %d}' % self.event.id
        session["riders_beginner"] = "[]"
        session["riders_20"] = "[]"
        session["riders_24"] = "[]"
        session.save()

        response = self.client.get(reverse("event:cancel") + "?source=entries")

        self.assertRedirects(response, reverse("event:confirm"), fetch_redirect_response=False)

    def test_checkout_allows_unregistration_after_registration_deadline_until_cancel_deadline(self):
        self.client.force_login(self.user)
        self.event.reg_open_to = timezone.now() - timedelta(hours=1)
        self.event.reg_cancel_to = timezone.now() + timedelta(days=2)
        self.event.save(update_fields=["reg_open_to", "reg_cancel_to"])
        Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )

        response = self.client.get(reverse("event:checkout"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Odhlásit")

    def test_checkout_shows_closed_unregistration_state_after_cancel_deadline(self):
        self.client.force_login(self.user)
        self.event.reg_cancel_to = timezone.now() - timedelta(hours=1)
        self.event.save(update_fields=["reg_cancel_to"])
        Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            payment_complete=True,
            is_20=True,
            class_20="Boys 6",
            fee_20=300,
        )

        response = self.client.get(reverse("event:checkout"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Odhlášení uzavřeno")
        self.assertContains(response, self.rider.last_name)

    def test_entry_page_closes_after_registration_deadline_even_if_cancel_deadline_is_later(self):
        self.client.force_login(self.user)
        self.event.reg_open_to = timezone.now() - timedelta(hours=1)
        self.event.reg_cancel_to = timezone.now() + timedelta(days=2)
        self.event.save(update_fields=["reg_open_to", "reg_cancel_to"])

        response = self.client.get(reverse("event:entry", kwargs={"pk": self.event.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "event/reg-close.html")

    def test_confirm_view_renders_checkout_summary_from_session(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["event"] = '{"event": %d}' % self.event.id
        session["riders_beginner"] = "[]"
        session["riders_20"] = '[{"model":"rider.rider","pk":1,"fields":{"uci_id":12345678999}}]'
        session["riders_24"] = "[]"
        session.save()

        response = self.client.get(reverse("event:confirm"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.event.name)
        self.assertContains(response, "Pokračovat k platbě")

    def test_entry_list_template_uses_data_hooks_instead_of_inline_handlers(self):
        template = (
            Path(settings.BASE_DIR) / "event" / "templates" / "event" / "entry-list.html"
        ).read_text(encoding="utf-8")

        self.assertIn("js/entries.js", template)
        self.assertNotIn("onclick=", template)
        self.assertNotIn("onkeyup=", template)
        self.assertIn("data-detail-url", template)

    @patch("event.views.views_entry.create_entry_checkout_session")
    def test_confirm_view_uses_checkout_session_service(self, create_checkout_session_mock):
        self.client.force_login(self.user)
        create_checkout_session_mock.return_value = SimpleNamespace(id="cs_test_entry")
        session = self.client.session
        session["event"] = '{"event": %d}' % self.event.id
        session["riders_beginner"] = "[]"
        session["riders_20"] = '[{"model":"rider.rider","pk":1,"fields":{"uci_id":12345678999}}]'
        session["riders_24"] = "[]"
        session.save()

        response = self.client.post(reverse("event:confirm"))

        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(response.content, {"id": "cs_test_entry"})
        create_checkout_session_mock.assert_called_once()

    @patch("event.views.views_entry.create_foreign_entry_checkout_session")
    def test_foreign_pay_view_uses_checkout_session_service(self, create_checkout_session_mock):
        create_checkout_session_mock.return_value = SimpleNamespace(
            id="cs_test_foreign",
            url="https://payments.example.test/checkout",
        )
        session = self.client.session
        session["foreign_summary_payload"] = json.dumps(
            {
                "customer_email": "foreign@example.com",
                "rows": [
                    {
                        "uci_id": "12345678901",
                        "first_name": "Foreign",
                        "last_name": "Rider",
                        "date_of_birth": "2010-01-01",
                        "sex": "Muž",
                        "plate": "11",
                        "nationality": "CZE",
                        "transponder_20": "123",
                        "transponder_24": "",
                        "challenge": True,
                        "championship": False,
                        "cruiser": False,
                    }
                ],
            }
        )
        session.save()

        response = self.client.post(
            reverse("event:entry-foreign-pay", kwargs={"pk": self.event.pk})
        )

        self.assertEqual(response.status_code, 303)
        self.assertEqual(response["Location"], "https://payments.example.test/checkout")
        create_checkout_session_mock.assert_called_once()

    def test_event_admin_templates_use_data_confirm_message(self):
        fees_template = (
            Path(settings.BASE_DIR) / "event" / "templates" / "event" / "fees-on-event.html"
        ).read_text(encoding="utf-8")
        receipts_template = (
            Path(settings.BASE_DIR) / "event" / "templates" / "event" / "cash-receipts-on-event.html"
        ).read_text(encoding="utf-8")

        self.assertIn("data-confirm-message", fees_template)
        self.assertNotIn("onclick=", fees_template)
        self.assertIn("data-confirm-message", receipts_template)
        self.assertNotIn("onclick=", receipts_template)

    def test_import_stats_template_uses_external_script_and_no_inline_handlers(self):
        template = (
            Path(settings.BASE_DIR) / "event" / "templates" / "event" / "import-stats.html"
        ).read_text(encoding="utf-8")

        self.assertIn("js/import_stats.js", template)
        self.assertIn("data-confirm-message", template)
        self.assertNotIn("onclick=", template)
        self.assertNotIn("<script>", template)

    def test_event_admin_ec_template_uses_external_script_and_no_inline_script(self):
        template = (
            Path(settings.BASE_DIR) / "event" / "templates" / "event" / "event-admin-ec.html"
        ).read_text(encoding="utf-8")

        self.assertIn("js/event_admin_ec.js", template)
        self.assertNotIn("<script>", template)
        self.assertIn('id="ec-results-add-input"', template)
        self.assertIn('id="ec-results-inputs"', template)

    def test_event_admin_template_uses_external_script_and_confirm_hooks(self):
        template = (
            Path(settings.BASE_DIR) / "event" / "templates" / "event" / "event-admin.html"
        ).read_text(encoding="utf-8")

        self.assertIn("js/event_admin.js", template)
        self.assertNotIn("<script>", template)
        self.assertIn("data-event-admin-form", template)
        self.assertIn("data-file-input", template)
        self.assertIn("data-file-selection", template)
        self.assertIn("data-confirm-message", template)

    def test_foreign_entry_page_uses_external_script_and_no_inline_handlers(self):
        response = self.client.get(
            reverse("event:entry-foreign", kwargs={"pk": self.event.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "js/entry_foreign.js")
        self.assertNotContains(response, "onsubmit=")
        self.assertNotContains(response, "onclick=")
        self.assertNotContains(response, "onchange=")
        self.assertNotContains(response, "onblur=")
        self.assertNotContains(response, "oninput=")
        self.assertContains(response, "data-add-rider")
        self.assertContains(response, "data-fetch-rider")

    def test_cancel_view_redirects_credit_checkout_back_to_credit(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("event:cancel") + "?source=credit")

        self.assertRedirects(response, reverse("event:credit"), fetch_redirect_response=False)

    def test_cancel_view_redirects_foreign_checkout_back_to_summary(self):
        session = self.client.session
        session["foreign_summary_payload"] = '{"customer_email":"foreign@example.com","rows":[{"uci_id":"12345678901","first_name":"Foreign","last_name":"Rider","date_of_birth":"2010-01-01","sex":"Muž","plate":"11","nationality":"CZE","transponder_20":"123","transponder_24":"","challenge":true,"championship":false,"cruiser":false}]}'
        session.save()

        response = self.client.get(
            reverse("event:cancel") + f"?source=foreign&event_id={self.event.id}"
        )

        self.assertRedirects(
            response,
            reverse("event:entry-foreign-summary", kwargs={"pk": self.event.id}),
            fetch_redirect_response=False,
        )

    def test_cancel_view_foreign_follow_redirect_keeps_summary_payload(self):
        session = self.client.session
        session["foreign_summary_payload"] = '{"customer_email":"foreign@example.com","rows":[{"uci_id":"12345678901","first_name":"Foreign","last_name":"Rider","date_of_birth":"2010-01-01","sex":"Muž","plate":"11","nationality":"CZE","transponder_20":"123","transponder_24":"","challenge":true,"championship":false,"cruiser":false}]}'
        session.save()

        response = self.client.get(
            reverse("event:cancel") + f"?source=foreign&event_id={self.event.id}",
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "foreign@example.com")
        self.assertContains(response, "Foreign")


class ForeignEntryHelperTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(team_name="Foreign Club")
        self.entry_classes = EntryClasses.objects.create(
            event_name="Foreign classes",
            boys_6="Boys 6",
            cr_boys_12_and_under="Cruiser 12 and under",
        )
        self.event = Event.objects.create(
            name="Foreign race",
            date=date.today() + timedelta(days=20),
            organizer=self.club,
            classes_and_fees_like=self.entry_classes,
            reg_open=True,
            reg_open_from=timezone.now() - timedelta(days=1),
            reg_open_to=timezone.now() + timedelta(days=1),
            type_for_ranking="Volný závod",
        )

    def test_validate_foreign_summary_payload_requires_date_of_birth(self):
        payload = {
            "customer_email": "foreign@example.com",
            "rows": [
                {
                    "first_name": "Anna",
                    "last_name": "Smith",
                    "uci_id": "12345678901",
                    "date_of_birth": "",
                    "plate": "12",
                    "category_20": True,
                    "category_24": False,
                }
            ],
        }

        self.assertFalse(validate_foreign_summary_payload(payload))

    def test_validate_foreign_summary_payload_rejects_championship_with_cruiser(self):
        payload = {
            "customer_email": "foreign@example.com",
            "rows": [
                {
                    "first_name": "Anna",
                    "last_name": "Smith",
                    "uci_id": "12345678901",
                    "date_of_birth": "2010-01-01",
                    "plate": "12",
                    "championship": True,
                    "cruiser": True,
                }
            ],
        }

        self.assertFalse(validate_foreign_summary_payload(payload))

    def test_enrich_foreign_summary_rows_supports_challenge_and_cruiser(self):
        summary_rows, total_fee = enrich_foreign_summary_rows(
            self.event,
            [
                {
                    "first_name": "Anna",
                    "last_name": "Smith",
                    "uci_id": "12345678901",
                    "date_of_birth": "2020-01-01",
                    "sex": "Muž",
                    "plate": "12",
                    "challenge": True,
                    "cruiser": True,
                }
            ],
        )

        self.assertEqual(len(summary_rows), 1)
        self.assertEqual(summary_rows[0]["selected_categories"], ["Challenge", "Cruiser"])
        self.assertEqual(summary_rows[0]["class_20"], "Boys 6")
        self.assertEqual(summary_rows[0]["class_24"], "Cruiser 12 and under")
        self.assertEqual(summary_rows[0]["fee_20"], 0)
        self.assertEqual(summary_rows[0]["fee_24"], 0)
        self.assertEqual(total_fee, 0)

    def test_build_public_entry_rows_marks_foreign_entries(self):
        foreign_entry = EntryForeign.objects.create(
            event=self.event,
            transaction_id="sess_123",
            first_name="Péter",
            last_name="Balogh",
            uci_id="10115844151",
            gender="Muž",
            nationality="HUN",
            plate="15",
            is_20=True,
            class_20="Elite 17+",
            fee_20=400,
            payment_complete=True,
        )

        rows = build_public_entry_rows([foreign_entry], is_foreign=True)

        self.assertEqual(len(rows), 1)
        self.assertTrue(rows[0].is_foreign)
        self.assertEqual(rows[0].club, "HUN")
        self.assertEqual(rows[0].category, "Elite 17+")

    def test_build_public_entry_rows_returns_both_categories_for_single_entry(self):
        foreign_entry = EntryForeign.objects.create(
            event=self.event,
            transaction_id="sess_multi",
            first_name="David",
            last_name="Průša",
            uci_id="10047037910",
            gender="Muž",
            nationality="CZE",
            plate="0",
            is_20=True,
            is_24=True,
            class_20="Master 30+",
            class_24="Cruiser",
            fee_20=400,
            fee_24=400,
            payment_complete=True,
        )

        rows = build_public_entry_rows([foreign_entry], is_foreign=True)

        self.assertEqual(len(rows), 2)
        self.assertEqual({row.category for row in rows}, {"Master 30+", "Cruiser"})

    def test_entry_foreign_save_normalizes_uci_id(self):
        entry = EntryForeign.objects.create(
            event=self.event,
            transaction_id="sess_normalized_uci",
            first_name="Péter",
            last_name="Balogh",
            uci_id="100 459 968",
            gender="Muž",
            nationality="HUN",
            plate="15",
            is_20=True,
            class_20="Elite 17+",
            fee_20=400,
            payment_complete=True,
        )

        self.assertEqual(entry.uci_id, "100459968")

    def test_sync_paid_foreign_riders_creates_missing_foreign_rider(self):
        EntryForeign.objects.create(
            event=self.event,
            transaction_id="sess_sync",
            first_name="Péter",
            last_name="Balogh",
            uci_id="10115844151",
            date_of_birth=date(1980, 7, 8),
            gender="Muž",
            nationality="HUN",
            plate="15",
            transponder_20="TR-20",
            transponder_24="TR-24",
            is_24=True,
            class_24="Cruiser",
            fee_24=400,
            payment_complete=True,
        )

        sync_paid_foreign_riders(self.event, "sess_sync")

        foreign_rider = ForeignRider.objects.get(uci_id=10115844151)
        self.assertEqual(foreign_rider.first_name, "Péter")
        self.assertEqual(foreign_rider.state, "HUN")
        self.assertEqual(foreign_rider.class_24, "Cruiser")

    def test_sync_paid_foreign_riders_handles_formatted_uci_id(self):
        EntryForeign.objects.create(
            event=self.event,
            transaction_id="sess_sync_spaced",
            first_name="Péter",
            last_name="Balogh",
            uci_id="100 459 968",
            date_of_birth=date(1980, 7, 8),
            gender="Muž",
            nationality="HUN",
            plate="15",
            transponder_20="TR-20",
            is_20=True,
            class_20="Boys 6",
            fee_20=400,
            payment_complete=True,
        )

        sync_paid_foreign_riders(self.event, "sess_sync_spaced")

        foreign_rider = ForeignRider.objects.get(uci_id=100459968)
        self.assertEqual(foreign_rider.first_name, "Péter")

    @patch("event.views.payment_helpers._construct_stripe_event")
    def test_webhook_marks_foreign_entries_paid_and_syncs_foreign_rider(self, construct_event_mock):
        foreign_entry = EntryForeign.objects.create(
            event=self.event,
            transaction_id="sess_foreign_webhook",
            first_name="Péter",
            last_name="Balogh",
            date_of_birth=date(1980, 7, 8),
            uci_id="10115844151",
            gender="Muž",
            nationality="HUN",
            club="",
            transponder="TR-20",
            is_20=True,
            class_20="Elite 17+",
            fee_20=400,
            payment_complete=False,
        )
        construct_event_mock.return_value = {
            "type": "checkout.session.completed",
            "data": {
                "object": {
                    "id": "sess_foreign_webhook",
                    "payment_status": "paid",
                    "payment_intent": "pi_foreign",
                    "customer_details": {
                        "name": "Foreign Buyer",
                        "email": "foreign@example.com",
                    },
                }
            },
        }

        response = self.client.post(
            reverse("event:stripe-credit-webhook"),
            data="{}",
            content_type="application/json",
            HTTP_STRIPE_SIGNATURE="sig_test",
        )

        self.assertEqual(response.status_code, 200)
        foreign_entry.refresh_from_db()
        self.assertTrue(foreign_entry.payment_complete)
        self.assertEqual(foreign_entry.customer_email, "foreign@example.com")
        foreign_rider = ForeignRider.objects.get(uci_id=10115844151)
        self.assertEqual(foreign_rider.first_name, "Péter")
        self.assertEqual(foreign_rider.state, "HUN")

    @patch("event.views.payment_helpers.stripe.checkout.Session.retrieve")
    def test_foreign_success_view_syncs_rider_even_when_entry_was_already_marked_paid(self, retrieve_mock):
        EntryForeign.objects.create(
            event=self.event,
            transaction_id="sess_foreign_success",
            first_name="Nina",
            last_name="Meyer",
            date_of_birth=date(2005, 5, 4),
            uci_id="10123456789",
            gender="Žena",
            nationality="GER",
            plate="G7",
            transponder_20="TX20",
            is_20=True,
            class_20="Boys 6",
            fee_20=400,
            payment_complete=True,
        )
        retrieve_mock.return_value = {
            "payment_status": "paid",
            "customer_details": {
                "name": "Foreign Buyer",
                "email": "foreign@example.com",
            },
        }

        response = self.client.get(
            reverse("event:entry-foreign-success", kwargs={"pk": self.event.pk})
            + "?session_id=sess_foreign_success"
        )

        self.assertEqual(response.status_code, 200)
        foreign_rider = ForeignRider.objects.get(uci_id=10123456789)
        self.assertEqual(foreign_rider.first_name, "Nina")
        self.assertEqual(foreign_rider.state, "GER")


class BalanceRecalculationViewTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user(
            first_name="Staff",
            last_name="Tester",
            username="staff_tester",
            email="staff_tester@example.com",
            password="StrongPass123!",
        )
        self.staff_user.is_active = True
        self.staff_user.is_staff = True
        self.staff_user.save()

    @patch("event.views.views_payment.recalculate_all_balances")
    def test_recalculate_balances_view_renders_status_page(self, recalculate_mock):
        self.client.force_login(self.staff_user)

        response = self.client.get(reverse("event:recalculate_balances_view"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Kontrola kreditu")
        self.assertContains(response, "Stavy kreditů byly úspěšně překontrolovány.")
        recalculate_mock.assert_called_once()


class EventAdminViewTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user(
            first_name="Admin",
            last_name="Tester",
            username="admin_tester",
            email="admin_tester@example.com",
            password="StrongPass123!",
        )
        self.staff_user.is_active = True
        self.staff_user.is_staff = True
        self.staff_user.save()

        self.club = Club.objects.create(team_name="Admin Club")
        self.event = Event.objects.create(
            name="Admin Race",
            date=date.today() + timedelta(days=7),
            organizer=self.club,
            reg_open=False,
            type_for_ranking="Volný závod",
        )
        self.rider = Rider.objects.create(
            uci_id=10000000010,
            first_name="Czech",
            last_name="Rider",
            gender="Muž",
            date_of_birth=date(2010, 1, 1),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            class_20="Boys 15",
        )

    def test_event_admin_counts_czech_and_foreign_riders_in_kpi(self):
        Entry.objects.create(
            event=self.event,
            rider=self.rider,
            is_20=True,
            class_20="Boys 15",
            fee_20=400,
            payment_complete=True,
        )
        EntryForeign.objects.create(
            event=self.event,
            transaction_id="sess_foreign_admin",
            first_name="Foreign",
            last_name="Rider",
            uci_id="10115844151",
            gender="Muž",
            nationality="HUN",
            club="Foreign Club",
            transponder="TR-20",
            is_20=True,
            class_20="Boys 15",
            fee_20=400,
            payment_complete=True,
        )

        self.client.force_login(self.staff_user)

        response = self.client.get(reverse("event:event-admin", kwargs={"pk": self.event.id}))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["sum_of_riders"], 2)
        self.assertEqual(response.context["sum_of_fees"], 800)
        self.assertEqual(response.context["asociation_fee"], 40)
        self.assertContains(response, ">2</p>", html=False)
        self.assertContains(response, "800 CZK")
        self.assertContains(response, "40 CZK")

    def test_event_admin_renders_when_legacy_event_has_missing_organizer_and_commission_fee(self):
        self.client.force_login(self.staff_user)
        self.event.organizer = None
        self.event.commission_fee = None

        with patch("event.views.views_admin.get_object_or_404", return_value=self.event):
            response = self.client.get(reverse("event:event-admin", kwargs={"pk": self.event.id}))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["asociation_fee"], 0)
        self.assertContains(response, "Volný závod")


class EntryForeignAdminTests(TestCase):
    def test_foreign_rider_link_handles_non_numeric_uci_id(self):
        entry = EntryForeign.objects.create(
            transaction_id="sess_invalid_uci",
            first_name="Foreign",
            last_name="Broken",
            uci_id="TEMP-UNKNOWN",
            gender="Muž",
            nationality="GER",
            club="Foreign Club",
            transponder="TR-20",
            is_20=True,
            class_20="Boys 15",
            fee_20=400,
            payment_complete=True,
        )

        admin_instance = EntryForeignAdmin(EntryForeign, admin.site)

        self.assertEqual(entry.uci_id, "")
        self.assertEqual(admin_instance.foreign_rider_link(entry), "Bez UCI ID")

    def test_changelist_renders_with_invalid_uci_id(self):
        staff_user = User.objects.create_user(
            first_name="Admin",
            last_name="Foreign",
            username="foreign_admin",
            email="foreign_admin@example.com",
            password="StrongPass123!",
        )
        staff_user.is_staff = True
        staff_user.is_superuser = True
        staff_user.save(update_fields=["is_staff", "is_superuser"])
        EntryForeign.objects.create(
            transaction_id="sess_invalid_uci_page",
            first_name="Foreign",
            last_name="Broken",
            uci_id="TEMP-UNKNOWN",
            gender="Muž",
            nationality="GER",
            club="Foreign Club",
            transponder="TR-24",
            is_24=True,
            class_24="Men 17-24",
            fee_24=400,
            payment_complete=True,
        )

        request = RequestFactory().get("/bmx-admin/event/entryforeign/")
        request.user = staff_user

        response = EntryForeignAdmin(EntryForeign, admin.site).changelist_view(request)
        response.render()

        self.assertEqual(response.status_code, 200)


class EntryAdminActionTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user(
            first_name="Admin",
            last_name="Checkout",
            username="checkout_admin",
            email="checkout_admin@example.com",
            password="StrongPass123!",
        )
        self.staff_user.is_active = True
        self.staff_user.is_staff = True
        self.staff_user.is_superuser = True
        self.staff_user.save(update_fields=["is_active", "is_staff", "is_superuser"])

        self.club = Club.objects.create(team_name="Checkout Admin Club")
        self.event = Event.objects.create(
            name="Checkout Admin Race",
            date=date.today() + timedelta(days=10),
            organizer=self.club,
            reg_open=False,
            type_for_ranking="Volný závod",
        )
        self.rider = Rider.objects.create(
            uci_id=10000012345,
            first_name="Admin",
            last_name="Rider",
            gender="Muž",
            date_of_birth=date(2010, 1, 1),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            class_20="Boys 15",
        )
        self.entry = Entry.objects.create(
            user=self.staff_user,
            event=self.event,
            rider=self.rider,
            is_20=True,
            class_20="Boys 15",
            fee_20=400,
            payment_complete=True,
        )

    def test_admin_checkout_action_shows_confirmation_page(self):
        self.client.force_login(self.staff_user)

        response = self.client.post(
            reverse("admin:event_entry_changelist"),
            {
                "action": "mark_checkout_with_refund",
                ACTION_CHECKBOX_NAME: [str(self.entry.pk)],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Potvrzení vrácení startovného")
        self.assertContains(response, self.rider.last_name)

    def test_admin_entry_change_page_loads_for_existing_entry(self):
        self.client.force_login(self.staff_user)

        response = self.client.get(reverse("admin:event_entry_change", args=[self.entry.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "transaction_date")

    def test_admin_checkout_action_confirmation_applies_refund(self):
        self.client.force_login(self.staff_user)

        response = self.client.post(
            reverse("admin:event_entry_changelist"),
            {
                "action": "mark_checkout_with_refund",
                "apply": "yes",
                ACTION_CHECKBOX_NAME: [str(self.entry.pk)],
            },
            follow=True,
        )

        self.entry.refresh_from_db()
        refund = CreditTransaction.objects.get(
            kind=CreditTransaction.Kind.CHECKOUT_REFUND,
            source_entry=self.entry,
        )
        audit = EntryAuditLog.objects.filter(entry=self.entry).latest("created_at")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(self.entry.checkout)
        self.assertEqual(refund.amount, 400)
        self.assertEqual(audit.source, "admin_action")
        self.assertEqual(audit.actor_id, self.staff_user.id)

    def test_admin_unmark_checkout_action_confirmation_removes_refund(self):
        self.client.force_login(self.staff_user)
        self.entry.checkout = True
        self.entry.save(update_fields=["checkout"])

        response = self.client.post(
            reverse("admin:event_entry_changelist"),
            {
                "action": "unmark_checkout_and_remove_refund",
                "apply": "yes",
                ACTION_CHECKBOX_NAME: [str(self.entry.pk)],
            },
            follow=True,
        )

        self.entry.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertFalse(self.entry.checkout)
        self.assertFalse(
            CreditTransaction.objects.filter(
                kind=CreditTransaction.Kind.CHECKOUT_REFUND,
                source_entry=self.entry,
            ).exists()
        )

    def test_admin_checkout_action_skips_invalid_entries_with_warning(self):
        self.client.force_login(self.staff_user)
        unpaid_entry = Entry.objects.create(
            user=self.staff_user,
            event=self.event,
            rider=self.rider,
            is_24=True,
            class_24="Cruiser",
            fee_24=250,
            payment_complete=False,
        )

        response = self.client.post(
            reverse("admin:event_entry_changelist"),
            {
                "action": "mark_checkout_with_refund",
                "apply": "yes",
                ACTION_CHECKBOX_NAME: [str(unpaid_entry.pk)],
            },
            follow=True,
        )

        unpaid_entry.refresh_from_db()
        messages = [message.message for message in get_messages(response.wsgi_request)]

        self.assertEqual(response.status_code, 200)
        self.assertFalse(unpaid_entry.checkout)
        self.assertIn("Checkout byl zapnut u 0 registrací.", messages)
        self.assertIn("1 registrací bylo přeskočeno kvůli nevalidnímu stavu.", messages)

    def test_entry_admin_detail_shows_refund_summary(self):
        self.entry.checkout = True
        self.entry.save(update_fields=["checkout"])
        admin_instance = admin.site._registry[Entry]

        summary = admin_instance.checkout_refund_summary(self.entry)

        self.assertIn("Částka:", str(summary))
        self.assertIn("400", str(summary))
        self.assertIn("Vrácení startovného za závod", str(summary))

    def test_entry_admin_detail_shows_checkout_audit_timeline(self):
        apply_entry_checkout(
            self.entry,
            checkout=True,
            actor=self.staff_user,
            source="admin_action",
            note="bulk action test",
        )
        admin_instance = admin.site._registry[Entry]

        timeline = admin_instance.checkout_audit_timeline(self.entry)

        self.assertIn("Změna checkout", str(timeline))
        self.assertIn("bulk action test", str(timeline))


class EventAdminAuditPanelTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user(
            first_name="Event",
            last_name="Admin",
            username="event_audit_admin",
            email="event_audit_admin@example.com",
            password="StrongPass123!",
        )
        self.staff_user.is_active = True
        self.staff_user.is_staff = True
        self.staff_user.is_superuser = True
        self.staff_user.save(update_fields=["is_active", "is_staff", "is_superuser"])

        self.club = Club.objects.create(team_name="Event Audit Club")
        self.event = Event.objects.create(
            name="Event Audit Race",
            date=date.today() + timedelta(days=10),
            organizer=self.club,
            reg_open=False,
            type_for_ranking="Volný závod",
        )
        self.rider = Rider.objects.create(
            uci_id=10000112345,
            first_name="Audit",
            last_name="Rider",
            gender="Muž",
            date_of_birth=date(2010, 1, 1),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            class_20="Boys 15",
        )
        self.entry = Entry.objects.create(
            user=self.staff_user,
            event=self.event,
            rider=self.rider,
            is_20=True,
            class_20="Boys 15",
            fee_20=400,
            payment_complete=True,
        )

    def test_event_admin_refund_summary_counts_event_refunds(self):
        apply_entry_checkout(
            self.entry,
            checkout=True,
            actor=self.staff_user,
            source="admin_action",
            note="event audit",
        )
        admin_instance = admin.site._registry[Event]

        summary = admin_instance.event_refund_summary(self.event)

        self.assertIn("Počet refundů:", str(summary))
        self.assertIn("1", str(summary))
        self.assertIn("400", str(summary))

    def test_event_admin_timeline_shows_event_checkout_audit(self):
        apply_entry_checkout(
            self.entry,
            checkout=True,
            actor=self.staff_user,
            source="admin_action",
            note="event audit timeline",
        )
        admin_instance = admin.site._registry[Event]

        timeline = admin_instance.event_checkout_audit_timeline(self.event)

        self.assertIn("Audit Rider", str(timeline))
        self.assertIn("Změna checkout", str(timeline))
        self.assertIn("event audit timeline", str(timeline))


class FinanceAdminAuditTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user(
            first_name="Finance",
            last_name="Auditor",
            username="finance_auditor",
            email="finance_auditor@example.com",
            password="StrongPass123!",
        )
        self.staff_user.is_active = True
        self.staff_user.is_staff = True
        self.staff_user.is_superuser = True
        self.staff_user.save(update_fields=["is_active", "is_staff", "is_superuser"])

        self.user = User.objects.create_user(
            first_name="Target",
            last_name="User",
            username="finance_target",
            email="finance_target@example.com",
            password="StrongPass123!",
        )
        self.user.is_active = True
        self.user.save(update_fields=["is_active"])

        self.club = Club.objects.create(team_name="Finance Audit Club")
        self.event = Event.objects.create(
            name="Finance Audit Race",
            date=date.today() + timedelta(days=10),
            organizer=self.club,
            reg_open=False,
            type_for_ranking="Volný závod",
        )
        self.rider = Rider.objects.create(
            uci_id=10000098765,
            first_name="Finance",
            last_name="Rider",
            gender="Muž",
            date_of_birth=date(2010, 1, 1),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            class_20="Boys 15",
        )
        self.entry = Entry.objects.create(
            user=self.user,
            event=self.event,
            rider=self.rider,
            is_20=True,
            class_20="Boys 15",
            fee_20=400,
            payment_complete=True,
        )

    def test_credit_transaction_admin_writes_persistent_audit_log(self):
        request = RequestFactory().post("/bmx-admin/event/credittransaction/")
        request.user = self.staff_user
        admin_instance = CreditTransactionAdmin(CreditTransaction, admin.site)

        credit = CreditTransaction(
            user=self.user,
            amount=500,
            payment_intent="manual topup",
            kind=CreditTransaction.Kind.TOPUP,
            payment_complete=True,
        )
        admin_instance.save_model(
            request,
            credit,
            SimpleNamespace(changed_data=["amount", "payment_complete"]),
            change=False,
        )

        audit = FinanceAuditLog.objects.get(target_model="CreditTransaction", target_object_id=credit.pk)
        self.assertEqual(audit.action, FinanceAuditLog.Action.CREATED)
        self.assertEqual(audit.actor_id, self.staff_user.id)
        self.assertEqual(audit.target_user_id_snapshot, self.user.id)
        self.assertEqual(audit.amount_snapshot, 500)
        self.assertEqual(audit.transaction_kind_snapshot, CreditTransaction.Kind.TOPUP)

    def test_credit_transaction_admin_marks_checkout_refund_as_readonly(self):
        refund = CreditTransaction.objects.create(
            user=self.user,
            source_entry=self.entry,
            amount=400,
            payment_complete=True,
            payment_intent=f"Vrácení startovného za závod {self.event.name}",
            kind=CreditTransaction.Kind.CHECKOUT_REFUND,
        )
        admin_instance = CreditTransactionAdmin(CreditTransaction, admin.site)

        readonly_fields = admin_instance.get_readonly_fields(None, refund)

        self.assertIn("amount", readonly_fields)
        self.assertIn("payment_intent", readonly_fields)
        self.assertIn("source_entry", readonly_fields)

    def test_credit_transaction_admin_disallows_deleting_checkout_refund(self):
        refund = CreditTransaction.objects.create(
            user=self.user,
            source_entry=self.entry,
            amount=400,
            payment_complete=True,
            payment_intent=f"Vrácení startovného za závod {self.event.name}",
            kind=CreditTransaction.Kind.CHECKOUT_REFUND,
        )
        request = RequestFactory().post("/bmx-admin/event/credittransaction/")
        request.user = self.staff_user
        admin_instance = CreditTransactionAdmin(CreditTransaction, admin.site)

        self.assertFalse(admin_instance.has_delete_permission(request, refund))

    def test_debet_transaction_admin_delete_writes_persistent_audit_log(self):
        request = RequestFactory().post("/bmx-admin/event/debettransaction/")
        request.user = self.staff_user
        admin_instance = DebetTransactionAdmin(DebetTransaction, admin.site)

        debet = DebetTransaction.objects.create(
            user=self.user,
            entry=self.entry,
            amount=400,
            payment_valid=True,
        )

        admin_instance.delete_model(request, debet)

        audit = FinanceAuditLog.objects.get(target_model="DebetTransaction", target_object_id=debet.pk)
        self.assertEqual(audit.action, FinanceAuditLog.Action.DELETED)
        self.assertEqual(audit.actor_id, self.staff_user.id)
        self.assertEqual(audit.amount_snapshot, 400)
        self.assertEqual(audit.payment_valid_snapshot, True)


class RemResultsImportTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(team_name="Import Club")
        self.event = Event.objects.create(
            name="REM Import Race",
            date=date(2025, 4, 13),
            organizer=self.club,
            reg_open=False,
            type_for_ranking="Český pohár",
        )
        self.rider = Rider.objects.create(
            uci_id=10125224253,
            first_name="Simon",
            last_name="Aksamit",
            gender="Muž",
            date_of_birth=date(2009, 8, 2),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            class_20="Boys 15",
            class_24="Boys 15 and 16",
        )

    def test_import_file_creates_only_results(self):
        rem_tsv = "\t".join(
            [
                "EVENT_NAME", "FIRST_NAME", "LAST_NAME", "CLUB", "CLASS", "REGISTRATION_CLASS", "TEAM", "SEX",
                "BIRTHDATE", "MAIL", "TYPE", "LICENCE_TYPE", "UCIID", "UCIID_EXP_DATE", "INTERNAL_IDENT", "STATUS",
                "PLATE", "TRANSPONDER", "ENTRY_PAYED", "PAYMENT_AMOUNT", "ADMIN_FEE", "TRANSPONDER_HIRE_PRICE",
                "TRANSPONDER_HIRE_FLAG", "PAYMENT_CURR", "PAYMENT_TYPE", "CLASS_RANKING", "MOTO1_GATE", "MOTO1_LANE",
                "MOTO1_PLACE", "MOTO1_RACE_POINTS", "MOTO1_MOTO_POINTS", "MOTO1_HILL_TIME", "MOTO1_INTER2_TIME", "MOTO1_TIME", "MOTO2_GATE", "MOTO2_LANE",
                "MOTO2_PLACE", "MOTO2_RACE_POINTS", "MOTO2_MOTO_POINTS", "MOTO2_TIME", "FINAL_SUBTYPE", "FINAL_GATE",
                "FINAL_LANE", "FINAL_PLACE", "FINAL_RACE_POINTS", "FINAL_MOTO_POINTS", "FINAL_INTER2_TIME", "FINAL_TIME", "F2_GATE",
                "F2_LANE", "F2_PLACE", "F2_RACE_POINTS", "F2_MOTO_POINTS", "F2_HILL_TIME", "F2_INTER2_TIME", "F2_TIME",
            ]
        )
        rem_row = "\t".join(
            [
                "2. zavod SAZKA Ceskeho poharu", "Simon", "Aksamit", "Import Club", "Boys 15-16", "Boys 15-16", "",
                "M", "02-08-2009", "simon@example.com", "C", "U", "10125224253", "31-12-2025", "", "S", "868",
                "HW-43726", "X", "500.00", "0.00", "0.00", "false", "CZK", "P", "7", "40", "1", "1st", "8", "1", "1.921",
                "12.441", "29.224", "83", "7", "2nd", "7", "2", "29.753", "A", "11", "3", "7th", "10", "7", "12.883",
                "30.798", "18", "4", "2nd", "0", "2", "1.877", "12.517", "29.446",
            ]
        )

        with tempfile.NamedTemporaryFile("w", suffix=".txt", delete=False, encoding="utf-8") as handle:
            handle.write(rem_tsv)
            handle.write("\n")
            handle.write(rem_row)
            file_path = handle.name

        try:
            SetResults.import_file(self.event.id, file_path)
        finally:
            os.unlink(file_path)

        result = Result.objects.get(event=self.event, rider=self.rider)
        self.assertEqual(result.place, 7)
        self.assertEqual(result.points, 75)
        self.assertFalse(RaceRun.objects.filter(result=result).exists())


class UciExportTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user(
            first_name="UCI",
            last_name="Staff",
            username="uci_staff",
            email="uci_staff@example.com",
            password="StrongPass123!",
        )
        self.staff_user.is_active = True
        self.staff_user.is_staff = True
        self.staff_user.save()

        self.club = Club.objects.create(team_name="UCI Club")
        self.event = Event.objects.create(
            name="UCI Combined Race",
            date=date(2025, 6, 1),
            organizer=self.club,
            reg_open=False,
            type_for_ranking="Český pohár",
            is_uci_race=True,
            uci_event_code="D2EV268824",
            uci_code_women_elite="WE123",
            uci_code_men_elite="ME123",
            uci_code_women_under_23="WU23123",
            uci_code_men_under_23="MU23123",
            uci_code_women_junior="WJ123",
            uci_code_men_junior="MJ123",
        )

        self.men_u23 = Rider.objects.create(
            uci_id=10000000001,
            first_name="Adam",
            last_name="U23",
            gender="Muž",
            date_of_birth=date(2004, 1, 1),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            is_elite=True,
            plate=101,
        )
        self.men_elite_best = Rider.objects.create(
            uci_id=10000000002,
            first_name="Boris",
            last_name="Elite",
            gender="Muž",
            date_of_birth=date(1998, 1, 1),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            is_elite=True,
            plate=202,
        )
        self.men_elite_second = Rider.objects.create(
            uci_id=10000000003,
            first_name="Cyril",
            last_name="Elite",
            gender="Muž",
            date_of_birth=date(1995, 1, 1),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            is_elite=True,
            plate=303,
        )

        self.u23_result = Result.objects.create(
            event=self.event,
            rider=self.men_u23,
            date=self.event.date,
            event_type=self.event.type_for_ranking,
            organizer=self.club.team_name,
            category="Combined Men",
            place=1,
            points=100,
            is_20=True,
        )
        self.elite_best_result = Result.objects.create(
            event=self.event,
            rider=self.men_elite_best,
            date=self.event.date,
            event_type=self.event.type_for_ranking,
            organizer=self.club.team_name,
            category="Combined Men",
            place=4,
            points=70,
            is_20=True,
        )
        self.elite_second_result = Result.objects.create(
            event=self.event,
            rider=self.men_elite_second,
            date=self.event.date,
            event_type=self.event.type_for_ranking,
            organizer=self.club.team_name,
            category="Combined Men",
            place=5,
            points=60,
            is_20=True,
        )

        RaceRun.objects.create(
            result=self.u23_result,
            event=self.event,
            rider=self.men_u23,
            is_20=True,
            is_beginner=False,
            round_type="FINAL",
            finish_time=31.111,
        )
        RaceRun.objects.create(
            result=self.elite_best_result,
            event=self.event,
            rider=self.men_elite_best,
            is_20=True,
            is_beginner=False,
            round_type="FINAL",
            finish_time=33.333,
        )
        RaceRun.objects.create(
            result=self.elite_second_result,
            event=self.event,
            rider=self.men_elite_second,
            is_20=True,
            is_beginner=False,
            round_type="FINAL",
            finish_time=34.444,
        )

    def _create_uci_template(self):
        temp_dir = tempfile.mkdtemp()
        template_path = os.path.join(temp_dir, "uci-template.xlsx")
        workbook = Workbook()
        general_ws = workbook.active
        general_ws.title = "General"
        general_ws["A4"] = "Competition Code"
        general_ws["A5"] = "Event Code"
        results_ws = workbook.create_sheet("Results")
        headers = ["Rank", "BIB", "UCI ID", "Last Name", "First Name", "Country", "Team", "Gender", "Phase", "Heat", "Result", "IRM"]
        for index, header in enumerate(headers, start=1):
            results_ws.cell(1, index, header)
        workbook.create_sheet("Reference")
        workbook.create_sheet("Country Reference")
        workbook.save(template_path)
        return temp_dir, template_path

    def test_uci_export_creates_six_files_and_recalculates_rank_within_each_uci_group(self):
        self.client.force_login(self.staff_user)
        temp_dir, template_path = self._create_uci_template()

        try:
            with patch("event.views.views_admin.UCI_EXPORT_TEMPLATE", template_path):
                response = self.client.get(reverse("event:export_uci_results", kwargs={"event_id": self.event.id}))
        finally:
            for filename in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, filename))
            os.rmdir(temp_dir)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")

        archive = zipfile.ZipFile(BytesIO(response.content))
        filenames = sorted(archive.namelist())
        self.assertEqual(len(filenames), 6)
        self.assertTrue(any("men_elite" in name for name in filenames))
        self.assertTrue(any("men_u23" in name for name in filenames))

        men_elite_name = next(name for name in filenames if "men_elite" in name)
        men_u23_name = next(name for name in filenames if "men_u23" in name)

        men_elite_workbook = load_workbook(BytesIO(archive.read(men_elite_name)))
        men_elite_results = men_elite_workbook["Results"]
        self.assertEqual(men_elite_results["A2"].value, 1)
        self.assertEqual(men_elite_results["C2"].value, self.men_elite_best.uci_id)
        self.assertEqual(men_elite_results["K2"].value, "1st, 33.333")
        self.assertEqual(men_elite_results["A3"].value, 2)
        self.assertEqual(men_elite_results["C3"].value, self.men_elite_second.uci_id)
        self.assertEqual(men_elite_results["K3"].value, "2nd, 34.444")

        men_u23_workbook = load_workbook(BytesIO(archive.read(men_u23_name)))
        men_u23_general = men_u23_workbook["General"]
        men_u23_results = men_u23_workbook["Results"]
        self.assertEqual(men_u23_general["B4"].value, "MU23123")
        self.assertEqual(men_u23_general["B5"].value, "D2EV268824")
        self.assertEqual(men_u23_results["A2"].value, 1)
        self.assertEqual(men_u23_results["C2"].value, self.men_u23.uci_id)

    def test_uci_export_redirects_with_flash_message_when_event_code_is_missing(self):
        self.client.force_login(self.staff_user)
        self.event.uci_event_code = ""
        self.event.save(update_fields=["uci_event_code"])
        temp_dir, template_path = self._create_uci_template()

        try:
            with patch("event.views.views_admin.UCI_EXPORT_TEMPLATE", template_path):
                response = self.client.get(
                    reverse("event:export_uci_results", kwargs={"event_id": self.event.id}),
                    follow=True,
                )
        finally:
            for filename in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, filename))
            os.rmdir(temp_dir)

        self.assertRedirects(response, reverse("event:event-admin", kwargs={"pk": self.event.id}))
        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertIn("U závodu chybí UCI_EVENT_CODE.", messages)

    def test_uci_export_service_builds_rows_and_zip_metadata(self):
        rows = build_uci_export_rows(self.event, "Men Elite")
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["rank"], 1)
        self.assertEqual(rows[0]["uci_id"], self.men_elite_best.uci_id)
        self.assertEqual(rows[0]["result"], "1st, 33.333")

        temp_dir, template_path = self._create_uci_template()
        try:
            zip_name, zip_bytes, export_metadata = generate_uci_export_zip(self.event, template_path)
        finally:
            for filename in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, filename))
            os.rmdir(temp_dir)

        self.assertTrue(zip_name.endswith(".zip"))
        self.assertEqual(len(export_metadata), 6)
        self.assertTrue(any(item["slug"] == "men_elite" and item["rows"] == 2 for item in export_metadata))

        archive = zipfile.ZipFile(BytesIO(zip_bytes))
        self.assertEqual(len(archive.namelist()), 6)


class EuropeanCupAdminTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user(
            first_name="EC",
            last_name="Admin",
            username="ec_admin",
            email="ec_admin@example.com",
            password="StrongPass123!",
        )
        self.staff_user.is_staff = True
        self.staff_user.is_active = True
        self.staff_user.save(update_fields=["is_staff", "is_active"])

        self.club = Club.objects.create(team_name="European Club")
        self.event = Event.objects.create(
            name="UEC BMX Racing European Cup - Round 1",
            date=date.today() + timedelta(days=10),
            organizer=self.club,
            type_for_ranking="Evropský pohár",
            reg_open=False,
        )

    def _make_pdf_upload(self, filename="results.pdf", text="results"):
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)
        pdf.drawString(100, 750, text)
        pdf.showPage()
        pdf.save()
        return SimpleUploadedFile(filename, buffer.getvalue(), content_type="application/pdf")

    def test_ec_admin_get_does_not_generate_exports(self):
        self.client.force_login(self.staff_user)

        with patch("event.views.views_admin._generate_ec_file") as generate_ec_file_mock, patch(
            "event.views.views_admin.generate_insurance_file"
        ) as generate_insurance_file_mock:
            response = self.client.get(reverse("event:event-admin", kwargs={"pk": self.event.id}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "btn-generate-ec-file")
        self.assertContains(response, "btn-generate-ec-insurance-file")
        generate_ec_file_mock.assert_not_called()
        generate_insurance_file_mock.assert_not_called()

    def test_ec_admin_generate_ec_file_runs_only_on_explicit_post(self):
        self.client.force_login(self.staff_user)

        with patch("event.views.views_admin._generate_ec_file") as generate_ec_file_mock:
            response = self.client.post(
                reverse("event:event-admin", kwargs={"pk": self.event.id}),
                {"btn-generate-ec-file": "1"},
            )

        self.assertEqual(response.status_code, 200)
        generate_ec_file_mock.assert_called_once_with(self.event)

    def test_ec_admin_rejects_non_pdf_results_upload(self):
        self.client.force_login(self.staff_user)

        response = self.client.post(
            reverse("event:event-admin", kwargs={"pk": self.event.id}),
            {
                "btn-upload-ec-results-pdf": "1",
                "results-pdf-files": SimpleUploadedFile(
                    "results.txt",
                    b"not-a-pdf",
                    content_type="text/plain",
                ),
            },
            follow=True,
        )

        self.assertRedirects(response, reverse("event:event-admin", kwargs={"pk": self.event.id}))
        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertTrue(any("mus\u00ed b\u00fdt PDF soubor" in message for message in messages))

    def test_ec_admin_accepts_single_pdf_results_upload(self):
        self.client.force_login(self.staff_user)

        response = self.client.post(
            reverse("event:event-admin", kwargs={"pk": self.event.id}),
            {
                "btn-upload-ec-results-pdf": "1",
                "results-pdf-files": self._make_pdf_upload(),
            },
            follow=True,
        )

        self.assertRedirects(response, reverse("event:event-admin", kwargs={"pk": self.event.id}))
        self.event.refresh_from_db()
        self.assertTrue(bool(self.event.full_results))
        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertIn("Výsledky byly nahrány a zveřejněny jako jedno sloučené PDF.", messages)

    def test_ec_admin_accepts_multiple_pdf_results_upload(self):
        self.client.force_login(self.staff_user)

        response = self.client.post(
            reverse("event:event-admin", kwargs={"pk": self.event.id}),
            {
                "btn-upload-ec-results-pdf": "1",
                "results-pdf-files": [
                    self._make_pdf_upload("results-1.pdf", "block-1"),
                    self._make_pdf_upload("results-2.pdf", "block-2"),
                    self._make_pdf_upload("results-3.pdf", "block-3"),
                ],
            },
            follow=True,
        )

        self.assertRedirects(response, reverse("event:event-admin", kwargs={"pk": self.event.id}))
        self.event.refresh_from_db()
        self.assertTrue(bool(self.event.full_results))
        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertIn("Výsledky byly nahrány a zveřejněny jako jedno sloučené PDF.", messages)


class EventPropositionEditorUploadTests(TestCase):
    def setUp(self):
        self.temp_media_dir = tempfile.TemporaryDirectory()
        self.media_override = override_settings(MEDIA_ROOT=self.temp_media_dir.name)
        self.media_override.enable()
        self.addCleanup(self.media_override.disable)
        self.addCleanup(self.temp_media_dir.cleanup)

        self.club = Club.objects.create(team_name="Upload Club")

        self.club_manager = User.objects.create_user(
            first_name="Club",
            last_name="Manager",
            username="club_manager",
            email="club_manager@example.com",
            password="StrongPass123!",
        )
        self.club_manager.is_active = True
        self.club_manager.is_club_manager = True
        self.club_manager.club = self.club
        self.club_manager.save(update_fields=["is_active", "is_club_manager", "club"])

        self.regular_user = User.objects.create_user(
            first_name="Regular",
            last_name="User",
            username="regular_user",
            email="regular_user@example.com",
            password="StrongPass123!",
        )
        self.regular_user.is_active = True
        self.regular_user.save(update_fields=["is_active"])

    def _make_image_upload(self, name="editor-image.png", image_format="PNG"):
        buffer = BytesIO()
        image = Image.new("RGB", (32, 32), color=(12, 120, 220))
        image.save(buffer, format=image_format)
        return SimpleUploadedFile(name, buffer.getvalue(), content_type="image/png")

    def test_club_manager_can_upload_editor_image(self):
        self.client.force_login(self.club_manager)

        response = self.client.post(
            reverse("event:proposition-editor-upload"),
            {"upload": self._make_image_upload()},
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("url", response.json())
        self.assertIn("/media/proposition_uploads/", response.json()["url"])

    def test_regular_authenticated_user_cannot_upload_editor_image(self):
        self.client.force_login(self.regular_user)

        response = self.client.post(
            reverse("event:proposition-editor-upload"),
            {"upload": self._make_image_upload()},
        )

        self.assertEqual(response.status_code, 403)
        self.assertIn("error", response.json())

    def test_non_image_upload_is_rejected(self):
        self.client.force_login(self.club_manager)

        response = self.client.post(
            reverse("event:proposition-editor-upload"),
            {
                "upload": SimpleUploadedFile(
                    "editor-file.txt",
                    b"not-an-image",
                    content_type="text/plain",
                )
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("error", response.json())

    @override_settings(CKEDITOR_5_MAX_FILE_SIZE=0)
    def test_editor_upload_rejects_file_over_size_limit(self):
        self.client.force_login(self.club_manager)

        response = self.client.post(
            reverse("event:proposition-editor-upload"),
            {"upload": self._make_image_upload()},
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("příliš velký", response.json()["error"]["message"])

    def test_editor_upload_rejects_disallowed_extension(self):
        self.client.force_login(self.club_manager)

        response = self.client.post(
            reverse("event:proposition-editor-upload"),
            {"upload": self._make_image_upload(name="editor-image.bmp")},
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("není povolena", response.json()["error"]["message"])

    @override_settings(CKEDITOR_5_MAX_IMAGE_WIDTH=16, CKEDITOR_5_MAX_IMAGE_HEIGHT=16)
    def test_editor_upload_rejects_image_over_dimension_limit(self):
        self.client.force_login(self.club_manager)

        response = self.client.post(
            reverse("event:proposition-editor-upload"),
            {"upload": self._make_image_upload()},
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("příliš rozměrný", response.json()["error"]["message"])


class EventPropositionSanitizationTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(team_name="Sanitize Club")
        self.user = Account.objects.create_user(
            first_name="Event",
            last_name="Editor",
            username="event_editor",
            email="event_editor@example.com",
            password="StrongPass123!",
        )
        self.event = Event.objects.create(
            name="Sanitized Proposition Event",
            date=date.today() + timedelta(days=20),
            organizer=self.club,
            type_for_ranking="Volný závod",
        )

    def test_event_proposition_save_sanitizes_html_and_preserves_safe_media(self):
        from event.models import EventProposition

        proposition = EventProposition.objects.create(
            event=self.event,
            created_by=self.user,
            updated_by=self.user,
            summary=(
                '<p>Summary</p><iframe src="https://www.youtube.com/embed/x"></iframe>'
                '<a href="https://www.youtube.com/watch?v=abc" target="_blank">Video</a>'
                '<img src="/media/proposition_uploads/test.png" onerror="alert(1)" alt="Track map">'
            ),
        )

        self.assertNotIn("<iframe", proposition.summary)
        self.assertIn('href="https://www.youtube.com/watch?v=abc"', proposition.summary)
        self.assertIn('rel="noopener noreferrer"', proposition.summary)
        self.assertIn('src="/media/proposition_uploads/test.png"', proposition.summary)
        self.assertNotIn("onerror", proposition.summary)


class EventPropositionFormTests(TestCase):
    def test_event_proposition_form_uses_ckeditor5_for_rich_text_fields(self):
        form = EventPropositionForm()

        for field_name in EventPropositionForm.RICH_TEXT_FIELDS:
            self.assertIsInstance(form.fields[field_name].widget, CKEditor5Widget)


class CommissarAssignmentsTests(TestCase):
    def setUp(self):
        self.commission_user = User.objects.create_user(
            first_name="Komise",
            last_name="User",
            username="commission_user",
            email="commission@example.com",
            password="StrongPass123!",
        )
        self.commission_user.is_active = True
        self.commission_user.is_commission = True
        self.commission_user.save(update_fields=["is_active", "is_commission"])

        self.staff_user = User.objects.create_user(
            first_name="Staff",
            last_name="Viewer",
            username="staff_viewer",
            email="staff_viewer@example.com",
            password="StrongPass123!",
        )
        self.staff_user.is_active = True
        self.staff_user.is_staff = True
        self.staff_user.save(update_fields=["is_active", "is_staff"])

        self.superuser = User.objects.create_user(
            first_name="Super",
            last_name="User",
            username="super_user",
            email="super_user@example.com",
            password="StrongPass123!",
        )
        self.superuser.is_active = True
        self.superuser.is_superuser = True
        self.superuser.save(update_fields=["is_active", "is_superuser"])

        self.club = Club.objects.create(team_name="Commissar Club")
        self.pcp = Commissar.objects.create(first_name="Petr", last_name="Hlavni", is_active=True)
        self.assist = Commissar.objects.create(first_name="Alena", last_name="Asistent", is_active=True)
        self.start = Commissar.objects.create(first_name="Start", last_name="Komisar", is_active=True)

        self.event = Event.objects.create(
            name="Commissar Race",
            date=date(date.today().year, 7, 15),
            organizer=self.club,
            type_for_ranking="Český pohár",
            reg_open=False,
        )

    def test_commission_user_sees_edit_button(self):
        self.client.force_login(self.commission_user)

        response = self.client.get(reverse("event:commissar-assignments"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Editovat")

    def test_staff_user_can_view_but_not_edit(self):
        self.client.force_login(self.staff_user)

        response = self.client.get(reverse("event:commissar-assignments"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Editovat")

    def test_superuser_sees_edit_button(self):
        self.client.force_login(self.superuser)

        response = self.client.get(reverse("event:commissar-assignments"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Editovat")

    def test_edit_mode_renders_sticky_change_bar(self):
        self.client.force_login(self.commission_user)

        response = self.client.get(
            f"{reverse('event:commissar-assignments')}?year={self.event.date.year}&edit=1"
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "commissar-edit-sticky-bar")
        self.assertContains(response, "Neuložené změny")

    def test_commission_user_can_update_assignments_and_audit_is_logged(self):
        self.client.force_login(self.commission_user)

        with self.assertLogs("audit", level="INFO") as audit_logs:
            response = self.client.post(
                reverse("event:commissar-assignments"),
                {
                    "year": str(self.event.date.year),
                    f"pcp_{self.event.id}": str(self.pcp.id),
                    f"pcp_assist_{self.event.id}": str(self.assist.id),
                    f"start_commissar_{self.event.id}": "",
                },
                follow=True,
            )

        self.assertRedirects(
            response,
            f"{reverse('event:commissar-assignments')}?year={self.event.date.year}",
            fetch_redirect_response=False,
        )
        self.event.refresh_from_db()
        self.assertEqual(self.event.pcp_id, self.pcp.id)
        self.assertEqual(self.event.pcp_assist_id, self.assist.id)
        self.assertTrue(any("commissar_assignment_updated" in line for line in audit_logs.output))

    def test_same_commissar_cannot_hold_two_roles_in_one_event(self):
        self.client.force_login(self.commission_user)

        response = self.client.post(
            reverse("event:commissar-assignments"),
            {
                "year": str(self.event.date.year),
                f"pcp_{self.event.id}": str(self.pcp.id),
                f"pcp_assist_{self.event.id}": str(self.pcp.id),
                f"start_commissar_{self.event.id}": "",
            },
            follow=True,
        )

        self.assertRedirects(
            response,
            f"{reverse('event:commissar-assignments')}?year={self.event.date.year}&edit=1",
            fetch_redirect_response=False,
        )
        self.event.refresh_from_db()
        self.assertIsNone(self.event.pcp_id)
        self.assertIsNone(self.event.pcp_assist_id)
        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertTrue(any("nemůže být jeden rozhodčí nasazen do více rolí" in message for message in messages))

    def test_conflicting_concurrent_edit_is_rejected(self):
        self.client.force_login(self.commission_user)
        self.event.pcp = self.start
        self.event.save(update_fields=["pcp"])

        response = self.client.post(
            reverse("event:commissar-assignments"),
            {
                "year": str(self.event.date.year),
                f"original_pcp_{self.event.id}": "",
                f"original_pcp_assist_{self.event.id}": "",
                f"original_start_commissar_{self.event.id}": "",
                f"pcp_{self.event.id}": str(self.pcp.id),
                f"pcp_assist_{self.event.id}": "",
                f"start_commissar_{self.event.id}": "",
            },
            follow=True,
        )

        self.assertRedirects(
            response,
            f"{reverse('event:commissar-assignments')}?year={self.event.date.year}&edit=1",
            fetch_redirect_response=False,
        )
        self.event.refresh_from_db()
        self.assertEqual(self.event.pcp_id, self.start.id)
        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertTrue(any("mezitím upravil někdo jiný" in message for message in messages))

    def test_event_model_rejects_duplicate_commissar_roles(self):
        self.event.pcp = self.pcp
        self.event.pcp_assist = self.pcp

        with self.assertRaises(ValidationError):
            self.event.full_clean()


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class RaceRunImportServiceTests(TestCase):
    def setUp(self):
        self.club = Club.objects.create(team_name="Import Club")
        self.event = Event.objects.create(
            name="Stats Import Race",
            date=date(2025, 10, 25),
            organizer=self.club,
            reg_open=False,
            type_for_ranking="Český pohár",
        )
        self.rider = Rider.objects.create(
            uci_id=10125224253,
            first_name="Simon",
            last_name="Aksamit",
            gender="Muž",
            date_of_birth=date(2009, 8, 2),
            club=self.club,
            is_active=True,
            is_approved=True,
            valid_licence=True,
            class_20="Boys 15",
            class_24="Boys 15 and 16",
            plate_text="868",
        )
        self.result = Result.objects.create(
            event=self.event,
            date=self.event.date,
            event_type=self.event.type_for_ranking,
            organizer=self.club.team_name,
            rider=self.rider,
            first_name=self.rider.first_name,
            last_name=self.rider.last_name,
            club=self.club.team_name,
            category="Boys 15-16",
            place=7,
            points=75,
            is_beginner=False,
            is_20=True,
        )

    def test_import_event_runs_builds_race_runs_from_html_stats(self):
        target_dir = os.path.join(settings.MEDIA_ROOT, "event_stats", str(self.event.pk))
        if os.path.isdir(target_dir):
            for filename in os.listdir(target_dir):
                os.remove(os.path.join(target_dir, filename))
        os.makedirs(target_dir, exist_ok=True)

        motos_html = """<html><body>
        <table class="gridtable">
        <caption>Boys 15-16 (1 Riders)</caption>
        <tr><th>Plate</th><th>Club</th><th>Name</th><th>Moto 1</th><th>Moto 2</th></tr>
        <tr><td>868</td><td>Import Club</td><td>Simon Aksamit</td><td>40 / 1</td><td>83 / 7</td></tr>
        </table>
        </body></html>"""
        motos_results_html = """<html><body>
        <table class="gridtable">
        <caption>Boys 15-16 (1 Riders)</caption>
        <tr><th>Ranking</th><th>Plate</th><th>Club</th><th>Name</th><th>Moto-Points</th><th>Moto 1</th><th>Moto 2</th></tr>
        <tr><td>7</td><td>868</td><td>Import Club</td><td>Simon Aksamit</td><td>3</td><td>1st 1.921 / 12.441 / 29.224</td><td>2nd 1.955 / 12.700 / 29.753</td></tr>
        </table>
        </body></html>"""
        final_html = """<html><body>
        <table class="gridtable">
        <caption>Boys 15-16 (1 Riders)</caption>
        <tr><th>Final</th><th>Lane</th><th>Plate</th><th>Name</th><th>Club</th></tr>
        <tr><td>F1 (A)</td><td>Pick 3</td><td>868</td><td>Simon Aksamit</td><td>Import Club</td></tr>
        </table>
        </body></html>"""

        files = {
            "motos__sample.html": motos_html,
            "motos_results__sample.html": motos_results_html,
            "final__sample.html": final_html,
        }
        for filename, content in files.items():
            with open(os.path.join(target_dir, filename), "w", encoding="utf-8") as handle:
                handle.write(content)

        imported_runs = RaceRunImportService().import_event_runs(self.event)

        self.assertEqual(imported_runs, 3)
        moto_1 = RaceRun.objects.get(result=self.result, round_type="MOTO", round_number=1)
        self.assertEqual(moto_1.heat_code, "40")
        self.assertEqual(moto_1.lane, 1)
        self.assertEqual(moto_1.place, "1st")
        self.assertEqual(moto_1.hill_time, 1.921)
        self.assertEqual(moto_1.split_1, 12.441)
        self.assertEqual(moto_1.finish_time, 29.224)
        self.assertEqual(moto_1.category, "Boys 15-16")
        self.assertEqual(moto_1.plate, "868")

        moto_2 = RaceRun.objects.get(result=self.result, round_type="MOTO", round_number=2)
        self.assertEqual(moto_2.heat_code, "83")
        self.assertEqual(moto_2.lane, 7)
        self.assertEqual(moto_2.place, "2nd")
        self.assertEqual(moto_2.finish_time, 29.753)

        final_run = RaceRun.objects.get(result=self.result, round_type="FINAL")
        self.assertEqual(final_run.heat_code, "F1 (A)")
        self.assertEqual(final_run.lane, 3)

    def test_import_event_runs_parses_rem_finish_and_hill_format(self):
        target_dir = os.path.join(settings.MEDIA_ROOT, "event_stats", str(self.event.pk))
        if os.path.isdir(target_dir):
            for filename in os.listdir(target_dir):
                os.remove(os.path.join(target_dir, filename))
        os.makedirs(target_dir, exist_ok=True)

        motos_html = """<html><body>
        <table class="gridtable">
        <caption>Boys 15-16 (1 Riders)</caption>
        <tr><th>Plate</th><th>Club</th><th>Name</th><th>Moto 1</th></tr>
        <tr><td>868</td><td>Import Club</td><td>Simon Aksamit</td><td>40 / 1</td></tr>
        </table>
        </body></html>"""
        motos_results_html = """<html><body>
        <table class="gridtable">
        <caption>Boys 15-16 (1 Riders)</caption>
        <tr><th>Ranking</th><th>Plate</th><th>Club</th><th>Name</th><th>Moto-Points</th><th>Moto 1</th></tr>
        <tr><td>7</td><td>868</td><td>Import Club</td><td>Simon Aksamit</td><td>3</td><td>3rd<br><small>39,205<br>{1.753}</small></td></tr>
        </table>
        </body></html>"""

        files = {
            "motos__sample.html": motos_html,
            "motos_results__sample.html": motos_results_html,
        }
        for filename, content in files.items():
            with open(os.path.join(target_dir, filename), "w", encoding="utf-8") as handle:
                handle.write(content)

        imported_runs = RaceRunImportService().import_event_runs(self.event)

        self.assertEqual(imported_runs, 1)
        moto_1 = RaceRun.objects.get(result=self.result, round_type="MOTO", round_number=1)
        self.assertEqual(moto_1.place, "3rd")
        self.assertEqual(moto_1.hill_time, 1.753)
        self.assertIsNone(moto_1.split_1)
        self.assertEqual(moto_1.finish_time, 39.205)

    def test_import_event_runs_parses_final_result_list_times_and_places(self):
        target_dir = os.path.join(settings.MEDIA_ROOT, "event_stats", str(self.event.pk))
        if os.path.isdir(target_dir):
            for filename in os.listdir(target_dir):
                os.remove(os.path.join(target_dir, filename))
        os.makedirs(target_dir, exist_ok=True)

        final_start_html = """<html><body>
        <table class="gridtable">
        <caption>Boys 15-16 (1 Riders)</caption>
        <tr><th>Final</th><th>Lane</th><th>Plate</th><th>Name</th><th>Club</th></tr>
        <tr><td>F1 (A)</td><td>Pick 3</td><td>868</td><td>Simon Aksamit</td><td>Import Club</td></tr>
        </table>
        </body></html>"""
        final_result_html = """<html><body>
        <table class="gridtable">
        <caption>Boys 15-16 (1 Riders)</caption>
        <tr><th>Final</th><th>Result</th><th>Plate</th><th>Club</th><th>Name</th></tr>
        <tr><td>F1 (A)</td><td>1st<br><small>33,380</small></td><td>868</td><td>Import Club</td><td>Simon Aksamit</td></tr>
        </table>
        </body></html>"""

        for filename, content in {
            "final__sample.html": final_start_html,
            "final_results__sample.html": final_result_html,
        }.items():
            with open(os.path.join(target_dir, filename), "w", encoding="utf-8") as handle:
                handle.write(content)

        imported_runs = RaceRunImportService().import_event_runs(self.event)

        self.assertEqual(imported_runs, 1)
        final_run = RaceRun.objects.get(result=self.result, round_type="FINAL")
        self.assertEqual(final_run.heat_code, "F1 (A)")
        self.assertEqual(final_run.lane, 3)
        self.assertEqual(final_run.place, "1st")
        self.assertEqual(final_run.finish_time, 33.380)

    def test_import_event_runs_creates_runs_without_result(self):
        self.result.delete()
        target_dir = os.path.join(settings.MEDIA_ROOT, "event_stats", str(self.event.pk))
        if os.path.isdir(target_dir):
            for filename in os.listdir(target_dir):
                os.remove(os.path.join(target_dir, filename))
        os.makedirs(target_dir, exist_ok=True)

        motos_html = """<html><body>
        <table class="gridtable">
        <caption>Boys 15-16 (1 Riders)</caption>
        <tr><th>Plate</th><th>Club</th><th>Name</th><th>Moto 1</th></tr>
        <tr><td>868</td><td>Import Club</td><td>Simon Aksamit</td><td>40 / 1</td></tr>
        </table>
        </body></html>"""
        motos_results_html = """<html><body>
        <table class="gridtable">
        <caption>Boys 15-16 (1 Riders)</caption>
        <tr><th>Ranking</th><th>Plate</th><th>Club</th><th>Name</th><th>Moto-Points</th><th>Moto 1</th></tr>
        <tr><td>7</td><td>868</td><td>Import Club</td><td>Simon Aksamit</td><td>3</td><td>3rd<br><small>39,205<br>{1.753}</small></td></tr>
        </table>
        </body></html>"""

        for filename, content in {
            "motos__sample.html": motos_html,
            "motos_results__sample.html": motos_results_html,
        }.items():
            with open(os.path.join(target_dir, filename), "w", encoding="utf-8") as handle:
                handle.write(content)

        imported_runs = RaceRunImportService().import_event_runs(self.event)

        self.assertEqual(imported_runs, 1)
        moto_1 = RaceRun.objects.get(event=self.event, rider=self.rider, round_type="MOTO", round_number=1)
        self.assertIsNone(moto_1.result)
        self.assertFalse(moto_1.is_beginner)
        self.assertTrue(moto_1.is_20)
        self.assertEqual(moto_1.finish_time, 39.205)


class EventFileReplacementSignalTests(TestCase):
    def setUp(self):
        self.temp_media_dir = tempfile.TemporaryDirectory()
        self.media_override = override_settings(MEDIA_ROOT=self.temp_media_dir.name)
        self.media_override.enable()
        self.addCleanup(self.media_override.disable)
        self.addCleanup(self.temp_media_dir.cleanup)

        self.club = Club.objects.create(team_name="Signal Club")
        self.event = Event.objects.create(
            name="Signal Race",
            date=date.today(),
            organizer=self.club,
            type_for_ranking="Volný závod",
            proposition=SimpleUploadedFile("old-proposition.pdf", b"old-file", content_type="application/pdf"),
        )

    def test_replaced_file_is_deleted_only_after_commit(self):
        old_file_name = self.event.proposition.name
        old_file_path = self.event.proposition.path

        with self.captureOnCommitCallbacks(execute=True):
            self.event.proposition = SimpleUploadedFile(
                "new-proposition.pdf",
                b"new-file",
                content_type="application/pdf",
            )
            self.event.save(update_fields=["proposition"])

        self.event.refresh_from_db()
        self.assertNotEqual(self.event.proposition.name, old_file_name)
        self.assertFalse(os.path.exists(old_file_path))

    def test_replaced_file_survives_transaction_rollback(self):
        old_file_name = self.event.proposition.name
        old_file_path = self.event.proposition.path

        with self.assertRaisesMessage(RuntimeError, "rollback"):
            with transaction.atomic():
                self.event.proposition = SimpleUploadedFile(
                    "rolled-back-proposition.pdf",
                    b"new-file",
                    content_type="application/pdf",
                )
                self.event.save(update_fields=["proposition"])
                raise RuntimeError("rollback")

        self.event.refresh_from_db()
        self.assertEqual(self.event.proposition.name, old_file_name)
        self.assertTrue(os.path.exists(old_file_path))


class RecalculateAllBalancesTests(TestCase):
    def setUp(self):
        self.active_user = User.objects.create_user(
            first_name="Active",
            last_name="User",
            username="active_user",
            email="active@example.com",
            password="StrongPass123!",
        )
        self.active_user.is_active = True
        self.active_user.credit = 999
        self.active_user.save(update_fields=["is_active", "credit"])

        self.inactive_user = User.objects.create_user(
            first_name="Inactive",
            last_name="User",
            username="inactive_user",
            email="inactive@example.com",
            password="StrongPass123!",
        )
        self.inactive_user.is_active = False
        self.inactive_user.save(update_fields=["is_active"])

        CreditTransaction.objects.create(user=self.active_user, amount=300, payment_complete=True)
        DebetTransaction.objects.create(user=self.active_user, amount=120, payment_valid=True)
        CreditTransaction.objects.create(user=self.inactive_user, amount=500, payment_complete=True)

        User.objects.filter(pk=self.active_user.pk).update(credit=999)
        User.objects.filter(pk=self.inactive_user.pk).update(credit=777)

    def test_recalculate_all_balances_updates_only_active_users(self):
        recalculate_all_balances()

        self.active_user.refresh_from_db()
        self.inactive_user.refresh_from_db()

        self.assertEqual(self.active_user.credit, 180)
        self.assertEqual(self.inactive_user.credit, 777)

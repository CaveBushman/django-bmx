import logging

from bmx.json_utils import html_safe_json
import os
import re
from collections import defaultdict
from statistics import mean, median, pstdev
from openpyxl import Workbook
from PIL import Image, UnidentifiedImageError
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.utils import timezone

logger = logging.getLogger(__name__)
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.translation import gettext as _, gettext_lazy as _gl, ngettext
from django.core.cache import cache
from django.contrib import messages
from django.core.paginator import Paginator
from django.views.decorators.cache import cache_control
from django.db.models import Count, Q
from django.contrib.admin.views.decorators import staff_member_required
from bmx import settings
from bmx.form_protection import build_security_context, protect_public_flow
from bmx.text_normalization import normalize_search_text
from accounts.models import AvatarChangeRequest
from rider.models import Rider, ForeignRider, RiderTransponderChange
from rider.rider import (
    two_years_inactive,
    start_licence_check,
    Participation,
    Cruiser,
    RiderSetClassesThread,
    first_line_riders_by_club_and_class,
    RiderQualifyToCNThread,
)
from club.models import Club
from event.models import Event, EventType, RaceRun, Result
from event.constants import event_type_color
from ranking.ranking import RANKING_RECOUNT_RUNNING_KEY, schedule_ranking_recount
import datetime
from datetime import date
from rider.rider import get_rider_data
from rider.subscriptions import (
    cancel_rider_stats_subscription,
    get_active_rider_stats_subscription,
    get_active_trainer_club_subscription,
    get_current_season_settings,
    has_active_trainer_club_extended_access,
    has_active_trainer_club_stats_access,
    purchase_rider_stats_subscription,
    resume_rider_stats_subscription,
)
from rider.models import MobileAppSubscription, RiderStatsSubscription, TrainerClubSubscription
from rider.mobile_subscriptions import (
    cancel_mobile_app_subscription,
    get_active_mobile_app_subscription,
    get_current_season_settings,
    purchase_mobile_app_subscription,
    resume_mobile_app_subscription,
)
from finance.models import EventInvoice, SubscriptionInvoice
from finance.subscription_invoices import SubscriptionInvoiceService
from rider.plates import display_plate, generate_available_plate_values, legacy_plate_int, normalize_plate_value
from rider.trainer_dashboard import (
    build_club_kpi_export_rows,
    build_club_riders_export_rows,
    build_trainer_dashboard_context,
    build_trainer_export_filename,
    export_rows_as_csv,
    export_rows_as_xlsx,
    get_exportable_trainer_club_or_403,
    normalize_export_format_or_404,
    handle_trainer_dashboard_post,
)
from rider.premium_stats_pdf import (
    build_rider_premium_stats_pdf,
    build_rider_premium_stats_pdf_filename,
)


from rider.views._common import *  # noqa: F401,F403
from rider.views._common import (  # podtržítkové helpery (import * je nepřenáší)
    _get_manageable_inactive_riders,
    _get_trainer_club_stats_subscription_for_rider,
    _get_premium_access_context,
    _validate_avatar_upload,
    _get_pending_avatar_request_for_target,
    _can_export_rider_premium_stats_pdf,
    _resolve_kpi_period,
    _build_rider_premium_stats_context,
    _parse_numeric_place,
    _safe_mean,
    _safe_median,
    _safe_stddev,
    _safe_rate,
    _clamp_score,
    _percentile_better_than,
    _get_table_column_key,
    _overall_result_places,
    _matches_wheel_filter,
    _build_chart_series,
    _build_track_options,
    _resolve_selected_track,
    _build_peer_context,
    _build_head_to_head,
    _build_track_stats,
    _rider_request_context,
    _render_rider_request,
    _get_event_year_options,
    _resolve_selected_year,
)


def riders_list_view(request):
    riders = (
        Rider.objects.filter(is_active=True, is_approved=True)
        .select_related("club")
        .only(
            "uci_id",
            "first_name",
            "last_name",
            "photo",
            "valid_licence",
            "is_qualify_to_cn_20",
            "is_qualify_to_cn_24",
            "club__team_name",
            "club__id",
            "is_20",
            "is_24",
            "class_20",
            "class_24",
            "plate",
            "plate_text",
            "plate_color_20",
        )
        .order_by("last_name", "first_name")
    )

    last_name_query = (request.GET.get("last_name") or "").strip()
    club_query = (request.GET.get("club") or "").strip()
    plate_query = (request.GET.get("plate") or "").strip()
    mcr_query = (request.GET.get("mcr") or "").strip()

    if last_name_query:
        riders = riders.filter(search_text_normalized__icontains=normalize_search_text(last_name_query))
    if club_query:
        riders = riders.filter(club__team_name__icontains=club_query)
    if plate_query:
        normalized_plate_query = normalize_plate_value(plate_query)
        plate_filter = Q(plate_text__iexact=normalized_plate_query)
        if normalized_plate_query.isdigit():
            plate_filter |= Q(plate=legacy_plate_int(normalized_plate_query))
        riders = riders.filter(plate_filter)
    if mcr_query == "qualified":
        riders = riders.filter(Q(is_qualify_to_cn_20=True) | Q(is_qualify_to_cn_24=True))

    paginator = Paginator(riders, 100)
    page_obj = paginator.get_page(request.GET.get("page"))

    metrics_queryset = Rider.objects.filter(is_active=True, is_approved=True)
    qualified_metrics_queryset = metrics_queryset.filter(
        Q(is_qualify_to_cn_20=True) | Q(is_qualify_to_cn_24=True)
    )
    data = {
        "riders": page_obj.object_list,
        "page_obj": page_obj,
        "total_riders": metrics_queryset.count(),
        "valid_licence_riders": metrics_queryset.filter(valid_licence=True).count(),
        "clubs_count": metrics_queryset.exclude(club__isnull=True).values("club").distinct().count(),
        "qualified_to_mcr_total": qualified_metrics_queryset.count(),
        "qualified_to_mcr_20": metrics_queryset.filter(is_qualify_to_cn_20=True).count(),
        "qualified_to_mcr_24": metrics_queryset.filter(is_qualify_to_cn_24=True).count(),
        "last_name_query": last_name_query,
        "club_query": club_query,
        "plate_query": plate_query,
        "mcr_query": mcr_query,
    }

    return render(request, "rider/riders-list.html", data)



def rider_detail_view(request, pk):
    rider = get_object_or_404(Rider, uci_id=pk)
    results = list(
        Result.objects.select_related("event")
        .filter(
            rider_id=rider.uci_id,
            date__gte=datetime.datetime.now() - datetime.timedelta(days=365),
        )
        .order_by("-date", "-id")
    )
    results_summary = {
        "points_20": sum(result.points or 0 for result in results if result.marked_20),
        "points_24": sum(result.points or 0 for result in results if result.marked_24),
        "outside_ranking_count": sum(
            1 for result in results if not result.marked_20 and not result.marked_24
        ),
    }
    season_settings = get_current_season_settings()
    premium_access_context = _get_premium_access_context(request.user, rider)
    premium_runs_count = RaceRun.objects.filter(rider=rider).count()
    data = {
        "rider": rider,
        "results": results,
        "results_summary": results_summary,
        **premium_access_context,
        "premium_runs_count": premium_runs_count,
        "premium_price": season_settings.rider_stats_monthly_price if season_settings else None,
        "can_manage_premium_stats": can_manage_premium_stats(request.user),
        "can_buy_premium_stats": request.user.is_authenticated and not premium_access_context["premium_access_via_admin"],
    }
    return render(request, "rider/rider-detail.html", data)



@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@staff_member_required
def free_plates_view(request):
    used_plates = [
        display_plate(plate_text, plate, fallback="")
        for plate_text, plate in Rider.objects.filter(is_active=True).values_list("plate_text", "plate")
    ]
    free_plates = generate_available_plate_values(used_plates)

    data = {
        "free_plates": free_plates,
        "free_plates_count": len(free_plates),
    }
    return render(request, "rider/rider-free-plates.html", data)



def rider_licence_lookup_view(request):
    uci_id = (request.GET.get("uci_id") or "").strip()
    if not uci_id.isdigit() or len(uci_id) != 11:
        return JsonResponse(
            {"ok": False, "message": _("UCI ID musí obsahovat přesně 11 číslic.")},
            status=400,
        )

    existing = Rider.objects.filter(uci_id=uci_id).first()
    if existing:
        return JsonResponse(
            {
                "ok": False,
                "message": _("Jezdec %(name)s už má přidělené startovní číslo.") % {'name': f"{existing.first_name} {existing.last_name}"},
            },
            status=409,
        )

    data_json, error_msg = get_rider_data(uci_id)
    if error_msg or not data_json:
        return JsonResponse(
            {"ok": False, "message": _("Licence nebyla nalezena.")},
            status=404,
        )

    first_name = data_json.get("firstName", "").strip()
    last_name = data_json.get("lastName", "").strip()
    birth = (data_json.get("birth", "") or "")[:10]
    gender_code = data_json.get("sex", {}).get("code", "M")
    gender = _("Žena") if gender_code == "F" else _("Muž")

    return JsonResponse(
        {
            "ok": True,
            "rider": {
                "uci_id": uci_id,
                "first_name": first_name,
                "last_name": last_name.capitalize() if last_name else "",
                "date_of_birth": birth,
                "gender": gender,
            },
        }
    )



@staff_member_required
def participation_riders_on_event(request):
    participation = Participation().count()

    file_path = os.path.join(settings.MEDIA_ROOT, "participation/participation.xlsx")
    if os.path.exists(file_path):
        with open(file_path, "rb") as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response["Content-Disposition"] = "inline; filename=" + os.path.basename(
                file_path
            )
            return response



@staff_member_required
def participation_stats_view(request):
    """Statistika účasti jezdců na závodech – grafy vývoje po letech."""
    EVENT_GROUPS = {
        "MČR":          [EventType.MCR_JEDNOTLIVCU],
        "Český pohár":  [EventType.CESKY_POHAR],
        "Česká liga":   [EventType.CESKA_LIGA],
        "Moravská liga": [EventType.MORAVSKA_LIGA],
        "Volné závody": [EventType.VOLNY_ZAVOD],
    }
    # Barvy série odvozené z kanonické mapy typů závodů (event/constants.py),
    # aby se shodovaly s odznaky v kalendáři a nedriftovaly.
    COLORS = {label: event_type_color(types[0]) for label, types in EVENT_GROUPS.items()}

    all_types = [t for types in EVENT_GROUPS.values() for t in types]

    # Unique riders per event (one row = one event)
    per_event = (
        Result.objects
        .filter(event__type_for_ranking__in=all_types, event__date__isnull=False)
        .values("event_id", "event__type_for_ranking", "event__date", "event__name")
        .annotate(rider_count=Count("rider_id", distinct=True))
        .filter(rider_count__gt=0)
        .order_by("event__date")
    )

    # Per group: collect individual events AND yearly aggregates
    events_by_group: dict[str, list] = {label: [] for label in EVENT_GROUPS}
    raw: dict[str, dict[int, list]] = {label: defaultdict(list) for label in EVENT_GROUPS}

    for row in per_event:
        year = row["event__date"].year
        date_str = row["event__date"].strftime("%d.%m.%Y")
        for label, types in EVENT_GROUPS.items():
            if row["event__type_for_ranking"] in types:
                events_by_group[label].append({
                    "date":  date_str,
                    "name":  row["event__name"] or "",
                    "count": row["rider_count"],
                })
                raw[label][year].append(row["rider_count"])
                break

    charts = {}
    for label, year_dict in raw.items():
        years = sorted(year_dict.keys())
        charts[label] = {
            "color":        COLORS[label],
            # yearly aggregates (pro 5 let a celou historii)
            "years":        years,
            "avg":          [round(sum(year_dict[y]) / len(year_dict[y])) for y in years],
            "max":          [max(year_dict[y]) for y in years],
            "event_count":  [len(year_dict[y]) for y in years],
            "total_unique": [sum(year_dict[y]) for y in years],
            # individual events (pro 1 rok a 2 roky)
            "events":       events_by_group[label],
        }

    return render(request, "rider/rider-participation-stats.html", {
        "charts_json": html_safe_json(charts),
        "chart_labels": list(EVENT_GROUPS.keys()),
    })



@staff_member_required
def riders_by_class_and_club(request):

    clubs = Club.objects.filter(is_active=True).order_by("team_name")
    file_name = "RIDERS_BY_CLUB_AND_CLASS.xlsx"

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    # Jeden DB dotaz místo 33 × počet klubů
    counts_qs = (
        Rider.objects.filter(is_active=True, is_approved=True)
        .values("club_id", "class_20")
        .annotate(count=Count("id"))
    )
    counts = {(row["club_id"], row["class_20"]): row["count"] for row in counts_qs}

    CLASSES = [
        "Boys 6", "Boys 7", "Girls 7", "Boys 8", "Girls 8",
        "Boys 9", "Girls 9", "Boys 10", "Girls 10", "Boys 11", "Girls 11",
        "Boys 12", "Girls 12", "Boys 13", "Girls 13", "Boys 14", "Girls 14",
        "Boys 15", "Girls 15", "Boys 16", "Girls 16",
        "Men 17-24", "Women 17-24", "Men 25-29", "Women 25 and over",
        "Men 30-34", "Men 35 and over", "Men Junior", "Women Junior",
        "Men Under 23", "Women Under 23", "Men Elite", "Women Elite",
    ]

    wb = Workbook()
    ws = wb.active

    first_line_riders_by_club_and_class(ws)

    row = 2
    for club in clubs:
        ws.cell(row, 1, club.team_name)
        for col, cls in enumerate(CLASSES, start=2):
            ws.cell(row, col, counts.get((club.id, cls), 0))
        row += 1
    wb.save(response)

    return response



def transponder_search_view(request):
    transponder_query = (request.GET.get("transponder") or "").strip().upper()
    results = []

    if transponder_query:
        czech_riders = Rider.objects.filter(
            Q(transponder_20__icontains=transponder_query)
            | Q(transponder_24__icontains=transponder_query)
        ).select_related("club")

        foreign_riders = ForeignRider.objects.filter(
            Q(transponder_20__icontains=transponder_query)
            | Q(transponder_24__icontains=transponder_query)
        )

        for rider in czech_riders:
            matched_in = []
            if (rider.transponder_20 or "").upper() == transponder_query:
                matched_in.append('20"')
            if (rider.transponder_24 or "").upper() == transponder_query:
                matched_in.append('24"')

            results.append(
                {
                    "first_name": rider.first_name,
                    "last_name": rider.last_name,
                    "club": rider.club.team_name if rider.club else "-",
                    "plate": rider.plate_display,
                    "matched_in": ", ".join(matched_in) or "-",
                    "status": _("Aktuální čip"),
                    "status_tone": "current",
                    "changed_at": None,
                }
            )

        for rider in foreign_riders:
            matched_in = []
            if (rider.transponder_20 or "").upper() == transponder_query:
                matched_in.append('20"')
            if (rider.transponder_24 or "").upper() == transponder_query:
                matched_in.append('24"')

            results.append(
                {
                    "first_name": rider.first_name,
                    "last_name": rider.last_name,
                    "club": rider.club or rider.state or "-",
                    "plate": rider.plate_display,
                    "matched_in": ", ".join(matched_in) or "-",
                    "status": _("Aktuální čip"),
                    "status_tone": "current",
                    "changed_at": None,
                }
            )

        current_chip_keys = {
            ("czech", result["first_name"], result["last_name"], result["matched_in"])
            for result in results
            if result["status_tone"] == "current"
        }

        historical_changes = (
            RiderTransponderChange.objects.filter(
                Q(old_transponder__iexact=transponder_query)
                | Q(new_transponder__iexact=transponder_query)
            )
            .select_related("rider__club", "changed_by")
            .order_by("-changed_at")
        )

        for change in historical_changes:
            rider = change.rider
            if not rider:
                continue

            slot_label = change.get_slot_display()
            is_current_now = (
                (slot_label == '20"' and (rider.transponder_20 or "").upper() == transponder_query)
                or (slot_label == '24"' and (rider.transponder_24 or "").upper() == transponder_query)
            )
            if is_current_now:
                continue

            dedupe_key = ("czech", rider.first_name, rider.last_name, slot_label)
            if dedupe_key in current_chip_keys:
                continue

            results.append(
                {
                    "first_name": rider.first_name,
                    "last_name": rider.last_name,
                    "club": rider.club.team_name if rider.club else "-",
                    "plate": rider.plate_display,
                    "matched_in": slot_label,
                    "status": _("Historický čip"),
                    "status_tone": "historical",
                    "changed_at": change.changed_at,
                }
            )

    data = {
        "transponder_query": transponder_query,
        "results": results,
        "has_search": bool(transponder_query),
    }
    return render(request, "rider/transponder-search.html", data)



def plate_search_view(request):
    plate_query = (request.GET.get("plate") or "").strip()
    results = []

    if plate_query:
        normalized_plate_query = normalize_plate_value(plate_query)
        plate_filter = Q(plate_text__iexact=normalized_plate_query)
        if normalized_plate_query.isdigit():
            plate_filter |= Q(plate=legacy_plate_int(normalized_plate_query))
        czech_riders = Rider.objects.filter(plate_filter, is_active=True).select_related("club")
        foreign_riders = ForeignRider.objects.filter(plate_filter)

        for rider in czech_riders:
            results.append(
                {
                    "plate": rider.plate_display,
                    "first_name": rider.first_name,
                    "last_name": rider.last_name,
                    "club": rider.club.team_name if rider.club else "-",
                }
            )

        for rider in foreign_riders:
            results.append(
                {
                    "plate": rider.plate_display,
                    "first_name": rider.first_name,
                    "last_name": rider.last_name,
                    "club": rider.club or rider.state or "-",
                }
            )

    data = {
        "plate_query": plate_query,
        "results": results,
        "has_search": bool(plate_query),
    }
    return render(request, "rider/plate-search.html", data)


__all__ = [
    'riders_list_view',
    'rider_detail_view',
    'free_plates_view',
    'rider_licence_lookup_view',
    'participation_riders_on_event',
    'participation_stats_view',
    'riders_by_class_and_club',
    'transponder_search_view',
    'plate_search_view',
]

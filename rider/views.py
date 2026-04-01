import logging
import os
import re
from collections import defaultdict
from io import BytesIO
from statistics import mean, median, pstdev
from openpyxl import Workbook
from PIL import Image, UnidentifiedImageError
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.utils import timezone

logger = logging.getLogger(__name__)
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.translation import gettext as _, gettext_lazy as _gl, ngettext
from django.core.cache import cache
from django.contrib import messages
from django.core.paginator import Paginator
from django.views.decorators.cache import cache_control
from django.db.models import Count, Q
from django.contrib.admin.views.decorators import staff_member_required
from bmx import settings
from accounts.models import AvatarChangeRequest
from .models import Rider, ForeignRider, RiderTransponderChange
from .rider import (
    two_years_inactive,
    start_licence_check,
    Participation,
    Cruiser,
    RiderSetClassesThread,
    first_line_riders_by_club_and_class,
    RiderQualifyToCNThread,
)
from rider.rider import set_all_riders_classes
from club.models import Club
from event.models import Event, RaceRun, Result
from ranking.ranking import RANKING_RECOUNT_RUNNING_KEY, schedule_ranking_recount
import datetime
from datetime import date
from rider.rider import get_rider_data
from rider.subscriptions import (
    cancel_rider_stats_subscription,
    cancel_trainer_club_subscription,
    get_active_rider_stats_subscription,
    get_active_trainer_club_subscription,
    get_active_trainer_extended_subscription,
    get_current_season_settings,
    has_active_trainer_club_extended_access,
    has_active_trainer_club_stats_access,
    purchase_trainer_club_subscription,
    purchase_rider_stats_subscription,
    resume_trainer_club_subscription,
    resume_rider_stats_subscription,
)
from rider.models import RiderStatsSubscription, TrainerClubSubscription
from finance.models import EventInvoice, SubscriptionInvoice
from finance.subscription_invoices import SubscriptionInvoiceService
from rider.plates import display_plate, generate_available_plate_values, legacy_plate_int, normalize_plate_value
from rider.trainer_dashboard import (
    build_club_kpi_export_rows,
    build_club_riders_export_rows,
    build_trainer_dashboard_context,
    build_trainer_export_filename,
    export_rows_as_csv,
    export_rows_as_xlsx,
    get_exportable_trainer_club_or_403,
    normalize_export_format_or_404,
    handle_trainer_dashboard_post,
)
from rider.premium_stats_pdf import (
    build_rider_premium_stats_pdf,
    build_rider_premium_stats_pdf_filename,
)

def can_manage_premium_stats(user):
    return user.is_authenticated and (user.is_admin or user.is_superuser or user.is_staff)


admin_required = user_passes_test(can_manage_premium_stats, login_url="/login/")


def can_access_rider_admin_only(user):
    return user.is_authenticated and user.is_admin


rider_admin_only_required = user_passes_test(can_access_rider_admin_only, login_url="/login/")


def can_manage_inactive_riders(user):
    return user.is_authenticated and (
        getattr(user, "is_admin", False)
        or getattr(user, "is_superuser", False)
        or getattr(user, "is_club_manager", False)
    )


inactive_riders_required = user_passes_test(can_manage_inactive_riders, login_url="/login/")


def _get_manageable_inactive_riders(user):
    inactive_riders = two_years_inactive()
    if getattr(user, "is_admin", False) or getattr(user, "is_superuser", False):
        return inactive_riders
    if getattr(user, "is_club_manager", False) and getattr(user, "club_id", None):
        return [rider for rider in inactive_riders if rider.club_id == user.club_id]
    return []


def can_access_trainer_dashboard(user):
    return user.is_authenticated and (
        getattr(user, "is_trainer", False) or user.is_admin or user.is_superuser or user.is_staff
    )


trainer_dashboard_required = user_passes_test(can_access_trainer_dashboard, login_url="/login/")
AVATAR_UPLOAD_MAX_BYTES = 5 * 1024 * 1024
AVATAR_ALLOWED_CONTENT_TYPES = {"image/jpeg", "image/png", "image/webp"}
FINAL_ROUND_TYPES = {"FINAL", "F2", "F4", "F8", "F16", "F32", "F64", "F128"}
TRACK_TABLE_COLUMNS = [
    ("M1", _gl("M1")),
    ("M2", _gl("M2")),
    ("M3", _gl("M3")),
    ("M4", _gl("M4")),
    ("M5", _gl("M5")),
    ("M6", _gl("M6")),
    ("M7", _gl("M7")),
    ("M8", _gl("M8")),
    ("F16", _gl("1/16")),
    ("F8", _gl("1/8")),
    ("F4", _gl("1/4")),
    ("F2", _gl("1/2")),
    ("FINAL", _gl("F")),
    ("DAY_MEDIAN", _gl("Medián dne")),
]
START_TABLE_COLUMNS = [
    ("M1", "M1"),
    ("M2", "M2"),
    ("M3", "M3"),
    ("M4", "M4"),
    ("M5", "M5"),
    ("M6", "M6"),
    ("M7", "M7"),
    ("M8", "M8"),
    ("F16", "1/16"),
    ("F8", "1/8"),
    ("F4", "1/4"),
    ("F2", "1/2"),
    ("FINAL", "F"),
    ("DAY_MEDIAN", _gl("Medián startu")),
]
SPLIT_TABLE_COLUMNS = [
    ("M1", "M1"),
    ("M2", "M2"),
    ("M3", "M3"),
    ("M4", "M4"),
    ("M5", "M5"),
    ("M6", "M6"),
    ("M7", "M7"),
    ("M8", "M8"),
    ("F16", "1/16"),
    ("F8", "1/8"),
    ("F4", "1/4"),
    ("F2", "1/2"),
    ("FINAL", "F"),
    ("DAY_MEDIAN", _gl("Medián Inter2")),
]
KPI_PERIOD_OPTIONS = [
    {"value": "1", "label": _gl("1 rok")},
    {"value": "2", "label": _gl("2 roky")},
    {"value": "3", "label": _gl("3 roky")},
    {"value": "5", "label": _gl("5 let")},
    {"value": "all", "label": _gl("Celá historie")},
]


def _get_trainer_club_stats_subscription_for_rider(user, rider, at_time=None):
    if not getattr(user, "is_authenticated", False):
        return None
    if rider.club_id is None:
        return None
    if not has_active_trainer_club_stats_access(user, rider.club, at_time=at_time):
        return None
    return get_active_trainer_club_subscription(
        user,
        rider.club,
        TrainerClubSubscription.PRODUCT_CLUB_STATS,
        at_time=at_time,
    )


def _get_premium_access_context(user, rider, at_time=None):
    individual_subscription = get_active_rider_stats_subscription(user, rider, at_time=at_time)
    trainer_club_subscription = _get_trainer_club_stats_subscription_for_rider(user, rider, at_time=at_time)
    premium_access_via_admin = can_manage_premium_stats(user)
    premium_access_via_individual = individual_subscription is not None
    premium_access_via_trainer_club = trainer_club_subscription is not None
    return {
        "premium_access": premium_access_via_admin or premium_access_via_individual or premium_access_via_trainer_club,
        "premium_access_via_admin": premium_access_via_admin,
        "premium_access_via_individual": premium_access_via_individual,
        "premium_access_via_trainer_club": premium_access_via_trainer_club,
        "active_subscription": individual_subscription,
        "trainer_club_subscription": trainer_club_subscription,
    }


def _validate_avatar_upload(uploaded_file):
    if not uploaded_file:
        raise ValueError(_("Vyber obrázek, který chceš nahrát."))
    if uploaded_file.size > AVATAR_UPLOAD_MAX_BYTES:
        raise ValueError(_("Avatar může mít maximálně 5 MB."))
    if uploaded_file.content_type not in AVATAR_ALLOWED_CONTENT_TYPES:
        raise ValueError(_("Povolené jsou pouze obrázky JPG, PNG nebo WEBP."))

    try:
        uploaded_file.seek(0)
        image = Image.open(uploaded_file)
        image.verify()
        uploaded_file.seek(0)
        checked = Image.open(uploaded_file)
        width, height = checked.size
        uploaded_file.seek(0)
    except (UnidentifiedImageError, OSError, ValueError):
        raise ValueError(_("Nahraný soubor není platný obrázek."))

    if width < 120 or height < 120:
        raise ValueError(_("Avatar musí mít alespoň 120 × 120 px."))
    if width > 6000 or height > 6000:
        raise ValueError(_("Avatar je příliš velký."))


def _get_pending_avatar_request_for_target(*, account=None, rider=None):
    AvatarChangeRequest.expire_stale_requests()
    queryset = AvatarChangeRequest.objects.filter(status=AvatarChangeRequest.STATUS_PENDING)
    if account is not None:
        return queryset.filter(target_account=account).first()
    if rider is not None:
        return queryset.filter(target_rider=rider).first()
    return None


def _can_export_rider_premium_stats_pdf(user, rider):
    if can_manage_premium_stats(user):
        return True
    if rider.club_id is None:
        return False
    return has_active_trainer_club_extended_access(user, rider.club)


def _resolve_kpi_period(period_value):
    if period_value == "all":
        return {
            "value": "all",
            "label": _("Celá historie"),
            "cutoff": None,
        }

    if period_value not in {"1", "2", "3", "5"}:
        period_value = "2"

    years = int(period_value)
    label = ngettext(
        "%(count)d rok",
        "%(count)d let",
        years
    ) % {'count': years}
    return {
        "value": period_value,
        "label": label,
        "cutoff": timezone.localdate() - datetime.timedelta(days=years * 365),
    }


def _build_rider_premium_stats_context(request, rider, premium_access_context):
    runs = list(
        RaceRun.objects.filter(rider=rider, is_beginner=False)
        .select_related("event", "event__organizer", "result")
        .order_by("-event__date", "round_type", "round_number", "id")
    )
    results = list(
        Result.objects.filter(rider_id=rider.uci_id, is_beginner=False)
        .select_related("event", "event__organizer")
        .order_by("-date", "-id")
    )
    kpi_period = _resolve_kpi_period(request.GET.get("years"))
    track_options = _build_track_options(results, runs, kpi_cutoff=kpi_period["cutoff"])
    selected_wheel = request.GET.get("wheel")
    if selected_wheel not in {"20", "24"}:
        selected_wheel = None
    selected_track = _resolve_selected_track(track_options, request.GET.get("track"))
    require_wheel_selection = False
    if selected_track is not None:
        track_results = [result for result in results if ((result.event and result.event.organizer_id) or 0) == selected_track["id"]]
        track_runs = [run for run in runs if ((run.event and run.event.organizer_id) or 0) == selected_track["id"]]
        timed_track_event_ids = {run.event_id for run in track_runs if run.finish_time is not None and run.event_id}
        timed_track_results = [result for result in track_results if result.event_id in timed_track_event_ids]
        timed_track_runs = [run for run in track_runs if run.finish_time is not None]
        available_wheels = []
        if any(result.is_20 for result in timed_track_results) or any(run.is_20 is True for run in timed_track_runs):
            available_wheels.append({"value": "20", "label": '20"'})
        if any(not result.is_20 for result in timed_track_results) or any(run.is_20 is False for run in timed_track_runs):
            available_wheels.append({"value": "24", "label": '24"'})
        if selected_wheel is None and len(available_wheels) == 1:
            selected_wheel = available_wheels[0]["value"]
        elif selected_wheel is None and len(available_wheels) > 1:
            require_wheel_selection = True
    else:
        available_wheels = []
    track_stats = None if require_wheel_selection else _build_track_stats(
        rider,
        selected_track,
        results,
        runs,
        wheel=selected_wheel,
        kpi_cutoff=kpi_period["cutoff"],
        kpi_label=kpi_period["label"],
    )
    compare_candidates = []
    if selected_track is not None and selected_wheel in {"20", "24"}:
        compare_candidates = _build_peer_context(
            rider,
            selected_track,
            selected_wheel,
            kpi_cutoff=kpi_period["cutoff"],
        )["candidates"]
    premium_runs_count = len([run for run in runs if run.finish_time is not None])
    return {
        "rider": rider,
        "runs": runs,
        "premium_runs_count": premium_runs_count,
        "subscription": premium_access_context["active_subscription"],
        "trainer_club_subscription": premium_access_context["trainer_club_subscription"],
        "has_admin_access": premium_access_context["premium_access_via_admin"],
        "has_trainer_club_access": premium_access_context["premium_access_via_trainer_club"],
        "kpi_period": kpi_period,
        "kpi_period_options": KPI_PERIOD_OPTIONS,
        "track_options": track_options,
        "selected_track": selected_track,
        "available_wheels": available_wheels,
        "selected_wheel": selected_wheel,
        "require_wheel_selection": require_wheel_selection,
        "track_stats": track_stats,
        "compare_candidates": compare_candidates,
        "premium_pdf_export_enabled": _can_export_rider_premium_stats_pdf(request.user, rider),
    }


@trainer_dashboard_required
def trainer_dashboard_view(request):
    if request.method == "POST":
        post_response = handle_trainer_dashboard_post(request)
        if post_response is not None:
            return post_response

    context = build_trainer_dashboard_context(request.user)
    return render(request, "rider/trainer-dashboard.html", context)


@trainer_dashboard_required
def trainer_club_riders_export_view(request, club_id, export_format):
    export_format = normalize_export_format_or_404(export_format)
    club = get_exportable_trainer_club_or_403(request.user, club_id)
    rows = build_club_riders_export_rows(club)
    headers = [
        ("uci_id", "UCI ID"),
        ("first_name", "Jméno"),
        ("last_name", "Příjmení"),
        ("class_20", "Kategorie 20\""),
        ("class_24", "Kategorie 24\""),
        ("plate", "Číslo"),
        ("transponder_20", "Transpondér 20\""),
        ("transponder_24", "Transpondér 24\""),
        ("valid_licence", "Platná licence"),
        ("email", "E-mail"),
        ("phone", "Telefon"),
    ]
    filename = build_trainer_export_filename(club, "riders")
    if export_format == "xlsx":
        return export_rows_as_xlsx(f"{filename}.xlsx", "Riders", headers, rows)
    return export_rows_as_csv(f"{filename}.csv", headers, rows)


@trainer_dashboard_required
def trainer_club_kpi_export_view(request, club_id, export_format):
    export_format = normalize_export_format_or_404(export_format)
    club = get_exportable_trainer_club_or_403(request.user, club_id)
    rows = build_club_kpi_export_rows(club)
    headers = [
        ("uci_id", "UCI ID"),
        ("first_name", "Jméno"),
        ("last_name", "Příjmení"),
        ("class_20", "Kategorie 20\""),
        ("class_24", "Kategorie 24\""),
        ("starts_total", "Starty celkem"),
        ("starts_last_2y", "Starty za 2 roky"),
        ("best_result", "Best result"),
        ("avg_place", "Průměrné pořadí"),
        ("median_finish", "Medián finish"),
        ("best_finish", "Best finish"),
        ("median_hill", "Medián hill"),
        ("median_split_1", "Medián split 1"),
    ]
    filename = build_trainer_export_filename(club, "kpi")
    if export_format == "xlsx":
        return export_rows_as_xlsx(f"{filename}.xlsx", "KPI", headers, rows)
    return export_rows_as_csv(f"{filename}.csv", headers, rows)


def riders_list_view(request):
    riders = (
        Rider.objects.filter(is_active=True, is_approved=True)
        .select_related("club")
        .only(
            "uci_id",
            "first_name",
            "last_name",
            "photo",
            "valid_licence",
            "is_qualify_to_cn_20",
            "is_qualify_to_cn_24",
            "club__team_name",
            "club__id",
            "is_20",
            "is_24",
            "class_20",
            "class_24",
            "plate",
            "plate_text",
            "plate_color_20",
        )
        .order_by("last_name", "first_name")
    )

    last_name_query = (request.GET.get("last_name") or "").strip()
    club_query = (request.GET.get("club") or "").strip()
    plate_query = (request.GET.get("plate") or "").strip()

    if last_name_query:
        riders = riders.filter(last_name__icontains=last_name_query)
    if club_query:
        riders = riders.filter(club__team_name__icontains=club_query)
    if plate_query:
        normalized_plate_query = normalize_plate_value(plate_query)
        plate_filter = Q(plate_text__iexact=normalized_plate_query)
        if normalized_plate_query.isdigit():
            plate_filter |= Q(plate=legacy_plate_int(normalized_plate_query))
        riders = riders.filter(plate_filter)

    paginator = Paginator(riders, 100)
    page_obj = paginator.get_page(request.GET.get("page"))

    metrics_queryset = Rider.objects.filter(is_active=True, is_approved=True)
    data = {
        "riders": page_obj.object_list,
        "page_obj": page_obj,
        "total_riders": metrics_queryset.count(),
        "valid_licence_riders": metrics_queryset.filter(valid_licence=True).count(),
        "clubs_count": metrics_queryset.exclude(club__isnull=True).values("club").distinct().count(),
        "last_name_query": last_name_query,
        "club_query": club_query,
        "plate_query": plate_query,
    }

    return render(request, "rider/riders-list.html", data)


def rider_detail_view(request, pk):
    rider = get_object_or_404(Rider, uci_id=pk)
    results = Result.objects.filter(
        rider_id=rider.uci_id,
        date__gte=datetime.datetime.now() - datetime.timedelta(days=365),
    ).order_by("date")
    season_settings = get_current_season_settings()
    premium_access_context = _get_premium_access_context(request.user, rider)
    premium_runs_count = RaceRun.objects.filter(rider=rider).count()
    data = {
        "rider": rider,
        "results": results,
        **premium_access_context,
        "premium_runs_count": premium_runs_count,
        "premium_price": season_settings.rider_stats_monthly_price if season_settings else None,
        "can_manage_premium_stats": can_manage_premium_stats(request.user),
        "can_buy_premium_stats": request.user.is_authenticated and not premium_access_context["premium_access_via_admin"],
    }
    return render(request, "rider/rider-detail.html", data)


@login_required
def rider_premium_stats_subscribe_view(request, pk):
    rider = get_object_or_404(Rider, uci_id=pk)
    if request.method != "POST":
        return redirect("rider:detail", pk=pk)

    try:
        subscription, created = purchase_rider_stats_subscription(request.user, rider)
    except ValueError as error:
        messages.error(request, str(error))
        return redirect("rider:detail", pk=pk)

    if created:
        messages.success(
            request,
            _("Prémiové statistiky jezdce %(name)s jsou aktivní do %(date)s.") % {'name': f"{rider.first_name} {rider.last_name}", 'date': timezone.localtime(subscription.expires_at).strftime("%d.%m.%Y %H:%M")}
        )
    else:
        messages.info(request, _("Prémiové statistiky tohoto jezdce už máš aktivní."))

    return redirect("rider:premium-stats", pk=pk)


def _parse_numeric_place(value):
    if value is None:
        return None
    match = re.search(r"\d+", str(value))
    return int(match.group()) if match else None


def _safe_mean(values):
    return round(mean(values), 2) if values else None


def _safe_median(values):
    return round(median(values), 2) if values else None


def _safe_stddev(values):
    if not values:
        return None
    if len(values) == 1:
        return 0.0
    return round(pstdev(values), 2)


def _safe_rate(numerator, denominator):
    if not denominator:
        return None
    return round((numerator / denominator) * 100, 1)


def _clamp_score(value):
    return max(0, min(int(round(value)), 100))


def _percentile_better_than(subject_value, peer_values):
    comparable = [value for value in peer_values if value is not None]
    if subject_value is None or not comparable:
        return None

    worse_count = sum(1 for value in comparable if value > subject_value)
    equal_count = sum(1 for value in comparable if value == subject_value)
    return round(((worse_count + (0.5 * equal_count)) / len(comparable)) * 100, 1)


def _get_table_column_key(run):
    if run.round_type == "MOTO":
        if not run.round_number or run.round_number > 8:
            return None
        return f"M{run.round_number}"
    if run.round_type in {"F16", "F8", "F4", "F2", "FINAL"}:
        return run.round_type
    return None


def _overall_result_places(results):
    per_event_places = {}
    standalone_places = []

    for result in results:
        if result.place is None:
            continue
        if result.event_id:
            current = per_event_places.get(result.event_id)
            if current is None or result.place < current:
                per_event_places[result.event_id] = result.place
        else:
            standalone_places.append(result.place)

    return list(per_event_places.values()) + standalone_places


def _matches_wheel_filter(result_or_run, wheel):
    if wheel not in {"20", "24"}:
        return True

    if hasattr(result_or_run, "result"):
        is_20 = bool(result_or_run.is_20)
    else:
        is_20 = bool(result_or_run.is_20)

    return is_20 if wheel == "20" else not is_20


def _build_chart_series(raw_event_rows, event_times, value_key):
    chart_events = []
    for item in raw_event_rows:
        event_id = item["event_id"]
        times = event_times.get(event_id, [])
        if times:
            chart_events.append(
                {
                    "date": item["date"],
                    "event_name": item["event_name"],
                    value_key: round(median(times), 3),
                }
            )

    chart_points = []
    chart_y_ticks = []
    if chart_events:
        min_time = min(item[value_key] for item in chart_events)
        max_time = max(item[value_key] for item in chart_events)
        time_span = max(max_time - min_time, 0.001)
        width = 1000
        height = 260
        left_pad = 56
        right_pad = 24
        top_pad = 24
        bottom_pad = 44
        usable_width = width - left_pad - right_pad
        usable_height = height - top_pad - bottom_pad

        for index, item in enumerate(chart_events):
            x = left_pad if len(chart_events) == 1 else left_pad + (usable_width * index / (len(chart_events) - 1))
            y = top_pad + ((max_time - item[value_key]) / time_span) * usable_height
            chart_points.append(
                {
                    "x": round(x, 2),
                    "y": round(y, 2),
                    "label": item["event_name"],
                    "short_label": item["date"].strftime("%d.%m.%Y") if item["date"] else f"Závod {index + 1}",
                    value_key: item[value_key],
                }
            )

        chart_polyline = " ".join(f"{point['x']},{point['y']}" for point in chart_points)
        for index in range(5):
            ratio = index / 4
            value = max_time - (time_span * ratio)
            y = top_pad + (usable_height * ratio)
            chart_y_ticks.append(
                {
                    "y": round(y, 2),
                    "label": round(value, 3),
                }
            )
    else:
        min_time = None
        max_time = None
        chart_polyline = ""

    return {
        "points": chart_points,
        "polyline": chart_polyline,
        "min": min_time,
        "max": max_time,
        "ticks": chart_y_ticks,
    }


def _build_track_options(results, runs, kpi_cutoff=None):
    track_map = {}

    def ensure_track(track_id, track_name):
        return track_map.setdefault(
            track_id,
            {
                "id": track_id,
                "name": track_name,
                "event_ids": set(),
                "timed_event_ids": set(),
                "results_count": 0,
                "runs_count": 0,
                "display_runs": set(),
                "recent_timed_runs_count": 0,
                "best_result": None,
            },
        )

    for result in results:
        organizer = result.event.organizer if result.event else None
        track_id = organizer.id if organizer else 0
        track_name = organizer.team_name if organizer else (result.organizer or "Neznámá trať")
        item = ensure_track(track_id, track_name)
        if result.event_id:
            item["event_ids"].add(result.event_id)
        item["results_count"] += 1
        if result.place is not None and (item["best_result"] is None or result.place < item["best_result"]):
            item["best_result"] = result.place

    for run in runs:
        organizer = run.event.organizer if run.event else None
        track_id = organizer.id if organizer else 0
        track_name = organizer.team_name if organizer else "Neznámá trať"
        item = ensure_track(track_id, track_name)
        if run.event_id:
            item["event_ids"].add(run.event_id)
        if run.finish_time is not None and run.event and run.event.date and (kpi_cutoff is None or run.event.date >= kpi_cutoff):
            item["runs_count"] += 1
            if run.event_id:
                item["timed_event_ids"].add(run.event_id)
            item["recent_timed_runs_count"] += 1

    result_places_by_track = defaultdict(list)
    for result in results:
        organizer = result.event.organizer if result.event else None
        track_id = organizer.id if organizer else 0
        result_places_by_track[track_id].append(result)

    track_options = []
    for item in track_map.values():
        overall_places = _overall_result_places(
            [
                result
                for result in result_places_by_track.get(item["id"], [])
                if result.event_id in item["timed_event_ids"]
            ]
        )
        item["best_result"] = min(overall_places) if overall_places else None
        item["starts_count"] = len(item["timed_event_ids"])
        item.pop("timed_event_ids", None)
        item.pop("display_runs", None)
        if item["runs_count"] <= 0:
            continue
        track_options.append(item)

    return sorted(track_options, key=lambda item: (-item["runs_count"], item["name"]))


def _resolve_selected_track(track_options, selected_track_id):
    if selected_track_id and selected_track_id.isdigit():
        selected_track = next((track for track in track_options if track["id"] == int(selected_track_id)), None)
        if selected_track is not None:
            return selected_track

    return track_options[0] if track_options else None


def _build_peer_context(rider, selected_track, wheel, kpi_cutoff=None):
    default_context = {
        "uci_category": None,
        "peer_label": None,
        "peer_results": Result.objects.none(),
        "peer_runs": RaceRun.objects.none(),
        "candidates": [],
    }

    if selected_track is None or wheel not in {"20", "24"}:
        return default_context

    if wheel == "20":
        uci_category = rider.class_20
        peer_label = f"{uci_category} (20\")" if uci_category else None
        peer_results = (
            Result.objects.filter(
                event__organizer_id=selected_track["id"],
                rider__class_20=uci_category,
                is_20=True,
                is_beginner=False,
            )
            .exclude(rider_id__isnull=True)
            .exclude(rider_id=rider.uci_id)
            .select_related("rider")
        )
        peer_runs = (
            RaceRun.objects.filter(
                event__organizer_id=selected_track["id"],
                is_20=True,
                is_beginner=False,
                rider__class_20=uci_category,
            )
            .exclude(rider_id__isnull=True)
            .exclude(rider_id=rider.id)
            .select_related("rider", "result")
        )
    else:
        uci_category = rider.class_24
        peer_label = f"{uci_category} (24\")" if uci_category else None
        peer_results = (
            Result.objects.filter(
                event__organizer_id=selected_track["id"],
                rider__class_24=uci_category,
                is_20=False,
                is_beginner=False,
            )
            .exclude(rider_id__isnull=True)
            .exclude(rider_id=rider.uci_id)
            .select_related("rider")
        )
        peer_runs = (
            RaceRun.objects.filter(
                event__organizer_id=selected_track["id"],
                is_20=False,
                is_beginner=False,
                rider__class_24=uci_category,
            )
            .exclude(rider_id__isnull=True)
            .exclude(rider_id=rider.id)
            .select_related("rider", "result")
        )

    if kpi_cutoff is not None:
        peer_results = peer_results.filter(date__gte=kpi_cutoff)
        peer_runs = peer_runs.filter(event__date__gte=kpi_cutoff)

    candidate_map = {}
    for result in peer_results:
        if result.rider_id and result.rider:
            candidate_map[result.rider.uci_id] = result.rider
    for run in peer_runs:
        if run.rider_id and run.rider:
            candidate_map[run.rider.uci_id] = run.rider

    candidates = sorted(
        candidate_map.values(),
        key=lambda item: ((item.last_name or "").lower(), (item.first_name or "").lower(), item.uci_id),
    )

    return {
        "uci_category": uci_category,
        "peer_label": peer_label,
        "peer_results": peer_results,
        "peer_runs": peer_runs,
        "candidates": candidates,
    }


def _build_head_to_head(base_results, opponent_results, base_runs, opponent_runs):
    def _build_run_map(runs, allowed_round_types):
        run_map = defaultdict(list)
        for run in runs:
            if run.round_type not in allowed_round_types:
                continue
            parsed_place = _parse_numeric_place(run.place)
            if parsed_place is None or not run.event_id:
                continue
            key = (
                run.event_id,
                run.round_type,
                run.round_number or 0,
                run.heat_code or "",
            )
            run_map[key].append(parsed_place)
        for values in run_map.values():
            values.sort()
        return run_map

    def _compare_run_maps(base_map, opponent_map):
        wins = losses = ties = 0
        for key in set(base_map.keys()) & set(opponent_map.keys()):
            for base_place, opponent_place in zip(base_map[key], opponent_map[key]):
                if base_place < opponent_place:
                    wins += 1
                elif base_place > opponent_place:
                    losses += 1
                else:
                    ties += 1
        return {
            "wins": wins,
            "losses": losses,
            "ties": ties,
            "shared": wins + losses + ties,
        }

    moto_h2h = _compare_run_maps(
        _build_run_map(base_runs, {"MOTO"}),
        _build_run_map(opponent_runs, {"MOTO"}),
    )
    final_h2h = _compare_run_maps(
        _build_run_map(base_runs, FINAL_ROUND_TYPES),
        _build_run_map(opponent_runs, FINAL_ROUND_TYPES),
    )

    base_results_map = {
        result.event_id: result.place
        for result in base_results
        if result.event_id and result.place is not None
    }
    opponent_results_map = {
        result.event_id: result.place
        for result in opponent_results
        if result.event_id and result.place is not None
    }
    event_wins = event_losses = event_ties = 0
    for event_id in set(base_results_map.keys()) & set(opponent_results_map.keys()):
        base_place = base_results_map[event_id]
        opponent_place = opponent_results_map[event_id]
        if base_place < opponent_place:
            event_wins += 1
        elif base_place > opponent_place:
            event_losses += 1
        else:
            event_ties += 1

    return {
        "moto": moto_h2h,
        "final": final_h2h,
        "event": {
            "wins": event_wins,
            "losses": event_losses,
            "ties": event_ties,
            "shared": event_wins + event_losses + event_ties,
        },
    }


def _build_track_stats(rider, selected_track, all_results, all_runs, wheel=None, kpi_cutoff=None, kpi_label="2 roky"):
    if selected_track is None:
        return None

    track_results = [result for result in all_results if ((result.event and result.event.organizer_id) or 0) == selected_track["id"]]
    track_runs = [run for run in all_runs if ((run.event and run.event.organizer_id) or 0) == selected_track["id"]]

    if wheel in {"20", "24"}:
        track_results = [result for result in track_results if _matches_wheel_filter(result, wheel)]
        track_runs = [run for run in track_runs if _matches_wheel_filter(run, wheel)]

    timed_track_runs = [run for run in track_runs if run.finish_time is not None]
    timed_track_event_ids = {run.event_id for run in timed_track_runs if run.event_id}
    timed_track_results = [result for result in track_results if result.event_id in timed_track_event_ids]
    track_has_20 = any(result.is_20 for result in timed_track_results) or any(run.is_20 is True for run in timed_track_runs)
    track_has_24 = any(not result.is_20 for result in timed_track_results) or any(run.is_20 is False for run in timed_track_runs)
    overall_timed_event_ids = {
        run.event_id
        for run in all_runs
        if run.finish_time is not None and _matches_wheel_filter(run, wheel) and run.event_id
    }
    overall_results = [
        result
        for result in all_results
        if _matches_wheel_filter(result, wheel) and result.event_id in overall_timed_event_ids
    ]
    recent_track_results = [result for result in track_results if result.date and (kpi_cutoff is None or result.date >= kpi_cutoff)]
    recent_track_runs = [
        run
        for run in track_runs
        if run.event and run.event.date and (kpi_cutoff is None or run.event.date >= kpi_cutoff)
    ]
    timed_recent_track_runs = [run for run in recent_track_runs if run.finish_time is not None]
    timed_recent_track_event_ids = {run.event_id for run in timed_recent_track_runs if run.event_id}
    recent_track_results = [result for result in recent_track_results if result.event_id in timed_recent_track_event_ids]
    recent_track_event_ids = timed_recent_track_event_ids
    event_metrics = defaultdict(
        lambda: {
            "moto_places": [],
            "final_places": [],
            "bad_runs": 0,
            "result_place": None,
        }
    )

    for result in recent_track_results:
        if result.event_id:
            event_metrics[result.event_id]["result_place"] = result.place

    finish_times = []
    hill_times = []
    split_times = []
    hill_ranks = []
    hill_rankable_runs = 0
    positions_gained_after_hill = []
    lane_hill_times = defaultdict(list)
    lane_result_places = defaultdict(list)
    moto_finish_times = []
    final_finish_times = []
    moto_places = []
    final_places = []
    bad_status_count = 0
    clean_runs = 0
    eligible_clean_runs = 0

    for run in timed_recent_track_runs:
        if run.lane and run.hill_time is not None:
            lane_hill_times[run.lane].append(float(run.hill_time))

        parsed_place = _parse_numeric_place(run.place)
        finish_time = run.finish_time
        run_metrics = event_metrics[run.event_id]

        if finish_time is not None:
            finish_times.append(float(finish_time))
            clean_runs += 1

        if run.hill_time is not None:
            hill_times.append(float(run.hill_time))

        if run.split_1 is not None:
            split_times.append(float(run.split_1))

        place_text = (run.place or "").upper()
        is_bad_status = any(flag in place_text for flag in ("DNF", "DNS", "DSQ", "REL"))
        if is_bad_status:
            bad_status_count += 1
            run_metrics["bad_runs"] += 1
            eligible_clean_runs += 1
        elif finish_time is not None:
            eligible_clean_runs += 1

        if run.round_type == "MOTO":
            if parsed_place is not None:
                moto_places.append(parsed_place)
                run_metrics["moto_places"].append(parsed_place)
            if finish_time is not None:
                moto_finish_times.append(float(finish_time))
        elif run.round_type in FINAL_ROUND_TYPES:
            if parsed_place is not None:
                final_places.append(parsed_place)
                run_metrics["final_places"].append(parsed_place)
            if finish_time is not None:
                final_finish_times.append(float(finish_time))

        if run.lane and parsed_place is not None:
            lane_result_places[run.lane].append(parsed_place)

    heat_runs = defaultdict(list)
    if recent_track_event_ids:
        comparable_runs_qs = RaceRun.objects.filter(
            event_id__in=recent_track_event_ids,
            is_beginner=False,
            hill_time__isnull=False,
        )
        if wheel in {"20", "24"}:
            comparable_runs_qs = comparable_runs_qs.filter(is_20=(wheel == "20"))

        for item in comparable_runs_qs.only("id", "event_id", "round_type", "round_number", "heat_code", "rider_id", "hill_time"):
            heat_key = (
                item.event_id,
                item.round_type,
                item.round_number or 0,
                item.heat_code or "",
            )
            heat_runs[heat_key].append(item)

    for run in timed_recent_track_runs:
        if run.hill_time is None or not run.event_id:
            continue
        heat_key = (
            run.event_id,
            run.round_type,
            run.round_number or 0,
            run.heat_code or "",
        )
        same_heat_runs = sorted(heat_runs.get(heat_key, []), key=lambda item: (float(item.hill_time), item.id))
        if not same_heat_runs:
            continue
        hill_rankable_runs += 1
        for index, item in enumerate(same_heat_runs, start=1):
            if item.id == run.id:
                hill_ranks.append(index)
                parsed_finish_place = _parse_numeric_place(run.place)
                if parsed_finish_place is not None:
                    positions_gained_after_hill.append(index - parsed_finish_place)
                break

    track_place_values = _overall_result_places(recent_track_results)
    overall_place_values = _overall_result_places(overall_results)
    events_count = len({result.event_id for result in recent_track_results if result.event_id})
    events_with_motos = sum(1 for item in event_metrics.values() if item["moto_places"])
    events_with_final_stage = sum(1 for item in event_metrics.values() if item["final_places"])
    events_with_final = len({run.event_id for run in timed_recent_track_runs if run.round_type == "FINAL"})
    events_with_moto_qualification_flag = len({
        run.event_id
        for run in timed_recent_track_runs
        if run.round_type == "MOTO" and run.qualified_to_next_round is True and run.event_id
    })
    events_with_any_moto_qualification_info = len({
        run.event_id
        for run in timed_recent_track_runs
        if run.round_type == "MOTO" and run.qualified_to_next_round is not None and run.event_id
    })
    progressed_events_count = (
        events_with_moto_qualification_flag
        if events_with_any_moto_qualification_info
        else events_with_final_stage
    )

    strong_start_events = [item for item in event_metrics.values() if item["moto_places"] and mean(item["moto_places"]) <= 3]
    weak_start_events = [item for item in event_metrics.values() if item["moto_places"] and mean(item["moto_places"]) > 3]

    converted_events = sum(
        1
        for item in strong_start_events
        if item["result_place"] is not None and item["result_place"] <= 3
    )
    recovered_events = sum(
        1
        for item in weak_start_events
        if item["result_place"] is not None and item["result_place"] < mean(item["moto_places"])
    )

    avg_moto_place = _safe_mean(moto_places)
    avg_final_place = _safe_mean(final_places)
    avg_track_place = _safe_mean(track_place_values)
    overall_avg_place = _safe_mean(overall_place_values)
    affinity_delta = round(overall_avg_place - avg_track_place, 2) if avg_track_place is not None and overall_avg_place is not None else None
    affinity_score = _clamp_score(50 + (affinity_delta or 0) * 20) if affinity_delta is not None else None
    if affinity_delta is None:
        affinity_label = None
    elif affinity_delta >= 1:
        affinity_label = _("Oblíbená trať")
    elif affinity_delta <= -1:
        affinity_label = _("Problémová trať")
    else:
        affinity_label = _("Neutrální trať")

    pressure_delta = round((avg_moto_place or 0) - (avg_final_place or 0), 2) if avg_moto_place is not None and avg_final_place is not None else None
    pressure_score = _clamp_score(50 + (pressure_delta or 0) * 15) if pressure_delta is not None else None
    stability_score = _clamp_score(
        100
        - ((_safe_stddev(track_place_values) or 0) * 12)
        - ((_safe_stddev(finish_times) or 0) * 8)
        - (_safe_rate(bad_status_count, len(timed_recent_track_runs)) or 0)
    ) if timed_recent_track_runs else None
    risk_score = _clamp_score(
        ((_safe_rate(bad_status_count, len(timed_recent_track_runs)) or 0) * 0.7)
        + ((_safe_stddev(track_place_values) or 0) * 10)
        + ((_safe_stddev(finish_times) or 0) * 6)
    ) if timed_recent_track_runs else None

    gate_avg = _safe_mean(hill_times)
    gate_std = _safe_stddev(hill_times)
    
    has_gate_data = gate_avg and gate_avg > 0 and len(hill_times) > 1
    gate_cv = (gate_std / gate_avg * 100) if has_gate_data else 0

    finish_avg = _safe_mean(finish_times)
    finish_std = _safe_stddev(finish_times)
    has_finish_data = finish_avg and finish_avg > 0 and len(finish_times) > 1
    finish_cv = (finish_std / finish_avg * 100) if has_finish_data else 0

    pos_std = _safe_stddev(track_place_values) or 0

    consistency_metric = None
    if has_gate_data and has_finish_data:
        consistency_metric = 0.4 * gate_cv + 0.4 * finish_cv + 0.2 * pos_std

    consistency_score = _clamp_score(100 - consistency_metric * 3) if (consistency_metric is not None and timed_recent_track_runs) else None

    consistency_label = None
    if consistency_score is not None:
        if consistency_score >= 90:
            consistency_label = _("Velmi stabilní")
        elif consistency_score >= 75:
            consistency_label = _("Stabilní")
        elif consistency_score >= 60:
            consistency_label = _("Kolísavý")
        else:
            consistency_label = _("Nevyrovnaný")

    best_hill_lane = None
    best_hill_lane_median = None
    if lane_hill_times:
        lane_medians = {
            lane: round(median(values), 3)
            for lane, values in lane_hill_times.items()
            if values
        }
        if lane_medians:
            best_hill_lane, best_hill_lane_median = min(lane_medians.items(), key=lambda item: item[1])

    best_result_lane = None
    best_result_lane_average = None
    if lane_result_places:
        lane_place_averages = {
            lane: round(mean(values), 2)
            for lane, values in lane_result_places.items()
            if values
        }
        if lane_place_averages:
            best_result_lane, best_result_lane_average = min(lane_place_averages.items(), key=lambda item: item[1])

    ordered_results = sorted(
        [result for result in recent_track_results if result.date],
        key=lambda item: (item.date, item.id),
    )
    recent_event_trends = {}
    for result in recent_track_results:
        if not result.event_id or not result.date:
            continue
        item = recent_event_trends.setdefault(
            result.event_id,
            {
                "date": result.date,
                "place": result.place,
                "times": [],
            },
        )
        item["place"] = result.place

    for run in timed_recent_track_runs:
        if run.event_id and run.event and run.event.date and run.finish_time is not None:
            item = recent_event_trends.setdefault(
                run.event_id,
                {
                    "date": run.event.date,
                    "place": None,
                    "times": [],
                },
            )
            item["times"].append(float(run.finish_time))

    recent_event_trend_rows = []
    for event_id, item in recent_event_trends.items():
        recent_event_trend_rows.append(
            {
                "event_id": event_id,
                "date": item["date"],
                "place": item["place"],
                "median_time": round(median(item["times"]), 3) if item["times"] else None,
            }
        )
    recent_event_trend_rows.sort(key=lambda item: (item["date"], item["event_id"]))

    trend_delta = None
    trend_detail = _("Potřebujeme víc startů")
    timed_trend_rows = [item for item in recent_event_trend_rows if item["median_time"] is not None]

    if len(timed_trend_rows) >= 2:
        earliest_event = timed_trend_rows[0]
        latest_event = timed_trend_rows[-1]
        time_delta = None
        time_signal = 0

        if earliest_event["median_time"] is not None and latest_event["median_time"] is not None:
            time_delta = round(earliest_event["median_time"] - latest_event["median_time"], 3)
            if time_delta > 0.05:
                time_signal = 1
            elif time_delta < -0.05:
                time_signal = -1

        if time_signal > 0:
            trend_label = _("Zlepšuje se")
        elif time_signal < 0:
            trend_label = _("Zhoršuje se")
        else:
            trend_label = _("Stagnuje")

        detail_parts = []
        if time_delta is not None:
            if time_delta > 0:
                detail_parts.append(_("čas o %(delta)s s rychlejší") % {'delta': abs(time_delta)})
            elif time_delta < 0:
                detail_parts.append(_("čas o %(delta)s s pomalejší") % {'delta': abs(time_delta)})
            else:
                detail_parts.append(_("čas beze změny"))

        trend_delta = time_delta
        trend_detail = _("Poslední vs první: ") + " | ".join(detail_parts) if detail_parts else trend_detail
    else:
        trend_label = _("Málo dat")

    event_rows_map = {}
    start_event_rows_map = {}
    split_event_rows_map = {}
    event_finish_times = defaultdict(list)
    event_hill_times = defaultdict(list)
    event_split_times = defaultdict(list)
    for run in sorted(
        timed_track_runs,
        key=lambda item: (
            item.event.date if item.event and item.event.date else date.min,
            item.event_id or 0,
            item.id,
        ),
        ):
        if run.finish_time is not None and run.event_id:
            event_finish_times[run.event_id].append(float(run.finish_time))
        if run.hill_time is not None and run.event_id:
            event_hill_times[run.event_id].append(float(run.hill_time))
        if run.split_1 is not None and run.event_id:
            event_split_times[run.event_id].append(float(run.split_1))

        column_key = _get_table_column_key(run)
        if not column_key:
            continue

        row = event_rows_map.setdefault(
            run.event_id,
            {
                "event_id": run.event_id,
                "date": run.event.date if run.event else None,
                "event_name": run.event.name if run.event else "-",
                "cells": {key: None for key, _ in TRACK_TABLE_COLUMNS},
            },
        )
        row["cells"][column_key] = {
            "place": run.place or "-",
            "time": run.finish_time,
        }
        if run.hill_time is not None:
            start_row = start_event_rows_map.setdefault(
                run.event_id,
                {
                    "event_id": run.event_id,
                    "date": run.event.date if run.event else None,
                    "event_name": run.event.name if run.event else "-",
                    "cells": {key: None for key, _ in START_TABLE_COLUMNS},
                },
            )
            start_row["cells"][column_key] = {
                "time": run.hill_time,
            }
        if run.split_1 is not None:
            split_row = split_event_rows_map.setdefault(
                run.event_id,
                {
                    "event_id": run.event_id,
                    "date": run.event.date if run.event else None,
                    "event_name": run.event.name if run.event else "-",
                    "cells": {key: None for key, _ in SPLIT_TABLE_COLUMNS},
                },
            )
            split_row["cells"][column_key] = {
                "time": run.split_1,
            }

    raw_event_rows = sorted(
        event_rows_map.values(),
        key=lambda item: (item["date"] or date.min, item["event_name"]),
    )
    event_rows = []
    for item in raw_event_rows:
        event_id = item["event_id"]
        day_times = event_finish_times.get(event_id, [])
        if day_times:
            item["cells"]["DAY_MEDIAN"] = {
                "place": "",
                "time": round(median(day_times), 3),
            }
        row_cells = []
        for key, label in TRACK_TABLE_COLUMNS:
            row_cells.append(
                {
                    "key": key,
                    "label": label,
                    "value": item["cells"].get(key),
                }
            )
        event_rows.append(
            {
                "event_id": item["event_id"],
                "date": item["date"],
                "event_name": item["event_name"],
                "cells": row_cells,
            }
        )

    finish_chart = _build_chart_series(raw_event_rows, event_finish_times, "median_time")

    raw_start_event_rows = sorted(
        start_event_rows_map.values(),
        key=lambda item: (item["date"] or date.min, item["event_name"]),
    )
    start_event_rows = []
    for item in raw_start_event_rows:
        event_id = item["event_id"]
        event_hill_time_list = event_hill_times.get(event_id, [])
        if event_hill_time_list:
            item["cells"]["DAY_MEDIAN"] = {
                "time": round(median(event_hill_time_list), 3),
            }
        row_cells = []
        for key, label in START_TABLE_COLUMNS:
            row_cells.append(
                {
                    "key": key,
                    "label": label,
                    "value": item["cells"].get(key),
                }
            )
        start_event_rows.append(
            {
                "event_id": item["event_id"],
                "date": item["date"],
                "event_name": item["event_name"],
                "cells": row_cells,
            }
        )

    start_chart = _build_chart_series(raw_start_event_rows, event_hill_times, "median_hill_time")

    raw_split_event_rows = sorted(
        split_event_rows_map.values(),
        key=lambda item: (item["date"] or date.min, item["event_name"]),
    )
    split_event_rows = []
    for item in raw_split_event_rows:
        event_id = item["event_id"]
        event_split_time_list = event_split_times.get(event_id, [])
        if event_split_time_list:
            item["cells"]["DAY_MEDIAN"] = {
                "time": round(median(event_split_time_list), 3),
            }
        row_cells = []
        for key, label in SPLIT_TABLE_COLUMNS:
            row_cells.append(
                {
                    "key": key,
                    "label": label,
                    "value": item["cells"].get(key),
                }
            )
        split_event_rows.append(
            {
                "event_id": item["event_id"],
                "date": item["date"],
                "event_name": item["event_name"],
                "cells": row_cells,
            }
        )

    split_chart = _build_chart_series(raw_split_event_rows, event_split_times, "median_split_time")

    peer_context = _build_peer_context(rider, selected_track, wheel, kpi_cutoff=kpi_cutoff)
    peer_place_percentile = None
    peer_finish_percentile = None
    peer_group_size = 0
    pace_index = None

    if peer_context["uci_category"]:
        peer_results = peer_context["peer_results"]
        peer_runs = peer_context["peer_runs"]
        peer_timed_event_ids = {run.event_id for run in peer_runs if run.event_id and run.finish_time is not None}
        peer_places_by_rider = defaultdict(list)
        for result in peer_results:
            if result.rider_id and result.place is not None and result.event_id in peer_timed_event_ids:
                peer_places_by_rider[result.rider_id].append(result.place)

        peer_finish_by_rider = defaultdict(list)
        for run in peer_runs:
            if run.rider_id and run.finish_time is not None:
                peer_finish_by_rider[run.rider_id].append(float(run.finish_time))

        peer_place_averages = [round(mean(values), 2) for values in peer_places_by_rider.values() if values]
        peer_finish_medians = [round(median(values), 3) for values in peer_finish_by_rider.values() if values]

        field_median_finish_time = median(peer_finish_medians) if peer_finish_medians else None
        rider_avg_finish_time = _safe_mean(finish_times)
        if field_median_finish_time and rider_avg_finish_time and rider_avg_finish_time > 0:
            pace_index = round(100 * (field_median_finish_time / rider_avg_finish_time), 1)

        peer_place_percentile = _percentile_better_than(avg_track_place, peer_place_averages)
        peer_finish_percentile = _percentile_better_than(_safe_median(finish_times), peer_finish_medians)
        peer_group_size = max(len(peer_place_averages), len(peer_finish_medians))

    return {
        "track_results": track_results,
        "track_runs": track_runs,
        "wheel": wheel,
        "track_has_20": track_has_20,
        "track_has_24": track_has_24,
        "kpi_cutoff": kpi_cutoff,
        "kpi_label": kpi_label,
        "total_starts_count": len(timed_track_event_ids),
        "recent_track_results_count": len(recent_track_results),
        "recent_track_runs_count": len(timed_recent_track_runs),
        "starts_count": events_count,
        "runs_count": len(timed_recent_track_runs),
        "final_runs_count": len([run for run in timed_recent_track_runs if run.round_type in FINAL_ROUND_TYPES]),
        "progressed_events_count": progressed_events_count,
        "progression_rate": _safe_rate(progressed_events_count, events_with_motos),
        "final_rate": _safe_rate(events_with_final, events_count),
        "average_moto_place": avg_moto_place,
        "average_final_place": avg_final_place,
        "best_result": min(track_place_values) if track_place_values else None,
        "median_result": _safe_median(track_place_values),
        "win_rate": _safe_rate(sum(1 for place in track_place_values if place == 1), len(track_place_values)),
        "podium_rate": _safe_rate(sum(1 for place in track_place_values if place <= 3), len(track_place_values)),
        "average_finish_time": _safe_mean(finish_times),
        "best_finish_time": round(min(finish_times), 3) if finish_times else None,
        "median_finish_time": _safe_median(finish_times),
        "best_hill_time": round(min(hill_times), 3) if hill_times else None,
        "median_hill_time": _safe_median(hill_times),
        "best_split_1": round(min(split_times), 3) if split_times else None,
        "median_split_1": _safe_median(split_times),
        "average_hill_rank": _safe_mean(hill_ranks),
        "hill_win_rate": _safe_rate(sum(1 for rank in hill_ranks if rank == 1), len(hill_ranks)),
        "hill_top3_rate": _safe_rate(sum(1 for rank in hill_ranks if rank <= 3), len(hill_ranks)),
        "average_positions_gained_after_hill": _safe_mean(positions_gained_after_hill),
        "hill_rankable_runs_count": hill_rankable_runs,
        "best_hill_lane": best_hill_lane,
        "best_hill_lane_median": best_hill_lane_median,
        "best_result_lane": best_result_lane,
        "best_result_lane_average": best_result_lane_average,
        "finish_time_stddev": _safe_stddev(finish_times),
        "conversion_score": _safe_rate(converted_events, len(strong_start_events)),
        "recovery_score": _safe_rate(recovered_events, len(weak_start_events)),
        "pressure_score": pressure_score,
        "pressure_delta": pressure_delta,
        "track_affinity_score": affinity_score,
        "track_affinity_delta": affinity_delta,
        "track_affinity_label": affinity_label,
        "stability_score": stability_score,
        "risk_score": risk_score,
        "consistency_score": consistency_score,
        "consistency_label": consistency_label,
        "bad_status_count": bad_status_count,
        "clean_run_rate": _safe_rate(clean_runs, eligible_clean_runs),
        "trend_label": trend_label,
        "trend_delta": trend_delta,
        "trend_detail": trend_detail,
        "recent_results": list(reversed(ordered_results[-8:])),
        "event_rows": event_rows,
        "table_columns": TRACK_TABLE_COLUMNS,
        "chart_points": finish_chart["points"],
        "chart_polyline": finish_chart["polyline"],
        "chart_min_time": finish_chart["min"],
        "chart_max_time": finish_chart["max"],
        "chart_y_ticks": finish_chart["ticks"],
        "start_event_rows": start_event_rows,
        "start_table_columns": START_TABLE_COLUMNS,
        "start_chart_points": start_chart["points"],
        "start_chart_polyline": start_chart["polyline"],
        "start_chart_min_time": start_chart["min"],
        "start_chart_max_time": start_chart["max"],
        "start_chart_y_ticks": start_chart["ticks"],
        "split_event_rows": split_event_rows,
        "split_table_columns": SPLIT_TABLE_COLUMNS,
        "split_chart_points": split_chart["points"],
        "split_chart_polyline": split_chart["polyline"],
        "split_chart_min_time": split_chart["min"],
        "split_chart_max_time": split_chart["max"],
        "split_chart_y_ticks": split_chart["ticks"],
        "peer_label": peer_context["peer_label"],
        "peer_group_size": peer_group_size,
        "peer_place_percentile": peer_place_percentile,
        "peer_finish_percentile": peer_finish_percentile,
        "pace_index": pace_index,
    }


@login_required
def rider_premium_stats_view(request, pk):
    rider = get_object_or_404(Rider, uci_id=pk)
    premium_access_context = _get_premium_access_context(request.user, rider)
    if not premium_access_context["premium_access"]:
        messages.error(
            request,
            _("Pro rozšířené časy z tratí potřebuješ aktivní předplatné tohoto jezdce nebo trenérský přístup přes klub."),
        )
        return redirect("rider:detail", pk=pk)
    return render(
        request,
        "rider/rider-premium-stats.html",
        _build_rider_premium_stats_context(request, rider, premium_access_context),
    )


@login_required
def rider_premium_stats_pdf_view(request, pk):
    rider = get_object_or_404(Rider, uci_id=pk)
    premium_access_context = _get_premium_access_context(request.user, rider)
    if not premium_access_context["premium_access"]:
        messages.error(
            request,
            _("Pro rozšířené časy z tratí potřebuješ aktivní předplatné tohoto jezdce nebo trenérský přístup přes klub."),
        )
        return redirect("rider:detail", pk=pk)
    if not _can_export_rider_premium_stats_pdf(request.user, rider):
        messages.error(
            request,
            _("PDF export premium statistik je dostupný jen pro aktivní trainer extended."),
        )
        return redirect("rider:premium-stats", pk=pk)

    context = _build_rider_premium_stats_context(request, rider, premium_access_context)
    if context["selected_track"] is None:
        messages.error(request, _("Nejdřív vyber trať nebo klub, který chceš exportovat do PDF."))
        return redirect("rider:premium-stats", pk=pk)
    if context["require_wheel_selection"]:
        messages.error(request, _("Pro PDF export nejdřív vyber disciplínu 20\" nebo 24\"."))
        return redirect("rider:premium-stats", pk=pk)
    if context["track_stats"] is None:
        messages.error(request, _("Pro vybranou trať zatím nejsou k dispozici detailní data pro PDF export."))
        return redirect("rider:premium-stats", pk=pk)

    pdf_bytes = build_rider_premium_stats_pdf(
        rider=rider,
        track=context["selected_track"],
        track_stats=context["track_stats"],
        kpi_period=context["kpi_period"],
    )
    filename = build_rider_premium_stats_pdf_filename(rider, context["selected_track"])
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


@login_required
def rider_compare_view(request, pk):
    rider = get_object_or_404(Rider, uci_id=pk)
    premium_access_context = _get_premium_access_context(request.user, rider)
    if not premium_access_context["premium_access"]:
        messages.error(
            request,
            _("Pro porovnání jezdců potřebuješ aktivní předplatné tohoto jezdce nebo trenérský přístup přes klub."),
        )
        return redirect("rider:detail", pk=pk)

    base_runs = list(
        RaceRun.objects.filter(rider=rider, is_beginner=False)
        .select_related("event", "event__organizer", "result")
        .order_by("-event__date", "round_type", "round_number", "id")
    )
    base_results = list(
        Result.objects.filter(rider_id=rider.uci_id, is_beginner=False)
        .select_related("event", "event__organizer")
        .order_by("-date", "-id")
    )
    kpi_period = _resolve_kpi_period(request.GET.get("years"))
    track_options = _build_track_options(base_results, base_runs, kpi_cutoff=kpi_period["cutoff"])
    selected_wheel = request.GET.get("wheel")
    if selected_wheel not in {"20", "24"}:
        selected_wheel = None
    selected_track = _resolve_selected_track(track_options, request.GET.get("track"))
    require_wheel_selection = False

    if selected_track is not None:
        selected_track_results = [result for result in base_results if ((result.event and result.event.organizer_id) or 0) == selected_track["id"]]
        selected_track_runs = [run for run in base_runs if ((run.event and run.event.organizer_id) or 0) == selected_track["id"]]
        timed_track_event_ids = {run.event_id for run in selected_track_runs if run.finish_time is not None and run.event_id}
        timed_track_results = [result for result in selected_track_results if result.event_id in timed_track_event_ids]
        timed_track_runs = [run for run in selected_track_runs if run.finish_time is not None]
        available_wheels = []
        if any(result.is_20 for result in timed_track_results) or any(run.is_20 is True for run in timed_track_runs):
            available_wheels.append({"value": "20", "label": '20"'})
        if any(not result.is_20 for result in timed_track_results) or any(run.is_20 is False for run in timed_track_runs):
            available_wheels.append({"value": "24", "label": '24"'})
        if selected_wheel is None and len(available_wheels) == 1:
            selected_wheel = available_wheels[0]["value"]
        elif selected_wheel is None and len(available_wheels) > 1:
            require_wheel_selection = True
    else:
        available_wheels = []

    base_track_stats = None if require_wheel_selection else _build_track_stats(
        rider,
        selected_track,
        base_results,
        base_runs,
        wheel=selected_wheel,
        kpi_cutoff=kpi_period["cutoff"],
        kpi_label=kpi_period["label"],
    )

    peer_context = _build_peer_context(rider, selected_track, selected_wheel, kpi_cutoff=kpi_period["cutoff"])
    opponent = None
    opponent_track_stats = None
    head_to_head = None
    opponent_id = request.GET.get("opponent")
    if opponent_id and selected_track is not None and selected_wheel in {"20", "24"}:
        opponent = next((item for item in peer_context["candidates"] if str(item.uci_id) == opponent_id), None)
        if opponent is not None:
            opponent_runs = list(
                RaceRun.objects.filter(rider=opponent, is_beginner=False)
                .select_related("event", "event__organizer", "result")
                .order_by("-event__date", "round_type", "round_number", "id")
            )
            opponent_results = list(
                Result.objects.filter(rider_id=opponent.uci_id, is_beginner=False)
                .select_related("event", "event__organizer")
                .order_by("-date", "-id")
            )
            opponent_track_stats = _build_track_stats(
                opponent,
                selected_track,
                opponent_results,
                opponent_runs,
                wheel=selected_wheel,
                kpi_cutoff=kpi_period["cutoff"],
                kpi_label=kpi_period["label"],
            )
            base_track_results = [
                result
                for result in base_results
                if ((result.event and result.event.organizer_id) or 0) == selected_track["id"] and _matches_wheel_filter(result, selected_wheel)
            ]
            base_track_runs = [
                run
                for run in base_runs
                if ((run.event and run.event.organizer_id) or 0) == selected_track["id"] and _matches_wheel_filter(run, selected_wheel)
            ]
            opponent_track_results = [
                result
                for result in opponent_results
                if ((result.event and result.event.organizer_id) or 0) == selected_track["id"] and _matches_wheel_filter(result, selected_wheel)
            ]
            opponent_track_runs = [
                run
                for run in opponent_runs
                if ((run.event and run.event.organizer_id) or 0) == selected_track["id"] and _matches_wheel_filter(run, selected_wheel)
            ]
            if kpi_period["cutoff"] is not None:
                base_track_results = [item for item in base_track_results if item.date and item.date >= kpi_period["cutoff"]]
                opponent_track_results = [item for item in opponent_track_results if item.date and item.date >= kpi_period["cutoff"]]
                base_track_runs = [item for item in base_track_runs if item.event and item.event.date and item.event.date >= kpi_period["cutoff"]]
                opponent_track_runs = [item for item in opponent_track_runs if item.event and item.event.date and item.event.date >= kpi_period["cutoff"]]
            head_to_head = _build_head_to_head(
                base_track_results,
                opponent_track_results,
                base_track_runs,
                opponent_track_runs,
            )

    data = {
        "rider": rider,
        "subscription": premium_access_context["active_subscription"],
        "trainer_club_subscription": premium_access_context["trainer_club_subscription"],
        "has_admin_access": premium_access_context["premium_access_via_admin"],
        "has_trainer_club_access": premium_access_context["premium_access_via_trainer_club"],
        "kpi_period": kpi_period,
        "kpi_period_options": KPI_PERIOD_OPTIONS,
        "track_options": track_options,
        "selected_track": selected_track,
        "available_wheels": available_wheels,
        "selected_wheel": selected_wheel,
        "require_wheel_selection": require_wheel_selection,
        "base_track_stats": base_track_stats,
        "opponent": opponent,
        "opponent_track_stats": opponent_track_stats,
        "compare_candidates": peer_context["candidates"],
        "head_to_head": head_to_head,
    }
    return render(request, "rider/rider-compare.html", data)


@login_required
def rider_premium_subscriptions_view(request):
    if request.method == "POST":
        subscription = get_object_or_404(
            RiderStatsSubscription.objects.select_related("rider"),
            pk=request.POST.get("subscription_id"),
            user=request.user,
        )
        action = request.POST.get("action")
        if action == "disable-renew":
            cancel_rider_stats_subscription(subscription)
            messages.success(
                request,
                _("Automatické obnovování pro %(name)s bylo vypnuto.") % {'name': f"{subscription.rider.first_name} {subscription.rider.last_name}"}
            )
        elif action == "enable-renew":
            resume_rider_stats_subscription(subscription)
            messages.success(
                request,
                _("Automatické obnovování pro %(name)s bylo zapnuto.") % {'name': f"{subscription.rider.first_name} {subscription.rider.last_name}"}
            )
        elif action == "delete":
            rider_name = f"{subscription.rider.first_name} {subscription.rider.last_name}"
            subscription.delete()
            messages.success(
                request,
                _("Předplatné jezdce %(name)s bylo smazáno z přehledu.") % {'name': rider_name}
            )
        return redirect("rider:premium-subscriptions")

    current_time = timezone.now()
    subscriptions = list(
        RiderStatsSubscription.objects.filter(user=request.user)
        .select_related("rider", "season")
        .order_by("-expires_at", "rider__last_name", "rider__first_name")
    )
    for subscription in subscriptions:
        subscription.can_delete = (
            not subscription.auto_renew and subscription.expires_at <= current_time
        )
    return render(
        request,
        "rider/rider-premium-subscriptions.html",
        {"subscriptions": subscriptions},
    )


@login_required
def account_settings_invoices_view(request):
    AvatarChangeRequest.expire_stale_requests()
    if request.method == "POST" and request.POST.get("action") == "submit-avatar-request":
        return submit_avatar_change_request_view(request)

    SubscriptionInvoiceService().ensure_for_user(request.user)
    invoices_qs = (
        SubscriptionInvoice.objects.filter(user=request.user)
        .order_by("-issue_date", "-created")
    )
    invoices_count = invoices_qs.count()
    invoices_page_obj = None
    if invoices_count > 25:
        invoices_page_obj = Paginator(invoices_qs, 25).get_page(request.GET.get("invoice_page"))
        invoices = list(invoices_page_obj.object_list)
    else:
        invoices = list(invoices_qs)
    individual_subscriptions = (
        RiderStatsSubscription.objects.filter(user=request.user)
        .select_related("rider")
        .order_by("-expires_at", "rider__last_name", "rider__first_name")[:5]
    )
    linked_riders = list(
        request.user.riders.select_related("club")
        .filter(is_active=True)
        .order_by("last_name", "first_name")
    )
    account_avatar_request_pending = _get_pending_avatar_request_for_target(account=request.user) is not None
    linked_rider_pending_request_ids = {
        request_obj.target_rider_id
        for request_obj in AvatarChangeRequest.objects.filter(
            status=AvatarChangeRequest.STATUS_PENDING,
            target_rider__in=linked_riders,
        ).only("target_rider_id")
    }
    trainer_subscriptions = (
        TrainerClubSubscription.objects.filter(user=request.user)
        .select_related("club")
        .order_by("-expires_at", "club__team_name")[:8]
    )
    subscriptions_summary_count = len(individual_subscriptions) + len(trainer_subscriptions)
    club_event_invoices = []
    club_event_invoices_count = 0
    club_event_invoices_page_obj = None
    managed_club = None
    if getattr(request.user, "is_club_manager", False) and request.user.club_id:
        managed_club = request.user.club
        club_event_invoices_qs = (
            EventInvoice.objects.filter(club=request.user.club)
            .select_related("event", "club")
            .order_by("-issue_date", "-created")
        )
        club_event_invoices_count = club_event_invoices_qs.count()
        if club_event_invoices_count > 25:
            club_event_invoices_page_obj = Paginator(club_event_invoices_qs, 25).get_page(
                request.GET.get("club_invoice_page")
            )
            club_event_invoices = list(club_event_invoices_page_obj.object_list)
        else:
            club_event_invoices = list(club_event_invoices_qs)
    return render(
        request,
        "rider/account-settings-invoices.html",
        {
            "invoices": invoices,
            "invoices_count": invoices_count,
            "invoices_page_obj": invoices_page_obj,
            "individual_subscriptions": individual_subscriptions,
            "linked_riders": linked_riders,
            "linked_riders_count": len(linked_riders),
            "account_avatar_request_pending": account_avatar_request_pending,
            "linked_rider_pending_request_ids": linked_rider_pending_request_ids,
            "trainer_subscriptions": trainer_subscriptions,
            "subscriptions_summary_count": subscriptions_summary_count,
            "can_manage_inactive_riders": can_manage_inactive_riders(request.user),
            "managed_club": managed_club,
            "club_event_invoices": club_event_invoices,
            "club_event_invoices_count": club_event_invoices_count,
            "club_event_invoices_page_obj": club_event_invoices_page_obj,
        },
    )


@login_required
def submit_avatar_change_request_view(request):
    if request.method != "POST":
        return redirect("rider:account")

    AvatarChangeRequest.expire_stale_requests()

    uploaded_file = request.FILES.get("avatar_image")
    target_type = request.POST.get("target_type")
    target_id = request.POST.get("target_id")

    try:
        _validate_avatar_upload(uploaded_file)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("rider:account")

    if target_type == "account":
        if str(request.user.pk) != str(target_id):
            messages.error(request, _("Nemůžeš žádat změnu avataru pro jiný účet."))
            return redirect("rider:account")
        pending_request = _get_pending_avatar_request_for_target(account=request.user)
        if pending_request:
            messages.error(request, _("Pro tvůj účet už čeká jedna žádost o změnu avataru na schválení."))
            return redirect("rider:account")
        AvatarChangeRequest.objects.create(
            uploaded_by=request.user,
            target_account=request.user,
            image=uploaded_file,
        )
        messages.success(request, _("Nový avatar účtu byl odeslán ke schválení administrátorem."))
        return redirect("rider:account")

    if target_type == "rider":
        rider = request.user.riders.filter(pk=target_id, is_active=True).first()
        if not rider:
            messages.error(request, _("Nemůžeš žádat změnu avataru pro cizího jezdce."))
            return redirect("rider:account")
        pending_request = _get_pending_avatar_request_for_target(rider=rider)
        if pending_request:
            messages.error(request, _("Pro tohoto jezdce už čeká jedna žádost o změnu avataru na schválení."))
            return redirect("rider:account")
        AvatarChangeRequest.objects.create(
            uploaded_by=request.user,
            target_rider=rider,
            image=uploaded_file,
        )
        messages.success(
            request,
            _("Nový avatar pro jezdce %(name)s byl odeslán ke schválení administrátorem.")
            % {"name": f"{rider.first_name} {rider.last_name}"},
        )
        return redirect("rider:account")

    messages.error(request, _("Neplatný cíl změny avataru."))
    return redirect("rider:account")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@staff_member_required
def rider_admin(request):
    AvatarChangeRequest.expire_stale_requests()
    total_riders = Rider.objects.filter().count()
    active_riders = Rider.objects.filter(is_active=True).count()
    pending_avatar_count = AvatarChangeRequest.objects.filter(
        status=AvatarChangeRequest.STATUS_PENDING
    ).count()
    data = {
        "total_riders": total_riders,
        "active_riders": active_riders,
        "pending_avatar_count": pending_avatar_count,
    }
    return render(request, "rider/rider-admin.html", data)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@staff_member_required
def free_plates_view(request):
    used_plates = [
        display_plate(plate_text, plate, fallback="")
        for plate_text, plate in Rider.objects.filter(is_active=True).values_list("plate_text", "plate")
    ]
    free_plates = generate_available_plate_values(used_plates)

    data = {
        "free_plates": free_plates,
        "free_plates_count": len(free_plates),
    }
    return render(request, "rider/rider-free-plates.html", data)


def _rider_request_context(**extra):
    clubs = Club.objects.filter(is_active=True)
    used_plates = [
        display_plate(plate_text, plate, fallback="")
        for plate_text, plate in Rider.objects.filter(is_active=True).values_list("plate_text", "plate")
    ]
    free_plates = generate_available_plate_values(used_plates)

    context = {
        "clubs": clubs,
        "free_plates": free_plates,
        "lookup_url": "/rider/new/licence-lookup",
    }
    context.update(extra)
    return context


def rider_licence_lookup_view(request):
    uci_id = (request.GET.get("uci_id") or "").strip()
    if not uci_id.isdigit() or len(uci_id) != 11:
        return JsonResponse(
            {"ok": False, "message": _("UCI ID musí obsahovat přesně 11 číslic.")},
            status=400,
        )

    existing = Rider.objects.filter(uci_id=uci_id).first()
    if existing:
        return JsonResponse(
            {
                "ok": False,
                "message": _("Jezdec %(name)s už má přidělené startovní číslo.") % {'name': f"{existing.first_name} {existing.last_name}"},
            },
            status=409,
        )

    data_json, error_msg = get_rider_data(uci_id)
    if error_msg or not data_json:
        return JsonResponse(
            {"ok": False, "message": _("Licence nebyla nalezena.")},
            status=404,
        )

    first_name = data_json.get("firstName", "").strip()
    last_name = data_json.get("lastName", "").strip()
    birth = (data_json.get("birth", "") or "")[:10]
    gender_code = data_json.get("sex", {}).get("code", "M")
    gender = _("Žena") if gender_code == "F" else _("Muž")

    return JsonResponse(
        {
            "ok": True,
            "rider": {
                "uci_id": uci_id,
                "first_name": first_name,
                "last_name": last_name.capitalize() if last_name else "",
                "date_of_birth": birth,
                "gender": gender,
            },
        }
    )


def rider_new_view(request):
    if request.method == "POST":
        context = _rider_request_context(
            first_name=request.POST.get("first_name", ""),
            last_name=request.POST.get("last_name", ""),
            date_of_birth=request.POST.get("date_of_birth", ""),
            gender=request.POST.get("gender", ""),
            uci_id=request.POST.get("uci_id", ""),
            lookup_confirmed=request.POST.get("lookup_confirmed", ""),
            selected_plate=request.POST.get("plate", ""),
            selected_club=request.POST.get("club", ""),
            is20_checked="is20" in request.POST,
            is24_checked="is24" in request.POST,
            elite_checked="elite" in request.POST,
            emergency_contact=request.POST.get("emergency-contact", ""),
            emergency_phone=request.POST.get("emergency-phone", ""),
        )

        required_fields = {
            "uci_id": "UCI ID",
            "first_name": "Jméno",
            "last_name": "Příjmení",
            "date_of_birth": "Datum narození",
            "plate": "Startovní číslo",
            "club": "Klub",
            "emergency-contact": "Nouzový kontakt",
            "emergency-phone": "Telefon nouzového kontaktu",
        }
        for field, label in required_fields.items():
            if not request.POST.get(field) or request.POST.get(field) in {"", "Vyber..."}:
                messages.error(request, _("Pole %(label)s je povinné.") % {'label': label})
                return render(request, "rider/rider-request.html", context)

        if request.POST.get("lookup_confirmed") != "1":
            messages.error(request, _("Nejprve ověř UCI ID proti licenci ČSC."))
            return render(request, "rider/rider-request.html", context)

        uci_id = request.POST["uci_id"].strip()
        if not uci_id.isdigit() or len(uci_id) != 11:
            messages.error(request, _("UCI ID musí obsahovat přesně 11 číslic."))
            return render(request, "rider/rider-request.html", context)

        existing = Rider.objects.filter(uci_id=uci_id).first()
        if existing:
            messages.error(
                request,
                _("Jezdec/jezdkyně %(name)s, UCI ID %(uci_id)s, již má přidělené číslo.") % {'name': f"{existing.first_name} {existing.last_name}", 'uci_id': uci_id}
            )
            return render(request, "rider/rider-request.html", context)

        data_json, error_msg = get_rider_data(uci_id)
        if error_msg or not data_json:
            messages.error(request, error_msg or _("Licence UCI ID nebyla nalezena."))
            return render(request, "rider/rider-request.html", context)

        if "is20" not in request.POST and "is24" not in request.POST:
            messages.error(request, _('Musíš vybrat, zda budeš jezdit 20" nebo 24" kolo.'))
            return render(request, "rider/rider-request.html", context)

        try:
            date_of_birth = datetime.datetime.strptime(
                request.POST["date_of_birth"], "%Y-%m-%d"
            )
        except (TypeError, ValueError):
            messages.error(request, _("Datum narození není ve správném formátu."))
            return render(request, "rider/rider-request.html", context)

        club = Club.objects.filter(id=request.POST.get("club")).first()
        if not club:
            messages.error(request, _("Vybraný klub nebyl nalezen."))
            return render(request, "rider/rider-request.html", context)

        Rider.objects.create(
            first_name=request.POST["first_name"].strip(),
            last_name=request.POST["last_name"].strip(),
            date_of_birth=date_of_birth,
            gender=request.POST.get("gender", "Muž"),
            uci_id=uci_id,
            is_20="is20" in request.POST,
            is_24="is24" in request.POST,
            is_elite="elite" in request.POST,
            plate_text=normalize_plate_value(request.POST["plate"]),
            plate=legacy_plate_int(request.POST["plate"]),
            club=club,
            is_active=True,
            is_approved=False,
            emergency_contact=request.POST["emergency-contact"],
            emergency_phone=request.POST["emergency-phone"],
        )
        return render(request, "rider/rider-request-success.html")

    return render(request, "rider/rider-request.html", _rider_request_context())


@login_required(login_url="/login/")
@inactive_riders_required
def inactive_riders_views(request):
    """ Function for views inactive riders, only request params"""
    inactive_riders = _get_manageable_inactive_riders(request.user)
    data = {
        "riders": inactive_riders,
        "sum": len(inactive_riders),
        "managed_club": request.user.club if getattr(request.user, "is_club_manager", False) and request.user.club_id else None,
    }
    return render(request, "rider/rider-inactive.html", data)


@login_required(login_url="/login/")
@inactive_riders_required
def deactivate_inactive_rider_view(request, rider_id):
    if request.method != "POST":
        return redirect("rider:inactive")

    inactive_rider_ids = {rider.pk for rider in _get_manageable_inactive_riders(request.user)}
    rider = get_object_or_404(Rider, pk=rider_id)

    if rider.pk not in inactive_rider_ids:
        messages.error(request, _("Jezdce nelze deaktivovat mimo seznam neaktivních jezdců."))
        return redirect("rider:inactive")

    rider.is_active = False
    rider.save(update_fields=["is_active"])
    logger.info(
        "Rider deactivated from inactive list by user_id=%s email=%s role_admin=%s role_club_manager=%s rider_id=%s rider_uci_id=%s club_id=%s",
        request.user.pk,
        request.user.email,
        getattr(request.user, "is_admin", False),
        getattr(request.user, "is_club_manager", False),
        rider.pk,
        rider.uci_id,
        rider.club_id,
    )
    messages.success(
        request,
        _("Jezdec %(name)s byl označen jako neaktivní.") % {'name': f"{rider.first_name} {rider.last_name}"}
    )
    return redirect("rider:inactive")


@staff_member_required
def licence_check_views(request):
    """ Function for checking valid licence"""
    started = start_licence_check()
    if started:
        data = {
            "status_eyebrow": _("Správa jezdců"),
            "status_title": _("Kontrola licencí byla spuštěna"),
            "status_description": _("Ověřování platnosti licencí nyní běží na pozadí. Po dokončení se stav propíše přímo v databázi jezdců."),
            "status_note": _("Není potřeba čekat na této stránce. Můžeš pokračovat v administraci nebo se vrátit zpět na přehled nástrojů."),
            "status_icon": "license",
            "primary_action_label": _("Zpět na administraci jezdců"),
            "primary_action_url": "rider:admin",
            "secondary_action_label": _("Seznam jezdců"),
            "secondary_action_url": "rider:list",
        }
    else:
        data = {
            "status_eyebrow": _("Správa jezdců"),
            "status_title": _("Kontrola licencí již probíhá"),
            "status_description": _("Ověřování platnosti licencí již běží na pozadí. Počkej na dokončení předchozího průchodu."),
            "status_note": _("Není potřeba spouštět kontrolu znovu. Výsledky se propíší do databáze automaticky."),
            "status_icon": "license",
            "primary_action_label": _("Zpět na administraci jezdců"),
            "primary_action_url": "rider:admin",
            "secondary_action_label": _("Seznam jezdců"),
            "secondary_action_url": "rider:list",
        }
    return render(request, "rider/rider-success.html", data)


@staff_member_required
def ranking_count_views(request):
    """ Function for recount ranking"""
    try:
        recount_already_running = bool(cache.get(RANKING_RECOUNT_RUNNING_KEY))
        schedule_ranking_recount()
        if recount_already_running:
            messages.info(request, _("Přepočet rankingu už běží. Další průchod byl zařazen do fronty."))
        else:
            messages.info(request, _("Přepočet rankingu byl spuštěn na pozadí."))
    except Exception as error:
        logger.exception("Naplánování přepočtu rankingu selhalo")
        messages.error(
            request,
            _("Při naplánování přepočtu rankingu došlo k následující chybě: %(message)s")
            % {"message": error},
        )
    return redirect("rider:admin")


@staff_member_required
def participation_riders_on_event(request):
    participation = Participation().count()

    file_path = os.path.join(settings.MEDIA_ROOT, "participation/participation.xlsx")
    if os.path.exists(file_path):
        with open(file_path, "rb") as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response["Content-Disposition"] = "inline; filename=" + os.path.basename(
                file_path
            )
            return response


@staff_member_required
def recalculate_riders_classes(request):
    # set_all_riders_classes()
    RiderSetClassesThread().start()
    data = {
        "status_eyebrow": _("Správa jezdců"),
        "status_title": _("Přepočet kategorií byl spuštěn"),
        "status_description": _("Kategorie jezdců se teď přepočítávají na pozadí podle aktuálních pravidel."),
        "status_note": _("Po dokončení se nové třídy propíšou do profilů jezdců. Můžeš se vrátit do administrace nebo pokračovat v jiné práci."),
        "status_icon": "classes",
        "primary_action_label": _("Zpět na administraci jezdců"),
        "primary_action_url": "rider:admin",
        "secondary_action_label": _("Seznam jezdců"),
        "secondary_action_url": "rider:list",
    }
    return render(request, "rider/rider-success.html", data)


@staff_member_required
def calculate_cruiser_median(request):
    year_options = list(_get_event_year_options())
    selected_year = _resolve_selected_year(request.GET.get("year"), year_options)

    cruiser = Cruiser()
    cruiser.year = int(selected_year)
    cruiser_data = cruiser.calculate_median()
    cruiser_results = cruiser_data["cruisers"]
    count_cruiser_results = len(cruiser_results)

    data = {
        "cruisers": cruiser_results,
        "sum": count_cruiser_results,
        "selected_year": int(selected_year),
        "year_options": year_options,
        "median_age": cruiser_data["median_age"],
        "cup_events": cruiser_data["cup_events"],
        "minimum_participations": cruiser_data["minimum_participations"],
    }
    return render(request, "rider/rider-cruiser.html", data)


@staff_member_required
def riders_by_class_and_club(request):

    clubs = Club.objects.filter(is_active=True).order_by("team_name")
    file_name = "RIDERS_BY_CLUB_AND_CLASS.xlsx"

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    # Jeden DB dotaz místo 33 × počet klubů
    counts_qs = (
        Rider.objects.filter(is_active=True, is_approved=True)
        .values("club_id", "class_20")
        .annotate(count=Count("id"))
    )
    counts = {(row["club_id"], row["class_20"]): row["count"] for row in counts_qs}

    CLASSES = [
        "Boys 6", "Boys 7", "Girls 7", "Boys 8", "Girls 8",
        "Boys 9", "Girls 9", "Boys 10", "Girls 10", "Boys 11", "Girls 11",
        "Boys 12", "Girls 12", "Boys 13", "Girls 13", "Boys 14", "Girls 14",
        "Boys 15", "Girls 15", "Boys 16", "Girls 16",
        "Men 17-24", "Women 17-24", "Men 25-29", "Women 25 and over",
        "Men 30-34", "Men 35 and over", "Men Junior", "Women Junior",
        "Men Under 23", "Women Under 23", "Men Elite", "Women Elite",
    ]

    wb = Workbook()
    ws = wb.active

    first_line_riders_by_club_and_class(ws)

    row = 2
    for club in clubs:
        ws.cell(row, 1, club.team_name)
        for col, cls in enumerate(CLASSES, start=2):
            ws.cell(row, col, counts.get((club.id, cls), 0))
        row += 1
    wb.save(response)

    return response


@staff_member_required
def qualify_to_cn(request):
    RiderQualifyToCNThread().start()
    data = {
        "status_eyebrow": _("Správa jezdců"),
        "status_title": _("Výpočet kvalifikace byl spuštěn"),
        "status_description": _("Kvalifikace na Mistrovství České republiky se právě počítá na pozadí."),
        "status_note": _("Po dokončení budou nové výsledky dostupné v příslušném přehledu. Není potřeba čekat na této stránce."),
        "status_icon": "qualify",
        "primary_action_label": _("Zpět na administraci jezdců"),
        "primary_action_url": "rider:admin",
        "secondary_action_label": _("Seznam jezdců"),
        "secondary_action_url": "rider:list",
    }
    return render(request, "rider/rider-success.html", data)
def _get_event_year_options():
    return Event.objects.exclude(date__isnull=True).dates("date", "year", order="DESC")


def _resolve_selected_year(selected_year, year_options):
    if selected_year is None:
        return str(year_options[0].year) if year_options else str(date.today().year)
    if not selected_year.isdigit():
        return str(year_options[0].year) if year_options else str(date.today().year)
    return selected_year


def transponder_search_view(request):
    transponder_query = (request.GET.get("transponder") or "").strip().upper()
    results = []

    if transponder_query:
        czech_riders = Rider.objects.filter(
            Q(transponder_20__icontains=transponder_query)
            | Q(transponder_24__icontains=transponder_query)
        ).select_related("club")

        foreign_riders = ForeignRider.objects.filter(
            Q(transponder_20__icontains=transponder_query)
            | Q(transponder_24__icontains=transponder_query)
        )

        for rider in czech_riders:
            matched_in = []
            if (rider.transponder_20 or "").upper() == transponder_query:
                matched_in.append('20"')
            if (rider.transponder_24 or "").upper() == transponder_query:
                matched_in.append('24"')

            results.append(
                {
                    "first_name": rider.first_name,
                    "last_name": rider.last_name,
                    "club": rider.club.team_name if rider.club else "-",
                    "plate": rider.plate_display,
                    "matched_in": ", ".join(matched_in) or "-",
                    "status": _("Aktuální čip"),
                    "status_tone": "current",
                    "changed_at": None,
                }
            )

        for rider in foreign_riders:
            matched_in = []
            if (rider.transponder_20 or "").upper() == transponder_query:
                matched_in.append('20"')
            if (rider.transponder_24 or "").upper() == transponder_query:
                matched_in.append('24"')

            results.append(
                {
                    "first_name": rider.first_name,
                    "last_name": rider.last_name,
                    "club": rider.club or rider.state or "-",
                    "plate": rider.plate_display,
                    "matched_in": ", ".join(matched_in) or "-",
                    "status": _("Aktuální čip"),
                    "status_tone": "current",
                    "changed_at": None,
                }
            )

        current_chip_keys = {
            ("czech", result["first_name"], result["last_name"], result["matched_in"])
            for result in results
            if result["status_tone"] == "current"
        }

        historical_changes = (
            RiderTransponderChange.objects.filter(
                Q(old_transponder__iexact=transponder_query)
                | Q(new_transponder__iexact=transponder_query)
            )
            .select_related("rider__club", "changed_by")
            .order_by("-changed_at")
        )

        for change in historical_changes:
            rider = change.rider
            if not rider:
                continue

            slot_label = change.get_slot_display()
            is_current_now = (
                (slot_label == '20"' and (rider.transponder_20 or "").upper() == transponder_query)
                or (slot_label == '24"' and (rider.transponder_24 or "").upper() == transponder_query)
            )
            if is_current_now:
                continue

            dedupe_key = ("czech", rider.first_name, rider.last_name, slot_label)
            if dedupe_key in current_chip_keys:
                continue

            results.append(
                {
                    "first_name": rider.first_name,
                    "last_name": rider.last_name,
                    "club": rider.club.team_name if rider.club else "-",
                    "plate": rider.plate_display,
                    "matched_in": slot_label,
                    "status": _("Historický čip"),
                    "status_tone": "historical",
                    "changed_at": change.changed_at,
                }
            )

    data = {
        "transponder_query": transponder_query,
        "results": results,
        "has_search": bool(transponder_query),
    }
    return render(request, "rider/transponder-search.html", data)


def plate_search_view(request):
    plate_query = (request.GET.get("plate") or "").strip()
    results = []

    if plate_query:
        normalized_plate_query = normalize_plate_value(plate_query)
        plate_filter = Q(plate_text__iexact=normalized_plate_query)
        if normalized_plate_query.isdigit():
            plate_filter |= Q(plate=legacy_plate_int(normalized_plate_query))
        czech_riders = Rider.objects.filter(plate_filter, is_active=True).select_related("club")
        foreign_riders = ForeignRider.objects.filter(plate_filter)

        for rider in czech_riders:
            results.append(
                {
                    "plate": rider.plate_display,
                    "first_name": rider.first_name,
                    "last_name": rider.last_name,
                    "club": rider.club.team_name if rider.club else "-",
                }
            )

        for rider in foreign_riders:
            results.append(
                {
                    "plate": rider.plate_display,
                    "first_name": rider.first_name,
                    "last_name": rider.last_name,
                    "club": rider.club or rider.state or "-",
                }
            )

    data = {
        "plate_query": plate_query,
        "results": results,
        "has_search": bool(plate_query),
    }
    return render(request, "rider/plate-search.html", data)

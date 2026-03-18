"""
rider/rider.py — pomocné funkce a třídy pro správu jezdců

Obsah:
  1. Vlákna (Threading) — operace na pozadí (licence, třídy, kvalifikace)
  2. ČSC API — získání tokenu a dat jezdce z API České cyklistiky
  3. Pomocné funkce — neaktivní jezdci, participace, Cruiser median
  4. Export Excel — záhlaví XLS exportů jezdců po klubech
  5. Správa transpondérů a tříd (hromadné utility)
  6. Notifikace — session příznaky pro admin panel
  7. Rozpoznání kategorie pro externí API
"""

import logging
from statistics import median
from django.utils import timezone
import requests

logger = logging.getLogger(__name__)
from decouple import config
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from rider.models import Rider
from event.models import Result, Event, Entry, SeasonSettings
import threading
from django.db.models import Q, Exists, OuterRef
from datetime import datetime, date, timedelta
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook


now = datetime.today().year


# ===========================================================================
# 1. VLÁKNA (THREADING) — operace na pozadí
# ===========================================================================

class CheckValidLicenceThread(threading.Thread):

    def __init__(self):
        threading.Thread.__init__(self)

    def run(self):
        refresh_valid_licences()


class RiderSetClassesThread(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)

    def run(self):
        riders = list(Rider.objects.filter(is_active=True))
        for rider in riders:
            rider.class_beginner = rider.set_class_beginner(rider)
            rider.class_20 = rider.set_class_20(rider)
            rider.class_24 = rider.set_class_24(rider)
        Rider.objects.bulk_update(riders, ['class_beginner', 'class_20', 'class_24'])


# ===========================================================================
# 2. ČSC API — přihlášení a načtení dat jezdce
# ===========================================================================

def get_api_token():
    """Získá access token z Czech Cycling Federation API (OAuth2 password flow)."""
    TOKEN_URL = "https://portal.api.czechcyclingfederation.com/connect/token"

    data = {
        "grant_type": "password",
        "username": config("LICENCE_USERNAME"),
        "password": config("LICENCE_PASSWORD"),
        "client_id": "CSSC_Blazor",
        "scope": "CSSC openid profile email roles"
    }

    headers = {
        "Content-Type": "application/x-www-form-urlencoded"
    }

    try:
        response = requests.post(
            TOKEN_URL,
            data=data,
            headers=headers,
            verify=True  # nebo False pokud testuješ
        )
        response.raise_for_status()
        token = response.json().get("access_token")
        if not token:
            raise ValueError("Token nebyl vrácen v odpovědi.")
        return token

    except Exception as e:
        logger.error(f"Chyba při získávání tokenu: {e}")
        return None


def get_rider_data(uci_id):
    logger.debug(f"Načítám data pro UCI ID: {uci_id}")

    token = get_api_token()
    if not token:
        logger.error("Nepodařilo se získat access token.")
        return None, "Nepodařilo se získat token k API ČSC."

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }

    url = f"https://portal.api.czechcyclingfederation.com/api/services/licenseinfo?uciId={uci_id}"

    logger.debug(f"Odesílám požadavek na: {url}")

    try:
        response = requests.get(url, headers=headers, verify=True)
        logger.debug(f"Status kód: {response.status_code}")

        if response.status_code == 404 or "Http_NotFound" in response.text:
            logger.warning(f"Licence UCI ID: {uci_id} nebyla nalezena v databázi ČSC.")
            return None, f"Licence UCI ID: {uci_id} nebyla nalezena."

        if not response.ok:
            logger.warning(f"Neočekávaná odpověď: {response.status_code}")
            return None, f"Nastala chyba: {response.status_code}"

        data = response.json()
        logger.debug(f"Úspěšně načteno pro UCI ID: {uci_id}")
        return data, None

    except Exception as e:
        logger.error(f"Výjimka při volání API ČSC: {e}")
        return None, f"Chyba při komunikaci s API ČSC: {e}"


def valid_licence(rider, token=None):
    """Ověření platnosti licence pomocí správného endpointu + access tokenu."""
    token = token or get_api_token()
    if not token:
        logger.warning("Nelze ověřit licenci – token se nezískal.")
        return None

    headers = {
        "Authorization": f"Bearer {token}"
    }

    year = date.today().year
    base_url = "https://portal.api.czechcyclingfederation.com"
    check_url = f"{base_url}/api/services/validuciid"

    params = {
        "year": year,
        "uciId": rider.uci_id
    }

    try:
        response = requests.get(check_url, headers=headers, params=params, verify=True)
        response.raise_for_status()

        data = response.json()
        is_valid = data.get("valid", False)

        rider.valid_licence = is_valid
        rider.save(update_fields=["valid_licence", "updated"])

        if is_valid:
            logger.info(f"{rider.uci_id} – {rider.first_name} {rider.last_name}: licence je platná.")
        else:
            logger.warning(f"{rider.uci_id} – {rider.first_name} {rider.last_name}: licence NENÍ platná.")

        return is_valid

    except Exception as e:
        logger.error(f"Chyba při ověřování licence {rider.uci_id}: {e}")
        return None


def refresh_valid_licences(riders=None):
    """Hromadně zkontroluje platnost licencí aktivních jezdců."""
    fixed_count = Rider.objects.filter(
        is_active=True,
        fix_valid_licence=True,
    ).update(valid_licence=True)

    queryset = riders or Rider.objects.filter(
        is_active=True,
        fix_valid_licence=False,
    ).only("id", "uci_id", "first_name", "last_name", "valid_licence", "updated")

    token = get_api_token()
    if not token:
        logger.warning("Hromadná kontrola licencí nebyla spuštěna, token se nezískal.")
        return {
            "checked": 0,
            "valid": 0,
            "invalid": 0,
            "failed": queryset.count(),
            "fixed": fixed_count,
        }

    checked = 0
    valid = 0
    invalid = 0
    failed = 0

    for rider in queryset.iterator(chunk_size=200):
        result = valid_licence(rider, token=token)
        if result is None:
            failed += 1
            continue

        checked += 1
        if result:
            valid += 1
        else:
            invalid += 1

    return {
        "checked": checked,
        "valid": valid,
        "invalid": invalid,
        "failed": failed,
        "fixed": fixed_count,
    }


# ===========================================================================
# 3. POMOCNÉ FUNKCE — neaktivní jezdci, statistiky
# ===========================================================================

def two_years_inactive():
    """Vrátí jezdce bez výsledku v předchozím kalendářním roce se startovním číslem starším než 2 roky."""
    previous_year = timezone.localdate().year - 1
    two_years_ago = timezone.now() - timedelta(days=365 * 2)

    # Filtrovat aktivní jezdce s přiděleným číslem a profilem starším než 2 roky.
    # Do doby, než se všechna ostrá data překlopí do plate_text, držíme fallback i na legacy plate.
    riders = Rider.objects.filter(
        is_active=True,
        is_approved=True,
    ).filter(
        (
            Q(plate_text__isnull=False)
            & ~Q(plate_text__exact="")
        )
        | Q(plate__gt=0)
    ).filter(
        Q(created__lte=two_years_ago) | Q(created__isnull=True)
    )

    # Subdotaz na kontrolu existence výsledků v předchozím kalendářním roce
    previous_year_results = Result.objects.filter(
        rider_id=OuterRef('uci_id'),
        event__date__year=previous_year
    )

    # Vrátí pouze jezdce, kteří nemají žádný výsledek v předchozím roce
    inactive_riders = riders.annotate(
        has_results=Exists(previous_year_results)
    ).filter(has_results=False).order_by('club')

    return list(inactive_riders)

class Participation:
    """ Class for count participation riders on events with export to excel file """

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def first_line(self):
        self.ws.cell(1, 1, "Last_name")
        self.ws.cell(1, 2, "First_name")
        self.ws.cell(1, 3, "UCI ID")
        self.ws.cell(1, 4, "Class_20")
        self.ws.cell(1, 5, "Class_24")
        self.ws.cell(1, 6, "CLub")
        self.ws.cell(1, 7, "MČR")
        self.ws.cell(1, 8, "ČP")
        self.ws.cell(1, 9, "Ostatní")

    def save(self):
        self.wb.save(filename='media/participation/participation.xlsx')

    def calculate(self):
        self.first_line()
        line = 2
        riders = Rider.objects.filter(is_active=True)
        for rider in riders:
            mcr: int = 0
            cp: int = 0
            others: int = 0
            participation = Result.objects.filter(rider_id=rider.uci_id, date__year=now)
            for part in participation:
                if part.event_type == "Mistrovství ČR jednotlivců":
                    mcr = 1
                elif part.event_type == "Český pohár":
                    cp += 1
                else:
                    others += 1
            self.ws.cell(line, 1, rider.last_name)
            self.ws.cell(line, 2, rider.first_name)
            self.ws.cell(line, 3, rider.uci_id)
            self.ws.cell(line, 4, rider.class_20)
            self.ws.cell(line, 5, rider.class_24)
            self.ws.cell(line, 6, str(rider.club))
            self.ws.cell(line, 7, mcr)
            self.ws.cell(line, 8, cp)
            self.ws.cell(line, 9, others)

            line += 1

    def count(self):
        self.first_line()
        self.calculate()
        self.save()


class Cruiser:
    def __init__(self):
        self.__NUMBER_OF_CUPS = 0
        self.__NUMBER_OF_PCS = 0
        self.year = now

    def set_number_of_cups(self, number):
        self.__NUMBER_OF_CUPS = number

    def set_number_of_peaces(self, number):
        self.__NUMBER_OF_PCS = number

    def calculate_median(self):
        cup_events = Event.objects.filter(
            type_for_ranking="Český pohár",
            date__year=self.year,
            canceled=False,
        ).count()
        minimum_participations = (cup_events // 2) + 1 if cup_events else 0

        entries = (
            Entry.objects.filter(
                event__type_for_ranking="Český pohár",
                event__date__year=self.year,
                event__canceled=False,
                is_24=True,
                payment_complete=True,
                checkout=False,
            )
            .select_related("rider")
            .order_by("-rider__date_of_birth")
        )

        cruisers_in_events = []
        for entry in entries:
            if entry.rider and entry.rider not in cruisers_in_events:
                cruisers_in_events.append(entry.rider)

        cruiser_results = []
        ages = []
        position = 1

        for cruiser in cruisers_in_events:
            participations = (
                Entry.objects.filter(
                    rider=cruiser,
                    event__type_for_ranking="Český pohár",
                    event__date__year=self.year,
                    event__canceled=False,
                    is_24=True,
                    payment_complete=True,
                    checkout=False,
                )
                .values("event_id")
                .distinct()
                .count()
            )
            if participations >= minimum_participations:
                cruiser_results.append(cruiser)
                age = self.year - cruiser.date_of_birth.year
                cruiser.age = age
                cruiser.position = position
                cruiser.participations = participations
                ages.append(age)
                position += 1

        return {
            "cruisers": cruiser_results,
            "median_age": median(ages) if ages else None,
            "cup_events": cup_events,
            "minimum_participations": minimum_participations,
        }


# ===========================================================================
# 4. EXPORT EXCEL — záhlaví XLS souborů
# ===========================================================================

def first_line_riders_by_club_and_class(ws):
    ws.cell(1, 1, "TEAM NAME")
    ws.cell(1, 2, "BOYS 6")
    ws.cell(1, 3, "BOYS 7")
    ws.cell(1, 4, "GIRLS 7")
    ws.cell(1, 5, "BOYS 8")
    ws.cell(1, 6, "GIRLS 8")
    ws.cell(1, 7, "BOYS 9")
    ws.cell(1, 8, "GIRLS 9")
    ws.cell(1, 9, "BOYS 10")
    ws.cell(1, 10, "GIRLS 10")
    ws.cell(1, 11, "BOYS 11")
    ws.cell(1, 12, "GIRLS 11")
    ws.cell(1, 13, "BOYS 12")
    ws.cell(1, 14, "GIRLS 12")
    ws.cell(1, 15, "BOYS 13")
    ws.cell(1, 16, "GIRLS 13")
    ws.cell(1, 17, "BOYS 14")
    ws.cell(1, 18, "GIRLS 14")
    ws.cell(1, 19, "BOYS 15")
    ws.cell(1, 20, "GIRLS 15")
    ws.cell(1, 21, "BOYS 16")
    ws.cell(1, 22, "GIRLS 16")
    ws.cell(1, 23, 'MEN 17-24')
    ws.cell(1, 24, 'WOMEN 17-24')
    ws.cell(1, 25, 'MEN 25-29')
    ws.cell(1, 26, 'WOMEN 25 AND OVER')
    ws.cell(1, 27, 'MEN 30-34')
    ws.cell(1, 28, 'MEN 35 AND OVER')
    ws.cell(1, 29, 'MEN JUNIOR')
    ws.cell(1, 30, 'WOMEN JUNIOR')
    ws.cell(1, 31, 'MEN UNDER 23')
    ws.cell(1, 32, 'WOMEN UNDER 23')
    ws.cell(1, 33, 'MEN ELITE')
    ws.cell(1, 34, 'WOMEN ELITE')


class RiderQualifyToCNThread(threading.Thread):
    """Class for set qualify to CN rider's status."""

    def __init__(self, year=None):
        threading.Thread.__init__(self)
        self.year = year

    def run(self):
        current_year = datetime.today().year
        year = self.year or current_year
        if year != current_year:
            logger.info(
                "Přepočet kvalifikace na MČR pro rok %s přeskočen. "
                "Pole is_qualify_to_cn_* reprezentují pouze aktuální rok %s.",
                year,
                current_year,
            )
            return

        settings = SeasonSettings.objects.filter(year=year).first()
        if not settings:
            logger.warning(f"Chybí SeasonSettings pro rok {year}, kvalifikace na MČR nebyla přepočítána.")
            return

        riders = Rider.objects.filter(is_active=True, is_approved=True)
        logger.info(f"Přepočítávám kvalifikaci na MČR pro rok {year}")

        for rider in riders:
            entries_20 = Entry.objects.filter(
                event__type_for_ranking="Český pohár",
                event__date__year=year,
                checkout=False,
                payment_complete=True,
                rider=rider,
                is_20=True,
                is_beginner=False,
            ).count()
            entries_24 = Entry.objects.filter(
                event__type_for_ranking="Český pohár",
                event__date__year=year,
                checkout=False,
                payment_complete=True,
                rider=rider,
                is_24=True,
            ).count()

            rider.is_qualify_to_cn_20 = entries_20 >= settings.qualify_to_cn
            rider.is_qualify_to_cn_24 = entries_24 >= settings.qualify_to_cn
            rider.save(update_fields=["is_qualify_to_cn_20", "is_qualify_to_cn_24"])


def should_recount_cn_qualification_for_event(event):
    """Vrátí True, pokud nahrání výsledků tohoto závodu má automaticky spustit přepočet kvalifikace na MČR."""
    if not event or event.type_for_ranking != "Český pohár" or not event.date:
        return False

    year = event.date.year
    current_year = datetime.today().year
    if year != current_year:
        return False

    settings = SeasonSettings.objects.filter(year=year).first()
    if not settings:
        return False

    championship = (
        Event.objects.filter(type_for_ranking="Mistrovství ČR jednotlivců", date__year=year)
        .exclude(date__isnull=True)
        .order_by("date")
        .first()
    )
    if not championship:
        return False

    cup_events_before_championship = list(
        Event.objects.filter(
            type_for_ranking="Český pohár",
            date__year=year,
            date__lt=championship.date,
        )
        .exclude(date__isnull=True)
        .order_by("date", "id")
    )
    if not cup_events_before_championship:
        return False

    event_ids = [cup_event.id for cup_event in cup_events_before_championship]
    if event.id not in event_ids:
        return False

    first_decisive_index = max(settings.qualify_to_cn - 1, 0)
    current_index = event_ids.index(event.id)
    return current_index >= first_decisive_index


def trigger_cn_qualification_recount_if_needed(event):
    """Spustí automatický přepočet kvalifikace na MČR na pozadí, pokud je závod v rozhodném okně."""
    if should_recount_cn_qualification_for_event(event):
        RiderQualifyToCNThread(year=event.date.year).start()
        logger.info(
            "Automatický přepočet kvalifikace na MČR spuštěn po uploadu výsledků závodu %s (%s).",
            event.id,
            event.name,
        )


def generate_insurance_file(event):

    """Vygeneruje stylovaný Excel pro pojištění jezdců s využitím dat z modelu a API"""

    entries = Entry.objects.filter(event=event, payment_complete=True, checkout=False).select_related('rider')

    wb = Workbook()
    ws = wb.active
    ws.title = "INSURANCE"

    # Styl záhlaví
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")  # indigo
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Kategorie", "Jméno", "Příjmení", "Datum narození", "Adresa"]
    ws.append(headers)

    # Styluj záhlaví
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    def process_entry(entry):
        rider = entry.rider
        try:
            rider_class = rider.class_20 if entry.is_20 else rider.class_24
            first_name = rider.first_name
            last_name = rider.last_name
            birth = str(rider.date_of_birth)
            street = rider.street or ""
            city = rider.city or ""
            zip_code = rider.zip or ""

            if not street or not city or not zip_code:
                api_data, error = get_rider_data(rider.uci_id)
                if api_data and not error:
                    street = street or api_data.get("street", "")
                    city = city or api_data.get("city", "")
                    zip_code = zip_code or api_data.get("postcode", "")
                    first_name = api_data.get("firstName", first_name)
                    last_name = api_data.get("lastName", last_name)
                    birth = api_data.get("birth", birth)[:10]

            address = f"{street}, {city}, PSČ: {zip_code}".strip().strip(',')
            return [rider_class, first_name, last_name, birth, address]

        except Exception as e:
            logger.error(f"Chyba u jezdce {rider.uci_id}: {e}")
            return None

    # Použij ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=10) as executor:
        rows = list(executor.map(process_entry, entries))

    # Přidej řádky a styluj je
    for row_idx, row in enumerate(rows, start=2):
        if row:
            ws.append(row)
            for col_idx, _ in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = center_align
                cell.border = border

    # Nastav šířky sloupců
    column_widths = [18, 16, 18, 16, 40]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Ulož soubor
    file_path = f"media/ec-files/INSURANCE_FOR_RACE_ID-{event.id}-{event.name}.xlsx"
    wb.save(file_path)

    # Ulož do eventu
    event.ec_insurance_file = file_path
    event.ec_insurance_file_created = timezone.now()
    event.save()

    return file_path
   

# ---------------------------------------------------------------------------
# SPRÁVA TRANSPONDÉRŮ A TŘÍD (hromadné utility)
# ---------------------------------------------------------------------------

def set_all_riders_classes():
    """Nastaví třídy všem aktivním jezdcům (hromadná varianta, nepoužívá threading).
    Upozornění: pro produkci použij raději RiderSetClassesThread, který provede
    bulk_update v jednom dotazu místo jednotlivých save().
    """
    riders = list(Rider.objects.filter(is_active=True))
    for rider in riders:
        rider.class_beginner = rider.set_class_beginner(rider)
        rider.class_20 = rider.set_class_20(rider)
        rider.class_24 = rider.set_class_24(rider)
    Rider.objects.bulk_update(riders, ['class_beginner', 'class_20', 'class_24'])


def clear_transponders():
    """Vyčistí pole transpondérů od hodnoty 'nan' (artefakt po importu z Excelu).
    Spouštět ručně pouze po importu dat.
    """
    riders = Rider.objects.filter(is_active=True)
    to_update = []
    for rider in riders:
        changed = False
        if rider.transponder_20 == "nan":
            rider.transponder_20 = ""
            changed = True
        if rider.transponder_24 == "nan":
            rider.transponder_24 = ""
            changed = True
        if changed:
            to_update.append(rider)
    if to_update:
        Rider.objects.bulk_update(to_update, ['transponder_20', 'transponder_24'])


def update_plate_notify(request):
    """Uloží do session příznak, zda existují jezdci čekající na schválení.
    Používá se v homepage_view pro zobrazení notifikace adminu.
    """
    request.session['plate'] = Rider.objects.filter(is_approved=False).exists()


# ---------------------------------------------------------------------------
# ROZPOZNÁNÍ KATEGORIE PRO EXTERNÍM API
# ---------------------------------------------------------------------------

def resolve_api_category_code(rider, is_20=False, is_24=False, is_beginner=False):
    """Vrací API kód kategorie závodníka podle pravidel ČSC a tříd v modelu Rider"""

    class20_to_api = {
        'Boys 6': 'B 6', 'Boys 7': 'B 7', 'Boys 8': 'B 8', 'Boys 9': 'B 9', 'Boys 10': 'B 10',
        'Boys 11': 'B 11', 'Boys 12': 'B 12', 'Boys 13': 'B 13', 'Boys 14': 'B 14',
        'Boys 15': 'B 15', 'Boys 16': 'B 16', 'Men 17-24': 'M 17/24', 'Men 25-29': 'M 25/29',
        'Men 30-34': 'M 30+', 'Men 35 and over': 'M 30+', 'Girls 6': 'G 6', 'Girls 7': 'G 7', 'Girls 8': 'G 8',
        'Girls 9': 'G 9', 'Girls 10': 'G 10', 'Girls 11': 'G 11', 'Girls 12': 'G 12',
        'Girls 13': 'G 13', 'Girls 14': 'G 14', 'Girls 15': 'G 15', 'Girls 16': 'G 16',
        'Women 17-24': 'WOMEN 17+', 'Women 25 and over': 'WOMEN 17+',
        'Men Junior': 'JUNIOR MEN', 'Women Junior': 'JUNIOR WOMEN',
        'Men Under 23': 'U23 MEN', 'Women Under 23': 'U23 WOMEN',
        'Men Elite': 'ELITE MEN', 'Women Elite': 'ELITE  WOMEN',
    }

    class24_to_api = {
        'Boys 12 and under': 'CRUISER', 'Boys 13 and 14': 'CRUISER', 'Boys 15 and 16': 'CRUISER',
        'Men 17-24': 'CRUISER', 'Men 25-29': 'CRUISER', 'Men 30-34': 'CRUISER',
        'Men 35-39': 'CRUISER', 'Men 40-44': 'CRUISER', 'Men 45-49': 'CRUISER',
        'Men 50 and over': 'CRUISER', 'Girls 12 and under': 'CRUISER',
        'Girls 13-16': 'CRUISER', 'Women 17-29': 'CRUISER',
        'Women 30-39': 'CRUISER', 'Women 40 and over': 'CRUISER'
    }

    beginners_to_api = {
        'Beginners 1': 'ČL1', 'Beginners 2': 'ČL1',
        'Beginners 3': 'ČL2', 'Beginners 4': 'ČL2',
    }

    if is_beginner:
        return beginners_to_api.get(rider.class_beginner, "")

    if is_20:
        return class20_to_api.get(rider.class_20, "")

    if is_24:
        return class24_to_api.get(rider.class_24, "CRUISER")

    return ""
    
